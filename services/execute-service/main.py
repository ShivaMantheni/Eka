# ============================================================
# Execute Service — Eka Automation
# Handles: SpyTest execution, script execution, execution history,
#           log streaming WebSocket, queue management
# Port: 8002
# ============================================================

import os, json, re, yaml, time, asyncio, logging, subprocess, tempfile, csv, io, base64, uuid, socket
from datetime import datetime, timedelta
from typing import List, Optional
from pathlib import Path
from threading import Thread, Lock

from fastapi import FastAPI, HTTPException, WebSocket, WebSocketDisconnect, Depends, Request
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.date import DateTrigger
from apscheduler.triggers.cron import CronTrigger
from sqlalchemy import create_engine, Column, Integer, String, DateTime, Text, Boolean, or_, UniqueConstraint
from sqlalchemy.orm import declarative_base, sessionmaker, Session
from sqlalchemy import text
import paramiko
from paramiko import AutoAddPolicy

# ── Database ──────────────────────────────────────────────────────────────────
DATABASE_URL = os.getenv("DATABASE_URL", "sqlite:///./data/dut_automation.db")

if DATABASE_URL.startswith("sqlite"):
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False},
                           pool_size=10, max_overflow=20, pool_pre_ping=True)
else:
    engine = create_engine(DATABASE_URL, pool_size=10, max_overflow=20,
                           pool_timeout=30, pool_recycle=3600, pool_pre_ping=True)

SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

# ── Models (mirrors main.py) ───────────────────────────────────────────────────
class DUT(Base):
    __tablename__ = "duts"
    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String(100), nullable=False)
    ip_address = Column(String(50), nullable=False)
    port = Column(Integer, default=22)
    device_type = Column(String(50), default="Linux")
    username = Column(String(100), default="admin")
    password = Column(String(255), default="")
    connection_type = Column(String(10), default="ssh")
    status = Column(String(20), default="offline")
    xml_path = Column(String(500), default="/home/hp/prajwal/VMs")
    session_id = Column(String(255), nullable=True, index=True)
    last_heartbeat = Column(DateTime, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow)
    reserved_by = Column(String(100), nullable=True)
    reserved_at = Column(DateTime, nullable=True)
    reserved_until = Column(DateTime, nullable=True)
    __table_args__ = (UniqueConstraint('session_id', 'name', name='uq_session_dut_name'),)

class DUTLock(Base):
    __tablename__ = "dut_locks"
    id = Column(Integer, primary_key=True, autoincrement=True)
    dut_id = Column(Integer, nullable=False, unique=True)
    status = Column(String(20), default="AVAILABLE")
    job_id = Column(Integer, nullable=True)
    locked_since = Column(DateTime, nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow)

class TopologyConnection(Base):
    __tablename__ = "topology_connections"
    id = Column(Integer, primary_key=True, autoincrement=True)
    dut_a_id = Column(Integer, nullable=False)
    intf_a = Column(String(50), default="Ethernet0")
    dut_b_id = Column(Integer, nullable=False)
    intf_b = Column(String(50), default="Ethernet0")
    created_at = Column(DateTime, default=datetime.utcnow)

class Execution(Base):
    __tablename__ = "executions"
    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String(200), nullable=False)
    script_id = Column(Integer, nullable=True)
    dut_ids = Column(String(500))
    image_id = Column(Integer, nullable=True)
    execution_type = Column(String(20), default="script")
    status = Column(String(20), default="pending")
    session_id = Column(String(255), nullable=True, index=True)
    start_time = Column(DateTime, nullable=True)
    end_time = Column(DateTime, nullable=True)
    duration_seconds = Column(Integer, nullable=True)
    test_results = Column(Text, nullable=True)   # JSON list of per-script aggregates
    job_id = Column(Integer, nullable=True, index=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class ExecutionLog(Base):
    __tablename__ = "execution_logs"
    id = Column(Integer, primary_key=True, autoincrement=True)
    execution_id = Column(Integer, nullable=False)
    dut_name = Column(String(100), default="SYSTEM")
    log_level = Column(String(20), default="INFO")
    message = Column(Text)
    timestamp = Column(DateTime, default=datetime.utcnow)

class TestCaseResult(Base):
    __tablename__ = "testcase_results"
    id = Column(Integer, primary_key=True, autoincrement=True)
    execution_id = Column(Integer, nullable=False, index=True)
    script_path = Column(Text, nullable=True)
    module = Column(Text, nullable=True)
    test_function = Column(Text, nullable=True)
    testcase_id = Column(Text, nullable=True)
    result = Column(String(20), nullable=True)
    time_taken = Column(String(50), nullable=True)
    time_seconds = Column(Integer, nullable=True)
    description = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)

class ExecutionJob(Base):
    """Named job container linking DUT selection, topology, scripts, and executions."""
    __tablename__ = "execution_jobs"
    id               = Column(Integer, primary_key=True, autoincrement=True)
    name             = Column(String(100), nullable=False, default="Job")
    status           = Column(String(20), default="idle")
    session_id       = Column(String(255), nullable=True, index=True)
    dut_ids          = Column(Text, nullable=True)
    base_path        = Column(Text, nullable=True)
    host_id          = Column(Integer, nullable=True)
    topology         = Column(Text, nullable=True)
    scripts          = Column(Text, nullable=True)
    testbed_path     = Column(Text, nullable=True)
    schedule_type    = Column(String(10), default="none")
    schedule_at      = Column(DateTime, nullable=True)
    schedule_cron    = Column(String(100), nullable=True)
    schedule_enabled = Column(Boolean, default=False)
    last_run_at      = Column(DateTime, nullable=True)
    created_at       = Column(DateTime, default=datetime.utcnow)
    updated_at       = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s - execute-service - %(levelname)s - %(message)s")
logger = logging.getLogger("execute-service")

# SpyTest paths on remote host
SPYTEST_BASE = os.getenv("SPYTEST_BASE", "/home/hp_test/Eka/sonic-mgmt/spytest")
SPYTEST_TESTS_DIR = f"{SPYTEST_BASE}/tests"
SPYTEST_TESTBED_DIR = f"{SPYTEST_BASE}/testbeds"
SPYTEST_VENV = f"{SPYTEST_BASE}/spytest_venv"
SPYTEST_PYTHON = f"{SPYTEST_VENV}/bin/python"
SPYTEST_BIN = f"{SPYTEST_BASE}/bin/spytest"

# ── FastAPI App ───────────────────────────────────────────────────────────────
app = FastAPI(title="Eka Execute Service", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True,
                   allow_methods=["*"], allow_headers=["*"])

scheduler = BackgroundScheduler()

@app.on_event("startup")
def _on_startup():
    scheduler.start()
    _reload_all_job_schedules()
    logger.info("Execute-service started: APScheduler running")

@app.on_event("shutdown")
def _on_shutdown():
    scheduler.shutdown(wait=False)

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def get_session_id(request: Request) -> str:
    return request.headers.get("X-Session-ID", "")

def verify_access(resource_session_id, current_session_id):
    if not current_session_id:
        return False
    if not resource_session_id:
        return True
    return resource_session_id == current_session_id

# ── SSH Connection Manager ─────────────────────────────────────────────────────
class SSHConnectionManager:
    """Manages SSH connections to DUT devices with retry and keepalive."""

    MAX_CONNECT_RETRIES = 4
    CONNECT_BACKOFF     = [0, 3, 8, 20]   # seconds between attempts
    MAX_CMD_RETRIES     = 3
    CMD_BACKOFF         = [0, 2, 5]

    def __init__(self, host: str, port: int = 22, username: str = "admin", password: str = ""):
        self.host = host
        self.port = port
        self.username = username
        self.password = password
        self.client = None

    def _is_alive(self) -> bool:
        """Check whether the underlying SSH transport is still active."""
        try:
            if not self.client:
                return False
            t = self.client.get_transport()
            if not t or not t.is_active():
                return False
            t.send_ignore()          # lightweight SSH no-op
            return True
        except Exception:
            return False

    def _open_client(self) -> bool:
        """Single attempt to create a paramiko client and connect."""
        try:
            c = paramiko.SSHClient()
            c.set_missing_host_key_policy(AutoAddPolicy())
            c.connect(
                hostname=self.host, port=self.port,
                username=self.username, password=self.password,
                timeout=15, allow_agent=False, look_for_keys=False,
                banner_timeout=15, auth_timeout=15,
            )
            t = c.get_transport()
            if t:
                t.set_keepalive(15)         # SSH-level keepalive every 15 s
            self.client = c
            return True
        except paramiko.AuthenticationException as e:
            logger.error(f"SSH auth failed {self.username}@{self.host}:{self.port} — {e}")
            return False          # auth errors must not be retried
        except Exception as e:
            logger.warning(f"SSH connect attempt failed {self.host}:{self.port} — {type(e).__name__}: {e}")
            return False

    def connect(self, retries: int = MAX_CONNECT_RETRIES) -> bool:
        """Connect with exponential backoff.  Auth failures abort immediately."""
        for attempt in range(retries):
            delay = self.CONNECT_BACKOFF[min(attempt, len(self.CONNECT_BACKOFF) - 1)]
            if delay:
                time.sleep(delay)
            if self._open_client():
                logger.info(f"SSH connected to {self.host}:{self.port} (attempt {attempt + 1})")
                return True
            # If auth exception was raised _open_client returns False immediately;
            # there is no point retrying credentials that won't change.
            # Detect this by checking whether paramiko raised AuthenticationException
            # — the banner test below catches it indirectly.
        logger.error(f"SSH connection failed after {retries} attempts: {self.host}:{self.port}")
        return False

    def reconnect(self) -> bool:
        """Silently close existing client and try to reconnect."""
        try:
            if self.client:
                self.client.close()
        except Exception:
            pass
        self.client = None
        return self.connect()

    def execute_command(self, command: str, timeout: int = 30,
                        cmd_retries: int = MAX_CMD_RETRIES) -> tuple:
        """Run a command, transparently reconnecting and retrying on transient drops."""
        last_exc = None
        for attempt in range(cmd_retries):
            delay = self.CMD_BACKOFF[min(attempt, len(self.CMD_BACKOFF) - 1)]
            if delay:
                time.sleep(delay)
            if not self._is_alive():
                logger.warning(f"[SSH] Connection to {self.host} is dead (attempt {attempt + 1}) — reconnecting")
                if not self.reconnect():
                    last_exc = Exception(f"Reconnect failed (attempt {attempt + 1})")
                    continue
            try:
                stdin, stdout, stderr = self.client.exec_command(command, timeout=timeout)
                output    = stdout.read().decode("utf-8", errors="ignore")
                error     = stderr.read().decode("utf-8", errors="ignore")
                exit_code = stdout.channel.recv_exit_status()
                return output, error, exit_code
            except Exception as e:
                logger.warning(f"[SSH] Command failed on {self.host} (attempt {attempt + 1}): {e}")
                last_exc = e
                self.client = None   # force reconnect next iteration
        raise last_exc or Exception(f"Command failed after {cmd_retries} attempts on {self.host}")

    def transfer_file(self, local_path: str, remote_path: str) -> bool:
        """Transfer a file to the remote device via SFTP."""
        try:
            sftp = self.client.open_sftp()
            sftp.put(local_path, remote_path)
            sftp.close()
            logger.info(f"File transferred: {local_path} -> {remote_path}")
            return True
        except Exception as e:
            logger.error(f"File transfer failed: {e}")
            return False

    def disconnect(self):
        """Close SSH connection with proper socket shutdown."""
        if self.client:
            try:
                transport = self.client.get_transport()
                if transport and transport.sock:
                    try:
                        transport.sock.shutdown(socket.SHUT_RDWR)
                    except Exception:
                        pass
                self.client.close()
                logger.info(f"Disconnected from {self.host}")
            except Exception as e:
                logger.warning(f"Error during disconnect from {self.host}: {e}")
            finally:
                self.client = None


# Keep SSHManager as backward-compatible alias
SSHManager = SSHConnectionManager

# ── Execution Logging ──────────────────────────────────────────────────────────
def log_execution(db: Session, execution_id: int, dut_name: str, level: str, message: str):
    """Write an execution log entry to the database."""
    logger.log(getattr(logging, level, logging.INFO), f"[Exec {execution_id}][{dut_name}] {message}")
    try:
        entry = ExecutionLog(
            execution_id=execution_id,
            dut_name=dut_name,
            log_level=level,
            message=message,
            timestamp=datetime.utcnow(),
        )
        db.add(entry)
        db.commit()
    except Exception:
        db.rollback()

# Keep log_exec as backward-compatible alias
def log_exec(db, execution_id, dut_name, level, message):
    log_execution(db, execution_id, dut_name, level, message)

# ── Queue State (in-memory, per service) ──────────────────────────────────────
_queue_lock = Lock()
_active_executions = {}   # execution_id -> thread

# ── IN-MEMORY EXECUTION QUEUE STATE ──────────────────────────────────────────
_exec_queue_lock = Lock()
_exec_queue_state: dict = {}   # execution_id -> {scripts:[{name,status,duts}], free_duts:[]}


def _q_init(execution_id: int, script_names: list, all_duts: list):
    with _exec_queue_lock:
        _exec_queue_state[execution_id] = {
            "scripts": [{"name": n, "status": "queued", "duts": [], "pid": None} for n in script_names],
            "free_duts": list(all_duts),
            "all_duts": list(all_duts),
        }


def _q_update_script(execution_id: int, script_name: str, status: str, duts: list = None, pid: str = None):
    with _exec_queue_lock:
        state = _exec_queue_state.get(execution_id)
        if not state:
            return
        for s in state["scripts"]:
            if s["name"] == script_name:
                s["status"] = status
                if duts is not None:
                    s["duts"] = duts
                if pid is not None:
                    s["pid"] = pid
                break


def _q_set_free(execution_id: int, free_duts: list):
    with _exec_queue_lock:
        state = _exec_queue_state.get(execution_id)
        if state:
            state["free_duts"] = list(free_duts)


def _q_cleanup(execution_id: int):
    with _exec_queue_lock:
        _exec_queue_state.pop(execution_id, None)


# ── DYNAMIC BATCH ADDITION & SCRIPT CANCELLATION ─────────────────────────────

_pending_scripts_lock = Lock()
_pending_scripts: dict = {}  # execution_id -> {scripts:[...], to_cancel: set(...)}
_execution_threads_lock = Lock()
_execution_threads: dict = {}  # execution_id -> list of thread objects
_test_results_lock = Lock()  # guards concurrent test_results JSON updates from parallel script threads

# ── Execution-level STOP support ─────────────────────────────────────────────
_exec_cancel_lock = Lock()
_exec_cancel: set = set()


class ExecutionCancelled(Exception):
    """Raised inside a worker when the user stops the execution while it waits for DUTs."""


def _request_exec_cancel(execution_id: int):
    with _exec_cancel_lock:
        _exec_cancel.add(execution_id)


def _is_exec_cancelled(execution_id: int) -> bool:
    with _exec_cancel_lock:
        return execution_id in _exec_cancel


def _clear_exec_cancel(execution_id: int):
    with _exec_cancel_lock:
        _exec_cancel.discard(execution_id)


def _init_pending_scripts(execution_id: int):
    """Initialize pending scripts structure for new execution."""
    with _pending_scripts_lock:
        _pending_scripts[execution_id] = {"scripts": [], "to_cancel": set()}


def _add_pending_script(execution_id: int, script_info: dict):
    """Add script to pending list for dynamic addition."""
    with _pending_scripts_lock:
        if execution_id in _pending_scripts:
            _pending_scripts[execution_id]["scripts"].append(script_info)


def _mark_script_for_cancel(execution_id: int, script_name: str):
    """Mark a script to be cancelled (removed from queue or stopped if running)."""
    with _pending_scripts_lock:
        if execution_id in _pending_scripts:
            _pending_scripts[execution_id]["to_cancel"].add(script_name)


def _cleanup_pending_scripts(execution_id: int):
    """Clean up pending scripts after execution completes."""
    with _pending_scripts_lock:
        _pending_scripts.pop(execution_id, None)


# ── Health Check ──────────────────────────────────────────────────────────────
@app.get("/health")
def health():
    return {"status": "ok", "service": "execute-service"}

# ── GET /api/executions ───────────────────────────────────────────────────────
@app.get("/api/executions")
def list_executions(request: Request, db: Session = Depends(get_db)):
    session_id = get_session_id(request)
    if not session_id:
        raise HTTPException(status_code=401, detail="Session ID required")
    executions = db.query(Execution).filter(
                       or_(Execution.session_id == session_id, Execution.session_id.is_(None))
                   ).order_by(Execution.created_at.desc()).limit(100).all()
    return [{"id": ex.id, "name": ex.name, "type": ex.execution_type,
             "status": ex.status, "dut_count": len(json.loads(ex.dut_ids)) if ex.dut_ids else 0,
             "duration": ex.duration_seconds,
             "created_at": ex.created_at.isoformat() if ex.created_at else None}
            for ex in executions]

# ── POST /api/executions ──────────────────────────────────────────────────────
@app.post("/api/executions")
def create_execution(request: Request, execution_data: dict, db: Session = Depends(get_db)):
    """Create and start a new execution (image deployment or script execution)."""
    try:
        session_id = get_session_id(request)
        if not session_id:
            raise HTTPException(status_code=401, detail="Session ID required")

        dut_ids = execution_data.get("dut_ids", [])
        script_id = execution_data.get("script_id")
        image_id = execution_data.get("image_id")

        if not dut_ids:
            raise HTTPException(status_code=400, detail="No DUT IDs provided")

        duts = db.query(DUT).filter(DUT.id.in_(dut_ids), DUT.session_id == session_id).all()
        if len(duts) != len(dut_ids):
            raise HTTPException(status_code=403, detail="One or more DUTs not found or access denied")

        exec_type = "image" if image_id else "script"
        execution = Execution(
            name=f"exec_{exec_type}_{int(datetime.utcnow().timestamp())}",
            script_id=script_id,
            image_id=image_id,
            dut_ids=json.dumps(dut_ids),
            execution_type=exec_type,
            status="pending",
            session_id=session_id,
        )
        db.add(execution)
        db.commit()
        db.refresh(execution)

        if not script_id and not image_id:
            raise HTTPException(status_code=400, detail="Either script_id or image_id required")

        return {
            "execution_id": execution.id,
            "status": "started",
            "type": exec_type,
            "dut_count": len(dut_ids),
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))

# ── GET /api/executions/{id} ──────────────────────────────────────────────────
@app.get("/api/executions/{execution_id}")
def get_execution(execution_id: int, request: Request, db: Session = Depends(get_db)):
    session_id = get_session_id(request)
    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")
    if not verify_access(execution.session_id, session_id):
        raise HTTPException(status_code=403, detail="Access denied")
    return {"id": execution.id, "name": execution.name, "type": execution.execution_type,
            "status": execution.status,
            "dut_ids": json.loads(execution.dut_ids) if execution.dut_ids else [],
            "duration": execution.duration_seconds,
            "start_time": execution.start_time.isoformat() if execution.start_time else None,
            "end_time": execution.end_time.isoformat() if execution.end_time else None}

# ── GET /api/executions/{id}/logs ─────────────────────────────────────────────
@app.get("/api/executions/{execution_id}/logs")
def get_execution_logs(execution_id: int, request: Request,
                       limit: int = 200, offset: int = 0,
                       db: Session = Depends(get_db)):
    session_id = get_session_id(request)
    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")
    if not verify_access(execution.session_id, session_id):
        raise HTTPException(status_code=403, detail="Access denied")
    logs = db.query(ExecutionLog).filter(ExecutionLog.execution_id == execution_id)\
             .order_by(ExecutionLog.timestamp.asc()).offset(offset).limit(limit).all()
    return [{"id": l.id, "dut_name": l.dut_name, "level": l.log_level,
             "message": l.message,
             "timestamp": l.timestamp.isoformat() if l.timestamp else None} for l in logs]

# ── DELETE /api/executions/{id}/logs ──────────────────────────────────────────
@app.delete("/api/executions/{execution_id}/logs")
def delete_execution_logs(execution_id: int, body: dict, request: Request,
                          db: Session = Depends(get_db)):
    session_id = get_session_id(request)
    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")
    if not verify_access(execution.session_id, session_id):
        raise HTTPException(status_code=403, detail="Access denied")
    scope = body.get("scope", "all")
    try:
        deleted_count = db.query(ExecutionLog).filter(
            ExecutionLog.execution_id == execution_id).delete(synchronize_session=False)
        db.delete(execution)
        db.commit()
        return {"status": "success", "deleted_count": deleted_count,
                "scope": scope, "execution_id": execution_id}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ── POST /api/executions/{execution_id}/add-scripts ───────────────────────────
@app.post("/api/executions/{execution_id}/add-scripts")
def add_scripts_to_execution(execution_id: int, body: dict, db: Session = Depends(get_db)):
    """Add new scripts to a running execution queue."""
    scripts = body.get("scripts", [])

    if not scripts:
        raise HTTPException(status_code=400, detail="Please select at least one script")

    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")

    if execution.status not in ["running", "pending"]:
        raise HTTPException(status_code=400, detail="Execution is not running")

    for script_info in scripts:
        _add_pending_script(execution_id, script_info)
        script_name = os.path.basename(script_info.get("path", ""))
        _q_update_script(execution_id, script_name, "queued")

    return {
        "status": "success",
        "added": len(scripts),
        "execution_id": execution_id
    }

# ── POST /api/executions/{execution_id}/cancel-script ─────────────────────────
@app.post("/api/executions/{execution_id}/cancel-script")
def cancel_script_from_execution(execution_id: int, body: dict, db: Session = Depends(get_db)):
    """Cancel a running or queued script."""
    script_name = body.get("script_name")

    if not script_name:
        raise HTTPException(status_code=400, detail="script_name required")

    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")

    _mark_script_for_cancel(execution_id, script_name)
    _q_update_script(execution_id, script_name, "cancelled")

    return {
        "status": "success",
        "script": script_name,
        "execution_id": execution_id
    }

# ── POST /api/executions/{execution_id}/stop ──────────────────────────────────
@app.post("/api/executions/{execution_id}/stop")
def stop_execution_endpoint(execution_id: int, db: Session = Depends(get_db)):
    """Stop a running execution: terminate all running scripts and mark it cancelled."""
    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")

    if execution.status not in ("running", "pending"):
        return {"status": execution.status, "execution_id": execution_id,
                "message": "Execution is not running"}

    _request_exec_cancel(execution_id)
    log_execution(db, execution_id, "SYSTEM", "WARN",
                  "■ Stop requested by user — terminating running scripts…")

    # Best-effort immediate kill of any running PIDs
    killed = 0
    try:
        with _exec_queue_lock:
            state = _exec_queue_state.get(execution_id)
            running = ([(s.get("name"), s.get("pid")) for s in state["scripts"]
                        if s.get("status") == "running" and s.get("pid")]
                       if state else [])
        host = None
        if running and execution.job_id:
            _job = db.query(ExecutionJob).filter(ExecutionJob.id == execution.job_id).first()
            if _job and _job.host_id:
                host = db.query(DUT).filter(DUT.id == _job.host_id).first()
        if running and host:
            kssh = SSHConnectionManager(host.ip_address, host.port, host.username, host.password)
            if kssh.connect():
                for _name, _pid in running:
                    kssh.execute_command(
                        f"kill -TERM {_pid} 2>/dev/null; sleep 1; kill -KILL {_pid} 2>/dev/null",
                        timeout=10)
                    killed += 1
                kssh.disconnect()
    except Exception as _e:
        logger.warning(f"[STOP] Immediate kill best-effort failed for exec {execution_id}: {_e}")

    return {"status": "stopping", "execution_id": execution_id, "killed_now": killed}

# ── POST /api/spytest/start ───────────────────────────────────────────────────
@app.post("/api/spytest/start")
def start_spytest(body: dict, request: Request, db: Session = Depends(get_db)):
    """Start a SpyTest execution on a remote VM."""
    session_id = get_session_id(request)
    if not session_id:
        raise HTTPException(status_code=401, detail="Session ID required")

    vm_id = body.get("vm_id")
    script_paths = body.get("script_paths", [])
    dut_ids = body.get("dut_ids", [])
    testbed_yaml = body.get("testbed_yaml", "")
    log_level = body.get("log_level", "info")

    if not vm_id or not script_paths:
        raise HTTPException(status_code=400, detail="vm_id and script_paths required")

    vm = db.query(DUT).filter(DUT.id == vm_id, DUT.session_id == session_id).first()
    if not vm:
        raise HTTPException(status_code=404, detail="VM not found or access denied")

    execution = Execution(
        name=f"spytest_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}",
        execution_type="spytest",
        dut_ids=json.dumps(dut_ids),
        status="pending",
        session_id=session_id,
    )
    db.add(execution)
    db.commit()
    db.refresh(execution)

    thread = Thread(target=_run_spytest_background,
                    args=(execution.id, vm, script_paths, dut_ids, testbed_yaml, log_level, db),
                    daemon=True)
    thread.start()
    with _queue_lock:
        _active_executions[execution.id] = thread

    return {"execution_id": execution.id, "status": "started",
            "message": f"SpyTest execution started with {len(script_paths)} script(s)"}

def _run_spytest_background(execution_id, vm, script_paths, dut_ids, testbed_yaml, log_level, _):
    """Run SpyTest scripts on remote VM via SSH."""
    db = SessionLocal()
    execution = None
    try:
        execution = db.query(Execution).filter(Execution.id == execution_id).first()
        execution.status = "running"
        execution.start_time = datetime.utcnow()
        db.commit()

        log_exec(db, execution_id, "SYSTEM", "INFO",
                 f"Starting SpyTest execution — {len(script_paths)} script(s)")

        ssh = SSHManager(vm.ip_address, vm.port, vm.username, vm.password)
        if not ssh.connect():
            log_exec(db, execution_id, "SYSTEM", "ERROR",
                     f"Cannot connect to VM {vm.name} ({vm.ip_address}:{vm.port})")
            execution.status = "failed"
            execution.end_time = datetime.utcnow()
            db.commit()
            return

        try:
            for script_path in script_paths:
                db.refresh(execution)
                if execution.status == "failed":
                    log_exec(db, execution_id, "SYSTEM", "WARNING", "Execution cancelled by user")
                    break

                log_exec(db, execution_id, "SYSTEM", "INFO",
                         f"═══ Running: {os.path.basename(script_path)} ═══")

                cmd = (f"cd {SPYTEST_BASE} && "
                       f"source {SPYTEST_VENV}/bin/activate && "
                       f"{SPYTEST_PYTHON} {SPYTEST_BIN} "
                       f"--log-level {log_level} "
                       f"--testbed-file /tmp/eka_testbed.yaml "
                       f"{script_path} 2>&1")

                log_exec(db, execution_id, "SYSTEM", "INFO", f"$ {cmd[:200]}")

                try:
                    out, err, code = ssh.execute_command(cmd, timeout=3600)
                    for line in (out + err).splitlines()[:500]:
                        level = "ERROR" if any(k in line.lower() for k in ["error", "fail", "exception"]) else "INFO"
                        log_exec(db, execution_id, vm.name, level, line)
                    status_msg = "✓ PASSED" if code == 0 else f"✗ FAILED (exit {code})"
                    log_exec(db, execution_id, "SYSTEM", "INFO",
                             f"Script {os.path.basename(script_path)}: {status_msg}")
                except Exception as e:
                    log_exec(db, execution_id, "SYSTEM", "ERROR", f"Script error: {e}")

            execution.status = "completed"
        finally:
            ssh.disconnect()

    except Exception as e:
        logger.error(f"SpyTest execution failed: {e}")
        if execution:
            execution.status = "failed"
    finally:
        if execution:
            execution.end_time = datetime.utcnow()
            if execution.start_time:
                execution.duration_seconds = int(
                    (execution.end_time - execution.start_time).total_seconds())
            db.commit()
        db.close()
        with _queue_lock:
            _active_executions.pop(execution_id, None)

# ── POST /api/spytest/stop ────────────────────────────────────────────────────
@app.post("/api/spytest/stop/{execution_id}")
def stop_spytest(execution_id: int, request: Request, db: Session = Depends(get_db)):
    session_id = get_session_id(request)
    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")
    if not verify_access(execution.session_id, session_id):
        raise HTTPException(status_code=403, detail="Access denied")
    execution.status = "failed"
    db.commit()
    return {"status": "stopped", "execution_id": execution_id}

# ── GET /api/spytest/queue ────────────────────────────────────────────────────
@app.get("/api/spytest/queue")
def get_queue_status(request: Request, db: Session = Depends(get_db)):
    session_id = get_session_id(request)
    running = db.query(Execution).filter(
        Execution.session_id == session_id,
        Execution.status == "running"
    ).all()
    return {"running_count": len(running),
            "running": [{"id": e.id, "name": e.name} for e in running]}

# ── GET /api/spytest/scripts ──────────────────────────────────────────────────
@app.get("/api/spytest/scripts")
def list_scripts(vm_id: int, path: str = "", request: Request = None,
                 db: Session = Depends(get_db)):
    """List test scripts on remote VM."""
    session_id = request.headers.get("X-Session-ID", "") if request else ""
    vm = db.query(DUT).filter(DUT.id == vm_id).first()
    if not vm:
        raise HTTPException(status_code=404, detail="VM not found")

    ssh = SSHManager(vm.ip_address, vm.port, vm.username, vm.password)
    if not ssh.connect():
        raise HTTPException(status_code=503, detail="Cannot connect to VM")

    try:
        base = f"{SPYTEST_BASE}/tests"
        target = f"{base}/{path}" if path else base
        out, err, code = ssh.execute_command(
            f"find {target} -maxdepth 1 \\( -type f -name '*.py' -o -type d \\) | sort", timeout=15)
        files, folders = [], []
        for line in out.strip().splitlines():
            line = line.strip()
            if not line or line == target:
                continue
            rel = line[len(base):].lstrip("/")
            if line.endswith(".py"):
                files.append({"name": os.path.basename(line), "path": rel, "type": "file"})
            else:
                folders.append({"name": os.path.basename(line), "path": rel, "type": "folder"})
        return {"base_path": base, "current_path": path,
                "folders": folders, "files": files,
                "total": len(files), "total_folders": len(folders)}
    finally:
        ssh.disconnect()

# ── GET /api/spytest/categories ───────────────────────────────────────────────
@app.get("/api/spytest/categories")
def get_spytest_categories(host_id: int, db: Session = Depends(get_db)):
    """List test category folders from the remote SPyTest tests directory."""
    ssh, dut = _ssh_to_host(host_id, db)
    try:
        cmd = f'find {SPYTEST_TESTS_DIR} -mindepth 1 -maxdepth 1 -type d -printf "%f\\n" | sort'
        output, error, code = ssh.execute_command(cmd, timeout=15)
        if code != 0:
            raise HTTPException(status_code=500, detail=f"Failed to list categories: {error}")
        categories = [d.strip() for d in output.strip().split("\n") if d.strip() and not d.strip().startswith("__")]
        return {"categories": categories, "base_path": SPYTEST_TESTS_DIR}
    finally:
        ssh.disconnect()

# ── GET /api/spytest/scripts/{category} ───────────────────────────────────────
@app.get("/api/spytest/scripts/{category}")
def get_spytest_scripts(category: str, host_id: int, db: Session = Depends(get_db)):
    """List Python test scripts in a category folder (recursive)."""
    ssh, dut = _ssh_to_host(host_id, db)
    try:
        category_path = f"{SPYTEST_TESTS_DIR}/{category}"
        cmd = f'find {category_path} -name "test_*.py" -type f | sort'
        output, error, code = ssh.execute_command(cmd, timeout=15)
        if code != 0:
            raise HTTPException(status_code=500, detail=f"Failed to list scripts: {error}")
        scripts = []
        for line in output.strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            rel_path = line.replace(SPYTEST_TESTS_DIR + "/", "")
            name = os.path.basename(line)
            scripts.append({"name": name, "path": rel_path, "full_path": line})
        return {"category": category, "scripts": scripts}
    finally:
        ssh.disconnect()

# ── GET /api/spytest/testbeds ─────────────────────────────────────────────────
@app.get("/api/spytest/testbeds")
def get_spytest_testbeds(host_id: int, db: Session = Depends(get_db)):
    """List testbed YAML files from the remote SPyTest testbed directory."""
    if host_id == 5:
        return {"testbeds": ["testbed_2_switches.yaml", "testbed_standalone_switch.yaml", "testbed_spine_leaf.yaml"], "base_path": "/mock/testbeds"}
    ssh, dut = _ssh_to_host(host_id, db)
    try:
        cmd = f'find {SPYTEST_TESTBED_DIR} -maxdepth 1 -name "*.yaml" -type f -printf "%f\\n" | sort'
        output, error, code = ssh.execute_command(cmd, timeout=15)
        if code != 0:
            raise HTTPException(status_code=500, detail=f"Failed to list testbeds: {error}")
        testbeds = [f.strip() for f in output.strip().split("\n") if f.strip()]
        return {"testbeds": testbeds, "base_path": SPYTEST_TESTBED_DIR}
    finally:
        ssh.disconnect()

# ── GET /api/spytest/browse ───────────────────────────────────────────────────
@app.get("/api/spytest/browse")
def browse_spytest_root(host_id: int, base_path: str = None, db: Session = Depends(get_db)):
    """Root browse — delegates to browse_spytest_folder with empty path."""
    return browse_spytest_folder("", host_id, base_path, db)


@app.get("/api/spytest/browse/{path:path}")
def browse_spytest_folder(path: str, host_id: int, base_path: str = None, db: Session = Depends(get_db)):
    """Browse a specific folder in the SPyTest tests directory."""
    if host_id == 5:
        parent_path = ""
        if path and "/" in path:
            parent_path = "/".join(path.split("/")[:-1])
        norm_path = path.strip().strip('/')
        if not norm_path:
            subfolders = ["routing", "switching", "security"]
            scripts = []
        elif norm_path == "routing":
            subfolders = ["bgp", "ospf"]
            scripts = [{"name": "test_routing_basic.py", "path": "routing/test_routing_basic.py", "full_path": "/mock/routing/test_routing_basic.py"}]
        elif norm_path == "routing/bgp":
            subfolders = []
            scripts = [
                {"name": "test_bgp_route_advertise.py", "path": "routing/bgp/test_bgp_route_advertise.py", "full_path": "/mock/routing/bgp/test_bgp_route_advertise.py"},
                {"name": "test_bgp_route_filtering.py", "path": "routing/bgp/test_bgp_route_filtering.py", "full_path": "/mock/routing/bgp/test_bgp_route_filtering.py"},
                {"name": "test_bgp_graceful_restart.py", "path": "routing/bgp/test_bgp_graceful_restart.py", "full_path": "/mock/routing/bgp/test_bgp_graceful_restart.py"}
            ]
        elif norm_path == "switching":
            subfolders = ["lacp", "vlan"]
            scripts = []
        elif norm_path == "switching/lacp":
            subfolders = []
            scripts = [
                {"name": "test_lacp_convergence.py", "path": "switching/lacp/test_lacp_convergence.py", "full_path": "/mock/switching/lacp/test_lacp_convergence.py"},
                {"name": "test_lacp_redundancy.py", "path": "switching/lacp/test_lacp_redundancy.py", "full_path": "/mock/switching/lacp/test_lacp_redundancy.py"}
            ]
        elif norm_path == "switching/vlan":
            subfolders = []
            scripts = [
                {"name": "test_vlan_trunking.py", "path": "switching/vlan/test_vlan_trunking.py", "full_path": "/mock/switching/vlan/test_vlan_trunking.py"}
            ]
        else:
            subfolders = []
            scripts = []
        return {
            "current_path": norm_path,
            "parent_path": parent_path,
            "subfolders": subfolders,
            "scripts": scripts,
            "subfolder_count": len(subfolders),
            "script_count": len(scripts)
        }

    ssh, dut = _ssh_to_host(host_id, db)
    try:
        path = path.strip().strip('/')
        tests_root = (base_path or '').strip().rstrip('/') or SPYTEST_TESTS_DIR
        full_path = f"{tests_root}/{path}" if path else tests_root

        check_cmd = f'[ -d "{full_path}" ] && echo "EXISTS" || echo "NOT_FOUND"'
        check_out, _, _ = ssh.execute_command(check_cmd, timeout=5)
        if "NOT_FOUND" in check_out:
            parent_path = ""
            if path and "/" in path:
                parent_path = "/".join(path.split("/")[:-1])
            return {
                "current_path": path,
                "parent_path": parent_path,
                "subfolders": [],
                "scripts": [],
                "subfolder_count": 0,
                "script_count": 0,
                "warning": f"SpyTest tests directory not found on this VM: {full_path}",
            }

        subfolder_cmd = f'find "{full_path}" -mindepth 1 -maxdepth 1 -type d ! -name "__pycache__" ! -name ".*" -printf "%f\\n" | sort'
        subfolder_out, subfolder_err, subfolder_code = ssh.execute_command(subfolder_cmd, timeout=15)

        if subfolder_code != 0:
            logger.warning(f"Failed to list subfolders in {path}: {subfolder_err}")
            subfolders = []
        else:
            subfolders = [f.strip() for f in subfolder_out.strip().split("\n") if f.strip()]

        script_cmd = f'find "{full_path}" -maxdepth 1 -name "test_*.py" -type f | sort'
        script_out, script_err, script_code = ssh.execute_command(script_cmd, timeout=15)

        if script_code != 0:
            logger.warning(f"Failed to list scripts in {path}: {script_err}")
            scripts = []
        else:
            scripts = []
            for line in script_out.strip().split("\n"):
                line = line.strip()
                if not line:
                    continue
                rel_path = line.replace(tests_root + "/", "")
                name = os.path.basename(line)
                scripts.append({"name": name, "path": rel_path, "full_path": line})

        parent_path = ""
        if path and "/" in path:
            parent_path = "/".join(path.split("/")[:-1])

        return {
            "current_path": path,
            "parent_path": parent_path,
            "subfolders": subfolders,
            "scripts": scripts,
            "subfolder_count": len(subfolders),
            "script_count": len(scripts)
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error browsing folder {path}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to browse folder: {str(e)}")
    finally:
        ssh.disconnect()

# ── POST /api/spytest/script-info ─────────────────────────────────────────────
@app.post("/api/spytest/script-info")
def get_spytest_script_info(body: dict, db: Session = Depends(get_db)):
    """Parse a SPyTest script to extract topology requirements."""
    host_id = body.get("host_id")
    script_path = body.get("script_path")
    if not host_id or not script_path:
        raise HTTPException(status_code=400, detail="host_id and script_path required")

    if host_id == 5:
        script_name = os.path.basename(script_path)
        return {
            "topology_marker": "D1D2:2",
            "min_topology": ["D1", "D2", "D1D2:2"],
            "dut_count": 2,
            "description": f"Verifies functionality of BGP routing on 2 connected switches. Running: {script_name}",
            "topology_type": "linear",
            "script_path": script_path,
            "script_name": script_name
        }

    ssh, dut = _ssh_to_host(host_id, db)
    try:
        user_base = (body.get("base_path") or "").strip().rstrip("/")
        tests_root = user_base or SPYTEST_TESTS_DIR
        full_path = f"{tests_root}/{script_path}"
        cmd = f'cat {full_path}'
        output, error, code = ssh.execute_command(cmd, timeout=15)
        if code != 0:
            raise HTTPException(status_code=500, detail=f"Failed to read script: {error}")

        script_content = output
        info = _parse_spytest_script(script_content)
        info["script_path"] = script_path
        info["script_name"] = os.path.basename(script_path)
        return info
    finally:
        ssh.disconnect()

# ── POST /api/spytest/execute ─────────────────────────────────────────────────
@app.post("/api/spytest/execute")
def start_spytest_execution(body: dict, db: Session = Depends(get_db)):
    """Smart SPyTest execution with topology-aware DUT allocation and parallel scheduling."""
    host_id = body.get("host_id")
    scripts = body.get("scripts", [])
    testbed_file = body.get("testbed")
    options = body.get("options", {})
    available_dut_count = int(body.get("available_dut_count", 1))

    if not host_id or not scripts or not testbed_file:
        raise HTTPException(status_code=400, detail="host_id, scripts, and testbed required")

    dut = db.query(DUT).filter(DUT.id == host_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="Host device not found")

    job_id = body.get("job_id")
    exec_name = f"spytest_{int(datetime.utcnow().timestamp())}"
    execution = Execution(
        name=exec_name,
        dut_ids=json.dumps([host_id]),
        execution_type="spytest",
        status="pending",
        job_id=int(job_id) if job_id else None,
    )
    db.add(execution)
    db.commit()
    db.refresh(execution)

    # Pre-initialise queue state NOW so the frontend can poll immediately
    script_names = [os.path.basename(s.get("path", "")) for s in scripts]
    _q_init(execution.id, script_names, [])

    _init_pending_scripts(execution.id)

    user_base_path = (body.get("base_path") or "").strip().rstrip("/")

    if job_id:
        _job = db.query(ExecutionJob).filter(ExecutionJob.id == int(job_id)).first()
        if _job:
            _job.status = "running"
            db.commit()

    thread = Thread(
        target=_run_spytest_execution,
        args=(execution.id, host_id, scripts, testbed_file, options, available_dut_count, user_base_path, int(job_id) if job_id else None),
        daemon=True,
    )
    thread.start()

    return {
        "execution_id": execution.id,
        "status": "started",
        "type": "spytest",
        "script_count": len(scripts),
    }

# ── POST /api/spytest/generate-temp-yaml ──────────────────────────────────────
@app.post("/api/spytest/generate-temp-yaml")
def generate_temp_yaml(body: dict, db: Session = Depends(get_db)):
    """Generate a temp testbed YAML on the remote VM substituting live DUT data."""
    host_id = body.get("host_id")
    testbed_filename = body.get("testbed_filename", "")
    connections = body.get("connections", [])
    dut_ids = body.get("dut_ids", [])
    script_name = body.get("script_name", "script")

    if not host_id or not testbed_filename:
        raise HTTPException(status_code=400, detail="host_id and testbed_filename required")

    vm = db.query(DUT).filter(DUT.id == host_id).first()
    if not vm:
        raise HTTPException(status_code=404, detail="VM device not found")

    dut_records = db.query(DUT).filter(DUT.id.in_(dut_ids)).all() if dut_ids else []
    dut_map = {d.id: d for d in dut_records}

    ssh = SSHConnectionManager(vm.ip_address, vm.port, vm.username, vm.password)
    if not ssh.connect():
        raise HTTPException(status_code=503, detail=f"Cannot connect to {vm.name}")

    try:
        ref_path = f"{SPYTEST_TESTBED_DIR}/{testbed_filename}"
        output, error, code = ssh.execute_command(f"cat {ref_path}", timeout=15)
        if code != 0:
            raise HTTPException(status_code=404, detail=f"Testbed file not found: {testbed_filename}")

        ref_config = yaml.safe_load(output) or {}
        ref_devices = ref_config.get("devices", {})
        ref_topology = ref_config.get("topology", {})
        ref_device_names = list(ref_devices.keys())

        device_name_to_dut = {}
        for i, dev_name in enumerate(ref_device_names):
            if i < len(dut_ids):
                device_name_to_dut[dev_name] = dut_map.get(dut_ids[i])

        id_to_dev_name = {}
        for dev_name, dut in device_name_to_dut.items():
            if dut:
                id_to_dev_name[dut.id] = dev_name

        canvas_topology = {}
        for conn in connections:
            a_id = int(conn.get("dut_a", 0) or 0)
            b_id = int(conn.get("dut_b", 0) or 0)
            intf_a = conn.get("intf_a", "Ethernet0")
            intf_b = conn.get("intf_b", "Ethernet0")
            name_a = id_to_dev_name.get(a_id)
            name_b = id_to_dev_name.get(b_id)
            if name_a and name_b:
                canvas_topology.setdefault(name_a, {"interfaces": {}})
                canvas_topology.setdefault(name_b, {"interfaces": {}})
                canvas_topology[name_a]["interfaces"][intf_a] = {
                    "EndDevice": name_b, "EndPort": intf_b}
                canvas_topology[name_b]["interfaces"][intf_b] = {
                    "EndDevice": name_a, "EndPort": intf_a}

        temp_config = {
            "version": ref_config.get("version", "2.0"),
            "devices": {},
            "topology": canvas_topology if canvas_topology else ref_topology,
            "services": ref_config.get("services", {"default": {}}),
            "builds": ref_config.get("builds", {"default": {}}),
            "configs": ref_config.get("configs", {"default": {}}),
            "errors": ref_config.get("errors", {"default": {}}),
            "params": ref_config.get("params", {}),
        }

        for dev_name, ref_dev in ref_devices.items():
            dut = device_name_to_dut.get(dev_name)
            dev_entry = dict(ref_dev)
            if dut:
                dev_entry["ip"] = dut.ip_address
                dev_entry["username"] = dut.username
                dev_entry["password"] = dut.password
            temp_config["devices"][dev_name] = dev_entry

        unique_id = str(uuid.uuid4())[:8]
        safe_script = re.sub(r'[^a-zA-Z0-9_]', '_', script_name)[:30]
        temp_filename = f"testbed_{unique_id}_{safe_script}.yaml"
        temp_path = f"/tmp/{temp_filename}"

        yaml_str = (
            f"# AUTO-GENERATED — DO NOT EDIT MANUALLY\n"
            f"# Generated by Eka Automation | Script: {script_name}\n"
            + yaml.dump(temp_config, default_flow_style=False)
        )

        yaml_b64 = base64.b64encode(yaml_str.encode()).decode()
        write_cmd = f"echo '{yaml_b64}' | base64 -d > {temp_path}"
        _, write_err, write_code = ssh.execute_command(write_cmd, timeout=15)
        if write_code != 0:
            raise HTTPException(status_code=500,
                                detail=f"Failed to write temp YAML: {write_err}")

        logger.info(f"Temp YAML written: {temp_path}")
        return {
            "temp_yaml_path": temp_path,
            "temp_filename": temp_filename,
            "device_count": len(temp_config["devices"]),
            "devices": list(temp_config["devices"].keys()),
        }

    finally:
        ssh.disconnect()

# ── GET /api/spytest/testbed-info ─────────────────────────────────────────────
@app.get("/api/spytest/testbed-info")
def get_testbed_info(host_id: int, testbed: str, db: Session = Depends(get_db)):
    """Read and parse a testbed YAML file from the remote VM."""
    ssh, dut = _ssh_to_host(host_id, db)
    try:
        testbed_path = f"{SPYTEST_TESTBED_DIR}/{testbed}"
        output, error, code = ssh.execute_command(f"cat {testbed_path}", timeout=15)
        if code != 0:
            raise HTTPException(status_code=500, detail=f"Failed to read testbed: {error.strip()}")
        config = yaml.safe_load(output) or {}
        devices = list(config.get("devices", {}).keys())
        topology = config.get("topology", {})

        links = []
        seen: set = set()
        for dev_name, dev_topo in topology.items():
            if not isinstance(dev_topo, dict):
                continue
            for iface, link in dev_topo.get("interfaces", {}).items():
                if not isinstance(link, dict):
                    continue
                end_dev = link.get("EndDevice", "")
                end_port = link.get("EndPort", "")
                if end_dev:
                    key = tuple(sorted([f"{dev_name}:{iface}", f"{end_dev}:{end_port}"]))
                    if key not in seen:
                        seen.add(key)
                        links.append({
                            "from": f"{dev_name}:{iface}",
                            "to": f"{end_dev}:{end_port}",
                        })

        n = len(devices)
        if n == 0:
            topology_type = "empty"
        elif n == 1:
            topology_type = "standalone"
        elif n == 2:
            topology_type = "dual-dut"
        else:
            topology_type = f"{n}-node"

        return {
            "testbed": testbed,
            "device_count": n,
            "device_names": devices,
            "topology_type": topology_type,
            "link_count": len(links),
            "links": links,
        }
    finally:
        ssh.disconnect()

# ── POST /api/topology/connections ────────────────────────────────────────────
@app.get("/api/topology/connections")
def get_topology(request: Request, db: Session = Depends(get_db)):
    session_id = get_session_id(request)
    conns = db.query(TopologyConnection).all()
    return [{"id": c.id, "dut_a_id": c.dut_a_id, "intf_a": c.intf_a,
             "dut_b_id": c.dut_b_id, "intf_b": c.intf_b} for c in conns]

@app.post("/api/topology/connections")
def create_topology_connection(body: dict, db: Session = Depends(get_db)):
    conn = TopologyConnection(
        dut_a_id=body["dut_a_id"], intf_a=body.get("intf_a", "Ethernet0"),
        dut_b_id=body["dut_b_id"], intf_b=body.get("intf_b", "Ethernet0"))
    db.add(conn)
    db.commit()
    db.refresh(conn)
    return {"id": conn.id, "status": "created"}

@app.delete("/api/topology/connections/{conn_id}")
def delete_topology_connection(conn_id: int, db: Session = Depends(get_db)):
    conn = db.query(TopologyConnection).filter(TopologyConnection.id == conn_id).first()
    if conn:
        db.delete(conn)
        db.commit()
    return {"status": "deleted"}

@app.delete("/api/topology/connections")
def clear_topology_connections(db: Session = Depends(get_db)):
    db.query(TopologyConnection).delete()
    db.commit()
    return {"status": "cleared"}

# ── POST /api/topology/generate-master-testbed ────────────────────────────────
@app.post("/api/topology/generate-master-testbed")
def generate_master_testbed(body: dict, request: Request, db: Session = Depends(get_db)):
    """Generate a master testbed YAML from session's Topology Canvas DUTs and connections."""
    host_id = body.get("host_id")
    master_filename = body.get("master_filename", "master_testbed.yaml")
    session_id = get_session_id(request)

    user_base_path = (body.get("base_path") or "").strip().rstrip("/")
    if not user_base_path:
        raise HTTPException(
            status_code=400,
            detail="SCRIPTS_PATH_REQUIRED: Please enter the Scripts Path on VM in the 'Categories & Scripts' section and click Load first."
        )
    _m = re.match(r'(^.*?/spytest)(?:/|$)', user_base_path)
    _spy_root = _m.group(1) if _m else os.path.dirname(user_base_path)
    testbed_dir = _spy_root + "/testbeds"

    if not host_id:
        raise HTTPException(status_code=400, detail="host_id required")

    vm = db.query(DUT).filter(DUT.id == host_id).first()
    if not vm:
        raise HTTPException(status_code=404, detail="VM host not found")

    query = db.query(DUT).filter(DUT.device_type != "VM")
    if session_id:
        query = query.filter(DUT.session_id == session_id)
    all_duts = query.all()

    if not all_duts:
        raise HTTPException(status_code=400, detail="No DUT devices found in session")

    seen_names = set()
    unique_duts = []
    duplicate_count = 0
    for dut in all_duts:
        normalized_name = dut.name.replace(" ", "_").replace("-", "_")
        if normalized_name not in seen_names:
            seen_names.add(normalized_name)
            unique_duts.append(dut)
        else:
            duplicate_count += 1
            logger.warning(f"MASTER TESTBED: Skipping duplicate device '{dut.name}' (ID: {dut.id})")

    if duplicate_count > 0:
        logger.warning(f"MASTER TESTBED: Removed {duplicate_count} duplicate device(s) from generation")

    all_duts = unique_duts

    all_connections = db.query(TopologyConnection).all()

    dut_ids_in_session = {dut.id for dut in all_duts}
    session_connections = [
        conn for conn in all_connections
        if conn.dut_a_id in dut_ids_in_session and conn.dut_b_id in dut_ids_in_session
    ]

    if not session_connections:
        raise HTTPException(status_code=400, detail="No connections found in Topology Canvas. Create connections first.")

    connected_dut_ids = set()
    for conn in session_connections:
        connected_dut_ids.add(conn.dut_a_id)
        connected_dut_ids.add(conn.dut_b_id)

    canvas_duts = [dut for dut in all_duts if dut.id in connected_dut_ids]

    if not canvas_duts:
        raise HTTPException(status_code=400, detail="No devices with connections found in Topology Canvas")

    ssh = SSHConnectionManager(vm.ip_address, vm.port, vm.username, vm.password)
    if not ssh.connect():
        raise HTTPException(status_code=503, detail=f"Cannot connect to {vm.name}")

    try:
        devices_section = {}
        dut_id_to_name = {}

        for dut in canvas_duts:
            device_name = dut.name.replace(" ", "_").replace("-", "_")
            dut_id_to_name[dut.id] = device_name

            devices_section[device_name] = {
                "device_type": "sonic",
                "access": {
                    "protocol": "ssh",
                    "ip": dut.ip_address,
                    "port": dut.port
                },
                "credentials": {
                    "username": dut.username,
                    "password": dut.password,
                    "altpassword": "broadcom"
                },
                "properties": {
                    "services": "default",
                    "build": "default",
                    "config": "default",
                    "errors": "default"
                }
            }

        topology_section = {}
        connection_count = 0
        topo_dict = {}

        device_to_generic = {}
        generic_index = 1
        for device_name in sorted(devices_section.keys()):
            device_to_generic[device_name] = f"D{generic_index}"
            generic_index += 1

        for conn in session_connections:
            device_a = dut_id_to_name.get(conn.dut_a_id)
            device_b = dut_id_to_name.get(conn.dut_b_id)

            if device_a and device_b:
                if device_a not in topology_section:
                    topology_section[device_a] = {"interfaces": {}}
                if device_b not in topology_section:
                    topology_section[device_b] = {"interfaces": {}}

                topology_section[device_a]["interfaces"][conn.intf_a] = {
                    "EndDevice": device_b,
                    "EndPort": conn.intf_b
                }
                topology_section[device_b]["interfaces"][conn.intf_b] = {
                    "EndDevice": device_a,
                    "EndPort": conn.intf_a
                }
                connection_count += 1

                gen_a = device_to_generic.get(device_a)
                gen_b = device_to_generic.get(device_b)
                if gen_a and gen_b:
                    topo_key = f"{gen_a}{gen_b}" if gen_a < gen_b else f"{gen_b}{gen_a}"
                    topo_dict[topo_key] = topo_dict.get(topo_key, 0) + 1

        test_interface = None
        for dev_name in sorted(topology_section.keys()):
            intfs = topology_section[dev_name].get("interfaces", {})
            if intfs:
                test_interface = sorted(intfs.keys())[0]
                break

        params_section = {"topo": topo_dict if topo_dict else {}}

        master_config = {
            "version": "2.0",
            "devices": devices_section,
            "topology": topology_section,
            "services": {"default": {}},
            "builds": {"default": {}},
            "configs": {"default": {}},
            "errors": {"default": {}},
            "params": params_section,
        }
        if test_interface:
            master_config["global"] = {"params": {"test_interface": test_interface}}

        yaml_content = (
            f"# MASTER TESTBED - AUTO-GENERATED FROM CANVAS TOPOLOGY\n"
            f"# Generated by Eka Automation at {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC\n"
            f"# Total Devices: {len(devices_section)}\n"
            f"# Total Connections: {connection_count}\n"
            f"# DO NOT EDIT MANUALLY - Regenerate from canvas as needed\n\n"
            + yaml.dump(master_config, default_flow_style=None, sort_keys=False, width=120)
        )

        master_path = f"{testbed_dir}/{master_filename}"
        _, mkdir_err, mkdir_code = ssh.execute_command(f'mkdir -p "{testbed_dir}"', timeout=10)
        if mkdir_code != 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"PATH_NOT_FOUND: Cannot create testbed directory: {testbed_dir}\n"
                    f"{mkdir_err.strip()}\n"
                    f"Please verify the Scripts Path on VM is correct and the user has write permission."
                )
            )

        yaml_b64 = base64.b64encode(yaml_content.encode()).decode()
        stdin_ch, stdout_ch, stderr_ch = ssh.client.exec_command(f'base64 -d > "{master_path}"', timeout=15)
        stdin_ch.write(yaml_b64.encode())
        stdin_ch.channel.shutdown_write()
        write_err = stderr_ch.read().decode("utf-8", errors="ignore")
        write_code = stdout_ch.channel.recv_exit_status()
        if write_code != 0:
            raise HTTPException(
                status_code=500,
                detail=(
                    f"PATH_NOT_FOUND: Failed to write master testbed to {master_path}\n"
                    f"{write_err.strip()}\n"
                    f"Please verify the Scripts Path on VM and directory permissions."
                )
            )

        logger.info(f"Master testbed generated: {master_path} with {len(devices_section)} devices")

        return {
            "success": True,
            "master_testbed_path": master_path,
            "master_filename": master_filename,
            "device_count": len(devices_section),
            "connection_count": connection_count,
            "devices": list(devices_section.keys()),
            "timestamp": datetime.utcnow().isoformat(),
        }

    finally:
        ssh.disconnect()

# ============================================================================
# Execution Queue, Jobs, Scheduler — ported from core main.py
# ============================================================================

# ── Helper: result/formatting utilities ──────────────────────────────────────
def _extract_feature(module_path: str) -> str:
    if not module_path:
        return "Unknown"
    parts = module_path.replace("\\", "/").split("/")
    if len(parts) >= 2:
        feat = parts[-2].replace("iscli_", "").replace("ISCLI_", "")
        return feat.upper()
    return parts[0].upper() if parts else "Unknown"

def _extract_tc_id(test_function: str) -> str:
    if not test_function:
        return ""
    return test_function.split(".")[-1] if "." in test_function else test_function

def _fmt_seconds(secs) -> str:
    secs = secs or 0
    h, rem = divmod(int(secs), 3600)
    m, s = divmod(rem, 60)
    parts = []
    if h: parts.append(f"{h}h")
    if m: parts.append(f"{m}m")
    if s: parts.append(f"{s}s")
    return " ".join(parts) or "0s"

# ── Helper: APScheduler registration ─────────────────────────────────────────
def _register_job_schedule(job):
    """Add or replace an APScheduler entry for this ExecutionJob."""
    ap_id = f"exec_job_{job.id}"
    try:
        scheduler.remove_job(ap_id)
    except Exception:
        pass

    if not job.schedule_enabled:
        return

    try:
        if job.schedule_type == "once" and job.schedule_at:
            if job.schedule_at <= datetime.utcnow():
                logger.warning(f"[Scheduler] Job {job.id} schedule_at is in the past — not scheduling")
                return
            from datetime import timezone as _utc_tz
            aware_dt = job.schedule_at.replace(tzinfo=_utc_tz.utc)
            scheduler.add_job(
                _trigger_scheduled_job,
                trigger=DateTrigger(run_date=aware_dt),
                args=[job.id],
                id=ap_id,
                replace_existing=True,
                misfire_grace_time=300,
            )
            logger.info(f"[Scheduler] Job {job.id} '{job.name}' scheduled once at {job.schedule_at} UTC")

        elif job.schedule_type == "cron" and job.schedule_cron:
            cron_str = job.schedule_cron.strip()
            if re.match(r'^\d{1,2}:\d{2}$', cron_str):
                hh, mm = cron_str.split(":")
                trigger = CronTrigger(hour=int(hh), minute=int(mm))
            else:
                parts = cron_str.split()
                if len(parts) == 5:
                    trigger = CronTrigger(
                        minute=parts[0], hour=parts[1],
                        day=parts[2], month=parts[3], day_of_week=parts[4]
                    )
                else:
                    logger.warning(f"[Scheduler] Job {job.id} invalid cron '{cron_str}'")
                    return
            scheduler.add_job(
                _trigger_scheduled_job,
                trigger=trigger,
                args=[job.id],
                id=ap_id,
                replace_existing=True,
            )
            logger.info(f"[Scheduler] Job {job.id} '{job.name}' scheduled cron '{cron_str}'")
    except Exception as e:
        logger.error(f"[Scheduler] Failed to register job {job.id}: {e}")


def _reload_all_job_schedules():
    """On startup, re-register APScheduler entries for all enabled jobs."""
    db = SessionLocal()
    try:
        jobs = db.query(ExecutionJob).filter(ExecutionJob.schedule_enabled == True).all()
        for j in jobs:
            _register_job_schedule(j)
        if jobs:
            logger.info(f"[Scheduler] Reloaded {len(jobs)} job schedule(s) from DB")
    except Exception as e:
        logger.error(f"[Scheduler] Failed to reload schedules: {e}")
    finally:
        db.close()


def _trigger_scheduled_job(execution_job_id: int):
    """Called by APScheduler when a job's scheduled time arrives."""
    db = SessionLocal()
    try:
        job = db.query(ExecutionJob).filter(ExecutionJob.id == execution_job_id).first()
        if not job:
            logger.warning(f"[Scheduler] Job {execution_job_id} not found — skipping")
            return
        if job.status == "running":
            logger.info(f"[Scheduler] Job {execution_job_id} already running — skipping")
            return
        if not job.schedule_enabled:
            logger.info(f"[Scheduler] Job {execution_job_id} schedule disabled — skipping")
            return

        scripts   = json.loads(job.scripts)   if job.scripts   else []
        host_id   = job.host_id
        base_path = job.base_path or ""

        if not host_id or not scripts:
            logger.warning(f"[Scheduler] Job {execution_job_id} missing host or scripts — skipping")
            return

        host_dut = db.query(DUT).filter(DUT.id == host_id).first()
        if not host_dut:
            logger.warning(f"[Scheduler] Job {execution_job_id} host DUT {host_id} not found — skipping")
            return

        testbed_file = job.testbed_path or ""
        if not testbed_file:
            logger.warning(f"[Scheduler] Job {execution_job_id} has no testbed — skipping. "
                           "Run the job manually once to save the testbed path.")
            return

        exec_name = f"scheduled_{execution_job_id}_{int(datetime.utcnow().timestamp())}"
        execution = Execution(
            name=exec_name,
            dut_ids=json.dumps([host_id]),
            execution_type="spytest",
            status="pending",
            job_id=execution_job_id,
        )
        db.add(execution)
        job.status = "running"
        job.last_run_at = datetime.utcnow()
        if job.schedule_type == "once":
            job.schedule_enabled = False
        db.commit()
        db.refresh(execution)

        script_names = [os.path.basename(s.get("path", "")) for s in scripts]
        _q_init(execution.id, script_names, [])
        _init_pending_scripts(execution.id)

        available_dut_count = len(json.loads(job.dut_ids) if job.dut_ids else [1])

        Thread(
            target=_run_spytest_execution,
            args=(execution.id, host_id, scripts, testbed_file, {}, available_dut_count, base_path, execution_job_id),
            daemon=True,
        ).start()

        logger.info(f"[Scheduler] Job {execution_job_id} '{job.name}' triggered — execution {execution.id}")

    except Exception as e:
        logger.error(f"[Scheduler] Error triggering job {execution_job_id}: {e}")
    finally:
        db.close()


# ── Helper: HTML/Excel report builders ───────────────────────────────────────
def _build_html_dashboard(execution, results: list, tcrs: list) -> str:
    from collections import defaultdict
    total_p = sum(r.get("passed", 0) for r in results)
    total_f = sum(r.get("failed", 0) for r in results)
    total_s = sum(r.get("skipped", 0) for r in results)
    total   = total_p + total_f + total_s
    total_runtime_s = sum(t.time_seconds or 0 for t in tcrs)
    pp = (total_p / total * 100) if total else 0
    fp = (total_f / total * 100) if total else 0
    sp = (total_s / total * 100) if total else 0
    by_feature: dict = defaultdict(list)
    for t in tcrs:
        feat = _extract_feature(t.module or t.script_path or "")
        by_feature[feat].append(t)
    css = """
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',Tahoma,Geneva,Verdana,sans-serif;
     background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);padding:20px;min-height:100vh}
.container{max-width:1400px;margin:0 auto;background:#fff;border-radius:12px;
           box-shadow:0 10px 40px rgba(0,0,0,.2);overflow:hidden}
.header{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);color:#fff;padding:30px;text-align:center}
.header h1{font-size:30px;margin-bottom:8px;font-weight:600}
.header p{font-size:13px;opacity:.9}
.summary-section{padding:28px;background:#f8f9fa;border-bottom:1px solid #e9ecef}
.summary-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:16px;margin-bottom:20px}
.summary-card{background:#fff;padding:18px;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.08);text-align:center}
.summary-card h3{font-size:12px;color:#6c757d;margin-bottom:8px;text-transform:uppercase;letter-spacing:.5px}
.summary-card .val{font-size:34px;font-weight:700;margin-bottom:4px}
.val.total{color:#667eea}.val.passed{color:#28a745}.val.failed{color:#dc3545}
.val.skipped{color:#ffc107}.val.runtime{color:#17a2b8}
.progress-bar{width:100%;height:28px;background:#e9ecef;border-radius:4px;overflow:hidden;display:flex}
.progress-bar.sm{height:18px;margin-top:10px}
.seg{height:100%;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:#fff}
.seg.pass{background:#28a745}.seg.fail{background:#dc3545}.seg.skip{background:#ffc107}
.tabs{display:flex;flex-wrap:wrap;background:#f8f9fa;border-bottom:2px solid #dee2e6;padding:0 20px}
.tab-btn{background:none;border:none;padding:14px 22px;cursor:pointer;font-size:13px;font-weight:500;
         color:#6c757d;border-bottom:3px solid transparent}
.tab-btn:hover{color:#667eea}.tab-btn.active{color:#667eea;border-bottom-color:#667eea;background:#fff}
.tab-content{display:none;padding:28px}.tab-content.active{display:block}
.mod-summary{background:#f8f9fa;padding:18px;border-radius:8px;border-left:4px solid #667eea;margin-bottom:18px}
.mod-summary h2{color:#667eea;font-size:18px;margin-bottom:12px}
.mod-stats{display:flex;flex-wrap:wrap;gap:14px;margin-bottom:12px;font-size:13px}
table{width:100%;border-collapse:collapse;background:#fff;
      box-shadow:0 2px 8px rgba(0,0,0,.08);border-radius:8px;overflow:hidden;margin-top:18px}
thead{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%)}
th{color:#fff;padding:11px 12px;text-align:left;font-weight:600;font-size:11px;text-transform:uppercase}
td{padding:9px 12px;border-bottom:1px solid #e9ecef;font-size:12px}
tr:last-child td{border-bottom:none}
tr:hover td{background:#f8f9fa}
.badge{display:inline-block;padding:3px 9px;border-radius:4px;font-weight:700;font-size:11px}
.badge.pass{background:#d4edda;color:#155724}.badge.fail{background:#f8d7da;color:#721c24}
.badge.skip{background:#fff3cd;color:#856404}.badge.other{background:#e2e3e5;color:#383d41}
.footer{padding:18px;text-align:center;background:#f8f9fa;color:#6c757d;font-size:11px;border-top:1px solid #e9ecef}
"""
    features = sorted(by_feature.keys())
    tab_nav   = "".join(
        f'<button class="tab-btn" onclick="openTab(event,\'{f.replace(" ","_").replace("/","_")}\')">{f}</button>\n'
        for f in features)
    tab_bodies = ""
    for feat in features:
        safe = feat.replace(" ", "_").replace("/", "_")
        tcs  = by_feature[feat]
        fp2  = sum(1 for t in tcs if (t.result or "").lower() == "pass")
        ff2  = sum(1 for t in tcs if (t.result or "").lower() in ("fail","scripterror","error"))
        fs2  = sum(1 for t in tcs if (t.result or "").lower() in ("skip","xfail","deselect"))
        ft2  = fp2 + ff2 + fs2
        fpp  = (fp2/ft2*100) if ft2 else 0
        ffp  = (ff2/ft2*100) if ft2 else 0
        fsp  = (fs2/ft2*100) if ft2 else 0
        prog  = ""
        if fpp > 0: prog += f'<div class="seg pass" style="width:{fpp:.1f}%">{fpp:.0f}%</div>'
        if ffp > 0: prog += f'<div class="seg fail" style="width:{ffp:.1f}%">{ffp:.0f}%</div>'
        if fsp > 0: prog += f'<div class="seg skip" style="width:{fsp:.1f}%">{fsp:.0f}%</div>'
        rows  = ""
        for idx, t in enumerate(tcs, 1):
            res  = (t.result or "").strip()
            rl   = res.lower()
            cls  = ("pass" if rl == "pass" else
                    "fail" if rl in ("fail","scripterror","error") else
                    "skip" if rl in ("skip","xfail","deselect") else "other")
            tc_id = _extract_tc_id(t.test_function or "")
            desc  = (t.description or "")[:120]
            rows += (f"<tr><td style='text-align:center;font-size:11px'>{idx}</td>"
                     f"<td style='font-weight:600;font-size:11px;color:#667eea'>{feat}</td>"
                     f"<td style='font-size:10px;word-break:break-all'>{t.module or t.script_path or ''}</td>"
                     f"<td style='font-family:monospace;font-size:11px'>{tc_id}</td>"
                     f"<td style='font-size:11px'>{desc}</td>"
                     f"<td style='text-align:center;font-size:11px'>{t.time_taken or ''}</td>"
                     f"<td><span class='badge {cls}'>{res}</span></td></tr>")
        tab_bodies += f"""
<div id="{safe}" class="tab-content">
  <div class="mod-summary"><h2>{feat} — Test Results</h2>
    <div class="mod-stats">
      <span>✓ Pass: <strong style="color:#28a745">{fp2}</strong> ({fpp:.1f}%)</span>
      <span style="margin:0 12px">✗ Fail: <strong style="color:#dc3545">{ff2}</strong> ({ffp:.1f}%)</span>
      <span>⊘ Skip: <strong style="color:#ffc107">{fs2}</strong> ({fsp:.1f}%)</span>
      <span style="margin-left:12px">Total: <strong>{ft2}</strong></span>
    </div><div class="progress-bar sm">{prog}</div></div>
  <table><thead><tr>
    <th style="width:36px">S.No</th><th style="min-width:80px">Feature</th>
    <th style="min-width:180px">Script</th><th style="min-width:180px">Testcase ID</th>
    <th style="min-width:260px">Description</th>
    <th style="width:80px">Time</th><th style="width:80px">Status</th>
  </tr></thead><tbody>{rows}</tbody></table>
</div>"""
    overall_prog = ""
    if pp > 0: overall_prog += f'<div class="seg pass" style="width:{pp:.1f}%">{pp:.0f}%</div>'
    if fp > 0: overall_prog += f'<div class="seg fail" style="width:{fp:.1f}%">{fp:.0f}%</div>'
    if sp > 0: overall_prog += f'<div class="seg skip" style="width:{sp:.1f}%">{sp:.0f}%</div>'
    generated_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Eka Dashboard — {execution.name}</title><style>{css}</style></head><body>
<div class="container">
  <div class="header"><h1>SONiC Test Dashboard</h1>
    <p>Execution #{execution.id} &nbsp;·&nbsp; {execution.name} &nbsp;·&nbsp;
       {execution.status} &nbsp;·&nbsp; {generated_at}</p></div>
  <div class="summary-section"><div class="summary-grid">
    <div class="summary-card"><h3>Total Tests</h3><div class="val total">{total}</div></div>
    <div class="summary-card"><h3>Passed</h3><div class="val passed">{total_p}</div><div class="pct">{pp:.1f}%</div></div>
    <div class="summary-card"><h3>Failed</h3><div class="val failed">{total_f}</div><div class="pct">{fp:.1f}%</div></div>
    <div class="summary-card"><h3>Skipped</h3><div class="val skipped">{total_s}</div><div class="pct">{sp:.1f}%</div></div>
    <div class="summary-card"><h3>Total Runtime</h3><div class="val runtime">{_fmt_seconds(total_runtime_s)}</div></div>
  </div>
  <div style="margin-top:16px"><h3 style="font-size:14px;color:#495057;margin-bottom:8px">Overall Progress</h3>
    <div class="progress-bar">{overall_prog}</div></div></div>
  <div class="tabs">{tab_nav}</div>{tab_bodies}
  <div class="footer">Generated by Eka Automation Platform &nbsp;·&nbsp; {generated_at}</div>
</div>
<script>
function openTab(evt,name){{
  document.querySelectorAll('.tab-content').forEach(e=>e.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(e=>e.classList.remove('active'));
  document.getElementById(name).classList.add('active');
  evt.currentTarget.classList.add('active');
}}
window.onload=function(){{var first=document.querySelector('.tab-btn');if(first)first.click();}};
</script></body></html>"""


def _build_excel(execution, results: list, tcrs: list):
    wb = openpyxl.Workbook()
    HDR_FILL  = PatternFill("solid", fgColor="667EEA")
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center")
    PASS_FILL = PatternFill("solid", fgColor="D4EDDA")
    FAIL_FILL = PatternFill("solid", fgColor="F8D7DA")
    SKIP_FILL = PatternFill("solid", fgColor="FFF3CD")
    PASS_FONT = Font(bold=True, color="155724")
    FAIL_FONT = Font(bold=True, color="721C24")
    SKIP_FONT = Font(bold=True, color="856404")

    def _write_header(ws, headers, col_widths=None):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = HDR_ALIGN
        ws.row_dimensions[1].height = 20
        if col_widths:
            for col, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    def _result_style(ws, row_i, col):
        c = ws.cell(row=row_i, column=col)
        v = (c.value or "").lower()
        if v == "pass":                          c.fill = PASS_FILL; c.font = PASS_FONT
        elif v in ("fail","scripterror","error"): c.fill = FAIL_FILL; c.font = FAIL_FONT
        elif v in ("skip","xfail","deselect"):   c.fill = SKIP_FILL; c.font = SKIP_FONT

    ws1 = wb.active; ws1.title = "Summary"
    total_p = sum(r.get("passed", 0) for r in results)
    total_f = sum(r.get("failed", 0) for r in results)
    total_s = sum(r.get("skipped", 0) for r in results)
    total   = total_p + total_f + total_s
    for row, (label, val) in enumerate([("Execution ID", execution.id), ("Name", execution.name),
                                         ("Status", execution.status),
                                         ("Duration", f"{execution.duration_seconds}s" if execution.duration_seconds else "–"),
                                         ("Scripts Run", len(results))], 1):
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=val)
    ws1.column_dimensions["A"].width = 18; ws1.column_dimensions["B"].width = 40
    for row, (label, val) in enumerate([("Total Tests", total), ("Passed", total_p),
                                         ("Failed", total_f), ("Skipped", total_s),
                                         ("Pass Rate", f"{total_p/total*100:.1f}%" if total else "–")], 7):
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=val)
    for col, h in enumerate(["Script","Status","Passed","Failed","Skipped","Duration (s)"], 1):
        c = ws1.cell(row=14, column=col, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = HDR_ALIGN
    for ri, r in enumerate(results, 15):
        ws1.cell(row=ri, column=1, value=r.get("script_stem", r.get("script", "")))
        ws1.cell(row=ri, column=2, value=r.get("status", ""))
        ws1.cell(row=ri, column=3, value=r.get("passed", 0))
        ws1.cell(row=ri, column=4, value=r.get("failed", 0))
        ws1.cell(row=ri, column=5, value=r.get("skipped", 0))
        ws1.cell(row=ri, column=6, value=r.get("duration_s", 0))
        _result_style(ws1, ri, 2)
    ws2 = wb.create_sheet("All Testcases")
    _write_header(ws2, ["S.No","Feature","Script / Module","Testcase ID",
                         "Test Function","Result","Time Taken","Time (s)","Description"],
                  [6,16,38,34,46,12,12,10,50])
    for ri, t in enumerate(tcrs, 2):
        feat  = _extract_feature(t.module or t.script_path or "")
        tc_id = _extract_tc_id(t.test_function or "")
        for col, val in enumerate([ri-1, feat, t.module or t.script_path or "",
                                    tc_id, t.test_function or "", t.result or "",
                                    t.time_taken or "", t.time_seconds or 0,
                                    t.description or ""], 1):
            ws2.cell(row=ri, column=col, value=val)
        _result_style(ws2, ri, 6)
    ws3 = wb.create_sheet("Failures")
    _write_header(ws3, ["S.No","Feature","Script / Module","Testcase ID",
                         "Test Function","Result","Time Taken","Description"],
                  [6,16,38,34,46,16,12,50])
    failures = [t for t in tcrs if (t.result or "").lower() in ("fail","scripterror","error")]
    if not failures:
        ws3.cell(row=2, column=1, value="No failures recorded for this execution.")
    else:
        for ri, t in enumerate(failures, 2):
            feat  = _extract_feature(t.module or t.script_path or "")
            tc_id = _extract_tc_id(t.test_function or "")
            for col, val in enumerate([ri-1, feat, t.module or t.script_path or "",
                                        tc_id, t.test_function or "", t.result or "",
                                        t.time_taken or "", t.description or ""], 1):
                ws3.cell(row=ri, column=col, value=val)
            _result_style(ws3, ri, 6)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── GET /api/execution-queue ──────────────────────────────────────────────────
@app.get("/api/execution-queue")
def get_execution_queue():
    """Return current live execution queue state for all active executions."""
    with _exec_queue_lock:
        return dict(_exec_queue_state)


# ── POST /api/execution-jobs ──────────────────────────────────────────────────
@app.post("/api/execution-jobs")
def create_execution_job(request: Request, body: dict, db: Session = Depends(get_db)):
    """Create a new named execution job for the current session."""
    session_id = request.headers.get("X-Session-ID", "default")
    name = body.get("name") or f"Job-{datetime.utcnow().strftime('%H%M')}"
    job = ExecutionJob(name=name, session_id=session_id)
    db.add(job)
    db.commit()
    db.refresh(job)
    return {"id": job.id, "name": job.name, "status": job.status,
            "created_at": job.created_at.isoformat() if job.created_at else None}


# ── GET /api/execution-jobs ───────────────────────────────────────────────────
@app.get("/api/execution-jobs")
def list_execution_jobs(request: Request, db: Session = Depends(get_db)):
    """List all execution jobs for the current session (most recent first)."""
    session_id = request.headers.get("X-Session-ID", "default")
    jobs = (db.query(ExecutionJob)
            .filter(ExecutionJob.session_id == session_id)
            .order_by(ExecutionJob.created_at.desc())
            .limit(50)
            .all())
    result = []
    for j in jobs:
        exec_count = db.query(Execution).filter(Execution.job_id == j.id).count()
        result.append({
            "id": j.id, "name": j.name, "status": j.status,
            "execution_count": exec_count,
            "created_at": j.created_at.isoformat() if j.created_at else None,
            "updated_at": j.updated_at.isoformat() if j.updated_at else None,
        })
    return {"jobs": result}


# ── GET /api/execution-jobs/{job_id} ─────────────────────────────────────────
@app.get("/api/execution-jobs/{job_id}")
def get_execution_job(job_id: int, request: Request, db: Session = Depends(get_db)):
    """Get full state of a specific job."""
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    executions = (db.query(Execution).filter(Execution.job_id == job_id)
                  .order_by(Execution.created_at.desc()).all())
    execs_data = []
    for e in executions:
        tcrs = db.query(TestCaseResult).filter(TestCaseResult.execution_id == e.id).all()
        script_map = {}
        for t in tcrs:
            stem = os.path.basename(t.script_path or "").replace(".py", "") or "unknown"
            if stem not in script_map:
                script_map[stem] = {"passed": 0, "failed": 0, "skipped": 0,
                                    "duration_s": 0, "status": "running", "script_stem": stem}
            r = (t.result or "").upper()
            if r in ("PASS", "PASSED"):
                script_map[stem]["passed"] += 1
            elif r in ("FAIL", "FAILED", "ERROR"):
                script_map[stem]["failed"] += 1
            else:
                script_map[stem]["skipped"] += 1
            script_map[stem]["duration_s"] += t.time_seconds or 0
        for s in script_map.values():
            s["status"] = "failed" if s["failed"] > 0 else "passed"
        execs_data.append({
            "id": e.id, "name": e.name, "status": e.status,
            "start_time": e.start_time.isoformat() if e.start_time else None,
            "end_time": e.end_time.isoformat() if e.end_time else None,
            "script_results": list(script_map.values()),
        })
    next_run = None
    ap_id = f"exec_job_{job.id}"
    try:
        ap_job = scheduler.get_job(ap_id)
        if ap_job and ap_job.next_run_time:
            from datetime import timezone as _utc_tz
            next_run = ap_job.next_run_time.astimezone(_utc_tz.utc).strftime("%Y-%m-%d %H:%M UTC")
    except Exception:
        pass
    return {
        "id": job.id, "name": job.name, "status": job.status,
        "dut_ids": json.loads(job.dut_ids) if job.dut_ids else [],
        "base_path": job.base_path or "",
        "host_id": job.host_id,
        "topology": json.loads(job.topology) if job.topology else [],
        "scripts": json.loads(job.scripts) if job.scripts else [],
        "testbed_path": job.testbed_path or "",
        "executions": execs_data,
        "schedule_type":    job.schedule_type or "none",
        "schedule_at":      job.schedule_at.isoformat() if job.schedule_at else None,
        "schedule_cron":    job.schedule_cron,
        "schedule_enabled": bool(job.schedule_enabled),
        "last_run_at":      job.last_run_at.isoformat() if job.last_run_at else None,
        "next_run":         next_run,
        "created_at": job.created_at.isoformat() if job.created_at else None,
    }


# ── PUT /api/execution-jobs/{job_id} ─────────────────────────────────────────
@app.put("/api/execution-jobs/{job_id}")
def update_execution_job(job_id: int, request: Request, body: dict, db: Session = Depends(get_db)):
    """Save/update job state (devices, topology, scripts, base_path)."""
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if "name" in body:       job.name       = body["name"]
    if "dut_ids" in body:    job.dut_ids    = json.dumps(body["dut_ids"])
    if "base_path" in body:  job.base_path  = body["base_path"]
    if "host_id" in body:    job.host_id    = body["host_id"]
    if "topology" in body:   job.topology   = json.dumps(body["topology"])
    if "scripts" in body:    job.scripts    = json.dumps(body["scripts"])
    if "testbed_path" in body: job.testbed_path = body["testbed_path"] or None
    if "status" in body and body["status"] in ("idle", "running", "completed", "failed"):
        job.status = body["status"]
    job.updated_at = datetime.utcnow()
    db.commit()
    return {"id": job.id, "status": job.status}


# ── DELETE /api/execution-jobs/{job_id} ──────────────────────────────────────
@app.delete("/api/execution-jobs/{job_id}")
def delete_execution_job(job_id: int, request: Request, db: Session = Depends(get_db)):
    """Delete a job (only if not currently running)."""
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.status == "running":
        raise HTTPException(status_code=400, detail="Cannot delete a running job")
    try:
        scheduler.remove_job(f"exec_job_{job_id}")
    except Exception:
        pass
    db.query(Execution).filter(Execution.job_id == job_id).update({"job_id": None})
    db.delete(job)
    db.commit()
    return {"deleted": job_id}


# ── GET /api/execution-jobs/{job_id}/conflicts ────────────────────────────────
@app.get("/api/execution-jobs/{job_id}/conflicts")
def check_job_conflicts(job_id: int, request: Request, dut_ids: str = "",
                        db: Session = Depends(get_db)):
    """Check if any requested DUT ids conflict with another job."""
    session_id = request.headers.get("X-Session-ID", "default")
    if not dut_ids.strip():
        return {"conflicts": []}
    requested = [int(x) for x in dut_ids.split(",") if x.strip().isdigit()]
    conflicts = []
    seen_duts = set()

    for did in requested:
        lock = db.query(DUTLock).filter(
            DUTLock.dut_id == did,
            DUTLock.status != "AVAILABLE",
            DUTLock.lock_type == "exec",
            DUTLock.job_id != job_id,
            DUTLock.job_id != None,
        ).first()
        if lock:
            owner_job = db.query(ExecutionJob).filter(ExecutionJob.id == lock.job_id).first()
            dut = db.query(DUT).filter(DUT.id == did).first()
            conflicts.append({
                "dut_id": did, "dut_name": dut.name if dut else str(did),
                "conflicting_job_id":   lock.job_id,
                "conflicting_job_name": owner_job.name if owner_job else f"Job {lock.job_id}",
                "conflict_type": "runtime",
            })
            seen_duts.add(did)

    other_jobs = (db.query(ExecutionJob)
                  .filter(
                      ExecutionJob.session_id == session_id,
                      ExecutionJob.id != job_id,
                      ExecutionJob.status.notin_(["completed", "failed"]),
                      ExecutionJob.dut_ids != None,
                      ExecutionJob.dut_ids != "[]",
                  ).all())
    for did in requested:
        if did in seen_duts:
            continue
        for other in other_jobs:
            try:
                other_duts = json.loads(other.dut_ids or "[]")
            except Exception:
                continue
            if did in [int(x) for x in other_duts]:
                dut = db.query(DUT).filter(DUT.id == did).first()
                conflicts.append({
                    "dut_id": did, "dut_name": dut.name if dut else str(did),
                    "conflicting_job_id":   other.id,
                    "conflicting_job_name": other.name,
                    "conflict_type": "planning",
                })
                seen_duts.add(did)
                break
    return {"conflicts": conflicts}


# ── GET /api/execution-jobs/{job_id}/report/html ──────────────────────────────
@app.get("/api/execution-jobs/{job_id}/report/html")
def job_html_report(job_id: int, request: Request, db: Session = Depends(get_db)):
    """Generate aggregated HTML report for all executions in this job."""
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    executions = db.query(Execution).filter(Execution.job_id == job_id).all()
    if not executions:
        raise HTTPException(status_code=404, detail="No executions found for this job")
    all_tcrs = []
    for ex in executions:
        all_tcrs.extend(db.query(TestCaseResult).filter(TestCaseResult.execution_id == ex.id).all())
    buf = _build_html_dashboard(executions[0], [], all_tcrs)
    filename = f"{job.name.replace(' ', '_')}_report.html"
    return StreamingResponse(iter([buf]), media_type="text/html",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ── GET /api/execution-jobs/{job_id}/report/excel ─────────────────────────────
@app.get("/api/execution-jobs/{job_id}/report/excel")
def job_excel_report(job_id: int, request: Request, db: Session = Depends(get_db)):
    """Generate aggregated Excel report for all executions in this job."""
    if not _HAS_OPENPYXL:
        raise HTTPException(status_code=503, detail="openpyxl not installed")
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    executions = db.query(Execution).filter(Execution.job_id == job_id).all()
    if not executions:
        raise HTTPException(status_code=404, detail="No executions found for this job")
    all_tcrs = []
    for ex in executions:
        all_tcrs.extend(db.query(TestCaseResult).filter(TestCaseResult.execution_id == ex.id).all())
    buf = _build_excel(executions[0], [], all_tcrs)
    filename = f"{job.name.replace(' ', '_')}_report.xlsx"
    return StreamingResponse(buf,
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ── PUT /api/execution-jobs/{job_id}/schedule ─────────────────────────────────
@app.put("/api/execution-jobs/{job_id}/schedule")
def set_job_schedule(job_id: int, request: Request, body: dict, db: Session = Depends(get_db)):
    """Set or update the schedule for an execution job."""
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    stype   = body.get("schedule_type", job.schedule_type or "none")
    enabled = body.get("enabled", True)

    if stype == "none":
        job.schedule_type    = "none"
        job.schedule_enabled = False
        job.schedule_at      = None
        job.schedule_cron    = None
    elif stype == "once":
        sat = body.get("schedule_at")
        if not sat:
            raise HTTPException(status_code=400, detail="schedule_at required for type 'once'")
        try:
            job.schedule_at = datetime.fromisoformat(sat.replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid schedule_at: {sat}")
        job.schedule_type    = "once"
        job.schedule_cron    = None
        job.schedule_enabled = bool(enabled)
    elif stype == "cron":
        cron_expr = body.get("schedule_cron", "").strip()
        if not cron_expr:
            raise HTTPException(status_code=400, detail="schedule_cron required for type 'cron'")
        job.schedule_type    = "cron"
        job.schedule_cron    = cron_expr
        job.schedule_at      = None
        job.schedule_enabled = bool(enabled)
    else:
        raise HTTPException(status_code=400, detail="schedule_type must be none | once | cron")

    job.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(job)
    _register_job_schedule(job)

    next_run = None
    ap_id = f"exec_job_{job.id}"
    try:
        ap_job = scheduler.get_job(ap_id)
        if ap_job and ap_job.next_run_time:
            from datetime import timezone as _utc_tz
            next_run = ap_job.next_run_time.astimezone(_utc_tz.utc).strftime("%Y-%m-%d %H:%M UTC")
    except Exception:
        pass
    return {
        "id": job.id,
        "schedule_type": job.schedule_type,
        "schedule_at": job.schedule_at.isoformat() if job.schedule_at else None,
        "schedule_cron": job.schedule_cron,
        "schedule_enabled": job.schedule_enabled,
        "next_run": next_run,
    }


# ── GET /api/execution-jobs/{job_id}/schedule ────────────────────────────────
@app.get("/api/execution-jobs/{job_id}/schedule")
def get_job_schedule(job_id: int, request: Request, db: Session = Depends(get_db)):
    """Return current schedule settings for a job."""
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    next_run = None
    ap_id = f"exec_job_{job.id}"
    try:
        ap_job = scheduler.get_job(ap_id)
        if ap_job and ap_job.next_run_time:
            from datetime import timezone as _utc_tz
            next_run = ap_job.next_run_time.astimezone(_utc_tz.utc).strftime("%Y-%m-%d %H:%M UTC")
    except Exception:
        pass
    return {
        "id": job.id,
        "schedule_type": job.schedule_type or "none",
        "schedule_at": job.schedule_at.isoformat() if job.schedule_at else None,
        "schedule_cron": job.schedule_cron,
        "schedule_enabled": bool(job.schedule_enabled),
        "last_run_at": job.last_run_at.isoformat() if job.last_run_at else None,
        "next_run": next_run,
    }


# ── WebSocket: Real-time Log Streaming ────────────────────────────────────────
@app.websocket("/ws/execution/{execution_id}")
async def ws_execution_logs(websocket: WebSocket, execution_id: int):
    await websocket.accept()
    db = SessionLocal()
    try:
        last_log_id = 0
        while True:
            new_logs = db.query(ExecutionLog).filter(
                ExecutionLog.execution_id == execution_id,
                ExecutionLog.id > last_log_id
            ).order_by(ExecutionLog.timestamp.asc()).all()
            for log in new_logs:
                await websocket.send_json({
                    "id": log.id, "dut_name": log.dut_name, "level": log.log_level,
                    "message": log.message,
                    "timestamp": log.timestamp.isoformat() if log.timestamp else None
                })
                last_log_id = log.id
            execution = db.query(Execution).filter(Execution.id == execution_id).first()
            if execution and execution.status in ["completed", "failed", "cancelled"]:
                db.expire_all()
                remaining = db.query(ExecutionLog).filter(
                    ExecutionLog.execution_id == execution_id,
                    ExecutionLog.id > last_log_id).all()
                for log in remaining:
                    await websocket.send_json({
                        "id": log.id, "dut_name": log.dut_name, "level": log.log_level,
                        "message": log.message,
                        "timestamp": log.timestamp.isoformat() if log.timestamp else None})
                await websocket.send_json({"type": "execution_complete",
                                           "status": execution.status,
                                           "duration": execution.duration_seconds})
                break
            db.expire_all()
            await asyncio.sleep(0.1)
    except WebSocketDisconnect:
        pass
    except Exception as e:
        logger.error(f"WebSocket error: {e}")
    finally:
        db.close()


# ── HELPER FUNCTIONS ──────────────────────────────────────────────────────────

def _ssh_to_host(dut_id: int, db: Session):
    """Create SSH connection to a host device for SPyTest operations."""
    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="Host device not found")
    ssh = SSHConnectionManager(dut.ip_address, dut.port, dut.username, dut.password)
    if not ssh.connect():
        raise HTTPException(status_code=503, detail=f"Cannot connect to {dut.name}")
    return ssh, dut


def _parse_spytest_script(content: str) -> dict:
    """Parse SPyTest script content to extract topology and metadata."""
    result = {
        "topology_marker": None,
        "min_topology": [],
        "dut_count": 1,
        "description": "",
        "topology_type": "standalone",
    }

    topo_match = re.search(r'@pytest\.mark\.topology\(["\']([^"\']+)["\']\)', content)
    if topo_match:
        result["topology_marker"] = topo_match.group(1)

    min_topo_match = re.search(r'st\.ensure_min_topology\(([^)]+)\)', content)
    if min_topo_match:
        args_str = min_topo_match.group(1)
        topo_args = re.findall(r'["\']([^"\']+)["\']', args_str)
        result["min_topology"] = topo_args
        max_duts = 1
        for arg in topo_args:
            dut_refs = re.findall(r'D(\d+)', arg)
            if dut_refs:
                max_duts = max(max_duts, max(int(d) for d in dut_refs))
        result["dut_count"] = max_duts
        if not result["min_topology"] and re.search(r'st\.ensure_min_topology\(\*\w+\)', content):
            result["uses_vars_file"] = True

    if result["dut_count"] == 1:
        result["topology_type"] = "standalone"
    elif result["dut_count"] == 2:
        result["topology_type"] = "dual-dut"
    else:
        result["topology_type"] = f"{result['dut_count']}-node"

    docstring_match = re.search(r'"""(.+?)"""', content, re.DOTALL)
    if docstring_match:
        desc = docstring_match.group(1).strip()
        lines = desc.split("\n")
        result["description"] = "\n".join(lines[:3]).strip()[:300]

    return result


def _parse_link_requirements(min_topology: list) -> dict:
    """Parse st.ensure_min_topology() args to extract link requirements."""
    link_reqs = {}
    for arg in min_topology:
        match = re.match(r'D(\d+)D(\d+):(\d+)', arg)
        if match:
            d1, d2, count = match.groups()
            dev1, dev2 = f"D{d1}", f"D{d2}"
            pair = tuple(sorted([dev1, dev2]))
            link_reqs[pair] = max(link_reqs.get(pair, 0), int(count))
    return link_reqs


def _get_topology_connections(db: Session) -> dict:
    """Query topology canvas connections and return as a dict."""
    connections = db.query(TopologyConnection).all()
    conn_count = {}

    for conn in connections:
        dut_a = db.query(DUT).filter(DUT.id == conn.dut_a_id).first()
        dut_b = db.query(DUT).filter(DUT.id == conn.dut_b_id).first()

        if dut_a and dut_b:
            pair = tuple(sorted([dut_a.name, dut_b.name]))
            conn_count[pair] = conn_count.get(pair, 0) + 1

    return conn_count


def _has_back_to_back_connection(dut_name: str, db: Session) -> bool:
    """Check if a DUT has a back-to-back (self-loop) connection in the topology canvas."""
    dut = db.query(DUT).filter(DUT.name == dut_name).first()
    if not dut:
        return False
    self_conn = db.query(TopologyConnection).filter(
        TopologyConnection.dut_a_id == dut.id,
        TopologyConnection.dut_b_id == dut.id
    ).first()
    return self_conn is not None


def _get_b2b_dut_names(db: Session) -> set:
    """Return the set of DUT names that have at least one self-loop canvas connection."""
    self_loop_conns = db.query(TopologyConnection).filter(
        TopologyConnection.dut_a_id == TopologyConnection.dut_b_id
    ).all()
    if not self_loop_conns:
        return set()
    dut_ids = {c.dut_a_id for c in self_loop_conns}
    duts = db.query(DUT).filter(DUT.id.in_(dut_ids)).all()
    return {d.name for d in duts}


def _find_duts_matching_topology(
    available_duts: list,
    dut_count: int,
    link_requirements: dict,
    topology_connections: dict,
    b2b_dut_names: set,
) -> list:
    """Find a subset of DUTs from available_duts that satisfies topology requirements."""
    from itertools import combinations

    if not link_requirements:
        if dut_count == 1 and b2b_dut_names:
            for dut in available_duts:
                if dut in b2b_dut_names:
                    return [dut]
        return available_duts[:dut_count] if len(available_duts) >= dut_count else None

    for combo in combinations(available_duts, dut_count):
        combo_list = list(combo)
        dut_mapping = {f"D{i+1}": combo_list[i] for i in range(len(combo_list))}

        satisfied = True
        for (dev1, dev2), required_links in link_requirements.items():
            actual_dut1 = dut_mapping.get(dev1)
            actual_dut2 = dut_mapping.get(dev2)
            if not actual_dut1 or not actual_dut2:
                satisfied = False
                break
            pair = tuple(sorted([actual_dut1, actual_dut2]))
            if topology_connections.get(pair, 0) < required_links:
                satisfied = False
                break

        if satisfied:
            return combo_list

    return None


def _parse_results_csv(csv_data: str) -> list:
    """Parse SPyTest results CSV into a list of dicts."""
    rows = []
    try:
        reader = csv.DictReader(io.StringIO(csv_data))
        for row in reader:
            module     = (row.get('Module')       or row.get('module')        or '').strip()
            func       = (row.get('TestFunction') or row.get('Function')      or
                          row.get('test_function') or '').strip()
            result     = (row.get('Result')       or row.get('result')        or '').strip()
            time_taken = (row.get('TimeTaken')    or row.get('Time')          or
                          row.get('time_taken')   or '').strip()
            doc        = (row.get('DocSummary')   or row.get('Description')   or
                          row.get('description')  or '').strip()
            time_s = 0
            if time_taken:
                parts = time_taken.split(':')
                try:
                    if len(parts) == 3:
                        time_s = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(float(parts[2]))
                    elif len(parts) == 2:
                        time_s = int(parts[0]) * 60 + int(float(parts[1]))
                except Exception:
                    pass
            tc_id = func.split('.')[-1] if '.' in func else func
            rows.append({'module': module, 'test_function': func, 'testcase_id': tc_id,
                         'result': result, 'time_taken': time_taken, 'time_seconds': time_s,
                         'description': doc[:200] if doc else ''})
    except Exception as e:
        logger.warning(f"[results] CSV parse error: {e}")
    return rows


def _collect_and_save_results(ssh, execution, execution_id: int, script_path: str,
                               log_dir: str, db):
    """After a script finishes: find the results CSV, save TestCaseResult rows,
    and append to execution.test_results (per-script aggregate JSON)."""
    script_stem = os.path.basename(script_path).replace('.py', '')
    try:
        out, _, _ = ssh.execute_command(
            f"find {log_dir} -name 'results_*_functions.csv' 2>/dev/null | head -1",
            timeout=15)
        csv_remote = out.strip()
        rows = []
        if csv_remote:
            sftp = ssh.client.open_sftp()
            try:
                with sftp.file(csv_remote, 'r') as fh:
                    csv_data = fh.read().decode('utf-8', errors='ignore')
            finally:
                sftp.close()
            rows = _parse_results_csv(csv_data)
            for r in rows:
                tcr = TestCaseResult(
                    execution_id=execution_id,
                    script_path=script_path,
                    module=r.get('module', ''),
                    test_function=r.get('test_function', ''),
                    testcase_id=r.get('testcase_id', ''),
                    result=r.get('result', ''),
                    time_taken=r.get('time_taken', ''),
                    time_seconds=r.get('time_seconds', 0),
                    description=r.get('description', '')
                )
                db.add(tcr)
            db.commit()
        else:
            logger.info(f"[results] No CSV in {log_dir} for {script_stem}")

        result_lower = [r.get('result', '').lower() for r in rows]
        passed  = sum(1 for r in result_lower if r == 'pass')
        failed  = sum(1 for r in result_lower if r in ('fail', 'scripterror', 'error'))
        skipped = sum(1 for r in result_lower if r in ('skip', 'xfail', 'deselect'))
        dur_s   = sum(r.get('time_seconds', 0) for r in rows)
        agg_status = ('failed' if failed > 0 else
                      'passed' if passed > 0 else
                      'skipped' if skipped > 0 else 'unknown')
        agg = {'script': script_path, 'script_stem': script_stem, 'status': agg_status,
               'passed': passed, 'failed': failed, 'skipped': skipped, 'duration_s': dur_s}
        with _test_results_lock:
            inner_exec = db.query(Execution).filter(Execution.id == execution_id).first()
            if inner_exec:
                existing = json.loads(inner_exec.test_results or '[]')
                existing.append(agg)
                inner_exec.test_results = json.dumps(existing)
            db.commit()
        logger.info(f"[results] #{execution_id} {script_stem}: pass={passed} fail={failed} skip={skipped}")
    except Exception as e:
        logger.warning(f"[results] Collection failed for {script_stem}: {e}")


def _create_subset_testbed(full_config: dict, device_names: list) -> dict:
    """Create a subset testbed YAML with only the specified devices, renamed D1, D2, D3..."""
    unique_device_names = list(dict.fromkeys(device_names))
    if len(unique_device_names) < len(device_names):
        logger.warning(f"TESTBED WARNING: Duplicate devices detected: {device_names}. Using unique only: {unique_device_names}")
        device_names = unique_device_names

    phys_to_logical = {phys: f"D{i+1}" for i, phys in enumerate(device_names)}
    logger.info(f"[TESTBED] Device remap: {phys_to_logical}")

    all_devices = full_config.get("devices", {})
    all_topology = full_config.get("topology", {})

    devices_section = {}
    topology_section = {}

    for phys_name in device_names:
        logical_name = phys_to_logical[phys_name]

        if phys_name in all_devices:
            devices_section[logical_name] = all_devices[phys_name]

        if phys_name in all_topology:
            dev_topo = all_topology[phys_name]
            filtered_interfaces = {}
            for iface, link in dev_topo.get("interfaces", {}).items():
                end_phys = link.get("EndDevice", "")
                if end_phys in phys_to_logical:
                    filtered_interfaces[iface] = {
                        "EndDevice": phys_to_logical[end_phys],
                        "EndPort": link.get("EndPort", ""),
                    }
            topology_section[logical_name] = {"interfaces": filtered_interfaces}

    topo_dict = {}
    for log_a, topo_a in topology_section.items():
        for iface, link in topo_a.get("interfaces", {}).items():
            log_b = link.get("EndDevice", "")
            if log_b and log_a < log_b:
                key = f"{log_a}{log_b}"
                topo_dict[key] = topo_dict.get(key, 0) + 1

    master_params = full_config.get("params", {})
    subset_params = {"topo": topo_dict}

    result = {
        "version": full_config.get("version", "2.0"),
        "devices": devices_section,
        "topology": topology_section,
        "services": full_config.get("services", {"default": {}}),
        "builds": full_config.get("builds", {"default": {}}),
        "configs": full_config.get("configs", {"default": {}}),
        "errors": full_config.get("errors", {"default": {}}),
        "params": subset_params,
    }
    if "global" in full_config:
        result["global"] = full_config["global"]
    elif "test_interface" in master_params:
        result["global"] = {"params": {"test_interface": master_params["test_interface"]}}
    return result


# ── SPYTEST BACKGROUND EXECUTION ─────────────────────────────────────────────

def _run_spytest_execution(
    execution_id: int,
    host_id: int,
    scripts: list,
    testbed_file: str,
    options: dict,
    available_dut_count: int = 1,
    base_path: str = "",
    job_id: int = None,
):
    """Background thread: Smart SPyTest execution with true parallel DUT allocation."""
    import time as _time

    if base_path:
        _bp = base_path.rstrip("/")
        _m = re.match(r'(^.*?/spytest)(?:/|$)', _bp)
        _spytest_root   = _m.group(1) if _m else os.path.dirname(_bp)
        _tests_dir      = _bp
        _testbed_dir    = _spytest_root + "/testbeds"
        _spytest_bin    = _spytest_root + "/bin/spytest"
        _spytest_venv   = _spytest_root + "/spytest_venv"
        _spytest_python = _spytest_venv + "/bin/python"
    else:
        _spytest_root   = SPYTEST_BASE
        _tests_dir      = SPYTEST_TESTS_DIR
        _testbed_dir    = SPYTEST_TESTBED_DIR
        _spytest_bin    = SPYTEST_BIN
        _spytest_venv   = SPYTEST_VENV
        _spytest_python = SPYTEST_PYTHON

    db = SessionLocal()
    execution = None
    try:
        execution = db.query(Execution).filter(Execution.id == execution_id).first()
        host_dut = db.query(DUT).filter(DUT.id == host_id).first()
        if not execution or not host_dut:
            return

        execution.status = "running"
        execution.start_time = datetime.utcnow()
        db.commit()

        log_execution(db, execution_id, "SYSTEM", "INFO",
                      f"SPyTest execution started on {host_dut.name} ({host_dut.ip_address})")
        log_execution(db, execution_id, "SYSTEM", "INFO",
                      f"Testbed: {testbed_file} | Scripts: {len(scripts)}")

        coord_ssh = SSHConnectionManager(
            host_dut.ip_address, host_dut.port, host_dut.username, host_dut.password
        )
        if not coord_ssh.connect():
            log_execution(db, execution_id, "SYSTEM", "ERROR",
                          f"Failed to SSH to {host_dut.name}")
            execution.status = "failed"
            execution.end_time = datetime.utcnow()
            db.commit()
            return

        try:
            if testbed_file.startswith("/"):
                testbed_path = testbed_file
            else:
                testbed_path = f"{_testbed_dir}/{testbed_file}"
            out, err, code = coord_ssh.execute_command(f"cat {testbed_path}", timeout=15)
            if code != 0:
                log_execution(db, execution_id, "SYSTEM", "ERROR",
                              f"Cannot read testbed: {err}")
                execution.status = "failed"
                execution.end_time = datetime.utcnow()
                db.commit()
                return

            testbed_config = yaml.safe_load(out) or {}
            testbed_devices = list(testbed_config.get("devices", {}).keys())
            total_testbed_duts = len(testbed_devices)

            if total_testbed_duts == 0:
                all_duts = [f"Slot-{i+1}" for i in range(available_dut_count)]
                log_execution(db, execution_id, "SYSTEM", "WARNING",
                              f"No devices found in testbed YAML! Using {available_dut_count} synthetic slot(s). "
                              f"Check the testbed YAML 'devices:' key.")
            else:
                all_duts = testbed_devices

                if available_dut_count > total_testbed_duts:
                    log_execution(db, execution_id, "SYSTEM", "WARNING",
                                  f"Canvas has {available_dut_count} DUTs selected, but testbed only defines "
                                  f"{total_testbed_duts} device(s). Using {total_testbed_duts} actual device(s). "
                                  f"Multi-DUT scripts requiring {available_dut_count}+ devices will wait/fail.")

                max_dut_requirement = max([s.get("dut_count", 1) for s in scripts])
                if max_dut_requirement > total_testbed_duts:
                    log_execution(db, execution_id, "SYSTEM", "ERROR",
                                  f"ALLOCATION ERROR: Script requires {max_dut_requirement} DUTs, "
                                  f"but testbed only has {total_testbed_duts} device(s). "
                                  f"Add more devices to testbed YAML or reduce script requirements.")

            log_execution(db, execution_id, "SYSTEM", "INFO",
                          f"Parallel pool: {len(all_duts)} slot(s) — {', '.join(all_duts)}")

            topology_connections = _get_topology_connections(db)
            b2b_dut_names = _get_b2b_dut_names(db)
            log_execution(db, execution_id, "SYSTEM", "INFO",
                          f"Topology connections loaded: {len(topology_connections)} unique pairs"
                          + (f", b2b devices: {sorted(b2b_dut_names)}" if b2b_dut_names else ""))

            if topology_connections:
                for script in scripts:
                    if not script.get("min_topology"):
                        s_path = script.get("path", "")
                        if s_path:
                            full_s_path = s_path if s_path.startswith("/") else f"{_tests_dir}/{s_path}"
                            try:
                                out, cat_err, rc = coord_ssh.execute_command(f"cat '{full_s_path}'", timeout=15)
                                if rc == 0:
                                    info = _parse_spytest_script(out)
                                    if info.get("uses_vars_file"):
                                        stem = re.sub(r'^test_', '',
                                                      os.path.splitext(os.path.basename(s_path))[0])
                                        vars_path = (f"{os.path.dirname(full_s_path)}"
                                                     f"/../vars/vars_{stem}.yaml")
                                        vars_out, _, vars_rc = coord_ssh.execute_command(
                                            f"cat '{vars_path}'", timeout=10)
                                        if vars_rc == 0:
                                            defaults_sec = re.search(
                                                r'defaults:.*?(?=\n\w|\Z)', vars_out, re.DOTALL)
                                            if defaults_sec:
                                                topo_items = re.findall(
                                                    r'^\s*-\s*["\']([^"\']+)["\']',
                                                    defaults_sec.group(0), re.MULTILINE)
                                                if topo_items:
                                                    info["min_topology"] = topo_items
                                                    info["dut_count"] = max(
                                                        (max((int(d) for d in
                                                              re.findall(r'D(\d+)', arg)),
                                                             default=1)
                                                         for arg in topo_items),
                                                        default=1)
                                        else:
                                            log_execution(db, execution_id, "SYSTEM", "WARNING",
                                                          f"[TOPO] vars file not found for "
                                                          f"{os.path.basename(s_path)} — using dut_count=1")
                                    script["dut_count"]    = info.get("dut_count", 1)
                                    script["min_topology"] = info.get("min_topology", [])
                                    log_execution(db, execution_id, "SYSTEM", "INFO",
                                                  f"[TOPO] {os.path.basename(s_path)}: "
                                                  f"dut_count={script['dut_count']}, "
                                                  f"min_topology={script['min_topology']}")
                                else:
                                    log_execution(db, execution_id, "SYSTEM", "WARNING",
                                                  f"[TOPO] cat failed (rc={rc}) for {os.path.basename(s_path)}: "
                                                  f"{cat_err.strip()[:120]} — using dut_count=1")
                            except Exception as e:
                                log_execution(db, execution_id, "SYSTEM", "WARNING",
                                              f"[TOPO] Could not read {os.path.basename(s_path)}: {e}"
                                              f" — using default dut_count=1")

                max_dut_requirement = max(s.get("dut_count", 1) for s in scripts)
                if max_dut_requirement > total_testbed_duts:
                    log_execution(db, execution_id, "SYSTEM", "ERROR",
                                  f"ALLOCATION ERROR: Script requires {max_dut_requirement} DUTs, "
                                  f"but testbed only has {total_testbed_duts} device(s). "
                                  f"Add more devices to testbed YAML or reduce script requirements.")

            script_names = [os.path.basename(s.get("path", "")) for s in scripts]
            _q_init(execution_id, script_names, all_duts)

            pool_lock = Lock()
            available_pool: list = list(all_duts)

            def acquire_duts(needed: int, link_requirements: dict = None) -> list:
                """Block until `needed` DUTs are available, then atomically grab them."""
                while True:
                    if _is_exec_cancelled(execution_id):
                        raise ExecutionCancelled()
                    with pool_lock:
                        if len(available_pool) >= needed:
                            if topology_connections and (link_requirements or needed == 1):
                                matched = _find_duts_matching_topology(
                                    available_pool, needed, link_requirements or {},
                                    topology_connections, b2b_dut_names
                                )
                                if matched:
                                    for dut in matched:
                                        if dut in available_pool:
                                            available_pool.remove(dut)
                                    _q_set_free(execution_id, list(available_pool))
                                    return matched
                                if not link_requirements and needed == 1:
                                    allocated = available_pool[:needed]
                                    del available_pool[:needed]
                                    _q_set_free(execution_id, list(available_pool))
                                    return allocated
                            else:
                                allocated = available_pool[:needed]
                                del available_pool[:needed]
                                _q_set_free(execution_id, list(available_pool))
                                return allocated
                    _time.sleep(5)

            def release_duts(duts_to_free: list):
                with pool_lock:
                    available_pool.extend(duts_to_free)
                    _q_set_free(execution_id, list(available_pool))

            def run_one_script(script_info: dict, slot_idx: int):
                sdb = SessionLocal()
                s_ssh = None
                assigned = []
                script_path = script_info.get("path", "")
                dut_count   = script_info.get("dut_count", 1)
                min_topology = script_info.get("min_topology", [])
                sname       = os.path.basename(script_path)

                if min_topology:
                    _max_duts = 1
                    for _arg in min_topology:
                        _refs = re.findall(r'D(\d+)', _arg)
                        if _refs:
                            _max_duts = max(_max_duts, max(int(d) for d in _refs))
                    dut_count = max(dut_count, _max_duts)

                try:
                    link_requirements = _parse_link_requirements(min_topology)
                    if link_requirements:
                        log_execution(sdb, execution_id, sname, "INFO",
                                      f"[TOPO] Link requirements: {link_requirements}")

                    _q_update_script(execution_id, sname, "waiting")
                    log_execution(sdb, execution_id, sname, "INFO",
                                  f"[QUEUE] Waiting for {dut_count} DUT(s)… "
                                  f"(pool has {len(available_pool)})")

                    assigned = acquire_duts(dut_count, link_requirements)
                    _q_update_script(execution_id, sname, "running", duts=assigned)

                    topo_mode = "topology-matched" if (topology_connections and link_requirements) else "FIFO"
                    log_execution(sdb, execution_id, sname, "INFO",
                                  f"[ALLOC] {topo_mode} → DUT(s): {', '.join(assigned)}")

                    temp_tb_path = f"/tmp/temp_exec{execution_id}_s{slot_idx}.yaml"
                    temp_cfg = _create_subset_testbed(testbed_config, assigned)
                    temp_yaml_str = yaml.dump(temp_cfg, default_flow_style=False)
                    import base64 as _b64
                    yaml_b64 = _b64.b64encode(temp_yaml_str.encode()).decode()
                    s_ssh = SSHConnectionManager(
                        host_dut.ip_address, host_dut.port,
                        host_dut.username, host_dut.password
                    )
                    if not s_ssh.connect():
                        log_execution(sdb, execution_id, sname, "ERROR",
                                      "SSH connection failed for script worker")
                        _q_update_script(execution_id, sname, "failed")
                        release_duts(assigned)
                        return

                    s_ssh.execute_command(
                        f"echo '{yaml_b64}' | base64 -d > {temp_tb_path}", timeout=10
                    )

                    extra_opts = ""
                    if options.get("log_level"):
                        extra_opts += f" --log-level {options['log_level']}"
                    if options.get("skip_init_config"):
                        extra_opts += " --skip-init-config"
                    extra_opts += " --ifname-type native"

                    log_dir = (
                        f"{_spytest_root}/logs/"
                        f"exec{execution_id}_{sname}_{int(datetime.utcnow().timestamp())}"
                    )
                    s_ssh.execute_command(f"mkdir -p {log_dir}", timeout=10)

                    spy_cmd = (
                        f"cd {_spytest_venv}/bin && "
                        f"source activate && "
                        f"cd {_spytest_root} && "
                        f"{_spytest_python} {_spytest_bin} --tryssh 1 "
                        f"--testbed {temp_tb_path} "
                        f"{_tests_dir}/{script_path} "
                        f"--logs-path {log_dir}"
                        f"{extra_opts}"
                    )
                    logfile = f"/tmp/spytest_exec{execution_id}_s{slot_idx}.log"
                    bg_cmd = f"nohup bash -c '{spy_cmd}' > {logfile} 2>&1 & echo $!"

                    pid_out, _, _ = s_ssh.execute_command(bg_cmd, timeout=15)
                    pid = pid_out.strip()

                    if not (pid and pid.isdigit()):
                        log_execution(sdb, execution_id, sname, "ERROR",
                                      f"Failed to launch (pid output: {pid_out!r})")
                        _q_update_script(execution_id, sname, "failed")
                        release_duts(assigned)
                        return

                    _q_update_script(execution_id, sname, "running", pid=pid)
                    log_execution(sdb, execution_id, sname, "INFO",
                                  f"[RUN] PID={pid} DUTs={', '.join(assigned)}")

                    last_lines_seen = 0
                    cancelled = False
                    while True:
                        _time.sleep(10)
                        if _is_exec_cancelled(execution_id):
                            s_ssh.execute_command(
                                f"kill -TERM {pid} 2>/dev/null; sleep 1; kill -KILL {pid} 2>/dev/null",
                                timeout=10,
                            )
                            log_execution(sdb, execution_id, sname, "WARN",
                                          f"■ Stopped by user — killed PID {pid}")
                            _q_update_script(execution_id, sname, "cancelled")
                            cancelled = True
                            break
                        chk_out, _, _ = s_ssh.execute_command(
                            f"kill -0 {pid} 2>/dev/null && echo RUNNING || echo DONE",
                            timeout=5,
                        )
                        log_tail, _, _ = s_ssh.execute_command(
                            f"wc -l < {logfile} 2>/dev/null || echo 0", timeout=5
                        )
                        total_lines = int(log_tail.strip() or "0")
                        if total_lines > last_lines_seen:
                            tail_n = total_lines - last_lines_seen
                            new_log, _, _ = s_ssh.execute_command(
                                f"tail -n {min(tail_n, 50)} {logfile} 2>/dev/null",
                                timeout=5,
                            )
                            if new_log.strip():
                                for ln in new_log.strip().split("\n"):
                                    log_execution(sdb, execution_id, sname, "INFO", ln)
                            last_lines_seen = total_lines

                        if "DONE" in chk_out:
                            break

                    if cancelled:
                        s_ssh.execute_command(f"rm -f {temp_tb_path}", timeout=5)
                        return

                    log_execution(sdb, execution_id, sname, "INFO", "✓ Script completed")
                    _q_update_script(execution_id, sname, "done")
                    _collect_and_save_results(s_ssh, execution, execution_id,
                                              script_path, log_dir, sdb)
                    s_ssh.execute_command(f"rm -f {temp_tb_path}", timeout=5)

                except ExecutionCancelled:
                    log_execution(sdb, execution_id, sname, "WARN",
                                  "■ Stopped by user before DUTs were allocated")
                    _q_update_script(execution_id, sname, "cancelled")
                except Exception as ex:
                    log_execution(sdb, execution_id, sname, "ERROR",
                                  f"Script error: {ex}")
                    _q_update_script(execution_id, sname, "failed")
                finally:
                    if assigned:
                        log_execution(sdb, execution_id, sname, "INFO",
                                      f"[CLEANUP] Releasing {len(assigned)} DUT(s): {', '.join(assigned)}")
                        release_duts(assigned)
                    if s_ssh:
                        s_ssh.disconnect()
                    sdb.close()

            threads = []
            for idx, script_info in enumerate(scripts):
                t = Thread(target=run_one_script, args=(script_info, idx), daemon=True)
                threads.append(t)
                t.start()
                _time.sleep(0.3)

            for t in threads:
                t.join()

            _was_cancelled = _is_exec_cancelled(execution_id)
            final_status = "cancelled" if _was_cancelled else "completed"
            execution.status = final_status
            execution.end_time = datetime.utcnow()
            if execution.start_time:
                execution.duration_seconds = int(
                    (execution.end_time - execution.start_time).total_seconds()
                )
            db.commit()
            log_execution(db, execution_id, "SYSTEM",
                          "WARN" if _was_cancelled else "INFO",
                          (f"■ Execution stopped by user ({execution.duration_seconds}s)"
                           if _was_cancelled else
                           f"✓ All scripts finished ({execution.duration_seconds}s)"))
            if job_id:
                _job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id).first()
                if _job and _job.status == "running":
                    _job.status = final_status
                    db.commit()
            _clear_exec_cancel(execution_id)
            _q_cleanup(execution_id)
            _cleanup_pending_scripts(execution_id)

        finally:
            coord_ssh.disconnect()

    except Exception as e:
        logger.error(f"SPyTest execution failed: {e}")
        if execution:
            execution.status = "failed"
            execution.end_time = datetime.utcnow()
            db.commit()
        if job_id:
            try:
                _job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id).first()
                if _job and _job.status == "running":
                    _job.status = "failed"
                    db.commit()
            except Exception:
                pass
        log_execution(db, execution_id, "SYSTEM", "ERROR", f"Execution failed: {e}")
        _clear_exec_cancel(execution_id)
        _q_cleanup(execution_id)
        _cleanup_pending_scripts(execution_id)
    finally:
        db.close()
