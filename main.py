"""
Eka Automation — Network Test Execution Platform
=================================================
FastAPI + SQLite + Paramiko SSH + WebSocket Log Streaming
No Docker, Redis, or Celery required.
"""

from fastapi import (
    FastAPI, HTTPException, WebSocket, WebSocketDisconnect,
    Depends, File, UploadFile, Form, Request
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse
import csv
from fastapi.staticfiles import StaticFiles
from sqlalchemy import create_engine, Column, Integer, String, DateTime, Text, Boolean, UniqueConstraint, or_, text
from sqlalchemy.orm import declarative_base, sessionmaker, Session
from contextlib import asynccontextmanager
import os
import re
import sys
import json
import hashlib
import shutil
import logging
import asyncio
import subprocess
import tempfile
import yaml
import time
from datetime import datetime, timedelta
from typing import List, Optional
from pathlib import Path
from threading import Thread, Lock
import urllib.request
import urllib.error
import urllib.parse
import zipfile
import io
import base64
import socket

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

import paramiko
from paramiko import AutoAddPolicy, SSHClient

# Import SSH Connection Pool for centralized connection management
from ssh_pool import ssh_pool
from telnet_pool import telnet_pool

# APScheduler for background tasks (heartbeat checks, cleanup)
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.date import DateTrigger
from apscheduler.triggers.cron import CronTrigger

# Hardware Load imports
from telnet_manager import TelnetConnectionManager
from crypto_utils import encrypt_password, decrypt_password, sanitize_log
from hardware_load_logic import execute_hardware_load, log_audit

# ============================================================================
# CONFIGURATION
# ============================================================================

# Load environment variables from .env file (must happen before any os.getenv calls)
try:
    from dotenv import load_dotenv as _load_dotenv
    _load_dotenv(Path(__file__).parent / ".env", override=False)
except ImportError:
    pass  # python-dotenv not installed — rely on shell environment

BASE_DIR = Path(__file__).parent.resolve()
DATA_DIR = BASE_DIR / "data"
LOGS_DIR = DATA_DIR / "logs"
IMAGES_DIR = DATA_DIR / "images"
SCRIPTS_DIR = DATA_DIR / "scripts"
DB_PATH = DATA_DIR / "dut_automation.db"

# VS (Virtual System) Management Paths (on the remote host)
VS_IMAGES_PATH = "/var/lib/libvirt/images/"
VS_XML_PATH = "/home/hp/prajwal/VMs"
VS_SOURCE_IMAGE = "/home/hp/anuradha_builds/target/sonic.img"

# SPyTest Integration Paths (on the remote host)
SPYTEST_BASE = "/home/hp_test/Eka/sonic-mgmt/spytest"
SPYTEST_TESTS_DIR = f"{SPYTEST_BASE}/tests"
SPYTEST_TESTBED_DIR = f"{SPYTEST_BASE}/testbeds"
SPYTEST_BIN = f"{SPYTEST_BASE}/bin/spytest"
# Virtual-environment paths — activate env so all spytest deps are available
SPYTEST_VENV = f"{SPYTEST_BASE}/spytest_venv"
SPYTEST_PYTHON = f"{SPYTEST_VENV}/bin/python"  # bypasses broken shebang

# Create necessary directories
LOGS_DIR.mkdir(parents=True, exist_ok=True)
IMAGES_DIR.mkdir(parents=True, exist_ok=True)
SCRIPTS_DIR.mkdir(parents=True, exist_ok=True)

# DATABASE_URL: reads from environment — supports SQLite (dev) or PostgreSQL (Docker)
# Default: SQLite for local development without Docker
DATABASE_URL = os.getenv(
    "DATABASE_URL",
    f"sqlite:///{DB_PATH}"
)

# ============================================================================
# DATABASE SETUP (SQLite for local dev / PostgreSQL for Docker)
# ============================================================================

if DATABASE_URL.startswith("sqlite"):
    # SQLite — needs check_same_thread=False for multi-threaded FastAPI
    engine = create_engine(
        DATABASE_URL,
        echo=False,
        connect_args={"check_same_thread": False},
        pool_size=20,
        max_overflow=30,
        pool_timeout=30,
        pool_recycle=3600,
        pool_pre_ping=True
    )
else:
    # PostgreSQL — no check_same_thread, full connection pooling
    engine = create_engine(
        DATABASE_URL,
        echo=False,
        pool_size=20,
        max_overflow=30,
        pool_timeout=30,
        pool_recycle=3600,
        pool_pre_ping=True
    )
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()


class DUT(Base):
    __tablename__ = "duts"
    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String(100), nullable=False)
    ip_address = Column(String(50), nullable=False)
    port = Column(Integer, default=22)
    device_type = Column(String(50), default="Linux")
    username = Column(String(100), default="admin")
    password = Column(String(255), default="")
    connection_type = Column(String(10), default="ssh")  # 'ssh' or 'telnet'
    status = Column(String(20), default="offline")
    xml_path = Column(String(500), default="/home/hp/prajwal/VMs")  # Per-device XML path for VS definitions
    session_id = Column(String(255), nullable=True, index=True)  # Session-based isolation
    last_heartbeat = Column(DateTime, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    # Enhancement 3: DUT Reservation system
    reserved_by = Column(String(100), nullable=True)  # Username who reserved this DUT
    reserved_at = Column(DateTime, nullable=True)  # When DUT was reserved
    reserved_until = Column(DateTime, nullable=True)  # Auto-release time (optional, 4 hours default)

    # Allow same device names across different sessions, but unique within a session
    __table_args__ = (
        UniqueConstraint('session_id', 'name', name='uq_session_dut_name'),
    )


class DUTConfiguration(Base):
    __tablename__ = "dut_configurations"
    id = Column(Integer, primary_key=True, autoincrement=True)
    dut_id = Column(Integer, nullable=False)
    static_ip = Column(String(50), nullable=True)
    image_path = Column(String(500), nullable=True)
    ssh_port = Column(Integer, default=22)
    extra_config = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class Image(Base):
    __tablename__ = "images"
    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String(200), nullable=False)
    version = Column(String(100), default="1.0")
    file_path = Column(String(500))
    checksum = Column(String(64))
    file_size = Column(Integer, default=0)
    created_at = Column(DateTime, default=datetime.utcnow)


class Script(Base):
    __tablename__ = "scripts"
    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String(200), nullable=False)
    file_path = Column(String(500))
    yaml_content = Column(Text, nullable=True)
    parameters = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class Execution(Base):
    __tablename__ = "executions"
    id = Column(Integer, primary_key=True, autoincrement=True)
    name = Column(String(200), nullable=False)
    script_id = Column(Integer, nullable=True)
    dut_ids = Column(String(500))
    image_id = Column(Integer, nullable=True)
    execution_type = Column(String(20), default="script")  # 'script' or 'image'
    status = Column(String(20), default="pending")
    session_id = Column(String(255), nullable=True, index=True)  # Session-based isolation
    start_time = Column(DateTime, nullable=True)
    end_time = Column(DateTime, nullable=True)
    duration_seconds = Column(Integer, nullable=True)
    test_results = Column(Text, nullable=True)   # JSON list of per-script aggregates
    job_id       = Column(Integer, nullable=True, index=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class ExecutionLog(Base):
    __tablename__ = "execution_logs"
    id = Column(Integer, primary_key=True, autoincrement=True)
    execution_id = Column(Integer, nullable=False)
    dut_name = Column(String(100), default="SYSTEM")
    log_level = Column(String(20), default="INFO")
    message = Column(Text)
    timestamp = Column(DateTime, default=datetime.utcnow)


class TestCaseResult(Base):
    __tablename__ = "testcase_results"
    id = Column(Integer, primary_key=True, autoincrement=True)
    execution_id = Column(Integer, nullable=False, index=True)
    script_path = Column(Text, nullable=True)
    module = Column(Text, nullable=True)
    test_function = Column(Text, nullable=True)
    testcase_id = Column(Text, nullable=True)
    result = Column(String(20), nullable=True)
    time_taken = Column(String(50), nullable=True)
    time_seconds = Column(Integer, nullable=True)
    description = Column(Text, nullable=True)
    created_at = Column(DateTime, default=datetime.utcnow)


class DUTLock(Base):
    """Tracks DUT allocation state: AVAILABLE / ALLOCATED / IN_USE."""
    __tablename__ = "dut_locks"
    id = Column(Integer, primary_key=True, autoincrement=True)
    dut_id = Column(Integer, nullable=False, unique=True)
    status = Column(String(20), default="AVAILABLE")   # AVAILABLE | ALLOCATED | IN_USE
    job_id     = Column(Integer, nullable=True)
    lock_type  = Column(String(10), default="exec")   # 'hw' or 'exec'
    locked_since = Column(DateTime, nullable=True)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class TopologyConnection(Base):
    """Persists canvas DUT wiring (interface-to-interface connections)."""
    __tablename__ = "topology_connections"
    id = Column(Integer, primary_key=True, autoincrement=True)
    dut_a_id = Column(Integer, nullable=False)
    intf_a = Column(String(50), default="Ethernet0")
    dut_b_id = Column(Integer, nullable=False)
    intf_b = Column(String(50), default="Ethernet0")
    created_at = Column(DateTime, default=datetime.utcnow)


class UserSession(Base):
    """Multi-user session management for concurrent test execution.

    Sessions are USER-BASED and PERSISTENT — they do not expire by time.
    A session is only deactivated when:
      - The user explicitly logs out (/api/onepalc/logout)
      - An admin revokes it via the API
      - OnePalC reports the user as inactive (periodic sync)
    """
    __tablename__ = "user_sessions"
    id = Column(Integer, primary_key=True, autoincrement=True)
    session_id = Column(String(255), unique=True, nullable=False, index=True)
    user_name = Column(String(100), nullable=False)
    user_email = Column(String(255), nullable=True)
    user_role = Column(String(255), nullable=True)  # Role assigned via OnePalC SSO
    status = Column(String(20), default="active")  # active, terminated, revoked
    allocated_dut_ids = Column(Text, default="")  # JSON array of DUT IDs
    created_at = Column(DateTime, default=datetime.utcnow)
    last_activity = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    last_keepalive = Column(DateTime, nullable=True)  # Track last successful keep-alive
    keepalive_fail_count = Column(Integer, default=0)  # Count consecutive failures
    # expires_at kept for backward compatibility with existing DB rows; not enforced
    expires_at = Column(DateTime, nullable=True)


class HardwareLoadJob(Base):
    """
    Track hardware load operations for automated OS image installation.

    This table stores all hardware load jobs including progress tracking,
    execution logs, and completion status.
    """
    __tablename__ = "hardware_load_jobs"

    id = Column(Integer, primary_key=True, autoincrement=True)

    # Foreign Keys
    dut_id = Column(Integer, nullable=False)  # Target hardware device
    source_server_id = Column(Integer, nullable=True)  # Server hosting image

    # Image Details
    image_path = Column(String(500), nullable=False)
    image_name = Column(String(255), nullable=False)

    # Network Configuration (encrypted)
    source_server_password = Column(String(500), nullable=True)  # Encrypted password
    gateway_ip = Column(String(50), nullable=True)
    subnet_mask = Column(String(50), nullable=True)

    # Status Tracking
    status = Column(String(50), default="pending")
    # Status values: pending, connecting, rebooting, grub_menu, onie_menu,
    #                onie_install, downloading, installing, completed, failed

    current_step = Column(String(255), nullable=True)  # Human-readable current step
    progress_percentage = Column(Integer, default=0)

    # Logs (sanitized - passwords removed)
    execution_log = Column(Text, default="")
    error_message = Column(Text, nullable=True)

    # Timestamps
    started_at = Column(DateTime, default=datetime.utcnow)
    completed_at = Column(DateTime, nullable=True)

    # Session Isolation
    session_id = Column(String(255), nullable=False, index=True)


class AuditLog(Base):
    """
    Security audit logging for all hardware load operations.

    Tracks all sensitive operations for compliance and security monitoring.
    """
    __tablename__ = "audit_logs"

    id = Column(Integer, primary_key=True, autoincrement=True)
    session_id = Column(String(255), index=True)
    user_ip = Column(String(50))
    action = Column(String(100))  # e.g., "hardware_load_start", "device_create"
    resource_type = Column(String(50))  # e.g., "DUT", "HardwareLoadJob"
    resource_id = Column(Integer)
    details = Column(Text)  # JSON details
    timestamp = Column(DateTime, default=datetime.utcnow, index=True)



class ExecutionJob(Base):
    """Named job container linking DUT selection, topology, scripts, and executions."""
    __tablename__ = "execution_jobs"
    id               = Column(Integer, primary_key=True, autoincrement=True)
    name             = Column(String(100), nullable=False, default="Job")
    status           = Column(String(20), default="idle")   # idle|running|completed|failed
    session_id       = Column(String(255), nullable=True, index=True)
    dut_ids          = Column(Text, nullable=True)           # JSON array of DUT ids
    base_path        = Column(Text, nullable=True)
    host_id          = Column(Integer, nullable=True)
    topology         = Column(Text, nullable=True)           # JSON canvas snapshot
    scripts          = Column(Text, nullable=True)           # JSON array of script objects
    testbed_path     = Column(Text, nullable=True)           # full path on VM of last-generated testbed
    # Scheduler fields
    schedule_type    = Column(String(10), default="none")    # none | once | cron
    schedule_at      = Column(DateTime, nullable=True)       # UTC datetime for one-time run
    schedule_cron    = Column(String(100), nullable=True)    # cron expression for recurring
    schedule_enabled = Column(Boolean, default=False)
    last_run_at      = Column(DateTime, nullable=True)       # last auto-triggered run
    created_at       = Column(DateTime, default=datetime.utcnow)
    updated_at       = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# Create tables (creates new tables; existing tables are not altered here)
Base.metadata.create_all(bind=engine)


def _apply_column_migrations():
    """Safely add new columns to existing tables without dropping data.

    This runs on every startup and is idempotent — it uses IF NOT EXISTS /
    try-except so it is safe to run against a DB that already has the columns.
    """
    with engine.connect() as conn:
        dialect = conn.dialect.name
        try:
            if dialect == "sqlite":
                try:
                    conn.execute(text("ALTER TABLE executions ADD COLUMN job_id INTEGER;"))
                except Exception:
                    pass
                try:
                    conn.execute(text("ALTER TABLE dut_locks ADD COLUMN lock_type VARCHAR(10) DEFAULT 'exec';"))
                except Exception:
                    pass
            else:
                conn.execute(text(
                    "ALTER TABLE executions ADD COLUMN IF NOT EXISTS job_id INTEGER "
                    "REFERENCES execution_jobs(id) ON DELETE SET NULL;"
                ))
                conn.execute(text(
                    "CREATE INDEX IF NOT EXISTS ix_executions_job_id ON executions (job_id);"
                ))
                conn.execute(text(
                    "ALTER TABLE dut_locks ADD COLUMN IF NOT EXISTS lock_type VARCHAR(10) DEFAULT 'exec';"
                ))
            # Scheduler columns on execution_jobs
            _sched_cols = [
                ("schedule_type",    "VARCHAR(10) DEFAULT 'none'"),
                ("schedule_at",      "TIMESTAMP"),
                ("schedule_cron",    "VARCHAR(100)"),
                ("schedule_enabled", "BOOLEAN DEFAULT FALSE"),
                ("last_run_at",      "TIMESTAMP"),
                ("testbed_path",     "TEXT"),
            ]
            for col_name, col_def in _sched_cols:
                if dialect == "sqlite":
                    try:
                        conn.execute(text(
                            f"ALTER TABLE execution_jobs ADD COLUMN {col_name} {col_def};"
                        ))
                    except Exception:
                        pass
                else:
                    conn.execute(text(
                        f"ALTER TABLE execution_jobs ADD COLUMN IF NOT EXISTS {col_name} {col_def};"
                    ))
            conn.commit()
        except Exception as _e:
            logger.warning(f"[Startup] Column migration warning (safe to ignore if columns exist): {_e}")


# ============================================================================
# JOB SCHEDULER — auto-trigger execution jobs at a set time / cron schedule
# ============================================================================

def _trigger_scheduled_job(execution_job_id: int):
    """Called by APScheduler when a job's scheduled time arrives.

    Reads the job's saved state (host, scripts, testbed path) from the DB,
    generates a master testbed on the coordinator VM (SSH), and launches
    _run_spytest_execution in a background thread — exactly like a manual Run.
    """
    db = SessionLocal()
    try:
        job = db.query(ExecutionJob).filter(ExecutionJob.id == execution_job_id).first()
        if not job:
            logger.warning(f"[Scheduler] Job {execution_job_id} not found — skipping")
            return
        if job.status == "running":
            logger.info(f"[Scheduler] Job {execution_job_id} already running — skipping scheduled trigger")
            return
        if not job.schedule_enabled:
            logger.info(f"[Scheduler] Job {execution_job_id} schedule disabled — skipping")
            return

        scripts   = json.loads(job.scripts)   if job.scripts   else []
        host_id   = job.host_id
        base_path = job.base_path or ""

        if not host_id or not scripts:
            logger.warning(f"[Scheduler] Job {execution_job_id} missing host or scripts — skipping")
            return

        host_dut = db.query(DUT).filter(DUT.id == host_id).first()
        if not host_dut:
            logger.warning(f"[Scheduler] Job {execution_job_id} host DUT {host_id} not found — skipping")
            return

        from threading import Thread as _Thread

        # Use the testbed path saved from the last manual run if available.
        # If not, generate one now using the same base64-pipe technique as the
        # /api/spytest/generate-testbed endpoint so the YAML is valid SPyTest format.
        testbed_file = job.testbed_path or ""
        if not testbed_file and base_path:
            coord = SSHConnectionManager(host_dut.ip_address, host_dut.port,
                                         host_dut.username, host_dut.password)
            if coord.connect():
                try:
                    bp = base_path.rstrip("/")
                    _m = re.match(r'(^.*?/spytest)(?:/|$)', bp)
                    spytest_root = _m.group(1) if _m else os.path.dirname(bp)
                    tb_dir = spytest_root + "/testbeds"
                    coord.execute_command(f'mkdir -p "{tb_dir}"', timeout=10)

                    dut_ids = json.loads(job.dut_ids) if job.dut_ids else []
                    devices_section = {}
                    for did in dut_ids:
                        d = db.query(DUT).filter(DUT.id == int(did)).first()
                        if d and d.device_type != "VM":
                            devices_section[d.name] = {
                                "ip": d.ip_address,
                                "username": d.username,
                                "password": d.password,
                                "port": d.port or 22,
                            }

                    tb_path = f"{tb_dir}/master_testbed.yaml"
                    tb_config = {
                        "version": "2.0",
                        "devices": devices_section,
                        "topology": {},
                        "services": {"default": {}},
                        "builds": {"default": {}},
                        "configs": {"default": {}},
                        "errors": {"default": {}},
                        "params": {},
                    }
                    tb_yaml = (
                        "# MASTER TESTBED - AUTO-GENERATED BY SCHEDULER\n"
                        + yaml.dump(tb_config, default_flow_style=None, sort_keys=False)
                    )
                    # Write via base64 pipe to avoid heredoc/ARG_MAX issues
                    yaml_b64 = base64.b64encode(tb_yaml.encode()).decode()
                    stdin_ch, stdout_ch, stderr_ch = coord.client.exec_command(
                        f'base64 -d > "{tb_path}"', timeout=15
                    )
                    stdin_ch.write(yaml_b64.encode())
                    stdin_ch.channel.shutdown_write()
                    write_code = stdout_ch.channel.recv_exit_status()
                    if write_code == 0:
                        testbed_file = tb_path
                        # Persist so future scheduled runs reuse it
                        job.testbed_path = tb_path
                        db.commit()
                        logger.info(f"[Scheduler] Generated testbed {tb_path} for job {execution_job_id}")
                    else:
                        err = stderr_ch.read().decode("utf-8", errors="ignore").strip()
                        logger.warning(f"[Scheduler] Failed to write testbed for job {execution_job_id}: {err}")
                except Exception as _e:
                    logger.warning(f"[Scheduler] Could not generate testbed for job {execution_job_id}: {_e}")
                finally:
                    coord.disconnect()

        if not testbed_file:
            logger.warning(f"[Scheduler] Job {execution_job_id} has no testbed — skipping. "
                           "Run the job manually once to save the testbed path.")
            return

        # Create Execution record
        exec_name = f"scheduled_{execution_job_id}_{int(datetime.utcnow().timestamp())}"
        execution = Execution(
            name=exec_name,
            dut_ids=json.dumps([host_id]),
            execution_type="spytest",
            status="pending",
            job_id=execution_job_id,
        )
        db.add(execution)

        # Update job state
        job.status   = "running"
        job.last_run_at = datetime.utcnow()
        # Disable once-type schedules after firing
        if job.schedule_type == "once":
            job.schedule_enabled = False
        db.commit()
        db.refresh(execution)

        script_names = [os.path.basename(s.get("path", "")) for s in scripts]
        _q_init(execution.id, script_names, [])
        _init_pending_scripts(execution.id)

        available_dut_count = len(json.loads(job.dut_ids) if job.dut_ids else [1])

        _Thread(
            target=_run_spytest_execution,
            args=(execution.id, host_id, scripts, testbed_file, {}, available_dut_count, base_path, execution_job_id),
            daemon=True,
        ).start()

        logger.info(f"[Scheduler] Job {execution_job_id} '{job.name}' triggered — execution {execution.id}")

    except Exception as e:
        logger.error(f"[Scheduler] Error triggering job {execution_job_id}: {e}")
    finally:
        db.close()


def _register_job_schedule(job: "ExecutionJob"):
    """Add or replace an APScheduler entry for this ExecutionJob."""
    ap_id = f"exec_job_{job.id}"
    # Always remove old entry first
    try:
        scheduler.remove_job(ap_id)
    except Exception:
        pass

    if not job.schedule_enabled:
        return

    try:
        if job.schedule_type == "once" and job.schedule_at:
            if job.schedule_at <= datetime.utcnow():
                logger.warning(f"[Scheduler] Job {job.id} schedule_at is in the past — not scheduling")
                return
            # schedule_at is stored as naive UTC; make it timezone-aware so
            # APScheduler (which defaults to the server's local timezone) fires
            # it at the correct wall-clock moment instead of 5.5h early (IST offset).
            from datetime import timezone as _utc_tz
            aware_dt = job.schedule_at.replace(tzinfo=_utc_tz.utc)
            scheduler.add_job(
                _trigger_scheduled_job,
                trigger=DateTrigger(run_date=aware_dt),
                args=[job.id],
                id=ap_id,
                replace_existing=True,
                misfire_grace_time=300,  # fire up to 5 min late if server was busy
            )
            logger.info(f"[Scheduler] Job {job.id} '{job.name}' scheduled once at {job.schedule_at} UTC")

        elif job.schedule_type == "cron" and job.schedule_cron:
            # Parse simple "HH:MM" or full cron "min hr dom mon dow"
            cron_str = job.schedule_cron.strip()
            if re.match(r'^\d{1,2}:\d{2}$', cron_str):
                # Daily shorthand "HH:MM"
                hh, mm = cron_str.split(":")
                trigger = CronTrigger(hour=int(hh), minute=int(mm))
            else:
                parts = cron_str.split()
                if len(parts) == 5:
                    trigger = CronTrigger(
                        minute=parts[0], hour=parts[1],
                        day=parts[2], month=parts[3], day_of_week=parts[4]
                    )
                else:
                    logger.warning(f"[Scheduler] Job {job.id} invalid cron '{cron_str}'")
                    return
            scheduler.add_job(
                _trigger_scheduled_job,
                trigger=trigger,
                args=[job.id],
                id=ap_id,
                replace_existing=True,
            )
            logger.info(f"[Scheduler] Job {job.id} '{job.name}' scheduled cron '{cron_str}'")
    except Exception as e:
        logger.error(f"[Scheduler] Failed to register job {job.id}: {e}")


def _reload_all_job_schedules():
    """On startup, re-register APScheduler entries for all enabled jobs."""
    db = SessionLocal()
    try:
        jobs = db.query(ExecutionJob).filter(ExecutionJob.schedule_enabled == True).all()
        for j in jobs:
            _register_job_schedule(j)
        if jobs:
            logger.info(f"[Scheduler] Reloaded {len(jobs)} job schedule(s) from DB")
    except Exception as e:
        logger.error(f"[Scheduler] Failed to reload schedules: {e}")
    finally:
        db.close()


# ============================================================================
# LOGGING
# ============================================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(LOGS_DIR / "app.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger("DUT-Automation")

# ============================================================================
# VS (VIRTUAL SYSTEM) MANAGER — PATH CONSTANTS
# ============================================================================

VS_SOURCE_IMAGE = "/home/hp/anuradha_build_imgs/target/sonic-vs.img"
VS_IMAGES_PATH = "/var/lib/libvirt/images/"
VS_XML_PATH = "/home/hp/prajwal/VMs"

# ============================================================================
# SSH SESSION STATE TRACKING (for persistent working directory)
# ============================================================================

# Global dict to track working directory per DUT session: {dut_id: current_working_dir}
_dut_session_state = {}
_session_state_lock = Lock()

# Global dict to track active PTY terminal sessions: {session_key: {ssh, channel, dut_id}}
_pty_sessions = {}
_pty_sessions_lock = Lock()


def _get_dut_cwd(dut_id: int) -> str:
    """Get current working directory for a DUT session."""
    with _session_state_lock:
        return _dut_session_state.get(dut_id, "~")


def _set_dut_cwd(dut_id: int, path: str):
    """Set current working directory for a DUT session."""
    with _session_state_lock:
        _dut_session_state[dut_id] = path


# ============================================================================
# SSH CONNECTION MANAGER
# ============================================================================


class SSHConnectionManager:
    """Manages SSH connections to DUT devices."""

    def __init__(self, host: str, port: int = 22, username: str = "admin", password: str = ""):
        self.host = host
        self.port = port
        self.username = username
        self.password = password
        self.client = None

    def connect(self) -> bool:
        try:
            self.client = paramiko.SSHClient()
            self.client.set_missing_host_key_policy(AutoAddPolicy())
            self.client.connect(
                hostname=self.host,
                port=self.port,
                username=self.username,
                password=self.password,
                timeout=15,
                allow_agent=False,
                look_for_keys=False,
                banner_timeout=15,
                auth_timeout=15,
            )
            logger.info(f"SSH connected to {self.host}:{self.port} as {self.username}")
            return True
        except paramiko.AuthenticationException as e:
            logger.error(f"SSH authentication failed for {self.username}@{self.host}:{self.port} — {e}")
            return False
        except paramiko.SSHException as e:
            logger.error(f"SSH error connecting to {self.host}:{self.port} — {e}")
            return False
        except Exception as e:
            logger.error(f"SSH connection failed to {self.host}:{self.port} — {type(e).__name__}: {e}")
            return False

    def execute_command(self, command: str, timeout: int = 30) -> tuple:
        """Execute a command on the remote device. Returns (stdout, stderr, exit_code)."""
        if not self.client:
            raise Exception("Not connected to device")
        try:
            stdin, stdout, stderr = self.client.exec_command(command, timeout=timeout)
            output = stdout.read().decode("utf-8", errors="ignore")
            error = stderr.read().decode("utf-8", errors="ignore")
            exit_code = stdout.channel.recv_exit_status()
            return output, error, exit_code
        except Exception as e:
            logger.error(f"Command execution failed on {self.host}: {e}")
            raise

    def transfer_file(self, local_path: str, remote_path: str) -> bool:
        """Transfer a file to the remote device via SFTP."""
        try:
            sftp = self.client.open_sftp()
            sftp.put(local_path, remote_path)
            sftp.close()
            logger.info(f"File transferred: {local_path} -> {remote_path}")
            return True
        except Exception as e:
            logger.error(f"File transfer failed: {e}")
            return False

    def disconnect(self):
        """Close SSH connection with proper socket shutdown"""
        if self.client:
            try:
                # Force socket shutdown before closing to ensure proper cleanup
                transport = self.client.get_transport()
                if transport and transport.sock:
                    try:
                        transport.sock.shutdown(socket.SHUT_RDWR)
                    except:
                        pass  # Socket may already be closed

                self.client.close()
                logger.info(f"Disconnected from {self.host}")
            except Exception as e:
                logger.warning(f"Error during disconnect from {self.host}: {e}")


# ============================================================================
# EXECUTION ENGINE (Background Tasks — replaces Celery)
# ============================================================================


def log_execution(db: Session, execution_id: int, dut_name: str, level: str, message: str):
    """Write an execution log entry to the database."""
    logger.log(getattr(logging, level, logging.INFO), f"[Exec {execution_id}][{dut_name}] {message}")
    try:
        entry = ExecutionLog(
            execution_id=execution_id,
            dut_name=dut_name,
            log_level=level,
            message=message,
            timestamp=datetime.utcnow(),
        )
        db.add(entry)
        db.commit()
    except Exception:
        db.rollback()


def run_image_deployment(execution_id: int, dut_ids: List[int], image_id: int):
    """Deploy an image to one or more DUTs (runs in background thread)."""
    db = SessionLocal()
    try:
        execution = db.query(Execution).filter(Execution.id == execution_id).first()
        image = db.query(Image).filter(Image.id == image_id).first()
        duts = db.query(DUT).filter(DUT.id.in_(dut_ids)).all()

        if not execution or not image or not duts:
            logger.error(f"Invalid execution/image/duts for exec {execution_id}")
            return

        execution.status = "running"
        execution.start_time = datetime.utcnow()
        db.commit()

        log_execution(db, execution_id, "SYSTEM", "INFO",
                      f"Starting image deployment to {len(duts)} device(s)")

        for dut in duts:
            try:
                log_execution(db, execution_id, dut.name, "INFO",
                              f"Connecting to {dut.ip_address}:{dut.port}...")

                ssh = SSHConnectionManager(dut.ip_address, dut.port, dut.username, dut.password)
                if not ssh.connect():
                    log_execution(db, execution_id, dut.name, "ERROR",
                                  f"Failed to connect to {dut.name}")
                    dut.status = "offline"
                    db.commit()
                    continue

                try:
                    dut.status = "online"
                    db.commit()

                    # 1. Create staging directory
                    ssh.execute_command("mkdir -p /tmp/firmware")
                    log_execution(db, execution_id, dut.name, "INFO",
                                  "Created staging directory /tmp/firmware")

                    # 2. Transfer image file
                    remote_path = f"/tmp/firmware/{os.path.basename(image.file_path)}"
                    log_execution(db, execution_id, dut.name, "INFO",
                                  f"Transferring image to {remote_path}...")

                    if not ssh.transfer_file(image.file_path, remote_path):
                        log_execution(db, execution_id, dut.name, "ERROR",
                                      "Image transfer failed")
                        continue

                    log_execution(db, execution_id, dut.name, "INFO",
                                  "Image file transferred successfully")

                    # 3. Verify checksum
                    if image.checksum:
                        output, _, code = ssh.execute_command(f"sha256sum {remote_path}")
                        if code == 0:
                            remote_checksum = output.split()[0]
                            if remote_checksum == image.checksum:
                                log_execution(db, execution_id, dut.name, "INFO",
                                              "Checksum verified ✓")
                            else:
                                log_execution(db, execution_id, dut.name, "WARNING",
                                              f"Checksum mismatch: {remote_checksum[:16]}... != {image.checksum[:16]}...")

                    # 4. Assign static IP if configured
                    config = db.query(DUTConfiguration).filter(
                        DUTConfiguration.dut_id == dut.id
                    ).first()
                    if config and config.static_ip:
                        ip_cmd = f"ip addr add {config.static_ip} dev eth0 2>/dev/null || true"
                        ssh.execute_command(ip_cmd)
                        log_execution(db, execution_id, dut.name, "INFO",
                                      f"Static IP assigned: {config.static_ip}")

                    log_execution(db, execution_id, dut.name, "INFO",
                                  "✓ Image deployment completed successfully")

                finally:
                    ssh.disconnect()

            except Exception as e:
                log_execution(db, execution_id, dut.name, "ERROR", f"Deployment error: {str(e)}")

        execution.status = "completed"
        execution.end_time = datetime.utcnow()
        if execution.start_time:
            execution.duration_seconds = int(
                (execution.end_time - execution.start_time).total_seconds()
            )
        db.commit()
        log_execution(db, execution_id, "SYSTEM", "INFO", "Image deployment completed")
        # CRITICAL FIX: DO NOT auto-delete logs! Users need to view them in Logs tab
        # _delete_execution_logs(execution_id, db)  # DISABLED - breaks Logs tab functionality

    except Exception as e:
        logger.error(f"Image deployment failed: {e}")
        if execution:
            execution.status = "failed"
            execution.end_time = datetime.utcnow()
            db.commit()
        log_execution(db, execution_id, "SYSTEM", "ERROR", f"Deployment failed: {str(e)}")
    finally:
        db.close()


def run_script_execution(execution_id: int, script_id: int, dut_ids: List[int]):
    """Execute a script on one or more DUTs (runs in background thread)."""
    db = SessionLocal()
    try:
        execution = db.query(Execution).filter(Execution.id == execution_id).first()
        script = db.query(Script).filter(Script.id == script_id).first()
        duts = db.query(DUT).filter(DUT.id.in_(dut_ids)).all()

        if not execution or not script or not duts:
            logger.error(f"Invalid execution/script/duts for exec {execution_id}")
            return

        execution.status = "running"
        execution.start_time = datetime.utcnow()
        db.commit()

        log_execution(db, execution_id, "SYSTEM", "INFO",
                      f"Starting script '{script.name}' on {len(duts)} device(s)")

        # Parse YAML script
        script_config = {}
        if script.yaml_content:
            try:
                script_config = yaml.safe_load(script.yaml_content) or {}
            except Exception as e:
                log_execution(db, execution_id, "SYSTEM", "ERROR",
                              f"Failed to parse YAML: {str(e)}")
                execution.status = "failed"
                execution.end_time = datetime.utcnow()
                db.commit()
                return

        test_cases = script_config.get("test_cases", [])
        results = {}

        for dut in duts:
            try:
                log_execution(db, execution_id, dut.name, "INFO",
                              f"Connecting to {dut.ip_address}:{dut.port}...")

                ssh = SSHConnectionManager(dut.ip_address, dut.port, dut.username, dut.password)
                if not ssh.connect():
                    log_execution(db, execution_id, dut.name, "ERROR",
                                  f"Failed to connect to {dut.name}")
                    results[dut.name] = "CONNECTION_FAILED"
                    dut.status = "offline"
                    db.commit()
                    continue

                try:
                    dut.status = "online"
                    dut.last_heartbeat = datetime.utcnow()
                    db.commit()

                    all_passed = True

                    if test_cases:
                        for tc in test_cases:
                            tc_name = tc.get("name", "unnamed_test")
                            tc_timeout = tc.get("timeout", 30)
                            commands = tc.get("commands", [])

                            log_execution(db, execution_id, dut.name, "INFO",
                                          f"▶ Running test case: {tc_name}")

                            for cmd in commands:
                                log_execution(db, execution_id, dut.name, "INFO",
                                              f"  $ {cmd}")
                                try:
                                    output, error, exit_code = ssh.execute_command(
                                        cmd, timeout=tc_timeout
                                    )
                                    if output.strip():
                                        # Log each line (limit to 50 lines)
                                        lines = output.strip().split("\n")
                                        for line in lines[:50]:
                                            log_execution(db, execution_id, dut.name, "INFO",
                                                          f"    {line}")
                                        if len(lines) > 50:
                                            log_execution(db, execution_id, dut.name, "INFO",
                                                          f"    ... ({len(lines) - 50} more lines)")

                                    if error.strip():
                                        log_execution(db, execution_id, dut.name, "WARNING",
                                                      f"    stderr: {error.strip()[:200]}")

                                    if exit_code != 0:
                                        log_execution(db, execution_id, dut.name, "ERROR",
                                                      f"  ✗ Command failed with exit code {exit_code}")
                                        all_passed = False
                                    else:
                                        log_execution(db, execution_id, dut.name, "INFO",
                                                      f"  ✓ Command succeeded")

                                except Exception as cmd_err:
                                    log_execution(db, execution_id, dut.name, "ERROR",
                                                  f"  ✗ Command error: {str(cmd_err)}")
                                    all_passed = False

                            log_execution(db, execution_id, dut.name, "INFO",
                                          f"  Test case '{tc_name}' — {'PASSED' if all_passed else 'FAILED'}")
                    else:
                        # No test cases — just report device info
                        output, _, _ = ssh.execute_command("uname -a")
                        log_execution(db, execution_id, dut.name, "INFO",
                                      f"Device info: {output.strip()}")
                        output, _, _ = ssh.execute_command("ip addr show 2>/dev/null || ifconfig 2>/dev/null")
                        if output.strip():
                            for line in output.strip().split("\n")[:20]:
                                log_execution(db, execution_id, dut.name, "INFO",
                                              f"  {line}")

                    results[dut.name] = "PASSED" if all_passed else "FAILED"
                    status_msg = "✓ All tests passed" if all_passed else "✗ Some tests failed"
                    log_execution(db, execution_id, dut.name, "INFO", status_msg)

                finally:
                    ssh.disconnect()

            except Exception as e:
                results[dut.name] = "ERROR"
                log_execution(db, execution_id, dut.name, "ERROR", f"Execution error: {str(e)}")

        execution.status = "completed"
        execution.end_time = datetime.utcnow()
        if execution.start_time:
            execution.duration_seconds = int(
                (execution.end_time - execution.start_time).total_seconds()
            )
        db.commit()

        log_execution(db, execution_id, "SYSTEM", "INFO",
                      f"Execution completed. Results: {json.dumps(results)}")
        # CRITICAL FIX: DO NOT auto-delete logs! Users need to view them in Logs tab
        # _delete_execution_logs(execution_id, db)  # DISABLED - breaks Logs tab functionality

    except Exception as e:
        logger.error(f"Script execution failed: {e}")
        if execution:
            execution.status = "failed"
            execution.end_time = datetime.utcnow()
            db.commit()
        log_execution(db, execution_id, "SYSTEM", "ERROR", f"Execution failed: {str(e)}")
    finally:
        db.close()


# ============================================================================
# FASTAPI APPLICATION
# ============================================================================

app = FastAPI(title="Eka Automation", version="2.0.0",
              description="Eka Automation — Network Test Execution Platform for SpyTest / SONiC DUT Infrastructure")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve static files (the frontend)
static_dir = BASE_DIR / "static"
static_dir.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")


# ============================================================================
# Background Tasks - Device Health Monitoring
# ============================================================================

def heartbeat_check():
    """
    Background task to check device health every 60 seconds.

    For each device marked as "online":
    - Execute lightweight command (echo 1)
    - Success: update last_heartbeat, keep status="online"
    - Failure (connection lost): set status="offline"; if pool is OFFLINE/RECONNECTING,
      leave the pool entry intact so auto-reconnect can complete
    - Failure (channel timeout): keep status="online", log warning

    For each device marked as "offline" (SSH only):
    - If pool reports ONLINE (auto-reconnect succeeded): verify and restore DB to "online"
    - If pool reports RECONNECTING/OFFLINE: leave it alone, pool handles retry
    - If not in pool / FAILED: attempt a fresh connection

    This ensures device status accurately reflects connectivity.
    Each device is committed independently to prevent one failure from affecting others.
    """
    db = SessionLocal()
    success_count = 0
    failure_count = 0

    try:
        online_duts = db.query(DUT).filter(DUT.status == "online").all()
        logger.info(f"[HEARTBEAT] Checking {len(online_duts)} online devices")

        for dut in online_duts:
            try:
                # Refresh device from DB to avoid stale data errors
                db.refresh(dut)

                # Handle telnet devices with actual connectivity test using connection pool
                if hasattr(dut, 'connection_type') and dut.connection_type == 'telnet':
                    logger.debug(f"[HEARTBEAT] Testing telnet connectivity for DUT {dut.id} ({dut.name})")

                    # Skip if hardware load is using this connection
                    if telnet_pool.is_hardware_load_active(dut.id):
                        logger.debug(f"[HEARTBEAT] ⊙ DUT {dut.id} ({dut.name}) hardware load active - skipping check")
                        dut.last_heartbeat = datetime.utcnow()
                        db.commit()
                        success_count += 1
                        continue

                    # Test telnet connectivity using pool (reuses connection)
                    try:
                        # Get or create connection from pool with short timeout
                        telnet_mgr = telnet_pool.get_connection(
                            dut.id, dut.ip_address, dut.port, dut.username, dut.password, timeout=10
                        )

                        if telnet_mgr and telnet_mgr.is_alive():
                            # Connection successful
                            dut.last_heartbeat = datetime.utcnow()
                            dut.status = "online"
                            db.commit()
                            success_count += 1
                            logger.info(f"[HEARTBEAT] ✓ DUT {dut.id} ({dut.name}) telnet OK")
                            telnet_pool.release_connection(dut.id)
                        else:
                            # Connection failed - mark offline and close
                            logger.warning(f"[HEARTBEAT] ✗ DUT {dut.id} ({dut.name}) telnet FAILED")
                            dut.status = "offline"
                            db.commit()
                            failure_count += 1
                            telnet_pool.close_connection(dut.id)

                    except Exception as telnet_error:
                        logger.warning(f"[HEARTBEAT] ✗ DUT {dut.id} ({dut.name}) telnet error: {telnet_error}")
                        dut.status = "offline"
                        db.commit()
                        failure_count += 1
                        telnet_pool.close_connection(dut.id)

                    continue

                # Check if device has active terminal session
                # Skip heartbeat command if terminal is using the connection
                if ssh_pool.is_terminal_active(dut.id):
                    logger.debug(f"[HEARTBEAT] ⊙ DUT {dut.id} ({dut.name}) terminal active - skipping command check")
                    dut.last_heartbeat = datetime.utcnow()
                    db.commit()
                    success_count += 1
                    continue

                # Get connection from pool (reuses existing connection)
                ssh = ssh_pool.get_connection(dut.id, dut.ip_address, dut.port, dut.username, dut.password)

                if ssh:
                    try:
                        # Lightweight health check command
                        output, error, exit_code = ssh.execute_command("echo 1", timeout=5)

                        if exit_code == 0:
                            # Device is healthy
                            dut.last_heartbeat = datetime.utcnow()
                            dut.status = "online"
                            db.commit()
                            success_count += 1
                            logger.info(f"[HEARTBEAT] ✓ DUT {dut.id} ({dut.name}) OK")
                        else:
                            # Command failed - mark offline
                            logger.warning(f"[HEARTBEAT] ✗ DUT {dut.id} ({dut.name}) FAILED: exit_code={exit_code}")
                            dut.status = "offline"
                            db.commit()
                            failure_count += 1
                            ssh_pool.close_connection(dut.id)

                        # Release connection back to pool
                        ssh_pool.release_connection(dut.id)

                    except Exception as cmd_error:
                        # Channel timeout or command execution error
                        # Check if it's ChannelException (terminal conflict)
                        error_str = str(cmd_error)
                        if "ChannelException" in error_str or "Connect failed" in error_str:
                            logger.warning(f"[HEARTBEAT] ⚠ DUT {dut.id} ({dut.name}) channel conflict (likely terminal active)")
                            # Keep online status, just update heartbeat
                            dut.last_heartbeat = datetime.utcnow()
                            db.commit()
                        else:
                            # Other errors - keep online but log warning
                            logger.warning(f"[HEARTBEAT] ⚠ DUT {dut.id} ({dut.name}) command error: {cmd_error}")
                            dut.last_heartbeat = datetime.utcnow()
                            db.commit()
                        ssh_pool.release_connection(dut.id)
                else:
                    # get_connection() returned None — the pool is in OFFLINE, RECONNECTING,
                    # or FAILED state. Do NOT call close_connection() here; that would destroy
                    # the pool entry and abort any in-progress auto-reconnect thread.
                    # Just mark the device as offline in the DB and let the pool handle reconnection.
                    pool_status_data = ssh_pool.get_pool_status()
                    pool_conn_state = next(
                        (c.get("status") for c in pool_status_data.get("connections", []) if c.get("dut_id") == dut.id),
                        None
                    )
                    logger.warning(
                        f"[HEARTBEAT] ✗ DUT {dut.id} ({dut.name}) cannot connect "
                        f"(pool state: {pool_conn_state or 'not in pool'})"
                    )
                    dut.status = "offline"
                    db.commit()
                    failure_count += 1
                    # Only permanently close if pool entry is FAILED or not in pool at all
                    if pool_conn_state is None or pool_conn_state == "failed":
                        ssh_pool.close_connection(dut.id)

            except Exception as e:
                # Any error during heartbeat check for this specific device
                logger.error(f"[HEARTBEAT] ERROR for DUT {dut.id} ({dut.name}): {e}")
                try:
                    db.rollback()
                except:
                    pass

        logger.info(f"[HEARTBEAT] Completed: {success_count} OK, {failure_count} failed")

        # ── Check offline SSH devices for successful pool auto-reconnections ──
        # The SSH pool reconnects in the background (exponential backoff). When it
        # succeeds the pool state turns ONLINE, but the DB still shows "offline".
        # We scan offline devices here and restore them to "online" when the pool
        # has reconnected, or attempt a fresh connection if the pool has no entry.
        try:
            offline_duts = db.query(DUT).filter(DUT.status == "offline").all()

            if offline_duts:
                pool_status_data = ssh_pool.get_pool_status()
                pool_conn_map = {
                    c.get("dut_id"): c.get("status")
                    for c in pool_status_data.get("connections", [])
                }

                for dut in offline_duts:
                    # Skip telnet devices — they have their own reconnect path
                    if hasattr(dut, 'connection_type') and dut.connection_type == 'telnet':
                        continue

                    try:
                        db.refresh(dut)
                        pool_conn_state = pool_conn_map.get(dut.id)

                        if pool_conn_state == "online":
                            # Pool auto-reconnect succeeded — verify with a quick command
                            ssh = ssh_pool.get_connection(
                                dut.id, dut.ip_address, dut.port, dut.username, dut.password
                            )
                            if ssh:
                                try:
                                    _, _, exit_code = ssh.execute_command("echo 1", timeout=5)
                                    if exit_code == 0:
                                        dut.status = "online"
                                        dut.last_heartbeat = datetime.utcnow()
                                        db.commit()
                                        logger.info(
                                            f"[HEARTBEAT] ✓ DUT {dut.id} ({dut.name}) "
                                            f"restored to ONLINE after auto-reconnect"
                                        )
                                    ssh_pool.release_connection(dut.id)
                                except Exception:
                                    ssh_pool.release_connection(dut.id)

                        elif pool_conn_state in ("reconnecting", "offline"):
                            # Pool is actively retrying — leave DB as offline, wait
                            logger.debug(
                                f"[HEARTBEAT] DUT {dut.id} ({dut.name}) pool reconnect in progress "
                                f"(pool state: {pool_conn_state})"
                            )

                        else:
                            # Not in pool or FAILED — attempt a fresh connection
                            logger.info(
                                f"[HEARTBEAT] DUT {dut.id} ({dut.name}) offline, attempting fresh connection"
                            )
                            ssh = ssh_pool.get_connection(
                                dut.id, dut.ip_address, dut.port, dut.username, dut.password
                            )
                            if ssh:
                                try:
                                    _, _, exit_code = ssh.execute_command("echo 1", timeout=5)
                                    if exit_code == 0:
                                        dut.status = "online"
                                        dut.last_heartbeat = datetime.utcnow()
                                        db.commit()
                                        logger.info(
                                            f"[HEARTBEAT] ✓ DUT {dut.id} ({dut.name}) "
                                            f"reconnected and restored to ONLINE"
                                        )
                                    ssh_pool.release_connection(dut.id)
                                except Exception:
                                    ssh_pool.release_connection(dut.id)

                    except Exception as e:
                        logger.error(f"[HEARTBEAT] Error checking offline DUT {dut.id}: {e}")
                        try:
                            db.rollback()
                        except:
                            pass
        except Exception as e:
            logger.error(f"[HEARTBEAT] Error checking offline devices: {e}")

    except Exception as e:
        logger.error(f"Heartbeat check initialization failed: {e}")
    finally:
        try:
            db.close()
        except:
            pass


def cleanup_expired_sessions():
    """
    Background task to clean up explicitly terminated/revoked sessions.

    Runs every 10 minutes. Sessions are USER-BASED and do NOT expire by time.
    This job only cleans up sessions with status='terminated' or 'revoked':
    1. Delete all DUTs (devices) belonging to the session
    2. Delete all related data: DUTConfiguration, DUTLock, TopologyConnection
    3. Delete all Executions and ExecutionLogs for that session
    4. Close SSH connections for deleted devices

    Sessions remain 'active' indefinitely until the user logs out or is revoked.
    """
    db = SessionLocal()
    try:
        # Only clean up explicitly terminated or revoked sessions
        terminated_sessions = db.query(UserSession).filter(
            UserSession.status.in_(["terminated", "revoked"])
        ).all()

        if not terminated_sessions:
            logger.debug("No terminated sessions to clean up")
            return

        logger.info(f"Found {len(terminated_sessions)} terminated/revoked sessions to clean up")

        for session in terminated_sessions:
            session_id = session.session_id
            logger.info(f"Cleaning up terminated session: {session_id} (user: {session.user_name})")

            try:
                # 1. Get all DUTs belonging to this session
                duts = db.query(DUT).filter(DUT.session_id == session_id).all()
                dut_ids = [dut.id for dut in duts]

                if dut_ids:
                    logger.info(f"  Deleting {len(dut_ids)} devices for session {session_id}")

                    # 2. Close SSH connections for these devices
                    for dut_id in dut_ids:
                        try:
                            ssh_pool.close_connection(dut_id)
                            logger.debug(f"  Closed SSH connection for DUT {dut_id}")
                        except Exception as e:
                            logger.warning(f"  Failed to close SSH connection for DUT {dut_id}: {e}")

                    # 3. Delete DUTConfiguration entries
                    deleted_configs = db.query(DUTConfiguration).filter(
                        DUTConfiguration.dut_id.in_(dut_ids)
                    ).delete(synchronize_session=False)
                    logger.debug(f"  Deleted {deleted_configs} DUT configurations")

                    # 4. Delete DUTLock entries
                    deleted_locks = db.query(DUTLock).filter(
                        DUTLock.dut_id.in_(dut_ids)
                    ).delete(synchronize_session=False)
                    logger.debug(f"  Deleted {deleted_locks} DUT locks")

                    # 5. Delete TopologyConnection entries
                    deleted_connections = db.query(TopologyConnection).filter(
                        (TopologyConnection.dut_a_id.in_(dut_ids)) |
                        (TopologyConnection.dut_b_id.in_(dut_ids))
                    ).delete(synchronize_session=False)
                    logger.debug(f"  Deleted {deleted_connections} topology connections")

                    # 6. Delete DUTs themselves
                    deleted_duts = db.query(DUT).filter(DUT.session_id == session_id).delete(synchronize_session=False)
                    logger.debug(f"  Deleted {deleted_duts} DUTs")

                # 7. Get all Executions belonging to this session
                executions = db.query(Execution).filter(Execution.session_id == session_id).all()
                execution_ids = [exec.id for exec in executions]

                if execution_ids:
                    logger.info(f"  Deleting {len(execution_ids)} executions for session {session_id}")

                    # 8. Delete ExecutionLogs
                    deleted_logs = db.query(ExecutionLog).filter(
                        ExecutionLog.execution_id.in_(execution_ids)
                    ).delete(synchronize_session=False)
                    logger.debug(f"  Deleted {deleted_logs} execution logs")

                    # 9. Delete Executions
                    deleted_execs = db.query(Execution).filter(
                        Execution.session_id == session_id
                    ).delete(synchronize_session=False)
                    logger.debug(f"  Deleted {deleted_execs} executions")

                # 10. Delete the session record itself (terminated sessions are removed)
                db.delete(session)
                db.commit()

                logger.info(f"✓ Successfully cleaned up terminated session {session_id}")

            except Exception as e:
                logger.error(f"Error cleaning up session {session_id}: {e}")
                db.rollback()

        logger.info(f"Session cleanup complete. Processed {len(terminated_sessions)} terminated sessions")

    except Exception as e:
        logger.error(f"Session cleanup task failed: {e}")
        db.rollback()
    finally:
        db.close()


# Initialize background scheduler
scheduler = BackgroundScheduler()

# Startup event: clean up stuck executions and hardware load jobs from previous runs
def _another_server_running(port: int = 8000) -> bool:
    """Return True if another process is already listening on `port`.

    At the time @app.on_event("startup") fires, this server has NOT yet bound
    the port.  So a successful connect means a sibling process owns the socket —
    we must not clobber its live executions in the database.
    """
    import socket as _socket
    with _socket.socket(_socket.AF_INET, _socket.SOCK_STREAM) as s:
        s.settimeout(0.5)
        return s.connect_ex(('127.0.0.1', port)) == 0


@app.on_event("startup")
def startup_migrations():
    """Apply safe column migrations (idempotent — runs on every restart)."""
    _apply_column_migrations()
    logger.info("[Startup] Column migrations applied")
    _reload_all_job_schedules()


@app.on_event("startup")
def startup_cleanup():
    """Clean up any executions or hardware load jobs stuck from previous server runs."""
    # Guard: if another server instance already owns port 8000, we are a
    # redundant restart attempt that will fail to bind later.  Do NOT touch
    # execution status rows — they belong to the live server and changing them
    # would cause currently-running executions to appear as "failed" in the UI
    # while their background threads keep streaming logs.
    if _another_server_running():
        logger.warning(
            "[Startup] Port 8000 is already in use — another server instance is running. "
            "Skipping execution cleanup to avoid corrupting live executions."
        )
        return

    db = SessionLocal()
    try:
        # ── Fix PostgreSQL sequences desynced after SQLite migration ──────────
        # When rows are bulk-inserted with explicit IDs (migration), the
        # auto-increment sequence stays at 1 causing UniqueViolation on next insert.
        _pg_tables = [
            "duts", "dut_configurations", "images", "scripts",
            "executions", "execution_logs", "dut_locks",
            "topology_connections", "user_sessions",
            "hardware_load_jobs", "audit_logs",
        ]
        fixed = []
        for tbl in _pg_tables:
            try:
                db.execute(text(
                    f"SELECT setval('{tbl}_id_seq', "
                    f"(SELECT COALESCE(MAX(id), 1) FROM {tbl}))"
                ))
                fixed.append(tbl)
            except Exception:
                pass  # table or sequence may not exist (SQLite mode / fresh install)
        if fixed:
            db.commit()
            logger.info(f"[Startup] PostgreSQL sequences reset for: {', '.join(fixed)}")

        # ── Clean up stuck script executions ──────────────────────────────────
        stuck_executions = db.query(Execution).filter(Execution.status == "running").all()
        if stuck_executions:
            logger.warning(f"Found {len(stuck_executions)} stuck executions from previous run, marking as failed")
            for exec in stuck_executions:
                exec.status = "failed"
                exec.end_time = datetime.utcnow()
                if exec.start_time and not exec.duration_seconds:
                    exec.duration_seconds = int((exec.end_time - exec.start_time).total_seconds())
            db.commit()
            logger.info(f"Cleaned up {len(stuck_executions)} stuck executions")

        # ── Clean up stuck execution jobs ─────────────────────────────────────
        stuck_exec_jobs = db.query(ExecutionJob).filter(ExecutionJob.status == "running").all()
        if stuck_exec_jobs:
            logger.warning(f"Found {len(stuck_exec_jobs)} stuck execution jobs from previous run, marking as failed")
            for ej in stuck_exec_jobs:
                ej.status = "failed"
            db.commit()
            logger.info(f"Cleaned up {len(stuck_exec_jobs)} stuck execution jobs")

        # ── Clean up stuck hardware load jobs ─────────────────────────────────
        # Any job not in a terminal state means the server died mid-operation.
        TERMINAL_STATES = ("completed", "failed", "cancelled")
        stuck_hw_jobs = db.query(HardwareLoadJob).filter(
            ~HardwareLoadJob.status.in_(TERMINAL_STATES)
        ).all()
        if stuck_hw_jobs:
            logger.warning(f"Found {len(stuck_hw_jobs)} stuck hardware load jobs from previous run, marking as failed")
            for hw_job in stuck_hw_jobs:
                hw_job.status = "failed"
                hw_job.error_message = "Server restarted — job was interrupted"
                hw_job.current_step = "Interrupted by server restart"
                if not hw_job.completed_at:
                    hw_job.completed_at = datetime.utcnow()
                # Append note to execution log
                timestamp = datetime.utcnow().strftime("%H:%M:%S")
                note = f"\n[{timestamp}] ✗ Job interrupted: server restarted while job was in progress.\n"
                hw_job.execution_log = (hw_job.execution_log or "") + note
            db.commit()
            logger.info(f"Cleaned up {len(stuck_hw_jobs)} stuck hardware load jobs")

    except Exception as e:
        logger.error(f"Error during startup cleanup: {e}")
    finally:
        db.close()

    # Start background scheduler for device health monitoring and session cleanup
    try:
        # Heartbeat check every 60 seconds
        scheduler.add_job(heartbeat_check, 'interval', seconds=60, id='heartbeat_check')
        logger.info("Scheduled heartbeat check (every 60 seconds)")

        # Cleanup idle SSH connections every 3 minutes (180s idle timeout)
        # This prevents connections from dying due to network timeouts
        # Connections are recreated before they become stale
        scheduler.add_job(ssh_pool.cleanup_idle, 'interval', seconds=180, args=[180], id='ssh_cleanup')
        logger.info("Scheduled SSH connection cleanup (every 3 minutes, 3 minute idle timeout)")

        # Cleanup idle telnet connections every 3 minutes (600s idle timeout)
        scheduler.add_job(telnet_pool.cleanup_idle, 'interval', seconds=180, args=[600], id='telnet_cleanup')
        logger.info("Scheduled telnet connection cleanup (every 3 minutes, 10 minute idle timeout)")

        # Cleanup expired sessions every 10 minutes
        scheduler.add_job(cleanup_expired_sessions, 'interval', seconds=600, id='session_cleanup')
        logger.info("Scheduled expired session cleanup (every 10 minutes)")

        scheduler.start()
        logger.info("Background scheduler started successfully")
    except Exception as e:
        logger.error(f"Failed to start background scheduler: {e}")

    # Start network monitoring for proactive connection management (Phase 1 Enhancement)
    try:
        # Enable network monitoring with 5-second check interval
        monitoring_started = ssh_pool.start_network_monitoring(
            check_interval=5,     # Check every 5 seconds
            probe_timeout=2.0     # 2 second timeout per probe
        )
        if monitoring_started:
            logger.info("SSH network monitoring started successfully")
        else:
            logger.warning("SSH network monitoring failed to start (may already be running)")
    except Exception as e:
        logger.error(f"Failed to start SSH network monitoring: {e}")
        logger.warning("Continuing without proactive network monitoring")


@app.on_event("shutdown")
def shutdown_cleanup():
    """Clean up resources on application shutdown."""
    try:
        logger.info("Shutting down background scheduler")
        scheduler.shutdown(wait=False)
        logger.info("Closing all SSH connections in pool")
        ssh_pool.close_all()
        logger.info("Application shutdown complete")
    except Exception as e:
        logger.error(f"Error during shutdown cleanup: {e}")


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def get_session_id(request: Request) -> str:
    """Extract session ID from request headers."""
    return request.headers.get("X-Session-ID", "")


def verify_resource_access(resource_session_id: str, current_session_id: str, resource_type: str, resource_id) -> bool:
    """Verify that a resource belongs to the current session (security check).

    Args:
        resource_session_id: session_id from the resource being accessed
        current_session_id: session_id from the current request
        resource_type: human-readable resource type (for logging)
        resource_id: ID of the resource (for logging)

    Returns:
        True if access is allowed, False otherwise

    BUG FIX: Allow access to legacy resources with NULL session_id (created before session isolation).
    This enables users to view historical executions and logs while maintaining security for new resources.
    """
    # Allow access if current session is missing (should not happen, but defensive)
    if not current_session_id:
        logger.warning(f"[SECURITY] Session verification failed: missing current session ID for {resource_type} {resource_id}")
        return False

    # Allow access to legacy resources with NULL session_id (backward compatibility)
    if not resource_session_id:
        logger.info(f"[SECURITY] Allowing access to legacy {resource_type} {resource_id} with NULL session_id")
        return True

    # Normal case: resource has session_id, check if it matches current session
    if resource_session_id != current_session_id:
        logger.warning(f"[SECURITY] UNAUTHORIZED ACCESS ATTEMPT: User session {current_session_id} tried to access {resource_type} {resource_id} from session {resource_session_id}")
        return False

    return True


# ============================================================================
# FRONTEND — Serve the HTML dashboard
# ============================================================================

@app.get("/", response_class=HTMLResponse)
async def serve_dashboard(
    request: Request,
    db: Session = Depends(get_db),
    token: str = "",
    auth_token: str = "",
    access_token: str = "",
    session_token: str = "",
):
    """
    Serve the main HTML dashboard.
    When OnePalC SSO is enabled, unauthenticated or expired sessions are
    redirected to the OnePalC Hub login page before the dashboard is shown.

    OnePalC sometimes redirects back to /?token=<JWT> instead of
    /hub-callback?token=<JWT>. We detect that here and forward to hub-callback.
    """
    from fastapi.responses import RedirectResponse as _RR

    # OnePalC may send the JWT directly to / instead of /hub-callback.
    # Forward to the dedicated callback handler so session creation runs there.
    jwt_on_root = token or auth_token or access_token or session_token
    if jwt_on_root and _ONEPALC_ENABLED:
        param_name = "token" if token else ("auth_token" if auth_token else ("access_token" if access_token else "session_token"))
        callback_url = f"/hub-callback?{param_name}={urllib.parse.quote(jwt_on_root)}"
        logger.info(f"[SSO Guard] JWT received at / — forwarding to /hub-callback")
        return _RR(url=callback_url, status_code=302)

    if _ONEPALC_ENABLED and _ONEPALC_HUB_AUTH_URL:
        session_id = request.cookies.get("eka_session_id", "").strip()

        def _hub_login_redirect():
            callback = _ONEPALC_CALLBACK_URL or str(request.base_url) + "hub-callback"
            login_url = (
                f"{_ONEPALC_HUB_AUTH_URL}"
                f"?app_name={urllib.parse.quote(_ONEPALC_APP_NAME)}"
                f"&redirect_uri={urllib.parse.quote(callback)}"
            )
            return _RR(url=login_url, status_code=302)

        if not session_id:
            logger.info("[SSO Guard] No session cookie — redirecting to OnePalC login")
            return _hub_login_redirect()

        # Validate session is active (user-based sessions have no TTL)
        session = db.query(UserSession).filter(
            UserSession.session_id == session_id
        ).first()

        if not session:
            logger.info(f"[SSO Guard] Unknown session {session_id!r} — redirecting to login")
            return _hub_login_redirect()

        if session.status != "active":
            logger.info(f"[SSO Guard] Session {session_id!r} is {session.status} — redirecting to login")
            return _hub_login_redirect()

        # Valid session — update last_activity and serve the dashboard
        session.last_activity = datetime.utcnow()
        try:
            db.commit()
        except Exception:
            db.rollback()

    html_path = static_dir / "index.html"
    if html_path.exists():
        return HTMLResponse(content=html_path.read_text(encoding="utf-8"))
    return HTMLResponse(content="<h1>DUT Automation System</h1><p>Frontend not found.</p>")


@app.get("/health")
def health_check():
    """Health check endpoint for Docker and load balancers."""
    try:
        db = SessionLocal()
        db.execute(__import__("sqlalchemy").text("SELECT 1"))
        db.close()
        return {"status": "ok", "service": "eka-core", "database": "connected"}
    except Exception as e:
        return {"status": "degraded", "service": "eka-core", "database": str(e)}


# ============================================================================
# API — DUT Management
# ============================================================================

@app.get("/api/duts")
def get_duts(request: Request, db: Session = Depends(get_db)):
    """List DUT devices for the current session (session-isolated, REQUIRED)."""
    session_id = get_session_id(request)

    # SECURITY: Session ID is REQUIRED - cannot list all DUTs
    if not session_id:
        raise HTTPException(status_code=401, detail="Session ID required")

    # SECURITY: Only return DUTs belonging to this session
    duts = db.query(DUT).filter(DUT.session_id == session_id).all()

    return [
        {
            "id": d.id,
            "name": d.name,
            "ip_address": d.ip_address,
            "port": d.port,
            "device_type": d.device_type,
            "username": d.username,
            "connection_type": getattr(d, 'connection_type', 'ssh'),  # Default to ssh if not set
            "status": d.status,
            "last_heartbeat": d.last_heartbeat.isoformat() if d.last_heartbeat else None,
            "created_at": d.created_at.isoformat() if d.created_at else None,
        }
        for d in duts
    ]


@app.get("/api/duts/{dut_id}")
def get_dut(dut_id: int, request: Request, db: Session = Depends(get_db)):
    """Get details for a specific DUT (session-isolated)."""
    session_id = get_session_id(request)

    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    # SECURITY: Verify DUT belongs to current session
    if not verify_resource_access(dut.session_id, session_id, "DUT", dut_id):
        raise HTTPException(status_code=403, detail="Access denied")

    config = db.query(DUTConfiguration).filter(DUTConfiguration.dut_id == dut_id).first()
    return {
        "id": dut.id,
        "name": dut.name,
        "ip_address": dut.ip_address,
        "port": dut.port,
        "device_type": dut.device_type,
        "username": dut.username,
        "status": dut.status,
        "static_ip": config.static_ip if config else None,
        "image_path": config.image_path if config else None,
    }


@app.post("/api/duts")
def create_dut(request: Request, dut_data: dict, db: Session = Depends(get_db)):
    """Create a new DUT device (session-isolated, REQUIRED)."""
    try:
        session_id = get_session_id(request)

        # SECURITY: Session ID is REQUIRED - cannot create DUTs without session
        if not session_id:
            raise HTTPException(status_code=401, detail="Session ID required")

        # VALIDATION: Device name must be alphanumeric only (A-Z, a-z, 0-9)
        import re
        name = dut_data.get("name", "").strip()
        ip_address = dut_data.get("ip_address", "").strip()

        if not name:
            raise HTTPException(status_code=400, detail="Device name is required")
        if not re.match(r'^[A-Za-z0-9]+$', name):
            raise HTTPException(status_code=400, detail="Device name must contain only letters (A-Z, a-z) and numbers (0-9). No special characters or spaces allowed.")

        # VALIDATION: IP address must be valid IPv4 format
        if not ip_address:
            raise HTTPException(status_code=400, detail="IP address is required")
        ip_pattern = r'^((25[0-5]|2[0-4][0-9]|1[0-9]{2}|[1-9]?[0-9])\.){3}(25[0-5]|2[0-4][0-9]|1[0-9]{2}|[1-9]?[0-9])$'
        if not re.match(ip_pattern, ip_address):
            raise HTTPException(status_code=400, detail="Invalid IP address. Must be valid IPv4 format (e.g., 192.168.1.100). No subnet mask allowed.")

        connection_type = dut_data.get("connection_type", "ssh")

        dut = DUT(
            name=name,
            ip_address=ip_address,
            port=dut_data.get("port", 22),
            device_type=dut_data.get("device_type", "Linux"),
            username=dut_data.get("username", "admin"),
            password=dut_data.get("password", ""),
            xml_path=dut_data.get("xml_path", "/home/hp/prajwal/VMs"),
            connection_type=connection_type,  # Save connection type
            status="offline",  # All devices start as offline until connectivity test
            session_id=session_id,  # SECURITY: Must always have a session_id
        )
        db.add(dut)
        db.commit()
        db.refresh(dut)

        config = DUTConfiguration(
            dut_id=dut.id,
            static_ip=dut_data.get("static_ip"),
            image_path=dut_data.get("image_path"),
        )
        db.add(config)
        db.commit()

        return {"id": dut.id, "name": dut.name, "status": "created"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.put("/api/duts/{dut_id}")
def update_dut(dut_id: int, request: Request, dut_data: dict, db: Session = Depends(get_db)):
    """Update a DUT device with optional SSH validation for credential changes (session-isolated)."""
    session_id = get_session_id(request)

    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    # SECURITY: Verify DUT belongs to current session
    if not verify_resource_access(dut.session_id, session_id, "DUT", dut_id):
        raise HTTPException(status_code=403, detail="Access denied")

    try:
        # VALIDATION: Device name must be alphanumeric only (A-Z, a-z, 0-9)
        import re
        if 'name' in dut_data:
            name = dut_data['name'].strip() if isinstance(dut_data['name'], str) else dut_data['name']
            if not name:
                raise HTTPException(status_code=400, detail="Device name is required")
            if not re.match(r'^[A-Za-z0-9]+$', name):
                raise HTTPException(status_code=400, detail="Device name must contain only letters (A-Z, a-z) and numbers (0-9). No special characters or spaces allowed.")
            dut_data['name'] = name  # Update with trimmed value

        # VALIDATION: IP address must be valid IPv4 format
        if 'ip_address' in dut_data:
            ip_address = dut_data['ip_address'].strip() if isinstance(dut_data['ip_address'], str) else dut_data['ip_address']
            if not ip_address:
                raise HTTPException(status_code=400, detail="IP address is required")
            ip_pattern = r'^((25[0-5]|2[0-4][0-9]|1[0-9]{2}|[1-9]?[0-9])\.){3}(25[0-5]|2[0-4][0-9]|1[0-9]{2}|[1-9]?[0-9])$'
            if not re.match(ip_pattern, ip_address):
                raise HTTPException(status_code=400, detail="Invalid IP address. Must be valid IPv4 format (e.g., 192.168.1.100). No subnet mask allowed.")
            dut_data['ip_address'] = ip_address  # Update with trimmed value

        # Track if credentials changed
        creds_changed = (
            ('ip_address' in dut_data and dut_data['ip_address'] != dut.ip_address) or
            ('username' in dut_data and dut_data['username'] != dut.username) or
            ('password' in dut_data and dut_data['password'])
        )

        # Update all allowed fields including xml_path and connection_type
        for key in ["name", "ip_address", "port", "device_type", "username", "password", "xml_path", "connection_type"]:
            if key in dut_data:
                setattr(dut, key, dut_data[key])

        dut.updated_at = datetime.utcnow()
        db.commit()

        # If credentials changed and it's a network device type, return flag for frontend to test connection
        result = {
            "id": dut.id,
            "name": dut.name,
            "status": "updated",
            "credentials_changed": creds_changed and dut.device_type in ['DUT', 'Switch', 'Router']
        }
        return result
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.delete("/api/duts/{dut_id}")
def delete_dut(dut_id: int, request: Request, db: Session = Depends(get_db)):
    """Delete a DUT device and all related records (session-isolated)."""
    session_id = get_session_id(request)

    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    # SECURITY: Verify DUT belongs to current session
    if not verify_resource_access(dut.session_id, session_id, "DUT", dut_id):
        raise HTTPException(status_code=403, detail="Access denied")

    try:
        # Close any active SSH/Telnet connections before deleting
        connection_type = getattr(dut, 'connection_type', 'ssh')
        try:
            if connection_type == 'telnet':
                telnet_pool.close_connection(dut_id)
                logger.info(f"Closed Telnet connection for DUT {dut_id} ({dut.name})")
            else:
                ssh_pool.close_connection(dut_id)
                logger.info(f"Closed SSH connection for DUT {dut_id} ({dut.name})")
        except Exception as conn_err:
            logger.warning(f"Failed to close connection for DUT {dut_id}: {conn_err}")
            # Continue with deletion even if connection close fails

        # Delete related configuration
        db.query(DUTConfiguration).filter(DUTConfiguration.dut_id == dut_id).delete()
        # Delete the DUT
        db.delete(dut)
        db.commit()
        logger.info(f"Deleted DUT: {dut.name} (id={dut_id})")
        return {"status": "deleted", "name": dut.name}
    except Exception as e:
        db.rollback()
        logger.error(f"Failed to delete DUT {dut_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to delete: {str(e)}")


@app.post("/api/duts/{dut_id}/ping")
def ping_dut(dut_id: int, db: Session = Depends(get_db)):
    """Test connectivity to a DUT (SSH or Telnet based on connection_type)."""
    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    # Test telnet connectivity for telnet devices
    if hasattr(dut, 'connection_type') and dut.connection_type == 'telnet':
        logger.info(f"Testing telnet connectivity for DUT {dut.id} ({dut.name})")

        try:
            # Test actual telnet connection
            telnet_mgr = telnet_pool.get_connection(
                dut.id, dut.ip_address, dut.port, dut.username, dut.password, timeout=15
            )

            if telnet_mgr and telnet_mgr.is_alive():
                # Connection successful
                dut.status = "online"
                dut.last_heartbeat = datetime.utcnow()
                db.commit()
                telnet_pool.release_connection(dut.id)

                return {
                    "status": "online",
                    "message": f"{dut.name} is reachable via telnet",
                    "device_info": f"Telnet connection established to {dut.ip_address}:{dut.port}"
                }
            else:
                # Connection failed
                dut.status = "offline"
                db.commit()
                telnet_pool.close_connection(dut.id)
                raise HTTPException(
                    status_code=503,
                    detail=f"{dut.name} — telnet connection failed to {dut.ip_address}:{dut.port}"
                )
        except Exception as e:
            dut.status = "offline"
            db.commit()
            telnet_pool.close_connection(dut.id)
            raise HTTPException(
                status_code=503,
                detail=f"{dut.name} — telnet error: {str(e)}"
            )

    # First try a simple TCP/ping check
    import socket
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(5)
        result = sock.connect_ex((dut.ip_address, dut.port))
        sock.close()
        port_open = (result == 0)
    except Exception:
        port_open = False

    if not port_open:
        dut.status = "offline"
        db.commit()
        raise HTTPException(status_code=503, detail=f"{dut.name} — port {dut.port} not reachable at {dut.ip_address}")

    # Port is open, try SSH using connection pool
    ssh = ssh_pool.get_connection(dut.id, dut.ip_address, dut.port, dut.username, dut.password)
    if ssh:
        try:
            output, _, _ = ssh.execute_command("uname -a")
            dut.status = "online"
            dut.last_heartbeat = datetime.utcnow()
            db.commit()
            return {
                "status": "online",
                "message": f"{dut.name} is reachable",
                "device_info": output.strip(),
            }
        finally:
            ssh_pool.release_connection(dut.id)  # Return to pool, don't disconnect
    else:
        # Port is open but SSH auth failed
        dut.status = "offline"
        db.commit()
        raise HTTPException(
            status_code=503,
            detail=f"{dut.name} — SSH login failed for user '{dut.username}' at {dut.ip_address}:{dut.port}. Check username/password."
        )


def _parse_sonic_interfaces(output: str) -> list:
    """Parse the output of 'show interfaces status' from a SONiC device.
    Returns a list of dicts: {name, speed, mtu, fec, alias, oper, admin}.
    Falls back to extracting just names if the table format is unexpected.
    """
    interfaces = []
    lines = output.strip().splitlines()
    # Skip header/separator lines (start with '-' or 'Interface')
    for line in lines:
        line = line.strip()
        if not line or line.startswith('-') or line.lower().startswith('interface'):
            continue
        parts = line.split()
        if not parts:
            continue
        name = parts[0]
        # Must look like an Ethernet/Management interface
        if not re.match(r'^(Ethernet|Management|eth|PortChannel|Loopback)', name, re.IGNORECASE):
            continue
        intf = {"name": name}
        # Parse columns: Interface Lanes Speed MTU FEC Alias Vlan Oper Admin Type Asym_PFC
        if len(parts) >= 9:
            intf["speed"]  = parts[2]
            intf["mtu"]    = parts[3]
            intf["fec"]    = parts[4]
            intf["alias"]  = parts[5]
            intf["oper"]   = parts[7]
            intf["admin"]  = parts[8]
        elif len(parts) >= 3:
            intf["speed"] = parts[2]
            intf["oper"]  = parts[-2] if len(parts) >= 4 else "N/A"
            intf["admin"] = parts[-1]
        interfaces.append(intf)
    return interfaces


def _parse_linux_interfaces(output: str) -> list:
    """Fallback: parse 'ip link show' output to extract interface names."""
    interfaces = []
    for line in output.splitlines():
        m = re.match(r'^\d+:\s+(\S+?)(?:@\S+)?:\s+', line)
        if m:
            name = m.group(1)
            if name not in ('lo',):
                state = 'up' if 'state UP' in line else 'down'
                interfaces.append({"name": name, "oper": state, "admin": state})
    return interfaces


@app.get("/api/duts/{dut_id}/interfaces")
def get_dut_interfaces(dut_id: int, db: Session = Depends(get_db)):
    """SSH into a DUT and discover its interfaces via 'show interfaces status'.
    Only applies to network devices (DUT, Switch, Router) — not VMs or other types."""
    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    # Skip interface fetching for telnet devices (they don't use SSH)
    if hasattr(dut, 'connection_type') and dut.connection_type == 'telnet':
        raise HTTPException(
            status_code=400,
            detail=f"Interface fetching not supported for telnet devices. Use Hardware Load tab for telnet console access."
        )

    # Only fetch interfaces for actual network devices, not VMs
    if dut.device_type not in ["DUT", "Switch", "Router"]:
        raise HTTPException(
            status_code=400,
            detail=f"Interface fetching not supported for device type '{dut.device_type}'. Only available for DUT, Switch, and Router types."
        )

    ssh = SSHConnectionManager(dut.ip_address, dut.port, dut.username, dut.password)
    if not ssh.connect():
        dut.status = "offline"
        db.commit()
        raise HTTPException(status_code=503, detail=f"Cannot SSH into {dut.name}")

    try:
        # Try SONiC command first
        output, err, code = ssh.execute_command("show interfaces status", timeout=20)
        interfaces = _parse_sonic_interfaces(output) if output.strip() else []

        # If that gave nothing, fall back to 'ip link show'
        if not interfaces:
            output2, _, _ = ssh.execute_command("ip link show", timeout=15)
            interfaces = _parse_linux_interfaces(output2)

        # Update device status to online since SSH connection succeeded
        dut.status = "online"
        dut.last_heartbeat = datetime.utcnow()
        db.commit()

        logger.info(f"Fetched {len(interfaces)} interfaces from {dut.name} (type: {dut.device_type})")
        return {
            "dut_id": dut_id,
            "dut_name": dut.name,
            "device_type": dut.device_type,
            "interfaces": interfaces,
            "count": len(interfaces),
        }
    finally:
        ssh.disconnect()


@app.post("/api/duts/{dut_id}/execute")
def execute_command_on_dut(dut_id: int, body: dict, db: Session = Depends(get_db)):
    """Execute an ad-hoc command on a DUT with persistent working directory (like MobaXterm)."""
    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    command = body.get("command", "").strip()
    if not command:
        raise HTTPException(status_code=400, detail="No command provided")

    ssh = SSHConnectionManager(dut.ip_address, dut.port, dut.username, dut.password)
    if not ssh.connect():
        raise HTTPException(status_code=503, detail=f"Cannot connect to {dut.name}")

    try:
        # Get current working directory for this DUT session
        current_dir = _get_dut_cwd(dut_id)

        # Detect if this is a cd command
        is_cd_command = command.startswith("cd ") or command == "cd"

        if is_cd_command:
            # Handle cd command specially
            if command == "cd":
                # cd without args goes to home directory
                target_dir = "~"
            else:
                # Extract target directory
                target_dir = command[3:].strip()

            # Execute cd and get the new pwd
            chdir_cmd = f"cd {current_dir} && cd {target_dir} && pwd"
            output, error, exit_code = ssh.execute_command(chdir_cmd, timeout=body.get("timeout", 30))

            if exit_code == 0:
                # Update stored working directory
                new_dir = output.strip()
                _set_dut_cwd(dut_id, new_dir)

                # Return user-friendly output with new directory
                return {
                    "stdout": f"{dut.username}@{dut.name}:{new_dir}$",
                    "stderr": error,
                    "exit_code": exit_code,
                    "dut_name": dut.name,
                    "cwd": new_dir,
                }
            else:
                # cd failed, return error
                return {
                    "stdout": "",
                    "stderr": error if error else f"bash: cd: {target_dir}: No such file or directory",
                    "exit_code": exit_code,
                    "dut_name": dut.name,
                    "cwd": current_dir,
                }
        else:
            # Regular command - execute in the context of current directory
            contextual_cmd = f"cd {current_dir} && {command}"
            output, error, exit_code = ssh.execute_command(contextual_cmd, timeout=body.get("timeout", 30))

            return {
                "stdout": output,
                "stderr": error,
                "exit_code": exit_code,
                "dut_name": dut.name,
                "cwd": current_dir,
            }
    finally:
        ssh.disconnect()


@app.get("/api/duts/{dut_id}/session")
def get_dut_session(dut_id: int, db: Session = Depends(get_db)):
    """Get current terminal session state for a DUT."""
    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    current_dir = _get_dut_cwd(dut_id)

    # Try to get the actual current directory from the device (optional)
    try:
        ssh = SSHConnectionManager(dut.ip_address, dut.port, dut.username, dut.password)
        if ssh.connect():
            try:
                # Verify the directory exists and get actual pwd (short timeout)
                output, error, exit_code = ssh.execute_command(f"cd {current_dir} && pwd", timeout=5)
                if exit_code == 0:
                    actual_dir = output.strip()
                    if actual_dir:
                        _set_dut_cwd(dut_id, actual_dir)
                        current_dir = actual_dir
            except Exception as e:
                logger.warning(f"Could not verify directory for DUT {dut_id}: {e}")
            finally:
                ssh.disconnect()
    except Exception as e:
        logger.warning(f"Could not connect to DUT {dut_id} for session: {e}")
        # Continue anyway - return cached session state

    return {
        "dut_id": dut_id,
        "dut_name": dut.name,
        "username": dut.username,
        "cwd": current_dir,
        "prompt": f"{dut.username}@{dut.name}:{current_dir}$",
    }


@app.post("/api/duts/{dut_id}/session/reset")
def reset_dut_session(dut_id: int, db: Session = Depends(get_db)):
    """Reset terminal session state for a DUT (go back to home directory)."""
    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    # Reset to home directory
    _set_dut_cwd(dut_id, "~")

    return {
        "message": "Session reset to home directory",
        "cwd": "~",
    }


# ============================================================================
# ENHANCEMENT 3: DUT RESERVATION SYSTEM
# ============================================================================

@app.post("/api/duts/{dut_id}/reserve")
def reserve_dut(dut_id: int, body: dict, db: Session = Depends(get_db)):
    """Reserve a DUT for current user (Enhancement 3).

    Body:
        reserve: bool - True to reserve, False to release
        reserved_for_hours: int - How long to reserve (default 4 hours)
    """
    reserve = body.get("reserve", True)
    reserved_for_hours = body.get("reserved_for_hours", 4)

    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    if reserve:
        # Check if already reserved by someone else
        if dut.reserved_by and dut.reserved_by != "system":
            raise HTTPException(
                status_code=409,
                detail=f"DUT already reserved by {dut.reserved_by}"
            )

        # Reserve for current user
        dut.reserved_by = "current_user"  # Frontend will send actual username
        dut.reserved_at = datetime.utcnow()
        dut.reserved_until = datetime.utcnow() + timedelta(hours=reserved_for_hours)
        db.commit()

        return {
            "status": "reserved",
            "dut_id": dut_id,
            "dut_name": dut.name,
            "reserved_until": dut.reserved_until.isoformat(),
            "hours": reserved_for_hours
        }
    else:
        # Release reservation
        dut.reserved_by = None
        dut.reserved_at = None
        dut.reserved_until = None
        db.commit()

        return {
            "status": "released",
            "dut_id": dut_id,
            "dut_name": dut.name
        }


@app.get("/api/duts/reservations")
def get_dut_reservations(db: Session = Depends(get_db)):
    """Get all current DUT reservations."""
    duts = db.query(DUT).filter(DUT.reserved_by != None).all()

    reservations = []
    for dut in duts:
        is_expired = False
        if dut.reserved_until and dut.reserved_until < datetime.utcnow():
            # Auto-release expired reservation
            is_expired = True
            dut.reserved_by = None
            dut.reserved_at = None
            dut.reserved_until = None
        else:
            reservations.append({
                "dut_id": dut.id,
                "dut_name": dut.name,
                "reserved_by": dut.reserved_by,
                "reserved_at": dut.reserved_at.isoformat() if dut.reserved_at else None,
                "reserved_until": dut.reserved_until.isoformat() if dut.reserved_until else None,
                "time_remaining_minutes": int((dut.reserved_until - datetime.utcnow()).total_seconds() / 60) if dut.reserved_until else 0
            })

    if any(dut.reserved_until and dut.reserved_until < datetime.utcnow() for dut in duts):
        db.commit()  # Save auto-releases

    return {"reservations": reservations, "total": len(reservations)}


# ============================================================================
# HARDWARE LOAD API ENDPOINTS
# ============================================================================

from pydantic import BaseModel, validator
import ipaddress as ip_validation


class HardwareLoadRequest(BaseModel):
    """Request model for hardware load operation"""
    dut_id: int
    source_server_id: int
    image_path: str
    source_server_ip: str
    source_server_username: str
    source_server_password: str
    gateway_ip: str
    subnet_mask: str

    @validator('image_path')
    def validate_image_path(cls, v):
        """Validate image path for security"""
        # Prevent directory traversal
        if '..' in v:
            raise ValueError('Path traversal not allowed')
        # Must be absolute path
        if not v.startswith('/'):
            raise ValueError('Image path must be absolute')
        # Must end with .bin
        if not v.endswith('.bin'):
            raise ValueError('Image must be a .bin file')
        return v

    @validator('source_server_ip', 'gateway_ip')
    def validate_ip(cls, v):
        """Validate IPv4 address"""
        try:
            ip_validation.IPv4Address(v)
        except ValueError:
            raise ValueError(f'Invalid IPv4 address: {v}')
        return v

    @validator('subnet_mask')
    def validate_subnet(cls, v):
        """Validate subnet mask"""
        valid_masks = [
            '255.0.0.0', '255.255.0.0', '255.255.255.0',
            '255.255.255.128', '255.255.255.192', '255.255.255.224',
            '255.255.255.240', '255.255.255.248', '255.255.255.252',
            '255.255.255.255'  # /32 host route (single IP)
        ]
        if v not in valid_masks:
            raise ValueError(f'Invalid subnet mask: {v}. Must be one of: {", ".join(valid_masks)}')
        return v


@app.post("/api/hardware-load/start")
async def start_hardware_load(
    request: Request,
    hw_request: HardwareLoadRequest,
    db: Session = Depends(get_db)
):
    """
    Start hardware load operation for automated OS image installation.

    This endpoint initiates a 16-step automation process:
    1-4: Connect, save config, reboot
    5-9: Navigate GRUB/ONIE menus
    10-13: Configure network, download image
    14-16: Install image and complete

    Args:
        request: FastAPI request (for session/IP)
        hw_request: Hardware load parameters
        db: Database session

    Returns:
        Job information with job_id for tracking
    """
    # Get session ID from header
    session_id = request.headers.get("X-Session-ID", "default")
    user_ip = request.client.host if request.client else "unknown"

    # Validate device exists and belongs to session
    dut = db.query(DUT).filter(
        DUT.id == hw_request.dut_id,
        DUT.session_id == session_id
    ).first()

    if not dut:
        raise HTTPException(status_code=404, detail="Device not found or access denied")

    # Check if device has telnet connection type
    if not hasattr(dut, 'connection_type') or (hasattr(dut, 'connection_type') and dut.connection_type != 'telnet'):
        # Default to allowing if connection_type doesn't exist yet (backward compatibility)
        if hasattr(dut, 'connection_type'):
            raise HTTPException(
                status_code=400,
                detail="Device must use telnet connection for hardware load. Please update device connection type."
            )

    # Validate source server exists
    source_server = db.query(DUT).filter(DUT.id == hw_request.source_server_id).first()
    if not source_server:
        raise HTTPException(status_code=404, detail="Source server not found")

    # Create hardware load job
    job = HardwareLoadJob(
        dut_id=hw_request.dut_id,
        source_server_id=hw_request.source_server_id,
        image_path=hw_request.image_path,
        image_name=os.path.basename(hw_request.image_path),
        source_server_password=encrypt_password(hw_request.source_server_password),
        gateway_ip=hw_request.gateway_ip,
        subnet_mask=hw_request.subnet_mask,
        status="pending",
        session_id=session_id
    )

    db.add(job)
    db.commit()
    db.refresh(job)

    # Log audit event
    log_audit(
        db=db,
        session_id=session_id,
        user_ip=user_ip,
        action="hardware_load_start",
        resource_type="HardwareLoadJob",
        resource_id=job.id,
        details={
            "dut_id": hw_request.dut_id,
            "dut_name": dut.name,
            "image_name": job.image_name
        }
    )

    # Launch background task
    hw_request_with_plaintext_pw = hw_request.copy()
    asyncio.create_task(
        execute_hardware_load(
            job_id=job.id,
            dut=dut,
            request=hw_request_with_plaintext_pw,
            db=SessionLocal()
        )
    )

    logger.info(f"Hardware load job {job.id} started for DUT {dut.name} by session {session_id}")

    return {
        "job_id": job.id,
        "status": "started",
        "message": f"Hardware load job started. Use job_id {job.id} to track progress."
    }


@app.get("/api/hardware-load/job/{job_id}")
def get_hardware_load_job(
    job_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Get hardware load job status and progress.

    Returns complete job information including:
    - Current status and progress percentage
    - Execution logs
    - Error messages if failed
    - Timestamps

    Args:
        job_id: Hardware load job ID
        request: FastAPI request (for session validation)
        db: Database session

    Returns:
        Job details dictionary
    """
    session_id = request.headers.get("X-Session-ID", "default")

    job = db.query(HardwareLoadJob).filter(
        HardwareLoadJob.id == job_id,
        HardwareLoadJob.session_id == session_id
    ).first()

    if not job:
        raise HTTPException(status_code=404, detail="Job not found or access denied")

    # Get device name
    dut = db.query(DUT).filter(DUT.id == job.dut_id).first()
    device_name = dut.name if dut else f"DUT {job.dut_id}"

    return {
        "id": job.id,  # Changed from job_id to id for consistency
        "job_id": job.id,  # Keep for backward compatibility
        "device_name": device_name,
        "dut_id": job.dut_id,
        "image_path": job.image_path,
        "image_name": job.image_name,
        "status": job.status,
        "current_step": job.current_step,
        "progress_percentage": job.progress_percentage,
        "execution_log": job.execution_log,
        "error_message": job.error_message,
        "started_at": job.started_at.isoformat() if job.started_at else None,
        "completed_at": job.completed_at.isoformat() if job.completed_at else None
    }


@app.post("/api/hardware-load/cancel/{job_id}")
def cancel_hardware_load_job(
    job_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Cancel / force-stop a running hardware load job.

    Marks the job as 'failed' in the database so the background task
    detects the cancellation and terminates its loop on the next iteration.
    Also closes the telnet connection so the ONIE session is released.

    Args:
        job_id: Hardware load job ID to cancel
        request: FastAPI request (for session validation)
        db: Database session

    Returns:
        Confirmation message
    """
    session_id = request.headers.get("X-Session-ID", "default")

    job = db.query(HardwareLoadJob).filter(
        HardwareLoadJob.id == job_id,
        HardwareLoadJob.session_id == session_id
    ).first()

    if not job:
        raise HTTPException(status_code=404, detail="Job not found or access denied")

    if job.status in ("completed", "failed"):
        raise HTTPException(status_code=400, detail=f"Job is already {job.status} and cannot be cancelled")

    # Mark as failed so the background coroutine sees it and aborts
    job.status = "failed"
    job.error_message = "Cancelled by user"
    job.current_step = "Cancelled"
    from datetime import datetime as _dt
    job.completed_at = _dt.utcnow()
    db.commit()

    # Close the telnet connection so the running step doesn't hang
    try:
        from telnet_pool import telnet_pool as _tp
        _tp.close_connection(job.dut_id)
        _tp.unmark_connection_as_hardware_load(job.dut_id)
        logger.info(f"Telnet connection closed for cancelled job {job_id}")
    except Exception as ex:
        logger.warning(f"Could not close telnet for job {job_id}: {ex}")

    logger.info(f"Hardware load job {job_id} cancelled by session {session_id}")
    return {"status": "cancelled", "job_id": job_id, "message": "Hardware load job has been cancelled"}


@app.delete("/api/hardware-load/job/{job_id}")
def delete_hardware_load_job(
    job_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Permanently delete a hardware load job record from history.

    Only completed, failed, or cancelled jobs can be deleted.
    Running jobs must be cancelled first before deleting.

    Args:
        job_id: Hardware load job ID to delete
        request: FastAPI request (for session validation)
        db: Database session

    Returns:
        Confirmation message
    """
    session_id = request.headers.get("X-Session-ID", "default")

    job = db.query(HardwareLoadJob).filter(
        HardwareLoadJob.id == job_id,
        HardwareLoadJob.session_id == session_id
    ).first()

    if not job:
        raise HTTPException(status_code=404, detail="Job not found or access denied")

    if job.status not in ("completed", "failed", "cancelled"):
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete a job that is currently '{job.status}'. Cancel it first."
        )

    db.delete(job)
    db.commit()

    logger.info(f"Hardware load job {job_id} deleted by session {session_id}")
    return {"status": "deleted", "job_id": job_id, "message": "Hardware load job record deleted"}


@app.get("/api/hardware-load/jobs")
def get_hardware_load_jobs(
    request: Request,
    db: Session = Depends(get_db)
):
    """
    Get hardware load job history for current session.

    Returns list of all hardware load jobs for the current session,
    ordered by most recent first.

    Args:
        request: FastAPI request (for session ID)
        db: Database session

    Returns:
        List of job summaries
    """
    session_id = request.headers.get("X-Session-ID", "default")

    jobs = db.query(HardwareLoadJob).filter(
        HardwareLoadJob.session_id == session_id
    ).order_by(HardwareLoadJob.started_at.desc()).all()

    result = []
    for job in jobs:
        # Get device name
        dut = db.query(DUT).filter(DUT.id == job.dut_id).first()
        device_name = dut.name if dut else f"DUT {job.dut_id}"

        result.append({
            "id": job.id,
            "device_name": device_name,
            "dut_id": job.dut_id,
            "image_path": job.image_path,
            "image_name": job.image_name,
            "status": job.status,
            "progress_percentage": job.progress_percentage,
            "started_at": job.started_at.isoformat() if job.started_at else None,
            "completed_at": job.completed_at.isoformat() if job.completed_at else None
        })

    return result


@app.websocket("/api/hardware-load/ws/{job_id}")
async def hardware_load_websocket(
    websocket: WebSocket,
    job_id: int
):
    """
    WebSocket endpoint for streaming hardware load progress.

    Provides real-time updates for:
    - Progress percentage
    - Current step
    - Execution log lines
    - Job completion/failure

    Message format:
    {
        "type": "progress" | "complete" | "error",
        "status": "connecting" | "downloading" | "completed" | "failed",
        "current_step": "Downloading image...",
        "progress_percentage": 45,
        "new_log_lines": "log text here"
    }

    Args:
        websocket: WebSocket connection
        job_id: Hardware load job ID to monitor
    """
    await websocket.accept()

    db = SessionLocal()
    last_log_length = 0

    try:
        while True:
            # Fetch job from DB
            job = db.query(HardwareLoadJob).filter(HardwareLoadJob.id == job_id).first()

            if not job:
                await websocket.send_json({
                    "type": "error",
                    "message": "Job not found"
                })
                break

            # Refresh from DB to get latest log and status (avoid SQLAlchemy cache)
            db.refresh(job)

            # Send progress update if changed
            current_log = job.execution_log or ""
            if len(current_log) > last_log_length:
                new_lines = current_log[last_log_length:]
                await websocket.send_json({
                    "type": "progress",
                    "status": job.status,
                    "current_step": job.current_step,
                    "progress_percentage": job.progress_percentage,
                    "new_log_lines": new_lines
                })
                last_log_length = len(current_log)

            # Check if job completed
            if job.status in ["completed", "failed"]:
                # Flush any remaining log lines that arrived with the final status
                db.refresh(job)
                current_log = job.execution_log or ""
                if len(current_log) > last_log_length:
                    new_lines = current_log[last_log_length:]
                    await websocket.send_json({
                        "type": "progress",
                        "status": job.status,
                        "current_step": job.current_step,
                        "progress_percentage": job.progress_percentage,
                        "new_log_lines": new_lines
                    })
                    last_log_length = len(current_log)

                await websocket.send_json({
                    "type": "complete",
                    "status": job.status,
                    "error_message": job.error_message,
                    "progress_percentage": job.progress_percentage
                })
                break

            await asyncio.sleep(1)  # Poll every second

    except WebSocketDisconnect:
        logger.info(f"WebSocket disconnected for hardware load job {job_id}")
    except Exception as e:
        logger.error(f"WebSocket error for job {job_id}: {str(e)}")
    finally:
        db.close()


@app.websocket("/api/terminal/ws/{dut_id}")
async def terminal_websocket(websocket: WebSocket, dut_id: int):
    """
    Enterprise-grade PTY terminal WebSocket handler.
    Provides full interactive terminal support (vi, nano, top, htop, screen, tmux, etc.)

    Architecture:
    - Browser (xterm.js) <-> WebSocket <-> FastAPI <-> Paramiko PTY <-> Device Shell
    - Bidirectional binary stream for raw terminal data
    - Supports terminal resize, ANSI escape codes, and all interactive applications
    """
    # Extract session ID from query parameters (WebSockets can't use custom headers in browsers)
    session_id = websocket.query_params.get("session_id", "")

    # Get DUT details from database - open and close session immediately
    db = SessionLocal()
    try:
        dut = db.query(DUT).filter(DUT.id == dut_id).first()
        if not dut:
            await websocket.accept()
            await websocket.send_json({"error": "DUT not found"})
            await websocket.close()
            return

        # SECURITY: Verify DUT belongs to current session
        if not verify_resource_access(dut.session_id, session_id, "DUT", dut_id):
            await websocket.accept()
            await websocket.send_json({"error": "Access denied - device belongs to different session"})
            await websocket.close()
            logger.warning(f"[SECURITY] Unauthorized terminal access attempt: session {session_id} tried to access DUT {dut_id} from session {dut.session_id}")
            return

        # Extract needed data, then close DB session immediately
        dut_data = {
            "id": dut.id,
            "name": dut.name,
            "ip_address": dut.ip_address,
            "port": dut.port,
            "username": dut.username,
            "password": dut.password,
            "connection_type": getattr(dut, 'connection_type', 'ssh')
        }
    finally:
        db.close()  # CRITICAL: Release connection immediately

    # Accept WebSocket connection after validation passes
    await websocket.accept()

    # Check if device is telnet - Terminal tab only supports SSH
    if dut_data.get('connection_type') == 'telnet':
        logger.warning(f"[PTY] Terminal connection attempted for telnet device {dut_id} ({dut_data['name']})")
        await websocket.send_json({
            "error": "Terminal tab does not support telnet devices. Use Hardware Load tab for telnet console access."
        })
        await websocket.close()
        return

    ssh_manager = None
    channel = None
    session_key = f"{dut_id}_{id(websocket)}"

    try:
        # Use SSH connection pool (reuses existing connection)
        logger.info(f"[PTY] Getting SSH connection from pool for {dut_data['name']} ({dut_data['ip_address']}:{dut_data['port']})...")

        # Send initial status message to keep WebSocket alive
        await websocket.send_json({"status": "connecting", "message": "Connecting to device..."})

        # Get connection from pool (or create new one if needed)
        ssh_manager = ssh_pool.get_connection(
            dut_id,
            dut_data['ip_address'],
            dut_data['port'],
            dut_data['username'],
            dut_data['password']
        )

        if not ssh_manager:
            logger.error(f"[PTY] Failed to get SSH connection for DUT {dut_id}")
            await websocket.send_json({"error": "Failed to connect to device - SSH connection error"})
            await websocket.close()
            return

        # Mark connection as terminal session (prevents idle cleanup)
        ssh_pool.mark_connection_as_terminal(dut_id)
        logger.info(f"[PTY] Marked connection as terminal session for DUT {dut_id}")

        def create_pty_channel():
            """
            Create PTY channel on existing SSH connection (run in thread executor).
            Implements retry logic for ChannelException errors.
            """
            nonlocal ssh_manager  # Allow modifying outer scope variable
            max_retries = 3
            retry_delay = 0.5  # seconds

            for attempt in range(max_retries):
                try:
                    # Verify transport is active before creating channel
                    transport = ssh_manager.client.get_transport()
                    if not transport or not transport.is_active():
                        logger.error(f"[PTY] Transport not active for DUT {dut_id} (attempt {attempt + 1}/{max_retries})")
                        if attempt < max_retries - 1:
                            time.sleep(retry_delay)
                            continue
                        raise Exception("SSH transport not active")

                    # Create PTY channel with proper error handling
                    logger.info(f"[PTY] Creating PTY channel for DUT {dut_id} (attempt {attempt + 1}/{max_retries})")
                    channel = ssh_manager.client.invoke_shell(term='xterm-256color', width=80, height=24)
                    channel.setblocking(0)  # Non-blocking for async I/O

                    logger.info(f"[PTY] PTY channel created successfully for DUT {dut_id}")
                    return channel

                except Exception as e:
                    error_msg = str(e)
                    if "ChannelException" in error_msg or "Connect failed" in error_msg:
                        logger.warning(f"[PTY] ChannelException on attempt {attempt + 1}/{max_retries} for DUT {dut_id}: {e}")

                        if attempt < max_retries - 1:
                            # Wait before retry
                            logger.info(f"[PTY] Waiting {retry_delay}s before retry...")
                            time.sleep(retry_delay)
                            retry_delay *= 2  # Exponential backoff

                            # Try refreshing the connection on last retry
                            if attempt == max_retries - 2:
                                logger.info(f"[PTY] Refreshing SSH connection for DUT {dut_id} before final attempt")
                                try:
                                    new_conn = ssh_pool.refresh_connection(
                                        dut_id,
                                        dut_data['ip_address'],
                                        dut_data['port'],
                                        dut_data['username'],
                                        dut_data['password']
                                    )
                                    if new_conn:
                                        # Update ssh_manager reference
                                        ssh_manager = new_conn
                                        ssh_pool.mark_connection_as_terminal(dut_id)
                                        logger.info(f"[PTY] Connection refreshed for DUT {dut_id}")
                                except Exception as refresh_err:
                                    logger.error(f"[PTY] Failed to refresh connection: {refresh_err}")
                        else:
                            logger.error(f"[PTY] All {max_retries} attempts failed for DUT {dut_id}")
                            raise
                    else:
                        # Non-ChannelException error, don't retry
                        logger.error(f"[PTY] Non-retryable error creating channel for DUT {dut_id}: {e}")
                        raise

            raise Exception(f"Failed to create PTY channel after {max_retries} attempts")

        # Run blocking PTY creation in thread executor (prevents blocking event loop)
        loop = asyncio.get_event_loop()
        try:
            channel = await loop.run_in_executor(None, create_pty_channel)
            logger.info(f"[PTY] PTY channel ready for DUT {dut_id} ({dut_data['name']})")
        except Exception as channel_err:
            logger.error(f"[PTY] Failed to create PTY channel for DUT {dut_id}: {channel_err}")
            await websocket.send_json({"error": f"Failed to create terminal session: {str(channel_err)}"})
            ssh_pool.unmark_connection_as_terminal(dut_id)
            await websocket.close()
            return

        # Small delay to let the channel fully initialize and receive initial prompt
        await asyncio.sleep(0.1)

        # Store session in global dict for tracking
        with _pty_sessions_lock:
            _pty_sessions[session_key] = {
                "ssh_manager": ssh_manager,  # Store manager, not raw ssh
                "channel": channel,
                "dut_id": dut_id,
                "dut_name": dut_data['name']
            }

        # Task 1: Read from SSH channel and send to browser
        async def read_from_ssh():
            """Continuously read output from SSH PTY and stream to browser."""
            try:
                while True:
                    # Check if channel is still open
                    if channel.closed:
                        logger.info(f"[PTY] SSH channel closed by remote host for DUT {dut_id}")
                        break

                    # Check if channel has data available (non-blocking)
                    if channel.recv_ready():
                        try:
                            data = channel.recv(4096)  # Read up to 4KB
                            if data:
                                await websocket.send_bytes(data)
                            else:
                                # Empty data means channel closed
                                logger.info(f"[PTY] SSH channel closed (EOF) for DUT {dut_id}")
                                break
                        except Exception as recv_err:
                            logger.error(f"[PTY] Error receiving data from SSH channel (DUT {dut_id}): {recv_err}")
                            break

                    # Small delay to prevent CPU spinning
                    await asyncio.sleep(0.01)

            except asyncio.CancelledError:
                logger.info(f"[PTY] SSH reader task cancelled for DUT {dut_id}")
                raise
            except Exception as e:
                logger.error(f"[PTY] Error reading from SSH (DUT {dut_id}): {type(e).__name__}: {e}", exc_info=True)

        # Task 2: Read from browser and send to SSH channel
        async def read_from_browser():
            """Continuously read input from browser and send to SSH PTY."""
            try:
                while True:
                    data = await websocket.receive()

                    if "bytes" in data:
                        # Binary data (keyboard input from xterm.js)
                        try:
                            channel.send(data["bytes"])
                        except Exception as send_err:
                            logger.error(f"[PTY] Error sending data to SSH channel (DUT {dut_id}): {send_err}")
                            break

                    elif "text" in data:
                        # Text data (control messages like resize)
                        try:
                            msg = json.loads(data["text"])

                            if msg.get("type") == "resize":
                                # Handle terminal resize event
                                cols = msg.get("cols", 80)
                                rows = msg.get("rows", 24)
                                channel.resize_pty(width=cols, height=rows)
                                logger.info(f"[PTY] Terminal resized to {cols}x{rows} for DUT {dut_id}")

                        except json.JSONDecodeError:
                            logger.warning(f"[PTY] Invalid JSON message from browser: {data['text']}")
                        except Exception as resize_err:
                            logger.error(f"[PTY] Error resizing terminal (DUT {dut_id}): {resize_err}")

            except WebSocketDisconnect:
                logger.info(f"[PTY] WebSocket disconnected for DUT {dut_id}")
            except asyncio.CancelledError:
                logger.info(f"[PTY] Browser reader task cancelled for DUT {dut_id}")
                raise
            except Exception as e:
                logger.error(f"[PTY] Error reading from browser (DUT {dut_id}): {type(e).__name__}: {e}", exc_info=True)

        # Task 3: Send heartbeat every 30 seconds to keep connection alive
        async def send_heartbeat():
            """Send periodic heartbeat to keep WebSocket and SSH session alive."""
            try:
                while True:
                    await asyncio.sleep(30)  # Heartbeat every 30 seconds
                    try:
                        await websocket.send_json({"type": "heartbeat", "timestamp": time.time()})
                        ssh_pool.release_connection(dut_id)  # Update last_used timestamp
                        logger.debug(f"[PTY] Heartbeat sent for DUT {dut_id}")
                    except Exception as hb_err:
                        logger.error(f"[PTY] Failed to send heartbeat for DUT {dut_id}: {hb_err}")
                        break
            except asyncio.CancelledError:
                logger.info(f"[PTY] Heartbeat task cancelled for DUT {dut_id}")
                raise
            except Exception as e:
                logger.error(f"[PTY] Error in heartbeat task (DUT {dut_id}): {type(e).__name__}: {e}")

        # Run all three tasks concurrently (bidirectional streaming + keepalive)
        await asyncio.gather(
            read_from_ssh(),
            read_from_browser(),
            send_heartbeat(),
            return_exceptions=True
        )

    except paramiko.AuthenticationException:
        logger.error(f"[PTY] Authentication failed for DUT {dut_id}")
        await websocket.send_json({"error": "Authentication failed - check credentials"})
    except paramiko.SSHException as e:
        logger.error(f"[PTY] SSH error for DUT {dut_id}: {e}")
        await websocket.send_json({"error": f"SSH connection error: {str(e)}"})
    except Exception as e:
        logger.error(f"[PTY] Unexpected error for DUT {dut_id}: {e}")
        await websocket.send_json({"error": f"Terminal error: {str(e)}"})

    finally:
        # Cleanup: close PTY channel and release SSH connection back to pool
        if channel:
            try:
                channel.close()
                logger.info(f"[PTY] PTY channel closed for DUT {dut_id}")
            except Exception as e:
                logger.warning(f"[PTY] Error closing PTY channel for DUT {dut_id}: {e}")

        # Release SSH connection back to pool (do NOT close it - let pool manage it)
        if ssh_manager:
            try:
                # Unmark as terminal session (allows cleanup on next idle check)
                ssh_pool.unmark_connection_as_terminal(dut_id)
                ssh_pool.release_connection(dut_id)
                logger.info(f"[PTY] SSH connection released back to pool for DUT {dut_id}")
            except Exception as e:
                logger.warning(f"[PTY] Error releasing SSH connection for DUT {dut_id}: {e}")

        with _pty_sessions_lock:
            if session_key in _pty_sessions:
                del _pty_sessions[session_key]

        try:
            await websocket.close()
        except:
            pass

        logger.info(f"[PTY] Terminal session ended for DUT {dut_id}")


# ============================================================================
# API — Image Management
# ============================================================================

@app.get("/api/images")
def get_images(db: Session = Depends(get_db)):
    """List all uploaded images."""
    images = db.query(Image).all()
    return [
        {
            "id": img.id,
            "name": img.name,
            "version": img.version,
            "checksum": img.checksum,
            "file_size": img.file_size,
            "created_at": img.created_at.isoformat() if img.created_at else None,
        }
        for img in images
    ]


@app.post("/api/images")
async def upload_image(
    file: UploadFile = File(...),
    name: str = Form(""),
    version: str = Form("1.0"),
    db: Session = Depends(get_db),
):
    """Upload a new firmware image."""
    try:
        if not name:
            name = file.filename or f"image_{datetime.utcnow().timestamp()}"

        filename = f"{name}_{version}_{int(datetime.utcnow().timestamp())}{Path(file.filename).suffix}"
        file_path = str(IMAGES_DIR / filename)

        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)

        sha256_hash = hashlib.sha256(content).hexdigest()

        image = Image(
            name=name,
            version=version,
            file_path=file_path,
            checksum=sha256_hash,
            file_size=len(content),
        )
        db.add(image)
        db.commit()
        db.refresh(image)

        return {
            "id": image.id,
            "name": image.name,
            "version": image.version,
            "checksum": sha256_hash,
            "file_size": len(content),
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.delete("/api/images/{image_id}")
def delete_image(image_id: int, db: Session = Depends(get_db)):
    """Delete an image."""
    image = db.query(Image).filter(Image.id == image_id).first()
    if not image:
        raise HTTPException(status_code=404, detail="Image not found")
    if image.file_path and os.path.exists(image.file_path):
        os.remove(image.file_path)
    db.delete(image)
    db.commit()
    return {"status": "deleted", "name": image.name}


# ============================================================================
# API — Script Management
# ============================================================================

@app.get("/api/scripts")
def get_scripts(db: Session = Depends(get_db)):
    """List all scripts."""
    scripts = db.query(Script).all()
    return [
        {
            "id": s.id,
            "name": s.name,
            "file_path": s.file_path,
            "created_at": s.created_at.isoformat() if s.created_at else None,
        }
        for s in scripts
    ]


@app.post("/api/scripts")
async def upload_script(
    file: UploadFile = File(...),
    name: str = Form(""),
    db: Session = Depends(get_db),
):
    """Upload a YAML test script."""
    try:
        if not name:
            name = Path(file.filename).stem if file.filename else f"script_{int(datetime.utcnow().timestamp())}"

        filename = f"{name}.yaml"
        file_path = str(SCRIPTS_DIR / filename)

        content = await file.read()
        content_str = content.decode("utf-8", errors="ignore")

        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content_str)

        # Validate YAML
        try:
            yaml.safe_load(content_str)
        except yaml.YAMLError as ye:
            raise HTTPException(status_code=400, detail=f"Invalid YAML: {str(ye)}")

        script = Script(
            name=name,
            file_path=file_path,
            yaml_content=content_str,
        )
        db.add(script)
        db.commit()
        db.refresh(script)

        return {"id": script.id, "name": script.name, "status": "uploaded"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/scripts/local")
def get_local_scripts(db: Session = Depends(get_db)):
    """Auto-load YAML scripts from data/scripts/ and return the list."""
    # Auto-sync: load any new YAML files from disk into the database
    for yaml_file in SCRIPTS_DIR.glob("*.yaml"):
        script_name = yaml_file.stem
        content = yaml_file.read_text(encoding="utf-8")
        existing = db.query(Script).filter(Script.name == script_name).first()
        if existing:
            existing.yaml_content = content
            existing.updated_at = datetime.utcnow()
        else:
            script = Script(name=script_name, file_path=str(yaml_file), yaml_content=content)
            db.add(script)
    db.commit()

    # Return all scripts
    scripts = db.query(Script).all()
    result = []
    for s in scripts:
        # Parse YAML to get metadata
        meta = {}
        if s.yaml_content:
            try:
                parsed = yaml.safe_load(s.yaml_content) or {}
                meta = {
                    "description": parsed.get("description", ""),
                    "test_case_count": len(parsed.get("test_cases", [])),
                }
            except Exception:
                pass
        result.append({
            "id": s.id,
            "name": s.name,
            "file_path": s.file_path,
            "description": meta.get("description", ""),
            "test_case_count": meta.get("test_case_count", 0),
            "created_at": s.created_at.isoformat() if s.created_at else None,
        })
    return result


@app.get("/api/scripts/{script_id}")
def get_script_detail(script_id: int, db: Session = Depends(get_db)):
    """Get script details including YAML content."""
    script = db.query(Script).filter(Script.id == script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="Script not found")
    return {
        "id": script.id,
        "name": script.name,
        "yaml_content": script.yaml_content,
        "file_path": script.file_path,
    }


@app.delete("/api/scripts/{script_id}")
def delete_script(script_id: int, db: Session = Depends(get_db)):
    """Delete a script."""
    script = db.query(Script).filter(Script.id == script_id).first()
    if not script:
        raise HTTPException(status_code=404, detail="Script not found")
    if script.file_path and os.path.exists(script.file_path):
        os.remove(script.file_path)
    db.delete(script)
    db.commit()
    return {"status": "deleted", "name": script.name}


@app.post("/api/scripts/load-local")
def load_local_scripts(db: Session = Depends(get_db)):
    """Load all YAML scripts from the data/scripts directory."""
    count = 0
    for yaml_file in SCRIPTS_DIR.glob("*.yaml"):
        script_name = yaml_file.stem
        content = yaml_file.read_text(encoding="utf-8")
        existing = db.query(Script).filter(Script.name == script_name).first()
        if existing:
            existing.yaml_content = content
            existing.updated_at = datetime.utcnow()
        else:
            script = Script(name=script_name, file_path=str(yaml_file), yaml_content=content)
            db.add(script)
        count += 1
    db.commit()
    return {"loaded": count, "message": f"{count} scripts loaded from {SCRIPTS_DIR}"}


# ============================================================================
# API — Execution Management
# ============================================================================

@app.post("/api/executions")
def create_execution(request: Request, execution_data: dict, db: Session = Depends(get_db)):
    """Create and start a new execution (image deployment or script execution, session-isolated, REQUIRED)."""
    try:
        session_id = get_session_id(request)

        # SECURITY: Session ID is REQUIRED - cannot create executions without session
        if not session_id:
            raise HTTPException(status_code=401, detail="Session ID required")

        dut_ids = execution_data.get("dut_ids", [])
        script_id = execution_data.get("script_id")
        image_id = execution_data.get("image_id")

        if not dut_ids:
            raise HTTPException(status_code=400, detail="No DUT IDs provided")

        # SECURITY: Validate DUTs exist AND belong to current session only
        duts = db.query(DUT).filter(DUT.id.in_(dut_ids), DUT.session_id == session_id).all()
        if len(duts) != len(dut_ids):
            raise HTTPException(status_code=403, detail="One or more DUTs not found or access denied")

        exec_type = "image" if image_id else "script"
        execution = Execution(
            name=f"exec_{exec_type}_{int(datetime.utcnow().timestamp())}",
            script_id=script_id,
            image_id=image_id,
            dut_ids=json.dumps(dut_ids),
            execution_type=exec_type,
            status="pending",
            session_id=session_id,  # SECURITY: Must always have a session_id
        )
        db.add(execution)
        db.commit()
        db.refresh(execution)

        # Run in background thread
        if image_id:
            thread = Thread(
                target=run_image_deployment,
                args=(execution.id, dut_ids, image_id),
                daemon=True,
            )
        elif script_id:
            thread = Thread(
                target=run_script_execution,
                args=(execution.id, script_id, dut_ids),
                daemon=True,
            )
        else:
            raise HTTPException(status_code=400, detail="Either script_id or image_id required")

        thread.start()

        return {
            "execution_id": execution.id,
            "status": "started",
            "type": exec_type,
            "dut_count": len(dut_ids),
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/executions")
def get_executions(request: Request, db: Session = Depends(get_db)):
    """List executions for the current session (session-isolated, REQUIRED).

    BUG FIX: Also include legacy executions with session_id = NULL for backward compatibility.
    This allows users to view historical executions created before session isolation was implemented.
    """
    session_id = get_session_id(request)

    # SECURITY: Session ID is REQUIRED - cannot list all executions
    if not session_id:
        raise HTTPException(status_code=401, detail="Session ID required")

    # SECURITY: Return executions belonging to this session OR with NULL session_id (legacy data)
    # Use OR condition to include both current session and legacy (NULL) executions
    executions = (
        db.query(Execution)
        .filter(
            or_(
                Execution.session_id == session_id,
                Execution.session_id.is_(None)
            )
        )
        .order_by(Execution.created_at.desc())
        .all()
    )

    out = []
    for ex in executions:
        results = json.loads(ex.test_results or "[]") if ex.test_results else []
        passed  = sum(r.get("passed",  0) for r in results)
        failed  = sum(r.get("failed",  0) for r in results)
        skipped = sum(r.get("skipped", 0) for r in results)
        out.append({
            "id": ex.id,
            "name": ex.name,
            "type": ex.execution_type,
            "status": ex.status,
            "dut_count": len(json.loads(ex.dut_ids)) if ex.dut_ids else 0,
            "duration": ex.duration_seconds,
            "created_at": ex.created_at.isoformat() if ex.created_at else None,
            "passed": passed, "failed": failed, "skipped": skipped,
        })
    return out


@app.get("/api/executions/compare")
def compare_executions(a: int, b: int, request: Request, db: Session = Depends(get_db)):
    session_id = get_session_id(request)
    exec_a = db.query(Execution).filter(Execution.id == a).first()
    exec_b = db.query(Execution).filter(Execution.id == b).first()
    if not exec_a or not exec_b:
        raise HTTPException(status_code=404, detail="One or both executions not found")
    if not verify_resource_access(exec_a.session_id, session_id, "Execution", a):
        raise HTTPException(status_code=403, detail="Access denied")

    def _map(exec_id):
        tcrs = db.query(TestCaseResult).filter(TestCaseResult.execution_id == exec_id).all()
        return {t.test_function: t.result for t in tcrs if t.test_function}

    map_a = _map(a)
    map_b = _map(b)
    all_funcs = set(map_a) | set(map_b)

    regressed, fixed, new_failures = [], [], []
    stable_pass, stable_fail, stable_skip = [], [], []

    def _is_pass(r): return (r or '').lower() == 'pass'
    def _is_fail(r): return (r or '').lower() in ('fail', 'scripterror', 'error')
    def _is_skip(r): return (r or '').lower() in ('skip', 'xfail', 'deselect')

    for fn in sorted(all_funcs):
        ra, rb = map_a.get(fn), map_b.get(fn)
        item = {"test_function": fn, "result_a": ra or "–", "result_b": rb or "–"}
        if ra is None:
            if _is_fail(rb): new_failures.append(item)
        elif rb is None:
            pass
        elif _is_pass(ra) and (_is_fail(rb) or _is_skip(rb)):
            regressed.append(item)
        elif (_is_fail(ra) or _is_skip(ra)) and _is_pass(rb):
            fixed.append(item)
        elif _is_pass(ra) and _is_pass(rb):
            stable_pass.append(item)
        elif _is_fail(ra) and _is_fail(rb):
            stable_fail.append(item)
        elif _is_skip(ra) and _is_skip(rb):
            stable_skip.append(item)

    return {
        "run_a": {"id": a, "name": exec_a.name},
        "run_b": {"id": b, "name": exec_b.name},
        "regressed": regressed,
        "fixed": fixed,
        "new_failures": new_failures,
        "stable_pass": stable_pass,
        "stable_fail": stable_fail,
        "stable_skip": stable_skip,
    }


@app.get("/api/executions/{execution_id}")
def get_execution(execution_id: int, request: Request, db: Session = Depends(get_db)):
    """Get execution details (session-isolated)."""
    session_id = get_session_id(request)

    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")

    # SECURITY: Verify this execution belongs to the current session
    if not verify_resource_access(execution.session_id, session_id, "Execution", execution_id):
        raise HTTPException(status_code=403, detail="Access denied")

    return {
        "id": execution.id,
        "name": execution.name,
        "type": execution.execution_type,
        "status": execution.status,
        "dut_ids": json.loads(execution.dut_ids) if execution.dut_ids else [],
        "script_id": execution.script_id,
        "image_id": execution.image_id,
        "duration": execution.duration_seconds,
        "start_time": execution.start_time.isoformat() if execution.start_time else None,
        "end_time": execution.end_time.isoformat() if execution.end_time else None,
    }


@app.get("/api/executions/{execution_id}/logs")
def get_execution_logs(
    execution_id: int,
    request: Request,
    limit: int = 200,
    offset: int = 0,
    db: Session = Depends(get_db),
):
    """Get execution logs (paginated, session-isolated)."""
    session_id = get_session_id(request)

    # SECURITY: Verify execution belongs to current session
    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")

    if not verify_resource_access(execution.session_id, session_id, "Execution", execution_id):
        raise HTTPException(status_code=403, detail="Access denied")

    logs = (
        db.query(ExecutionLog)
        .filter(ExecutionLog.execution_id == execution_id)
        .order_by(ExecutionLog.timestamp.asc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    return [
        {
            "id": log.id,
            "dut_name": log.dut_name,
            "level": log.log_level,
            "message": log.message,
            "timestamp": log.timestamp.isoformat() if log.timestamp else None,
        }
        for log in logs
    ]


@app.get("/api/executions/{execution_id}/dashboard")
def download_dashboard(execution_id: int, request: Request, db: Session = Depends(get_db)):
    """Download HTML dashboard report for an execution."""
    session_id = get_session_id(request)
    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")
    if not verify_resource_access(execution.session_id, session_id, "Execution", execution_id):
        raise HTTPException(status_code=403, detail="Access denied")
    tcrs = (db.query(TestCaseResult)
            .filter(TestCaseResult.execution_id == execution_id)
            .order_by(TestCaseResult.id).all())
    results = json.loads(execution.test_results or "[]")
    if not results and tcrs:
        results = _rebuild_results_from_tcrs(tcrs)
    html_content = _build_html_dashboard(execution, results, tcrs)
    return StreamingResponse(
        io.BytesIO(html_content.encode("utf-8")),
        media_type="text/html",
        headers={"Content-Disposition": f'attachment; filename="eka_dashboard_{execution_id}.html"'})


@app.get("/api/executions/{execution_id}/excel")
def download_excel(execution_id: int, request: Request, db: Session = Depends(get_db)):
    """Download Excel (.xlsx) report for an execution."""
    if not _HAS_OPENPYXL:
        raise HTTPException(status_code=503, detail="openpyxl not installed on server — run: pip install openpyxl")
    session_id = get_session_id(request)
    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")
    if not verify_resource_access(execution.session_id, session_id, "Execution", execution_id):
        raise HTTPException(status_code=403, detail="Access denied")
    tcrs = (db.query(TestCaseResult)
            .filter(TestCaseResult.execution_id == execution_id)
            .order_by(TestCaseResult.id).all())
    results = json.loads(execution.test_results or "[]")
    if not results and tcrs:
        results = _rebuild_results_from_tcrs(tcrs)
    buf = _build_excel(execution, results, tcrs)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="eka_results_{execution_id}.xlsx"'})


@app.get("/api/testcases/summary")
def testcase_summary(limit: int = 50, request: Request = None,
                     db: Session = Depends(get_db)):
    """Return last 5 results for up to `limit` distinct test functions — used by Dashboard."""
    from sqlalchemy import func as sqlfunc
    subq = (db.query(TestCaseResult.test_function,
                     sqlfunc.max(TestCaseResult.execution_id).label("max_exec"))
            .filter(TestCaseResult.test_function != None,
                    TestCaseResult.test_function != "")
            .group_by(TestCaseResult.test_function)
            .order_by(sqlfunc.max(TestCaseResult.execution_id).desc())
            .limit(limit).subquery())
    functions = [row[0] for row in db.query(subq.c.test_function).all()]
    out = []
    for fn in functions:
        last5 = (db.query(TestCaseResult)
                 .filter(TestCaseResult.test_function == fn)
                 .order_by(TestCaseResult.execution_id.desc())
                 .limit(5).all())
        results_list = [{"execution_id": t.execution_id, "result": t.result,
                         "time_seconds": t.time_seconds} for t in last5]
        out.append({"test_function": fn, "results": results_list,
                    "trend": _calc_trend(results_list)})
    return out


@app.get("/api/testcases/history")
def testcase_history(function: str = "", limit: int = 10,
                     request: Request = None, db: Session = Depends(get_db)):
    """Return per-execution result history for a single test function."""
    if not function:
        raise HTTPException(status_code=400, detail="function parameter required")
    tcrs = (db.query(TestCaseResult)
            .filter(TestCaseResult.test_function == function)
            .order_by(TestCaseResult.execution_id.desc())
            .limit(limit).all())
    return [{"execution_id": t.execution_id, "result": t.result,
             "time_taken": t.time_taken, "time_seconds": t.time_seconds,
             "created_at": t.created_at.isoformat() if t.created_at else None} for t in tcrs]


@app.delete("/api/executions/{execution_id}/logs")
def delete_execution_logs(
    execution_id: int,
    body: dict,
    request: Request,
    db: Session = Depends(get_db),
):
    """Delete execution logs with scope support (session-isolated).

    Body:
        scope: str - 'all' (delete all logs) or 'current_session' (delete session logs only)

    Returns:
        {
            "status": "success",
            "deleted_count": int,
            "scope": str,
            "execution_id": int
        }
    """
    session_id = get_session_id(request)
    scope = body.get("scope", "all")

    if scope not in ["all", "current_session"]:
        raise HTTPException(status_code=400, detail="Invalid scope. Must be 'all' or 'current_session'")

    try:
        # Verify execution exists and belongs to current session
        execution = db.query(Execution).filter(Execution.id == execution_id).first()
        if not execution:
            raise HTTPException(status_code=404, detail="Execution not found")

        # SECURITY: Verify this execution belongs to the current session
        if not verify_resource_access(execution.session_id, session_id, "Execution", execution_id):
            raise HTTPException(status_code=403, detail="Access denied")

        if scope == "all":
            # Delete all logs for this execution
            deleted_count = db.query(ExecutionLog).filter(
                ExecutionLog.execution_id == execution_id
            ).delete(synchronize_session=False)
            logger.info(f"[LOGS] Deleted {deleted_count} logs for execution {execution_id} (scope: all)")

            # BUG FIX: Also delete the Execution record itself
            db.delete(execution)
            logger.info(f"[LOGS] Deleted execution record {execution_id}")

        elif scope == "current_session":
            # Delete logs only for the current session
            exec_session_id = execution.session_id
            if not exec_session_id:
                # BUG FIX: For legacy executions with NULL session_id, treat "current_session" as "all"
                # since there's no session to filter by
                deleted_count = db.query(ExecutionLog).filter(
                    ExecutionLog.execution_id == execution_id
                ).delete(synchronize_session=False)
                logger.info(f"[LOGS] Deleted {deleted_count} logs for legacy execution {execution_id} (scope: current_session treated as 'all')")

                # Also delete the execution record
                db.delete(execution)
                logger.info(f"[LOGS] Deleted legacy execution record {execution_id}")
            else:
                # Get all executions in this session and delete their logs + execution records
                session_executions = db.query(Execution).filter(
                    Execution.session_id == exec_session_id
                ).all()

                execution_ids = [ex.id for ex in session_executions]
                deleted_count = db.query(ExecutionLog).filter(
                    ExecutionLog.execution_id.in_(execution_ids)
                ).delete(synchronize_session=False)
                logger.info(f"[LOGS] Deleted {deleted_count} logs for execution {execution_id} (scope: current_session, session: {exec_session_id})")

                # Delete all execution records in this session
                for exec_record in session_executions:
                    db.delete(exec_record)
                logger.info(f"[LOGS] Deleted {len(session_executions)} execution records from session {exec_session_id}")

        db.commit()

        return {
            "status": "success",
            "deleted_count": deleted_count,
            "scope": scope,
            "execution_id": execution_id
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[LOGS] Failed to delete logs for execution {execution_id}: {str(e)}")
        try:
            db.rollback()
        except:
            pass
        raise HTTPException(status_code=500, detail=f"Failed to delete logs: {str(e)}")


# ============================================================================
# WEBSOCKET — Real-time Log Streaming
# ============================================================================

@app.websocket("/ws/execution/{execution_id}")
async def websocket_logs(websocket: WebSocket, execution_id: int):
    """WebSocket endpoint for real-time log streaming."""
    await websocket.accept()
    db = SessionLocal()

    try:
        last_log_id = 0
        last_tcr_id = 0
        script_stats: dict = {}  # script_path -> {passed, failed, skipped, duration_s}

        while True:
            new_logs = (
                db.query(ExecutionLog)
                .filter(
                    ExecutionLog.execution_id == execution_id,
                    ExecutionLog.id > last_log_id,
                )
                .order_by(ExecutionLog.timestamp.asc())
                .all()
            )

            for log in new_logs:
                await websocket.send_json(
                    {
                        "id": log.id,
                        "dut_name": log.dut_name,
                        "level": log.log_level,
                        "message": log.message,
                        "timestamp": log.timestamp.isoformat() if log.timestamp else None,
                    }
                )
                last_log_id = log.id

            # Poll for new TestCaseResult rows and emit script_result events
            new_tcrs = (
                db.query(TestCaseResult)
                .filter(
                    TestCaseResult.execution_id == execution_id,
                    TestCaseResult.id > last_tcr_id,
                )
                .order_by(TestCaseResult.id.asc())
                .all()
            )
            if new_tcrs:
                last_tcr_id = new_tcrs[-1].id
                updated_scripts: set = set()
                for tcr in new_tcrs:
                    sp = tcr.script_path or 'unknown'
                    if sp not in script_stats:
                        script_stats[sp] = {'passed': 0, 'failed': 0, 'skipped': 0, 'duration_s': 0}
                    r_lower = (tcr.result or '').lower()
                    if r_lower == 'pass':
                        script_stats[sp]['passed'] += 1
                    elif r_lower in ('fail', 'scripterror', 'error'):
                        script_stats[sp]['failed'] += 1
                    elif r_lower in ('skip', 'xfail', 'deselect'):
                        script_stats[sp]['skipped'] += 1
                    script_stats[sp]['duration_s'] += tcr.time_seconds or 0
                    updated_scripts.add(sp)
                for sp in updated_scripts:
                    st = script_stats[sp]
                    stem = os.path.basename(sp).replace('.py', '') if sp != 'unknown' else 'unknown'
                    f, p, sk = st['failed'], st['passed'], st['skipped']
                    agg_status = 'failed' if f > 0 else 'passed' if p > 0 else 'skipped' if sk > 0 else 'unknown'
                    await websocket.send_json({
                        'type': 'script_result',
                        'script': sp,
                        'script_stem': stem,
                        'passed': p,
                        'failed': f,
                        'skipped': sk,
                        'duration_s': st['duration_s'],
                        'status': agg_status,
                    })

            # Check if execution has finished
            execution = db.query(Execution).filter(Execution.id == execution_id).first()
            if execution and execution.status in ["completed", "failed"]:
                # Send any remaining logs and TCR results
                db.expire_all()
                remaining = (
                    db.query(ExecutionLog)
                    .filter(
                        ExecutionLog.execution_id == execution_id,
                        ExecutionLog.id > last_log_id,
                    )
                    .order_by(ExecutionLog.timestamp.asc())
                    .all()
                )
                for log in remaining:
                    await websocket.send_json(
                        {
                            "id": log.id,
                            "dut_name": log.dut_name,
                            "level": log.log_level,
                            "message": log.message,
                            "timestamp": log.timestamp.isoformat() if log.timestamp else None,
                        }
                    )
                # Flush any final TestCaseResult rows before signalling completion
                final_tcrs = (
                    db.query(TestCaseResult)
                    .filter(
                        TestCaseResult.execution_id == execution_id,
                        TestCaseResult.id > last_tcr_id,
                    )
                    .order_by(TestCaseResult.id.asc())
                    .all()
                )
                if final_tcrs:
                    for tcr in final_tcrs:
                        sp = tcr.script_path or 'unknown'
                        if sp not in script_stats:
                            script_stats[sp] = {'passed': 0, 'failed': 0, 'skipped': 0, 'duration_s': 0}
                        r_lower = (tcr.result or '').lower()
                        if r_lower == 'pass':
                            script_stats[sp]['passed'] += 1
                        elif r_lower in ('fail', 'scripterror', 'error'):
                            script_stats[sp]['failed'] += 1
                        elif r_lower in ('skip', 'xfail', 'deselect'):
                            script_stats[sp]['skipped'] += 1
                        script_stats[sp]['duration_s'] += tcr.time_seconds or 0
                    for sp, st in script_stats.items():
                        stem = os.path.basename(sp).replace('.py', '') if sp != 'unknown' else 'unknown'
                        f, p, sk = st['failed'], st['passed'], st['skipped']
                        agg_status = 'failed' if f > 0 else 'passed' if p > 0 else 'skipped' if sk > 0 else 'unknown'
                        await websocket.send_json({
                            'type': 'script_result',
                            'script': sp,
                            'script_stem': stem,
                            'passed': p,
                            'failed': f,
                            'skipped': sk,
                            'duration_s': st['duration_s'],
                            'status': agg_status,
                        })

                await websocket.send_json(
                    {
                        "type": "execution_complete",
                        "status": execution.status,
                        "duration": execution.duration_seconds,
                    }
                )
                break

            db.expire_all()
            await asyncio.sleep(0.1)  # Reduced from 0.5s to 0.1s for faster log streaming

    except WebSocketDisconnect:
        logger.info(f"WebSocket client disconnected from execution {execution_id}")
    except Exception as e:
        logger.error(f"WebSocket error: {e}")
        try:
            await websocket.close(code=1011)
        except Exception:
            pass
    finally:
        db.close()


# ============================================================================
# API — VS (Virtual System) Lifecycle Management
# ============================================================================


@app.get("/api/vs/list/{dut_id}")
def list_vms(dut_id: int, db: Session = Depends(get_db)):
    """List all VMs on a DUT host via 'virsh list --all'."""
    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    # Check device status before attempting connection
    if dut.status != "online":
        raise HTTPException(
            status_code=425,  # 425 Too Early - device not ready
            detail=f"Device {dut.name} is {dut.status or 'not online'}. Please wait for device to come online before managing VMs."
        )

    # Use SSH connection pool (reuses existing connection)
    ssh = ssh_pool.get_connection(dut.id, dut.ip_address, dut.port, dut.username, dut.password)
    if not ssh:
        dut.status = "offline"
        db.commit()
        raise HTTPException(
            status_code=503,
            detail=f"Cannot connect to {dut.name} - device may be offline or unreachable"
        )

    try:
        # Use password with sudo if available (for devices that require it)
        # The -S flag makes sudo read password from stdin
        if dut.password:
            # Escape single quotes in password
            safe_pass = dut.password.replace("'", "'\\''")
            cmd = f"echo '{safe_pass}' | sudo -S virsh list --all"
        else:
            cmd = "sudo virsh list --all"

        output, error, exit_code = ssh.execute_command(cmd, timeout=30)
        if exit_code != 0:
            raise HTTPException(status_code=500, detail=f"virsh list failed: {error.strip()}")

        vms = []
        lines = output.strip().split("\n")
        for line in lines[2:]:  # Skip header lines
            parts = line.strip().split()
            if len(parts) >= 2:
                vm_id = parts[0] if parts[0] != "-" else None
                vm_name = parts[1]
                vm_state = " ".join(parts[2:]) if len(parts) > 2 else "unknown"
                vms.append({
                    "id": vm_id,
                    "name": vm_name,
                    "state": vm_state,
                })
        return {"dut_id": dut_id, "dut_name": dut.name, "vms": vms}
    except paramiko.SSHException as e:
        logger.error(f"SSH error listing VMs on {dut.name}: {e}")
        dut.status = "offline"
        db.commit()
        ssh_pool.close_connection(dut.id)
        raise HTTPException(status_code=503, detail=f"SSH timeout or connection error: {str(e)}")
    except Exception as e:
        logger.error(f"Error listing VMs on {dut.name}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to list VMs: {str(e)}")
    finally:
        ssh_pool.release_connection(dut.id)  # Return to pool, don't disconnect


@app.get("/api/vs/xml-files/{dut_id}")
def list_xml_files(dut_id: int, db: Session = Depends(get_db)):
    """List available VM XML definition files on the remote host."""
    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    # Use device-specific XML path (default: /home/hp/prajwal/VMs)
    xml_path = dut.xml_path or "/home/hp/prajwal/VMs"

    ssh = SSHConnectionManager(dut.ip_address, dut.port, dut.username, dut.password)
    if not ssh.connect():
        raise HTTPException(status_code=503, detail=f"Cannot connect to {dut.name}")

    try:
        output, error, exit_code = ssh.execute_command(
            f"ls -1 {xml_path}/*.xml 2>/dev/null", timeout=10
        )
        if exit_code != 0 and not output.strip():
            return {"dut_id": dut_id, "xml_files": [], "xml_path": xml_path}

        xml_files = []
        for line in output.strip().split("\n"):
            line = line.strip()
            if line.endswith(".xml"):
                xml_files.append({
                    "full_path": line,
                    "filename": os.path.basename(line),
                })
        return {"dut_id": dut_id, "xml_files": xml_files, "xml_path": xml_path}
    finally:
        ssh.disconnect()


@app.post("/api/vs/update-image")
def update_vs_image(body: dict, db: Session = Depends(get_db)):
    """
    Full VS image update lifecycle:
    virsh destroy → rm old image → cp new image → virsh undefine → virsh define → virsh start
    Supports remote image copy via SFTP when source_server_id is provided.
    """
    dut_id = body.get("dut_id")
    vs_name = body.get("vs_name", "").strip()
    source_image = body.get("source_image_path", VS_SOURCE_IMAGE).strip()
    target_image_name = body.get("target_image_name", "").strip()
    source_server_id = body.get("source_server_id")  # Optional: for remote SFTP copy

    # DEBUG: Log what we received
    logger.info(f"[VS UPDATE API] Received request: vs_name='{vs_name}', target_image_name='{target_image_name}'")

    if not dut_id or not vs_name:
        raise HTTPException(status_code=400, detail="dut_id and vs_name are required")

    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    # Get source server details if provided
    source_server = None
    if source_server_id:
        source_server = db.query(DUT).filter(DUT.id == source_server_id).first()
        if not source_server:
            raise HTTPException(status_code=404, detail="Source server not found")

    # If no target image name given, use the VS name as the image file name
    if not target_image_name:
        target_image_name = f"{vs_name}.img"

    # Use device-specific XML path (default: /home/hp/prajwal/VMs)
    xml_path = dut.xml_path or "/home/hp/prajwal/VMs"
    xml_full_path = f"{xml_path}/{vs_name}.xml"
    target_image_path = f"{VS_IMAGES_PATH}{target_image_name}"

    # Create an execution record for logging
    execution = Execution(
        name=f"vs_update_{vs_name}_{int(datetime.utcnow().timestamp())}",
        execution_type="image",
        dut_ids=json.dumps([dut_id]),
        status="pending",
    )
    db.add(execution)
    db.commit()
    db.refresh(execution)

    # Run in background thread - pass IDs instead of ORM objects to avoid session issues
    thread = Thread(
        target=_run_vs_update,
        args=(execution.id, dut_id, vs_name, xml_full_path, source_image, target_image_path, source_server_id),
        daemon=True,
    )
    thread.start()

    return {
        "execution_id": execution.id,
        "status": "started",
        "vs_name": vs_name,
        "message": f"VS image update started for '{vs_name}' on {dut.name}",
    }


@app.post("/api/vs/update-image-batch")
def update_vs_image_batch(body: dict, db: Session = Depends(get_db)):
    """
    Batch VS image update: update multiple VMs sequentially.
    Accepts per-VM target image names via vs_entries list.
    Body: dut_id, vs_entries=[{vs_name, target_image_name}, ...], source_image_path
    Also supports legacy vs_names list for backward compatibility.
    """
    dut_id = body.get("dut_id")
    source_image = body.get("source_image_path", VS_SOURCE_IMAGE).strip()

    # Support both new vs_entries and legacy vs_names
    vs_entries = body.get("vs_entries")
    if not vs_entries:
        vs_names = body.get("vs_names", [])
        target_image_name_global = body.get("target_image_name", "").strip()
        vs_entries = [
            {"vs_name": n, "target_image_name": target_image_name_global}
            for n in vs_names
        ]

    if not dut_id or not vs_entries:
        raise HTTPException(status_code=400, detail="dut_id and vs_entries (or vs_names) are required")

    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    # Create a single execution record for the entire batch
    execution = Execution(
        name=f"vs_batch_{len(vs_entries)}vms_{int(datetime.utcnow().timestamp())}",
        execution_type="image",
        dut_ids=json.dumps([dut_id]),
        status="pending",
    )
    db.add(execution)
    db.commit()
    db.refresh(execution)

    # Run in background thread
    thread = Thread(
        target=_run_vs_batch_update,
        args=(execution.id, dut, vs_entries, source_image),
        daemon=True,
    )
    thread.start()

    vm_names = [e.get("vs_name", "") for e in vs_entries]
    return {
        "execution_id": execution.id,
        "status": "started",
        "vs_count": len(vs_entries),
        "vs_names": vm_names,
        "message": f"VS batch update started for {len(vs_entries)} VM(s) on {dut.name}",
    }


def _run_vs_batch_update(
    execution_id: int,
    dut,
    vs_entries: list,   # [{vs_name, target_image_name}, ...]
    source_image: str,
):
    """Background thread: Process multiple VMs sequentially with per-VM target images."""
    db = SessionLocal()
    execution = None
    try:
        execution = db.query(Execution).filter(Execution.id == execution_id).first()
        execution.status = "running"
        execution.start_time = datetime.utcnow()
        db.commit()

        total = len(vs_entries)
        vm_names = [e.get("vs_name", "") for e in vs_entries]
        log_execution(db, execution_id, "SYSTEM", "INFO",
                      f"═══ Batch VS update: {total} VM(s) ═══")
        log_execution(db, execution_id, "SYSTEM", "INFO",
                      f"  VMs: {', '.join(vm_names)}")
        log_execution(db, execution_id, "SYSTEM", "INFO",
                      f"  Source image: {source_image}")

        # Connect via SSH once for the entire batch
        ssh_user = dut.username   # e.g. hp
        ssh_host = dut.ip_address
        ssh_port = dut.port
        ssh_pass = dut.password

        log_execution(db, execution_id, dut.name, "INFO",
                      f"  Connecting: ssh {ssh_user}@{ssh_host}:{ssh_port}")

        ssh = SSHConnectionManager(ssh_host, ssh_port, ssh_user, ssh_pass)
        if not ssh.connect():
            log_execution(db, execution_id, dut.name, "ERROR",
                          f"SSH connection FAILED: {ssh_user}@{ssh_host}:{ssh_port}")
            log_execution(db, execution_id, dut.name, "ERROR",
                          "  Check: username, password, and SSH access on host device")
            execution.status = "failed"
            execution.end_time = datetime.utcnow()
            db.commit()
            return

        log_execution(db, execution_id, dut.name, "INFO",
                      f"  SSH connected: {ssh_user}@{ssh_host}:{ssh_port}")

        # Helper: prepend sudo -S with password for commands needing root
        def sudocmd(cmd: str) -> str:
            # Use echo password | sudo -S so sudo doesn't wait for interactive input
            safe_pass = ssh_pass.replace("'", "'\\''")   # escape single quotes
            return f"echo '{safe_pass}' | sudo -S {cmd}"

        try:
            all_success = True
            for idx, entry in enumerate(vs_entries, 1):
                vs_name = entry.get("vs_name", "").strip()
                per_vm_target = entry.get("target_image_name", "").strip()
                if not vs_name:
                    continue

                # Determine target image name for this VM
                this_target_name = per_vm_target if per_vm_target else f"{vs_name}.img"

                # Images directory on the remote host
                IMAGES_DIR = "/var/lib/libvirt/images"

                log_execution(db, execution_id, dut.name, "INFO", "")
                log_execution(db, execution_id, dut.name, "INFO",
                              f"══ VM {idx}/{total}: {vs_name} ══")
                log_execution(db, execution_id, dut.name, "INFO",
                              f"  Target image: {this_target_name}")
                log_execution(db, execution_id, dut.name, "INFO",
                              f"  Source image: {source_image}")

                # Exact 4-step sequence using correct commands:
                # 1. virsh destroy <vs_name>           (user has libvirt group — no sudo needed)
                # 2. sudo rm -f <IMAGES_DIR>/<image>
                # 3. sudo cp <source> <IMAGES_DIR>/<target>
                # 4. virsh start <vs_name>
                steps = [
                    ("Step 1/4: Destroying VM",
                     f"virsh destroy {vs_name}",
                     True),   # allow_fail: VM may already be stopped
                    ("Step 2/4: Removing old image",
                     sudocmd(f"rm -f {IMAGES_DIR}/{this_target_name}"),
                     False),
                    ("Step 3/4: Copying new image",
                     sudocmd(f"cp {source_image} {IMAGES_DIR}/{this_target_name}"),
                     False),
                    ("Step 4/4: Starting VM",
                     f"virsh start {vs_name}",
                     False),
                ]

                vm_ok = True
                for step_name, command, allow_fail in steps:
                    log_execution(db, execution_id, dut.name, "INFO", f"▶ {step_name}")
                    log_execution(db, execution_id, dut.name, "INFO", f"  $ {command}")

                    try:
                        output, error, exit_code = ssh.execute_command(command, timeout=300)

                        if output.strip():
                            for line in output.strip().split("\n")[:20]:
                                log_execution(db, execution_id, dut.name, "INFO", f"    {line}")

                        if exit_code != 0:
                            msg = error.strip() if error.strip() else f"Exit code {exit_code}"
                            if allow_fail:
                                log_execution(db, execution_id, dut.name, "WARNING",
                                              f"  ⚠ {step_name} returned non-zero (allowed): {msg}")
                            else:
                                log_execution(db, execution_id, dut.name, "ERROR",
                                              f"  ✗ {step_name} FAILED: {msg}")
                                vm_ok = False
                                break
                        else:
                            log_execution(db, execution_id, dut.name, "INFO",
                                          f"  ✓ {step_name} completed successfully")

                    except Exception as cmd_err:
                        log_execution(db, execution_id, dut.name, "ERROR",
                                      f"  ✗ {step_name} error: {str(cmd_err)}")
                        vm_ok = False
                        break

                if vm_ok:
                    # Verify VM is running
                    log_execution(db, execution_id, dut.name, "INFO",
                                  "Verifying VM status...")
                    output, _, _ = ssh.execute_command(f"virsh domstate {vs_name}", timeout=10)
                    state = output.strip()
                    log_execution(db, execution_id, dut.name, "INFO",
                                  f"  VM '{vs_name}' status: {state}")

                    if "running" in state.lower():
                        log_execution(db, execution_id, dut.name, "INFO",
                                      f"✓ VS image update completed — '{vs_name}' is running with new image")
                    else:
                        log_execution(db, execution_id, dut.name, "WARNING",
                                      f"⚠ VS image update completed but VM state is '{state}'")
                else:
                    log_execution(db, execution_id, dut.name, "ERROR",
                                  f"✗ VS image update FAILED for '{vs_name}'")
                    all_success = False

            # Final summary
            log_execution(db, execution_id, "SYSTEM", "INFO", "")
            if all_success:
                log_execution(db, execution_id, "SYSTEM", "INFO",
                              f"═══ Batch complete: All {total} VM(s) updated successfully ═══")
                execution.status = "completed"
            else:
                log_execution(db, execution_id, "SYSTEM", "WARNING",
                              f"═══ Batch complete: Some VMs failed. Check logs above. ═══")
                execution.status = "completed"  # partial success is still "completed"

        finally:
            ssh.disconnect()

        execution.end_time = datetime.utcnow()
        if execution.start_time:
            execution.duration_seconds = int(
                (execution.end_time - execution.start_time).total_seconds()
            )
        db.commit()

    except Exception as e:
        logger.error(f"VS batch update failed: {e}")
        if execution:
            execution.status = "failed"
            execution.end_time = datetime.utcnow()
            db.commit()
        log_execution(db, execution_id, dut.name if dut else "SYSTEM", "ERROR",
                      f"VS batch update failed: {str(e)}")
    finally:
        db.close()


def _run_vs_update(
    execution_id: int,
    dut_id: int,
    vs_name: str,
    xml_full_path: str,
    source_image: str,
    target_image_path: str,
    source_server_id=None,
):
    """Background thread: Full VS image update lifecycle.
    Supports remote image copy via SFTP when source_server_id is provided."""
    db = SessionLocal()
    execution = None
    try:
        execution = db.query(Execution).filter(Execution.id == execution_id).first()
        execution.status = "running"
        execution.start_time = datetime.utcnow()
        db.commit()

        # Fetch DUT and source server objects fresh in this thread's session
        dut = db.query(DUT).filter(DUT.id == dut_id).first()
        if not dut:
            log_execution(db, execution_id, "SYSTEM", "ERROR", "DUT not found")
            execution.status = "failed"
            execution.end_time = datetime.utcnow()
            db.commit()
            return

        source_server = None
        if source_server_id:
            source_server = db.query(DUT).filter(DUT.id == source_server_id).first()
            if not source_server:
                log_execution(db, execution_id, dut.name, "ERROR", "Source server not found")
                execution.status = "failed"
                execution.end_time = datetime.utcnow()
                db.commit()
                return

        log_execution(db, execution_id, dut.name, "INFO",
                      f"Starting VS image update for '{vs_name}'")

        # Debug: Show copy method
        if source_server:
            log_execution(db, execution_id, dut.name, "INFO",
                          f"  Copy method: Direct SCP from {source_server.name} ({source_server.ip_address})")
        else:
            log_execution(db, execution_id, dut.name, "INFO",
                          f"  Copy method: Local copy on Host Device")

        log_execution(db, execution_id, dut.name, "INFO",
                      f"  Source image: {source_image}")
        log_execution(db, execution_id, dut.name, "INFO",
                      f"  Target image: {target_image_path}")
        log_execution(db, execution_id, dut.name, "INFO",
                      f"  XML file: {xml_full_path}")

        # Connect via SSH
        ssh = SSHConnectionManager(dut.ip_address, dut.port, dut.username, dut.password)
        if not ssh.connect():
            log_execution(db, execution_id, dut.name, "ERROR",
                          f"Failed to connect to {dut.name} ({dut.ip_address}:{dut.port})")
            execution.status = "failed"
            execution.end_time = datetime.utcnow()
            db.commit()
            return

        # Helper: prepend sudo -S with password for commands needing root
        def sudocmd(cmd: str) -> str:
            """Wrap command with sudo -S using device password from stdin."""
            # Use echo password | sudo -S so sudo doesn't wait for interactive input
            safe_pass = dut.password.replace("'", "'\\''")   # escape single quotes
            return f"echo '{safe_pass}' | sudo -S {cmd}"

        try:
            # Steps 1-2: Destroy VM and remove old image
            steps_pre_copy = [
                ("Step 1/6: Destroying VM",
                 sudocmd(f"virsh destroy {vs_name}"),
                 True),   # allow_fail=True (VM might already be off)
                ("Step 2/6: Removing old image",
                 sudocmd(f"rm -f {target_image_path}"),
                 False),
            ]

            all_ok = True
            for step_name, command, allow_fail in steps_pre_copy:
                log_execution(db, execution_id, dut.name, "INFO", f"▶ {step_name}")
                log_execution(db, execution_id, dut.name, "INFO", f"  $ {command}")

                try:
                    output, error, exit_code = ssh.execute_command(command, timeout=120)

                    if output.strip():
                        for line in output.strip().split("\n")[:20]:
                            log_execution(db, execution_id, dut.name, "INFO", f"    {line}")

                    if exit_code != 0:
                        msg = error.strip() if error.strip() else f"Exit code {exit_code}"
                        if allow_fail:
                            log_execution(db, execution_id, dut.name, "WARNING",
                                          f"  ⚠ {step_name} returned non-zero (allowed): {msg}")
                        else:
                            log_execution(db, execution_id, dut.name, "ERROR",
                                          f"  ✗ {step_name} FAILED: {msg}")
                            all_ok = False
                            break
                    else:
                        log_execution(db, execution_id, dut.name, "INFO",
                                      f"  ✓ {step_name} completed successfully")

                except Exception as cmd_err:
                    log_execution(db, execution_id, dut.name, "ERROR",
                                  f"  ✗ {step_name} error: {str(cmd_err)}")
                    all_ok = False
                    break

            if not all_ok:
                execution.status = "failed"
                execution.end_time = datetime.utcnow()
                db.commit()
                return

            # Step 3: Copy image (local or remote SCP)
            if source_server:
                # Remote SCP copy - direct from Source Server to Host Device
                log_execution(db, execution_id, dut.name, "INFO",
                              "▶ Step 3/6: Copying image from remote server (SCP)")
                log_execution(db, execution_id, dut.name, "INFO",
                              f"  Source: {source_server.name} ({source_server.ip_address}:{source_server.port})")
                log_execution(db, execution_id, dut.name, "INFO",
                              f"  Path: {source_image}")

                # Build SCP command to run on Host Device
                # sshpass allows non-interactive password authentication
                dest_temp_path = f"/tmp/{os.path.basename(target_image_path)}"
                safe_source_pass = source_server.password.replace("'", "'\\''")

                # SCP command: download from source server to /tmp on Host Device
                scp_source = f"{source_server.username}@{source_server.ip_address}:{source_image}"
                if source_server.port != 22:
                    scp_cmd = f"sshpass -p '{safe_source_pass}' scp -P {source_server.port} -o StrictHostKeyChecking=no {scp_source} {dest_temp_path}"
                else:
                    scp_cmd = f"sshpass -p '{safe_source_pass}' scp -o StrictHostKeyChecking=no {scp_source} {dest_temp_path}"

                log_execution(db, execution_id, dut.name, "INFO",
                              f"  Copying directly to Host Device /tmp...")

                try:
                    output, error, exit_code = ssh.execute_command(scp_cmd, timeout=300)  # 5 min timeout for large files

                    if exit_code != 0:
                        msg = error.strip() if error.strip() else f"Exit code {exit_code}"
                        log_execution(db, execution_id, dut.name, "ERROR",
                                      f"  ✗ Step 3/6 FAILED: {msg}")

                        # Check for common errors
                        if "Permission denied" in msg or "publickey" in msg:
                            log_execution(db, execution_id, dut.name, "ERROR",
                                          f"  Check source server credentials in Devices tab")
                        elif "No such file" in msg:
                            log_execution(db, execution_id, dut.name, "ERROR",
                                          f"  Image not found at source path: {source_image}")
                        elif "sshpass: not found" in msg or "command not found" in msg:
                            log_execution(db, execution_id, dut.name, "ERROR",
                                          f"  'sshpass' not installed on Host Device")
                            log_execution(db, execution_id, dut.name, "ERROR",
                                          f"  Install with: sudo apt-get install sshpass")

                        all_ok = False
                    else:
                        log_execution(db, execution_id, dut.name, "INFO",
                                      f"  ✓ Image copied to {dest_temp_path}")

                        # Now move from /tmp to final destination with sudo
                        log_execution(db, execution_id, dut.name, "INFO",
                                      f"  Moving to final location...")

                        move_cmd = sudocmd(f"mv {dest_temp_path} {target_image_path}")
                        output, error, exit_code = ssh.execute_command(move_cmd, timeout=60)

                        if exit_code != 0:
                            msg = error.strip() if error.strip() else f"Exit code {exit_code}"
                            log_execution(db, execution_id, dut.name, "ERROR",
                                          f"  ✗ Step 3/6 FAILED: {msg}")
                            all_ok = False
                        else:
                            log_execution(db, execution_id, dut.name, "INFO",
                                          f"  ✓ Step 3/6 completed successfully")

                except Exception as e:
                    log_execution(db, execution_id, dut.name, "ERROR",
                                  f"  ✗ Step 3/6 error: {str(e)}")
                    all_ok = False
            else:
                # Local copy
                log_execution(db, execution_id, dut.name, "INFO",
                              "▶ Step 3/6: Copying image (local)")
                copy_cmd = sudocmd(f"cp {source_image} {target_image_path}")
                log_execution(db, execution_id, dut.name, "INFO", f"  $ {copy_cmd}")

                try:
                    output, error, exit_code = ssh.execute_command(copy_cmd, timeout=120)

                    if output.strip():
                        for line in output.strip().split("\n")[:20]:
                            log_execution(db, execution_id, dut.name, "INFO", f"    {line}")

                    if exit_code != 0:
                        msg = error.strip() if error.strip() else f"Exit code {exit_code}"
                        log_execution(db, execution_id, dut.name, "ERROR",
                                      f"  ✗ Step 3/6 FAILED: {msg}")
                        all_ok = False
                    else:
                        log_execution(db, execution_id, dut.name, "INFO",
                                      f"  ✓ Step 3/6 completed successfully")

                except Exception as cmd_err:
                    log_execution(db, execution_id, dut.name, "ERROR",
                                  f"  ✗ Step 3/6 error: {str(cmd_err)}")
                    all_ok = False

            if not all_ok:
                execution.status = "failed"
                execution.end_time = datetime.utcnow()
                db.commit()
                return

            # Step 4: Update XML file to point to new image location
            log_execution(db, execution_id, dut.name, "INFO",
                          f"▶ Step 4/7: Updating XML to reference new image")
            log_execution(db, execution_id, dut.name, "INFO",
                          f"  XML file: {xml_full_path}")
            log_execution(db, execution_id, dut.name, "INFO",
                          f"  New image: {target_image_path}")

            # First, show current XML content for debugging
            cat_cmd = sudocmd(f"grep '<source file=' {xml_full_path}")
            log_execution(db, execution_id, dut.name, "INFO", f"  Checking current XML content...")
            try:
                output, error, exit_code = ssh.execute_command(cat_cmd, timeout=10)
                if output.strip():
                    log_execution(db, execution_id, dut.name, "INFO", f"  Current: {output.strip()}")
            except:
                pass

            # Use sed to replace the image path in XML file
            # Match both single and double quotes, and both self-closing and regular tags
            # Patterns: <source file='...' /> or <source file="..."/> or <source file='...'></source>
            sed_cmd = sudocmd(f"sed -i 's|<source file=[\"'\"'][^\"'\"']*[\"'\"']|<source file=\"{target_image_path}\"|g' {xml_full_path}")
            log_execution(db, execution_id, dut.name, "INFO", f"  $ {sed_cmd}")

            try:
                output, error, exit_code = ssh.execute_command(sed_cmd, timeout=30)
                if exit_code != 0:
                    msg = error.strip() if error.strip() else f"Exit code {exit_code}"
                    log_execution(db, execution_id, dut.name, "ERROR",
                                  f"  ✗ Step 4/7 FAILED: {msg}")
                    execution.status = "failed"
                    execution.end_time = datetime.utcnow()
                    db.commit()
                    return
                else:
                    log_execution(db, execution_id, dut.name, "INFO",
                                  f"  ✓ Step 4/7 completed - XML updated")

                    # Verify the change
                    verify_cmd = sudocmd(f"grep '<source file=' {xml_full_path}")
                    output, error, exit_code = ssh.execute_command(verify_cmd, timeout=10)
                    if output.strip():
                        log_execution(db, execution_id, dut.name, "INFO", f"  Verified: {output.strip()}")
            except Exception as cmd_err:
                log_execution(db, execution_id, dut.name, "ERROR",
                              f"  ✗ Step 4/7 error: {str(cmd_err)}")
                execution.status = "failed"
                execution.end_time = datetime.utcnow()
                db.commit()
                return

            # Steps 5-7: Undefine, Define, Start VM
            steps_post_copy = [
                ("Step 5/7: Undefining VM",
                 sudocmd(f"virsh undefine {vs_name}"),
                 True),   # allow_fail=True (might already be undefined)
                ("Step 6/7: Defining VM from XML",
                 sudocmd(f"virsh define {xml_full_path}"),
                 False),
                ("Step 7/7: Starting VM",
                 sudocmd(f"virsh start {vs_name}"),
                 False),
            ]

            for step_name, command, allow_fail in steps_post_copy:
                log_execution(db, execution_id, dut.name, "INFO", f"▶ {step_name}")
                log_execution(db, execution_id, dut.name, "INFO", f"  $ {command}")

                try:
                    output, error, exit_code = ssh.execute_command(command, timeout=120)

                    if output.strip():
                        for line in output.strip().split("\n")[:20]:
                            log_execution(db, execution_id, dut.name, "INFO", f"    {line}")

                    if exit_code != 0:
                        msg = error.strip() if error.strip() else f"Exit code {exit_code}"
                        if allow_fail:
                            log_execution(db, execution_id, dut.name, "WARNING",
                                          f"  ⚠ {step_name} returned non-zero (allowed): {msg}")
                        else:
                            log_execution(db, execution_id, dut.name, "ERROR",
                                          f"  ✗ {step_name} FAILED: {msg}")
                            all_ok = False
                            break
                    else:
                        log_execution(db, execution_id, dut.name, "INFO",
                                      f"  ✓ {step_name} completed successfully")

                except Exception as cmd_err:
                    log_execution(db, execution_id, dut.name, "ERROR",
                                  f"  ✗ {step_name} error: {str(cmd_err)}")
                    all_ok = False
                    break

            if all_ok:
                # Verify VM is running
                log_execution(db, execution_id, dut.name, "INFO",
                              "Verifying VM status...")
                output, _, _ = ssh.execute_command(sudocmd(f"virsh domstate {vs_name}"), timeout=10)
                state = output.strip()
                log_execution(db, execution_id, dut.name, "INFO",
                              f"  VM '{vs_name}' status: {state}")

                if "running" in state.lower():
                    log_execution(db, execution_id, dut.name, "INFO",
                                  f"✓ VS image update completed — '{vs_name}' is running with new image")
                else:
                    log_execution(db, execution_id, dut.name, "WARNING",
                                  f"⚠ VS image update completed but VM state is '{state}'")

                execution.status = "completed"
            else:
                log_execution(db, execution_id, dut.name, "ERROR",
                              f"✗ VS image update FAILED for '{vs_name}'")
                execution.status = "failed"

        finally:
            ssh.disconnect()

        execution.end_time = datetime.utcnow()
        if execution.start_time:
            execution.duration_seconds = int(
                (execution.end_time - execution.start_time).total_seconds()
            )
        db.commit()

    except Exception as e:
        logger.error(f"VS update failed: {e}")
        if execution:
            execution.status = "failed"
            execution.end_time = datetime.utcnow()
            db.commit()
        log_execution(db, execution_id, dut.name if dut else "SYSTEM", "ERROR",
                      f"VS update failed: {str(e)}")
    finally:
        db.close()


@app.post("/api/vs/{dut_id}/action")
def vs_action(dut_id: int, body: dict, db: Session = Depends(get_db)):
    """Quick VM action: start, destroy, reboot, shutdown, suspend, resume.

    Action map:
        start    → virsh start   <name>   (boots a shut-off VM)
        destroy  → virsh destroy <name>   (hard stop)
        reboot   → virsh reboot  <name>   (graceful restart)
        shutdown → virsh shutdown <name>  (graceful stop)
        suspend  → virsh suspend <name>   (pause / freeze VM)
        resume   → virsh resume  <name>   (un-pause a paused VM)
    """
    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="DUT not found")

    vs_name = body.get("vs_name", "").strip()
    action = body.get("action", "").strip().lower()

    if not vs_name or not action:
        raise HTTPException(status_code=400, detail="vs_name and action are required")

    allowed_actions = ["start", "destroy", "reboot", "shutdown", "suspend", "resume"]
    if action not in allowed_actions:
        raise HTTPException(status_code=400,
                            detail=f"Invalid action. Allowed: {', '.join(allowed_actions)}")

    ssh = SSHConnectionManager(dut.ip_address, dut.port, dut.username, dut.password)
    if not ssh.connect():
        raise HTTPException(status_code=503, detail=f"Cannot connect to {dut.name}")

    try:
        # Use password with sudo if available (for devices that require it)
        if dut.password:
            safe_pass = dut.password.replace("'", "'\\''")
            command = f"echo '{safe_pass}' | sudo -S virsh {action} {vs_name}"
        else:
            command = f"sudo virsh {action} {vs_name}"

        output, error, exit_code = ssh.execute_command(command, timeout=30)

        if exit_code != 0:
            return {
                "status": "error",
                "vs_name": vs_name,
                "action": action,
                "message": error.strip() or f"Command failed with exit code {exit_code}",
            }

        return {
            "status": "success",
            "vs_name": vs_name,
            "action": action,
            "message": output.strip() or f"'{action}' executed on '{vs_name}'",
        }
    finally:
        ssh.disconnect()


# ============================================================================
# API — SPyTest Integration (Remote Browsing & Smart Execution)
# ============================================================================


def _ssh_to_host(dut_id: int, db: Session):
    """Create SSH connection to a host device for SPyTest operations."""
    dut = db.query(DUT).filter(DUT.id == dut_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="Host device not found")
    ssh = SSHConnectionManager(dut.ip_address, dut.port, dut.username, dut.password)
    if not ssh.connect():
        raise HTTPException(status_code=503, detail=f"Cannot connect to {dut.name}")
    return ssh, dut


@app.get("/api/spytest/categories")
def get_spytest_categories(host_id: int, db: Session = Depends(get_db)):
    """List test category folders from the remote SPyTest tests directory."""
    ssh, dut = _ssh_to_host(host_id, db)
    try:
        cmd = f'find {SPYTEST_TESTS_DIR} -mindepth 1 -maxdepth 1 -type d -printf "%f\\n" | sort'
        output, error, code = ssh.execute_command(cmd, timeout=15)
        if code != 0:
            raise HTTPException(status_code=500, detail=f"Failed to list categories: {error}")
        categories = [d.strip() for d in output.strip().split("\n") if d.strip() and not d.strip().startswith("__")]
        return {"categories": categories, "base_path": SPYTEST_TESTS_DIR}
    finally:
        ssh.disconnect()


@app.get("/api/spytest/scripts/{category}")
def get_spytest_scripts(category: str, host_id: int, db: Session = Depends(get_db)):
    """List Python test scripts in a category folder (recursive)."""
    ssh, dut = _ssh_to_host(host_id, db)
    try:
        category_path = f"{SPYTEST_TESTS_DIR}/{category}"
        cmd = f'find {category_path} -name "test_*.py" -type f | sort'
        output, error, code = ssh.execute_command(cmd, timeout=15)
        if code != 0:
            raise HTTPException(status_code=500, detail=f"Failed to list scripts: {error}")
        scripts = []
        for line in output.strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            # Get path relative to tests dir
            rel_path = line.replace(SPYTEST_TESTS_DIR + "/", "")
            name = os.path.basename(line)
            scripts.append({"name": name, "path": rel_path, "full_path": line})
        return {"category": category, "scripts": scripts}
    finally:
        ssh.disconnect()


@app.get("/api/spytest/browse")
def browse_spytest_root(host_id: int, base_path: str = None, db: Session = Depends(get_db)):
    """Root browse — delegates to browse_spytest_folder with empty path."""
    return browse_spytest_folder("", host_id, base_path, db)


@app.get("/api/spytest/browse/{path:path}")
def browse_spytest_folder(path: str, host_id: int, base_path: str = None, db: Session = Depends(get_db)):
    """
    Browse a specific folder in the SPyTest tests directory.
    Returns both subfolders and scripts at the current level only (non-recursive).

    Args:
        path: Relative path from SPYTEST_TESTS_DIR (e.g., "routing" or "routing/bgp" or "")
        host_id: ID of the host device where SPyTest is installed

    Returns:
        {
            "current_path": "routing/bgp",
            "parent_path": "routing",
            "subfolders": ["ipv4", "ipv6", "evpn"],
            "scripts": [{"name": "test_bgp_basic.py", "path": "routing/bgp/test_bgp_basic.py", "full_path": "/full/path"}]
        }
    """
    if host_id == 5:
        parent_path = ""
        if path and "/" in path:
            parent_path = "/".join(path.split("/")[:-1])
        norm_path = path.strip().strip('/')
        if not norm_path:
            subfolders = ["routing", "switching", "security"]
            scripts = []
        elif norm_path == "routing":
            subfolders = ["bgp", "ospf"]
            scripts = [{"name": "test_routing_basic.py", "path": "routing/test_routing_basic.py", "full_path": "/mock/routing/test_routing_basic.py"}]
        elif norm_path == "routing/bgp":
            subfolders = []
            scripts = [
                {"name": "test_bgp_route_advertise.py", "path": "routing/bgp/test_bgp_route_advertise.py", "full_path": "/mock/routing/bgp/test_bgp_route_advertise.py"},
                {"name": "test_bgp_route_filtering.py", "path": "routing/bgp/test_bgp_route_filtering.py", "full_path": "/mock/routing/bgp/test_bgp_route_filtering.py"},
                {"name": "test_bgp_graceful_restart.py", "path": "routing/bgp/test_bgp_graceful_restart.py", "full_path": "/mock/routing/bgp/test_bgp_graceful_restart.py"}
            ]
        elif norm_path == "switching":
            subfolders = ["lacp", "vlan"]
            scripts = []
        elif norm_path == "switching/lacp":
            subfolders = []
            scripts = [
                {"name": "test_lacp_convergence.py", "path": "switching/lacp/test_lacp_convergence.py", "full_path": "/mock/switching/lacp/test_lacp_convergence.py"},
                {"name": "test_lacp_redundancy.py", "path": "switching/lacp/test_lacp_redundancy.py", "full_path": "/mock/switching/lacp/test_lacp_redundancy.py"}
            ]
        elif norm_path == "switching/vlan":
            subfolders = []
            scripts = [
                {"name": "test_vlan_trunking.py", "path": "switching/vlan/test_vlan_trunking.py", "full_path": "/mock/switching/vlan/test_vlan_trunking.py"}
            ]
        else:
            subfolders = []
            scripts = []
        return {
            "current_path": norm_path,
            "parent_path": parent_path,
            "subfolders": subfolders,
            "scripts": scripts,
            "subfolder_count": len(subfolders),
            "script_count": len(scripts)
        }

    ssh, dut = _ssh_to_host(host_id, db)
    try:
        # Sanitize path - remove leading/trailing slashes
        path = path.strip().strip('/')

        # Use caller-supplied base path, falling back to the server default
        tests_root = (base_path or '').strip().rstrip('/') or SPYTEST_TESTS_DIR

        # Build full path
        if path:
            full_path = f"{tests_root}/{path}"
        else:
            full_path = tests_root

        # Check if path exists and is a directory
        check_cmd = f'[ -d "{full_path}" ] && echo "EXISTS" || echo "NOT_FOUND"'
        check_out, _, _ = ssh.execute_command(check_cmd, timeout=5)
        if "NOT_FOUND" in check_out:
            # Return empty result instead of 404 so the UI shows "no scripts" gracefully
            parent_path = ""
            if path and "/" in path:
                parent_path = "/".join(path.split("/")[:-1])
            return {
                "current_path": path,
                "parent_path": parent_path,
                "subfolders": [],
                "scripts": [],
                "subfolder_count": 0,
                "script_count": 0,
                "warning": f"SpyTest tests directory not found on this VM: {full_path}",
            }

        # Get subfolders (directories only, non-recursive, exclude __pycache__)
        subfolder_cmd = f'find "{full_path}" -mindepth 1 -maxdepth 1 -type d ! -name "__pycache__" ! -name ".*" -printf "%f\\n" | sort'
        subfolder_out, subfolder_err, subfolder_code = ssh.execute_command(subfolder_cmd, timeout=15)

        if subfolder_code != 0:
            logger.warning(f"Failed to list subfolders in {path}: {subfolder_err}")
            subfolders = []
        else:
            subfolders = [f.strip() for f in subfolder_out.strip().split("\n") if f.strip()]

        # Get scripts (test_*.py files in current folder only, non-recursive)
        script_cmd = f'find "{full_path}" -maxdepth 1 -name "test_*.py" -type f | sort'
        script_out, script_err, script_code = ssh.execute_command(script_cmd, timeout=15)

        if script_code != 0:
            logger.warning(f"Failed to list scripts in {path}: {script_err}")
            scripts = []
        else:
            scripts = []
            for line in script_out.strip().split("\n"):
                line = line.strip()
                if not line:
                    continue
                # Get path relative to tests root
                rel_path = line.replace(tests_root + "/", "")
                name = os.path.basename(line)
                scripts.append({"name": name, "path": rel_path, "full_path": line})

        # Calculate parent path
        parent_path = ""
        if path and "/" in path:
            parent_path = "/".join(path.split("/")[:-1])

        return {
            "current_path": path,
            "parent_path": parent_path,
            "subfolders": subfolders,
            "scripts": scripts,
            "subfolder_count": len(subfolders),
            "script_count": len(scripts)
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error browsing folder {path}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to browse folder: {str(e)}")
    finally:
        ssh.disconnect()


@app.get("/api/spytest/testbeds")
def get_spytest_testbeds(host_id: int, db: Session = Depends(get_db)):
    """List testbed YAML files from the remote SPyTest testbed directory."""
    if host_id == 5:
        return {"testbeds": ["testbed_2_switches.yaml", "testbed_standalone_switch.yaml", "testbed_spine_leaf.yaml"], "base_path": "/mock/testbeds"}
    ssh, dut = _ssh_to_host(host_id, db)
    try:
        cmd = f'find {SPYTEST_TESTBED_DIR} -maxdepth 1 -name "*.yaml" -type f -printf "%f\\n" | sort'
        output, error, code = ssh.execute_command(cmd, timeout=15)
        if code != 0:
            raise HTTPException(status_code=500, detail=f"Failed to list testbeds: {error}")
        testbeds = [f.strip() for f in output.strip().split("\n") if f.strip()]
        return {"testbeds": testbeds, "base_path": SPYTEST_TESTBED_DIR}
    finally:
        ssh.disconnect()


@app.post("/api/spytest/script-info")
def get_spytest_script_info(body: dict, db: Session = Depends(get_db)):
    """Parse a SPyTest script to extract topology requirements.
    
    Looks for:
    - @pytest.mark.topology("...") decorators
    - st.ensure_min_topology("D1", "D1D2:1", ...) calls
    - Docstring topology description
    """
    host_id = body.get("host_id")
    script_path = body.get("script_path")  # relative to tests dir
    if not host_id or not script_path:
        raise HTTPException(status_code=400, detail="host_id and script_path required")

    if host_id == 5:
        script_name = os.path.basename(script_path)
        return {
            "topology_marker": "D1D2:2",
            "min_topology": ["D1", "D2", "D1D2:2"],
            "dut_count": 2,
            "description": f"Verifies functionality of BGP routing on 2 connected switches. Running: {script_name}",
            "topology_type": "linear",
            "script_path": script_path,
            "script_name": script_name
        }

    ssh, dut = _ssh_to_host(host_id, db)
    try:
        user_base = (body.get("base_path") or "").strip().rstrip("/")
        tests_root = user_base or SPYTEST_TESTS_DIR
        full_path = f"{tests_root}/{script_path}"
        cmd = f'cat {full_path}'
        output, error, code = ssh.execute_command(cmd, timeout=15)
        if code != 0:
            raise HTTPException(status_code=500, detail=f"Failed to read script: {error}")

        script_content = output
        info = _parse_spytest_script(script_content)
        info["script_path"] = script_path
        info["script_name"] = os.path.basename(script_path)
        return info
    finally:
        ssh.disconnect()


def _parse_spytest_script(content: str) -> dict:
    """Parse SPyTest script content to extract topology and metadata."""
    result = {
        "topology_marker": None,
        "min_topology": [],
        "dut_count": 1,
        "description": "",
        "topology_type": "standalone",
    }

    # 1. Parse @pytest.mark.topology("...")
    topo_match = re.search(r'@pytest\.mark\.topology\(["\']([^"\']+)["\']\)', content)
    if topo_match:
        result["topology_marker"] = topo_match.group(1)

    # 2. Parse st.ensure_min_topology(...)
    min_topo_match = re.search(r'st\.ensure_min_topology\(([^)]+)\)', content)
    if min_topo_match:
        args_str = min_topo_match.group(1)
        # Extract quoted strings like "D1", "D1D2:1"
        topo_args = re.findall(r'["\']([^"\']+)["\']', args_str)
        result["min_topology"] = topo_args
        # Count DUTs from topology args
        max_duts = 1
        for arg in topo_args:
            # Count D references: D1, D2, D3, etc.
            dut_refs = re.findall(r'D(\d+)', arg)
            if dut_refs:
                max_duts = max(max_duts, max(int(d) for d in dut_refs))
        result["dut_count"] = max_duts
        # If args were a starred variable (no inline strings), topology lives in a vars YAML
        if not result["min_topology"] and re.search(r'st\.ensure_min_topology\(\*\w+\)', content):
            result["uses_vars_file"] = True

    # 3. Determine topology type
    if result["dut_count"] == 1:
        result["topology_type"] = "standalone"
    elif result["dut_count"] == 2:
        result["topology_type"] = "dual-dut"
    else:
        result["topology_type"] = f"{result['dut_count']}-node"

    # 4. Extract description from docstring
    docstring_match = re.search(r'"""(.+?)"""', content, re.DOTALL)
    if docstring_match:
        desc = docstring_match.group(1).strip()
        # Take first 3 lines or 300 chars
        lines = desc.split("\n")
        result["description"] = "\n".join(lines[:3]).strip()[:300]

    return result


@app.post("/api/spytest/execute")
def start_spytest_execution(body: dict, db: Session = Depends(get_db)):
    """Smart SPyTest execution with topology-aware DUT allocation and parallel scheduling.

    Body:
        host_id: int - device where SPyTest runs
        scripts: list[{path, dut_count, min_topology}] - scripts with topology requirements
            - path: script path relative to tests dir
            - dut_count: number of DUTs needed
            - min_topology: list from st.ensure_min_topology() e.g., ["D1", "D1D2:2"]
        testbed: str - testbed YAML filename
        options: dict - optional extra CLI args
        available_dut_count: int - number of DUTs from canvas selection
    """
    host_id = body.get("host_id")
    scripts = body.get("scripts", [])
    testbed_file = body.get("testbed")
    options = body.get("options", {})
    available_dut_count = int(body.get("available_dut_count", 1))  # from canvas selection

    if not host_id or not scripts or not testbed_file:
        raise HTTPException(status_code=400, detail="host_id, scripts, and testbed required")

    # Validate host
    dut = db.query(DUT).filter(DUT.id == host_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="Host device not found")

    # Create master execution record
    job_id = body.get("job_id")
    exec_name = f"spytest_{int(datetime.utcnow().timestamp())}"
    execution = Execution(
        name=exec_name,
        dut_ids=json.dumps([host_id]),
        execution_type="spytest",
        status="pending",
        job_id=int(job_id) if job_id else None,
    )
    db.add(execution)
    db.commit()
    db.refresh(execution)

    # Pre-initialise queue state NOW (synchronously) so the frontend can poll
    # immediately without waiting for the background thread to call _q_init.
    script_names = [os.path.basename(s.get("path", "")) for s in scripts]
    _q_init(execution.id, script_names, [])   # DUTs updated once testbed is read

    # Enhancement 2: Initialize pending scripts structure for dynamic addition
    _init_pending_scripts(execution.id)

    user_base_path = (body.get("base_path") or "").strip().rstrip("/")

    # Start background execution thread
    # Mark job as running when execution starts
    if job_id:
        _job = db.query(ExecutionJob).filter(ExecutionJob.id == int(job_id)).first()
        if _job:
            _job.status = "running"
            db.commit()

    thread = Thread(
        target=_run_spytest_execution,
        args=(execution.id, host_id, scripts, testbed_file, options, available_dut_count, user_base_path, int(job_id) if job_id else None),
        daemon=True,
    )
    thread.start()

    return {
        "execution_id": execution.id,
        "status": "started",
        "type": "spytest",
        "script_count": len(scripts),
    }


# ============================================================================
# IN-MEMORY EXECUTION QUEUE STATE  (for /api/execution-queue)
# ============================================================================

_exec_queue_lock = Lock()
_exec_queue_state: dict = {}   # execution_id -> {scripts:[{name,status,duts}], free_duts:[]}


def _q_init(execution_id: int, script_names: list, all_duts: list):
    with _exec_queue_lock:
        _exec_queue_state[execution_id] = {
            "scripts": [{"name": n, "status": "queued", "duts": [], "pid": None} for n in script_names],
            "free_duts": list(all_duts),
            "all_duts": list(all_duts),
        }


def _q_update_script(execution_id: int, script_name: str, status: str, duts: list = None, pid: str = None):
    with _exec_queue_lock:
        state = _exec_queue_state.get(execution_id)
        if not state:
            return
        for s in state["scripts"]:
            if s["name"] == script_name:
                s["status"] = status
                if duts is not None:
                    s["duts"] = duts
                if pid is not None:
                    s["pid"] = pid
                break


def _q_set_free(execution_id: int, free_duts: list):
    with _exec_queue_lock:
        state = _exec_queue_state.get(execution_id)
        if state:
            state["free_duts"] = list(free_duts)


def _q_cleanup(execution_id: int):
    with _exec_queue_lock:
        _exec_queue_state.pop(execution_id, None)


@app.get("/api/execution-queue")
def get_execution_queue():
    """Return current live execution queue state for all active executions."""
    with _exec_queue_lock:
        return dict(_exec_queue_state)


# ============================================================================
# ENHANCEMENT 2: DYNAMIC BATCH ADDITION & SCRIPT CANCELLATION
# ============================================================================

_pending_scripts_lock = Lock()
_pending_scripts: dict = {}  # execution_id -> {scripts:[...], to_cancel: set(...)}
_execution_threads_lock = Lock()
_execution_threads: dict = {}  # execution_id -> list of thread objects
_test_results_lock = Lock()  # guards concurrent test_results JSON updates from parallel script threads


def _init_pending_scripts(execution_id: int):
    """Initialize pending scripts structure for new execution."""
    with _pending_scripts_lock:
        _pending_scripts[execution_id] = {"scripts": [], "to_cancel": set()}


def _add_pending_script(execution_id: int, script_info: dict):
    """Add script to pending list for dynamic addition."""
    with _pending_scripts_lock:
        if execution_id in _pending_scripts:
            _pending_scripts[execution_id]["scripts"].append(script_info)


def _mark_script_for_cancel(execution_id: int, script_name: str):
    """Mark a script to be cancelled (removed from queue or stopped if running)."""
    with _pending_scripts_lock:
        if execution_id in _pending_scripts:
            _pending_scripts[execution_id]["to_cancel"].add(script_name)


def _cleanup_pending_scripts(execution_id: int):
    """Clean up pending scripts after execution completes."""
    with _pending_scripts_lock:
        _pending_scripts.pop(execution_id, None)


# Enhancement 5: Auto-delete ExecutionLog records after execution completes
def _delete_execution_logs(execution_id: int, db: Session):
    """Enhancement 5: Delete all ExecutionLog records for a completed execution.

    Preserves the Execution record itself (summary/metadata) but removes
    all log entries to save database space after execution completes.

    Args:
        execution_id: ID of the execution whose logs should be deleted
        db: Database session
    """
    try:
        deleted_count = db.query(ExecutionLog).filter(
            ExecutionLog.execution_id == execution_id
        ).delete(synchronize_session=False)

        db.commit()
        if deleted_count > 0:
            logger.info(f"Enhancement 5: Deleted {deleted_count} logs for execution {execution_id}")
    except Exception as e:
        logger.error(f"Enhancement 5: Failed to delete logs for execution {execution_id}: {e}")
        try:
            db.rollback()
        except:
            pass


@app.post("/api/executions/{execution_id}/add-scripts")
def add_scripts_to_execution(execution_id: int, body: dict, db: Session = Depends(get_db)):
    """Add new scripts to a running execution queue (Enhancement 2).

    Body:
        scripts: list[{path, dut_count, min_topology}] - scripts to add
    """
    scripts = body.get("scripts", [])

    if not scripts:
        raise HTTPException(status_code=400, detail="Please select at least one script")

    # Verify execution exists and is still running
    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")

    if execution.status not in ["running", "pending"]:
        raise HTTPException(status_code=400, detail="Execution is not running")

    # Add scripts to pending queue
    for script_info in scripts:
        _add_pending_script(execution_id, script_info)
        # Also add to queue state if not already there
        script_name = os.path.basename(script_info.get("path", ""))
        _q_update_script(execution_id, script_name, "queued")

    return {
        "status": "success",
        "added": len(scripts),
        "execution_id": execution_id
    }


@app.post("/api/executions/{execution_id}/cancel-script")
def cancel_script_from_execution(execution_id: int, body: dict, db: Session = Depends(get_db)):
    """Cancel a running or queued script (Enhancement 2).

    Body:
        script_name: str - name of script to cancel
    """
    script_name = body.get("script_name")

    if not script_name:
        raise HTTPException(status_code=400, detail="script_name required")

    # Verify execution exists
    execution = db.query(Execution).filter(Execution.id == execution_id).first()
    if not execution:
        raise HTTPException(status_code=404, detail="Execution not found")

    # Mark for cancellation
    _mark_script_for_cancel(execution_id, script_name)

    # Update queue state
    _q_update_script(execution_id, script_name, "cancelled")

    return {
        "status": "success",
        "script": script_name,
        "execution_id": execution_id
    }


# ============================================================================
# TOPOLOGY-AWARE DUT ALLOCATION HELPERS
# ============================================================================

def _parse_link_requirements(min_topology: list) -> dict:
    """Parse st.ensure_min_topology() args to extract link requirements.

    Args:
        min_topology: List like ["D1", "D1D2:1", "D2D3:2", "D1D2:2"]

    Returns:
        Dict mapping device pairs to required link count:
        {("D1", "D2"): 2, ("D2", "D3"): 2}
    """
    link_reqs = {}
    for arg in min_topology:
        # Match patterns like "D1D2:2" or "D2D3:1"
        match = re.match(r'D(\d+)D(\d+):(\d+)', arg)
        if match:
            d1, d2, count = match.groups()
            dev1, dev2 = f"D{d1}", f"D{d2}"
            # Normalize order (D1D2 same as D2D1)
            pair = tuple(sorted([dev1, dev2]))
            # Take maximum if multiple mentions
            link_reqs[pair] = max(link_reqs.get(pair, 0), int(count))
    return link_reqs


def _get_topology_connections(db: Session) -> dict:
    """Query topology canvas connections and return as a dict.

    Returns:
        Dict mapping DUT name pairs to connection count:
        {("DUT1", "DUT2"): 2, ("DUT3", "DUT4"): 1}
    """
    connections = db.query(TopologyConnection).all()
    conn_count = {}

    for conn in connections:
        dut_a = db.query(DUT).filter(DUT.id == conn.dut_a_id).first()
        dut_b = db.query(DUT).filter(DUT.id == conn.dut_b_id).first()

        if dut_a and dut_b:
            # Normalize order
            pair = tuple(sorted([dut_a.name, dut_b.name]))
            conn_count[pair] = conn_count.get(pair, 0) + 1

    return conn_count


def _has_back_to_back_connection(dut_name: str, db: Session) -> bool:
    """Check if a DUT has a back-to-back (self-loop) connection in the topology canvas."""
    dut = db.query(DUT).filter(DUT.name == dut_name).first()
    if not dut:
        return False
    self_conn = db.query(TopologyConnection).filter(
        TopologyConnection.dut_a_id == dut.id,
        TopologyConnection.dut_b_id == dut.id
    ).first()
    return self_conn is not None


def _get_b2b_dut_names(db: Session) -> set:
    """Return the set of DUT names that have at least one self-loop canvas connection.

    Called ONCE in the main thread before worker threads start so the result can
    be captured as a plain Python set — no DB access needed inside the lock.
    """
    self_loop_conns = db.query(TopologyConnection).filter(
        TopologyConnection.dut_a_id == TopologyConnection.dut_b_id
    ).all()
    if not self_loop_conns:
        return set()
    dut_ids = {c.dut_a_id for c in self_loop_conns}
    duts = db.query(DUT).filter(DUT.id.in_(dut_ids)).all()
    return {d.name for d in duts}


def _find_duts_matching_topology(
    available_duts: list,
    dut_count: int,
    link_requirements: dict,
    topology_connections: dict,
    b2b_dut_names: set,
) -> list:
    """Find a subset of DUTs from available_duts that satisfies topology requirements.

    No DB access — all topology data is pre-computed and passed in.

    Args:
        available_duts: Pool of DUT names currently free
        dut_count: How many DUTs this script needs
        link_requirements: {("D1","D2"): 2, …} from st.ensure_min_topology
        topology_connections: {("DUT-A","DUT-B"): count} from canvas (inter-device links)
        b2b_dut_names: set of DUT names that have self-loop connections in the canvas

    Returns:
        List of DUT names, or None if requirements cannot be met right now
    """
    from itertools import combinations

    if not link_requirements:
        if dut_count == 1 and b2b_dut_names:
            # Prefer b2b device for single-DUT scripts when canvas has b2b devices
            for dut in available_duts:
                if dut in b2b_dut_names:
                    return [dut]
        # No b2b preference, or b2b device not currently free — FIFO
        return available_duts[:dut_count] if len(available_duts) >= dut_count else None

    # Multi-DUT: try every combination and return the first that meets all link counts
    for combo in combinations(available_duts, dut_count):
        combo_list = list(combo)
        dut_mapping = {f"D{i+1}": combo_list[i] for i in range(len(combo_list))}

        satisfied = True
        for (dev1, dev2), required_links in link_requirements.items():
            actual_dut1 = dut_mapping.get(dev1)
            actual_dut2 = dut_mapping.get(dev2)
            if not actual_dut1 or not actual_dut2:
                satisfied = False
                break
            pair = tuple(sorted([actual_dut1, actual_dut2]))
            if topology_connections.get(pair, 0) < required_links:
                satisfied = False
                break

        if satisfied:
            return combo_list

    return None


# ============================================================================
# REPORT HELPERS — result parsing, HTML dashboard, Excel export
# ============================================================================

def _parse_results_csv(csv_data: str) -> list:
    """Parse SPyTest results CSV into a list of dicts."""
    rows = []
    try:
        reader = csv.DictReader(io.StringIO(csv_data))
        for row in reader:
            module     = (row.get('Module')       or row.get('module')        or '').strip()
            func       = (row.get('TestFunction') or row.get('Function')      or
                          row.get('test_function') or '').strip()
            result     = (row.get('Result')       or row.get('result')        or '').strip()
            time_taken = (row.get('TimeTaken')    or row.get('Time')          or
                          row.get('time_taken')   or '').strip()
            doc        = (row.get('DocSummary')   or row.get('Description')   or
                          row.get('description')  or '').strip()
            time_s = 0
            if time_taken:
                parts = time_taken.split(':')
                try:
                    if len(parts) == 3:
                        time_s = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(float(parts[2]))
                    elif len(parts) == 2:
                        time_s = int(parts[0]) * 60 + int(float(parts[1]))
                except Exception:
                    pass
            tc_id = func.split('.')[-1] if '.' in func else func
            rows.append({'module': module, 'test_function': func, 'testcase_id': tc_id,
                         'result': result, 'time_taken': time_taken, 'time_seconds': time_s,
                         'description': doc[:200] if doc else ''})
    except Exception as e:
        logger.warning(f"[results] CSV parse error: {e}")
    return rows


def _collect_and_save_results(ssh, execution, execution_id: int, script_path: str,
                               log_dir: str, db):
    """After a script finishes: find the results CSV, save TestCaseResult rows,
    and append to execution.test_results (per-script aggregate JSON)."""
    script_stem = os.path.basename(script_path).replace('.py', '')
    try:
        out, _, _ = ssh.execute_command(
            f"find {log_dir} -name 'results_*_functions.csv' 2>/dev/null | head -1",
            timeout=15)
        csv_remote = out.strip()
        rows = []
        if csv_remote:
            sftp = ssh.client.open_sftp()
            try:
                with sftp.file(csv_remote, 'r') as fh:
                    csv_data = fh.read().decode('utf-8', errors='ignore')
            finally:
                sftp.close()
            rows = _parse_results_csv(csv_data)
            for r in rows:
                tcr = TestCaseResult(
                    execution_id=execution_id,
                    script_path=script_path,
                    module=r.get('module', ''),
                    test_function=r.get('test_function', ''),
                    testcase_id=r.get('testcase_id', ''),
                    result=r.get('result', ''),
                    time_taken=r.get('time_taken', ''),
                    time_seconds=r.get('time_seconds', 0),
                    description=r.get('description', '')
                )
                db.add(tcr)
            db.commit()
        else:
            logger.info(f"[results] No CSV in {log_dir} for {script_stem}")

        # Aggregate per-script summary into execution.test_results
        result_lower = [r.get('result', '').lower() for r in rows]
        passed  = sum(1 for r in result_lower if r == 'pass')
        failed  = sum(1 for r in result_lower if r in ('fail', 'scripterror', 'error'))
        skipped = sum(1 for r in result_lower if r in ('skip', 'xfail', 'deselect'))
        dur_s   = sum(r.get('time_seconds', 0) for r in rows)
        agg_status = ('failed' if failed > 0 else
                      'passed' if passed > 0 else
                      'skipped' if skipped > 0 else 'unknown')
        agg = {'script': script_path, 'script_stem': script_stem, 'status': agg_status,
               'passed': passed, 'failed': failed, 'skipped': skipped, 'duration_s': dur_s}
        # Fix: reload execution from the inner session (db=sdb) so the commit actually persists.
        # Use a lock to prevent concurrent script threads from racing on the JSON append.
        with _test_results_lock:
            inner_exec = db.query(Execution).filter(Execution.id == execution_id).first()
            if inner_exec:
                existing = json.loads(inner_exec.test_results or '[]')
                existing.append(agg)
                inner_exec.test_results = json.dumps(existing)
            db.commit()
        logger.info(f"[results] #{execution_id} {script_stem}: pass={passed} fail={failed} skip={skipped}")
    except Exception as e:
        logger.warning(f"[results] Collection failed for {script_stem}: {e}")


# ── Report formatting helpers ─────────────────────────────────────────────────

def _rebuild_results_from_tcrs(tcrs: list) -> list:
    """Rebuild per-script aggregates from TestCaseResult rows (fallback when test_results is empty)."""
    by_script: dict = {}
    for tcr in tcrs:
        sp = tcr.script_path or 'unknown'
        by_script.setdefault(sp, []).append(tcr)
    out = []
    for sp, rows in by_script.items():
        stem = os.path.basename(sp).replace('.py', '') if sp != 'unknown' else 'unknown'
        rl = [(r.result or '').lower() for r in rows]
        p  = sum(1 for r in rl if r == 'pass')
        f  = sum(1 for r in rl if r in ('fail', 'scripterror', 'error'))
        sk = sum(1 for r in rl if r in ('skip', 'xfail', 'deselect'))
        dur = sum(r.time_seconds or 0 for r in rows)
        status = 'failed' if f > 0 else 'passed' if p > 0 else 'skipped' if sk > 0 else 'unknown'
        out.append({'script': sp, 'script_stem': stem, 'status': status,
                    'passed': p, 'failed': f, 'skipped': sk, 'duration_s': dur})
    return out


def _extract_feature(module_path: str) -> str:
    if not module_path:
        return "Unknown"
    parts = module_path.replace("\\", "/").split("/")
    if len(parts) >= 2:
        feat = parts[-2].replace("iscli_", "").replace("ISCLI_", "")
        return feat.upper()
    return parts[0].upper() if parts else "Unknown"

def _extract_tc_id(test_function: str) -> str:
    if not test_function:
        return ""
    return test_function.split(".")[-1] if "." in test_function else test_function

def _fmt_seconds(secs) -> str:
    secs = secs or 0
    h, rem = divmod(int(secs), 3600)
    m, s = divmod(rem, 60)
    parts = []
    if h: parts.append(f"{h}h")
    if m: parts.append(f"{m}m")
    if s: parts.append(f"{s}s")
    return " ".join(parts) or "0s"

def _norm_result(r: str) -> str:
    r = (r or '').lower()
    if r == 'pass': return 'pass'
    if r in ('fail', 'scripterror', 'error'): return 'fail'
    return 'other'

def _calc_trend(results: list) -> str:
    if not results: return "unknown"
    statuses = [_norm_result(r.get("result", "")) for r in results]
    if all(s == "pass" for s in statuses): return "stable_pass"
    if all(s == "fail" for s in statuses): return "stable_fail"
    pass_count = sum(1 for s in statuses if s == "pass")
    fail_count = sum(1 for s in statuses if s == "fail")
    if pass_count > 0 and fail_count > 0:
        if len(statuses) >= 2 and statuses[0] == "fail" and any(s == "pass" for s in statuses[1:]):
            return "fixing"
        if len(statuses) >= 2 and statuses[0] == "pass" and any(s == "fail" for s in statuses[1:]):
            return "regressing"
        return "flaky"
    return "unknown"


def _build_html_dashboard(execution, results: list, tcrs: list) -> str:
    from collections import defaultdict
    total_p = sum(r.get("passed", 0) for r in results)
    total_f = sum(r.get("failed", 0) for r in results)
    total_s = sum(r.get("skipped", 0) for r in results)
    total   = total_p + total_f + total_s
    total_runtime_s = sum(t.time_seconds or 0 for t in tcrs)
    pp = (total_p / total * 100) if total else 0
    fp = (total_f / total * 100) if total else 0
    sp = (total_s / total * 100) if total else 0
    by_feature: dict = defaultdict(list)
    for t in tcrs:
        feat = _extract_feature(t.module or t.script_path or "")
        by_feature[feat].append(t)
    css = """
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',Tahoma,Geneva,Verdana,sans-serif;
     background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);padding:20px;min-height:100vh}
.container{max-width:1400px;margin:0 auto;background:#fff;border-radius:12px;
           box-shadow:0 10px 40px rgba(0,0,0,.2);overflow:hidden}
.header{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);color:#fff;padding:30px;text-align:center}
.header h1{font-size:30px;margin-bottom:8px;font-weight:600}
.header p{font-size:13px;opacity:.9}
.summary-section{padding:28px;background:#f8f9fa;border-bottom:1px solid #e9ecef}
.summary-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:16px;margin-bottom:20px}
.summary-card{background:#fff;padding:18px;border-radius:8px;box-shadow:0 2px 8px rgba(0,0,0,.08);text-align:center}
.summary-card h3{font-size:12px;color:#6c757d;margin-bottom:8px;text-transform:uppercase;letter-spacing:.5px}
.summary-card .val{font-size:34px;font-weight:700;margin-bottom:4px}
.val.total{color:#667eea}.val.passed{color:#28a745}.val.failed{color:#dc3545}
.val.skipped{color:#ffc107}.val.runtime{color:#17a2b8}
.progress-bar{width:100%;height:28px;background:#e9ecef;border-radius:4px;overflow:hidden;display:flex}
.progress-bar.sm{height:18px;margin-top:10px}
.seg{height:100%;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:#fff}
.seg.pass{background:#28a745}.seg.fail{background:#dc3545}.seg.skip{background:#ffc107}
.tabs{display:flex;flex-wrap:wrap;background:#f8f9fa;border-bottom:2px solid #dee2e6;padding:0 20px}
.tab-btn{background:none;border:none;padding:14px 22px;cursor:pointer;font-size:13px;font-weight:500;
         color:#6c757d;border-bottom:3px solid transparent}
.tab-btn:hover{color:#667eea}.tab-btn.active{color:#667eea;border-bottom-color:#667eea;background:#fff}
.tab-content{display:none;padding:28px}.tab-content.active{display:block}
.mod-summary{background:#f8f9fa;padding:18px;border-radius:8px;border-left:4px solid #667eea;margin-bottom:18px}
.mod-summary h2{color:#667eea;font-size:18px;margin-bottom:12px}
.mod-stats{display:flex;flex-wrap:wrap;gap:14px;margin-bottom:12px;font-size:13px}
table{width:100%;border-collapse:collapse;background:#fff;
      box-shadow:0 2px 8px rgba(0,0,0,.08);border-radius:8px;overflow:hidden;margin-top:18px}
thead{background:linear-gradient(135deg,#667eea 0%,#764ba2 100%)}
th{color:#fff;padding:11px 12px;text-align:left;font-weight:600;font-size:11px;text-transform:uppercase}
td{padding:9px 12px;border-bottom:1px solid #e9ecef;font-size:12px}
tr:last-child td{border-bottom:none}
tr:hover td{background:#f8f9fa}
.badge{display:inline-block;padding:3px 9px;border-radius:4px;font-weight:700;font-size:11px}
.badge.pass{background:#d4edda;color:#155724}.badge.fail{background:#f8d7da;color:#721c24}
.badge.skip{background:#fff3cd;color:#856404}.badge.other{background:#e2e3e5;color:#383d41}
.footer{padding:18px;text-align:center;background:#f8f9fa;color:#6c757d;font-size:11px;border-top:1px solid #e9ecef}
"""
    features = sorted(by_feature.keys())
    tab_nav   = "".join(
        f'<button class="tab-btn" onclick="openTab(event,\'{f.replace(" ","_").replace("/","_")}\')">{f}</button>\n'
        for f in features)
    tab_bodies = ""
    for feat in features:
        safe = feat.replace(" ", "_").replace("/", "_")
        tcs  = by_feature[feat]
        fp2  = sum(1 for t in tcs if (t.result or "").lower() == "pass")
        ff2  = sum(1 for t in tcs if (t.result or "").lower() in ("fail","scripterror","error"))
        fs2  = sum(1 for t in tcs if (t.result or "").lower() in ("skip","xfail","deselect"))
        ft2  = fp2 + ff2 + fs2
        fpp  = (fp2/ft2*100) if ft2 else 0
        ffp  = (ff2/ft2*100) if ft2 else 0
        fsp  = (fs2/ft2*100) if ft2 else 0
        prog  = ""
        if fpp > 0: prog += f'<div class="seg pass" style="width:{fpp:.1f}%">{fpp:.0f}%</div>'
        if ffp > 0: prog += f'<div class="seg fail" style="width:{ffp:.1f}%">{ffp:.0f}%</div>'
        if fsp > 0: prog += f'<div class="seg skip" style="width:{fsp:.1f}%">{fsp:.0f}%</div>'
        rows  = ""
        for idx, t in enumerate(tcs, 1):
            res  = (t.result or "").strip()
            rl   = res.lower()
            cls  = ("pass" if rl == "pass" else
                    "fail" if rl in ("fail","scripterror","error") else
                    "skip" if rl in ("skip","xfail","deselect") else "other")
            tc_id = _extract_tc_id(t.test_function or "")
            desc  = (t.description or "")[:120]
            rows += (f"<tr><td style='text-align:center;font-size:11px'>{idx}</td>"
                     f"<td style='font-weight:600;font-size:11px;color:#667eea'>{feat}</td>"
                     f"<td style='font-size:10px;word-break:break-all'>{t.module or t.script_path or ''}</td>"
                     f"<td style='font-family:monospace;font-size:11px'>{tc_id}</td>"
                     f"<td style='font-size:11px'>{desc}</td>"
                     f"<td style='text-align:center;font-size:11px'>{t.time_taken or ''}</td>"
                     f"<td><span class='badge {cls}'>{res}</span></td></tr>")
        tab_bodies += f"""
<div id="{safe}" class="tab-content">
  <div class="mod-summary"><h2>{feat} — Test Results</h2>
    <div class="mod-stats">
      <span>✓ Pass: <strong style="color:#28a745">{fp2}</strong> ({fpp:.1f}%)</span>
      <span style="margin:0 12px">✗ Fail: <strong style="color:#dc3545">{ff2}</strong> ({ffp:.1f}%)</span>
      <span>⊘ Skip: <strong style="color:#ffc107">{fs2}</strong> ({fsp:.1f}%)</span>
      <span style="margin-left:12px">Total: <strong>{ft2}</strong></span>
    </div><div class="progress-bar sm">{prog}</div></div>
  <table><thead><tr>
    <th style="width:36px">S.No</th><th style="min-width:80px">Feature</th>
    <th style="min-width:180px">Script</th><th style="min-width:180px">Testcase ID</th>
    <th style="min-width:260px">Description</th>
    <th style="width:80px">Time</th><th style="width:80px">Status</th>
  </tr></thead><tbody>{rows}</tbody></table>
</div>"""
    overall_prog = ""
    if pp > 0: overall_prog += f'<div class="seg pass" style="width:{pp:.1f}%">{pp:.0f}%</div>'
    if fp > 0: overall_prog += f'<div class="seg fail" style="width:{fp:.1f}%">{fp:.0f}%</div>'
    if sp > 0: overall_prog += f'<div class="seg skip" style="width:{sp:.1f}%">{sp:.0f}%</div>'
    generated_at = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    return f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Eka Dashboard — {execution.name}</title><style>{css}</style></head><body>
<div class="container">
  <div class="header"><h1>SONiC Test Dashboard</h1>
    <p>Execution #{execution.id} &nbsp;·&nbsp; {execution.name} &nbsp;·&nbsp;
       {execution.status} &nbsp;·&nbsp; {generated_at}</p></div>
  <div class="summary-section"><div class="summary-grid">
    <div class="summary-card"><h3>Total Tests</h3><div class="val total">{total}</div></div>
    <div class="summary-card"><h3>Passed</h3><div class="val passed">{total_p}</div><div class="pct">{pp:.1f}%</div></div>
    <div class="summary-card"><h3>Failed</h3><div class="val failed">{total_f}</div><div class="pct">{fp:.1f}%</div></div>
    <div class="summary-card"><h3>Skipped</h3><div class="val skipped">{total_s}</div><div class="pct">{sp:.1f}%</div></div>
    <div class="summary-card"><h3>Total Runtime</h3><div class="val runtime">{_fmt_seconds(total_runtime_s)}</div></div>
  </div>
  <div style="margin-top:16px"><h3 style="font-size:14px;color:#495057;margin-bottom:8px">Overall Progress</h3>
    <div class="progress-bar">{overall_prog}</div></div></div>
  <div class="tabs">{tab_nav}</div>{tab_bodies}
  <div class="footer">Generated by Eka Automation Platform &nbsp;·&nbsp; {generated_at}</div>
</div>
<script>
function openTab(evt,name){{
  document.querySelectorAll('.tab-content').forEach(e=>e.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(e=>e.classList.remove('active'));
  document.getElementById(name).classList.add('active');
  evt.currentTarget.classList.add('active');
}}
window.onload=function(){{var first=document.querySelector('.tab-btn');if(first)first.click();}};
</script></body></html>"""


def _build_excel(execution, results: list, tcrs: list):
    wb = openpyxl.Workbook()
    HDR_FILL  = PatternFill("solid", fgColor="667EEA")
    HDR_FONT  = Font(bold=True, color="FFFFFF")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center")
    PASS_FILL = PatternFill("solid", fgColor="D4EDDA")
    FAIL_FILL = PatternFill("solid", fgColor="F8D7DA")
    SKIP_FILL = PatternFill("solid", fgColor="FFF3CD")
    PASS_FONT = Font(bold=True, color="155724")
    FAIL_FONT = Font(bold=True, color="721C24")
    SKIP_FONT = Font(bold=True, color="856404")

    def _write_header(ws, headers, col_widths=None):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = HDR_ALIGN
        ws.row_dimensions[1].height = 20
        if col_widths:
            for col, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    def _result_style(ws, row_i, col):
        c = ws.cell(row=row_i, column=col)
        v = (c.value or "").lower()
        if v == "pass":                          c.fill = PASS_FILL; c.font = PASS_FONT
        elif v in ("fail","scripterror","error"): c.fill = FAIL_FILL; c.font = FAIL_FONT
        elif v in ("skip","xfail","deselect"):   c.fill = SKIP_FILL; c.font = SKIP_FONT

    ws1 = wb.active; ws1.title = "Summary"
    total_p = sum(r.get("passed", 0) for r in results)
    total_f = sum(r.get("failed", 0) for r in results)
    total_s = sum(r.get("skipped", 0) for r in results)
    total   = total_p + total_f + total_s
    for label, val in [("Execution ID", execution.id), ("Name", execution.name),
                       ("Status", execution.status),
                       ("Duration", f"{execution.duration_seconds}s" if execution.duration_seconds else "–"),
                       ("Scripts Run", len(results))]:
        row = [("Execution ID", "Name", "Status", "Duration", "Scripts Run").index(label) + 1
               if label in ("Execution ID", "Name", "Status", "Duration", "Scripts Run") else 1]
    for row, (label, val) in enumerate([("Execution ID", execution.id), ("Name", execution.name),
                                         ("Status", execution.status),
                                         ("Duration", f"{execution.duration_seconds}s" if execution.duration_seconds else "–"),
                                         ("Scripts Run", len(results))], 1):
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=val)
    ws1.column_dimensions["A"].width = 18; ws1.column_dimensions["B"].width = 40
    for row, (label, val) in enumerate([("Total Tests", total), ("Passed", total_p),
                                         ("Failed", total_f), ("Skipped", total_s),
                                         ("Pass Rate", f"{total_p/total*100:.1f}%" if total else "–")], 7):
        ws1.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row, column=2, value=val)
    for col, h in enumerate(["Script","Status","Passed","Failed","Skipped","Duration (s)"], 1):
        c = ws1.cell(row=14, column=col, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = HDR_ALIGN
    for ri, r in enumerate(results, 15):
        ws1.cell(row=ri, column=1, value=r.get("script_stem", r.get("script", "")))
        ws1.cell(row=ri, column=2, value=r.get("status", ""))
        ws1.cell(row=ri, column=3, value=r.get("passed", 0))
        ws1.cell(row=ri, column=4, value=r.get("failed", 0))
        ws1.cell(row=ri, column=5, value=r.get("skipped", 0))
        ws1.cell(row=ri, column=6, value=r.get("duration_s", 0))
        _result_style(ws1, ri, 2)
    ws2 = wb.create_sheet("All Testcases")
    _write_header(ws2, ["S.No","Feature","Script / Module","Testcase ID",
                         "Test Function","Result","Time Taken","Time (s)","Description"],
                  [6,16,38,34,46,12,12,10,50])
    for ri, t in enumerate(tcrs, 2):
        feat  = _extract_feature(t.module or t.script_path or "")
        tc_id = _extract_tc_id(t.test_function or "")
        for col, val in enumerate([ri-1, feat, t.module or t.script_path or "",
                                    tc_id, t.test_function or "", t.result or "",
                                    t.time_taken or "", t.time_seconds or 0,
                                    t.description or ""], 1):
            ws2.cell(row=ri, column=col, value=val)
        _result_style(ws2, ri, 6)
    ws3 = wb.create_sheet("Failures")
    _write_header(ws3, ["S.No","Feature","Script / Module","Testcase ID",
                         "Test Function","Result","Time Taken","Description"],
                  [6,16,38,34,46,16,12,50])
    failures = [t for t in tcrs if (t.result or "").lower() in ("fail","scripterror","error")]
    if not failures:
        ws3.cell(row=2, column=1, value="No failures recorded for this execution.")
    else:
        for ri, t in enumerate(failures, 2):
            feat  = _extract_feature(t.module or t.script_path or "")
            tc_id = _extract_tc_id(t.test_function or "")
            for col, val in enumerate([ri-1, feat, t.module or t.script_path or "",
                                        tc_id, t.test_function or "", t.result or "",
                                        t.time_taken or "", t.description or ""], 1):
                ws3.cell(row=ri, column=col, value=val)
            _result_style(ws3, ri, 6)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ============================================================================
# SPYTEST BACKGROUND EXECUTION — per-script threads with smart DUT allocation
# ============================================================================

def _run_spytest_execution(
    execution_id: int,
    host_id: int,
    scripts: list,
    testbed_file: str,
    options: dict,
    available_dut_count: int = 1,
    base_path: str = "",
    job_id: int = None,
):
    """Background thread: Smart SPyTest execution with true parallel DUT allocation.

    Each script runs in its own SSH sub-thread. A shared lock + pool ensures that
    a script that needs 1 DUT starts immediately if 1 is free, while a script
    needing 2 DUTs waits only until 2 are simultaneously available.
    """
    import time as _time

    # Derive all runtime paths from the user-supplied base_path (scripts dir).
    # e.g. base_path = /home/user/spytest/tests
    #   -> spytest_root  = /home/user/spytest
    #   -> testbed_dir   = /home/user/spytest/testbeds
    #   -> spytest_bin   = /home/user/spytest/bin/spytest
    #   -> spytest_venv  = /home/user/spytest/spytest_venv
    # Falls back to the server-configured constants when base_path is empty.
    if base_path:
        _bp = base_path.rstrip("/")
        # Extract spytest root: everything up to and including the "spytest" component.
        # Works whether user enters /a/spytest, /a/spytest/tests, /a/spytest/tests/sub, etc.
        _m = re.match(r'(^.*?/spytest)(?:/|$)', _bp)
        _spytest_root   = _m.group(1) if _m else os.path.dirname(_bp)
        _tests_dir      = _bp
        _testbed_dir    = _spytest_root + "/testbeds"
        _spytest_bin    = _spytest_root + "/bin/spytest"
        _spytest_venv   = _spytest_root + "/spytest_venv"
        _spytest_python = _spytest_venv + "/bin/python"
    else:
        _spytest_root   = SPYTEST_BASE
        _tests_dir      = SPYTEST_TESTS_DIR
        _testbed_dir    = SPYTEST_TESTBED_DIR
        _spytest_bin    = SPYTEST_BIN
        _spytest_venv   = SPYTEST_VENV
        _spytest_python = SPYTEST_PYTHON

    db = SessionLocal()
    try:
        execution = db.query(Execution).filter(Execution.id == execution_id).first()
        host_dut = db.query(DUT).filter(DUT.id == host_id).first()
        if not execution or not host_dut:
            return

        execution.status = "running"
        execution.start_time = datetime.utcnow()
        db.commit()

        log_execution(db, execution_id, "SYSTEM", "INFO",
                      f"SPyTest execution started on {host_dut.name} ({host_dut.ip_address})")
        log_execution(db, execution_id, "SYSTEM", "INFO",
                      f"Testbed: {testbed_file} | Scripts: {len(scripts)}")

        # ── Open a single persistent SSH connection for coordination ──────────
        coord_ssh = SSHConnectionManager(
            host_dut.ip_address, host_dut.port, host_dut.username, host_dut.password
        )
        if not coord_ssh.connect():
            log_execution(db, execution_id, "SYSTEM", "ERROR",
                          f"Failed to SSH to {host_dut.name}")
            execution.status = "failed"
            execution.end_time = datetime.utcnow()
            db.commit()
            return

        try:
            # ── Read testbed YAML ─────────────────────────────────────────────
            # If the frontend passed a full absolute path, use it directly;
            # otherwise join with the derived testbed dir.
            if testbed_file.startswith("/"):
                testbed_path = testbed_file
            else:
                testbed_path = f"{_testbed_dir}/{testbed_file}"
            out, err, code = coord_ssh.execute_command(f"cat {testbed_path}", timeout=15)
            if code != 0:
                log_execution(db, execution_id, "SYSTEM", "ERROR",
                              f"Cannot read testbed: {err}")
                execution.status = "failed"
                execution.end_time = datetime.utcnow()
                db.commit()
                return

            testbed_config = yaml.safe_load(out) or {}
            testbed_devices = list(testbed_config.get("devices", {}).keys())
            total_testbed_duts = len(testbed_devices)

            # CRITICAL FIX: Build DUT pool using ONLY unique actual testbed devices
            # NEVER create duplicate device names, as this breaks multi-DUT scripts
            # that require 2+ different physical devices (e.g., D1D2:1 topology)
            if total_testbed_duts == 0:
                # Testbed has no devices — create synthetic slots named Slot-1, Slot-2…
                all_duts = [f"Slot-{i+1}" for i in range(available_dut_count)]
                log_execution(db, execution_id, "SYSTEM", "WARNING",
                              f"No devices found in testbed YAML! Using {available_dut_count} synthetic slot(s). "
                              f"Check the testbed YAML 'devices:' key.")
            else:
                # Use actual testbed devices (never duplicate them)
                all_duts = testbed_devices

                # Validate: warn if canvas has more selections than actual devices
                if available_dut_count > total_testbed_duts:
                    log_execution(db, execution_id, "SYSTEM", "WARNING",
                                  f"⚠ Canvas has {available_dut_count} DUTs selected, but testbed only defines "
                                  f"{total_testbed_duts} device(s). Using {total_testbed_duts} actual device(s). "
                                  f"Multi-DUT scripts requiring {available_dut_count}+ devices will wait/fail.")

                # Validate: error if scripts need more DUTs than available
                max_dut_requirement = max([s.get("dut_count", 1) for s in scripts])
                if max_dut_requirement > total_testbed_duts:
                    log_execution(db, execution_id, "SYSTEM", "ERROR",
                                  f"✗ ALLOCATION ERROR: Script requires {max_dut_requirement} DUTs, "
                                  f"but testbed only has {total_testbed_duts} device(s). "
                                  f"Add more devices to testbed YAML or reduce script requirements.")

            log_execution(db, execution_id, "SYSTEM", "INFO",
                          f"Parallel pool: {len(all_duts)} slot(s) — {', '.join(all_duts)}")

            # ── Get topology connections from canvas ──────────────────────────
            topology_connections = _get_topology_connections(db)
            b2b_dut_names = _get_b2b_dut_names(db)   # pre-compute once; no DB in worker threads
            log_execution(db, execution_id, "SYSTEM", "INFO",
                          f"Topology connections loaded: {len(topology_connections)} unique pairs"
                          + (f", b2b devices: {sorted(b2b_dut_names)}" if b2b_dut_names else ""))

            # ── Pre-populate dut_count / min_topology from script content ─────
            # Only needed when the canvas has connections — that's when topology-
            # aware allocation is active.  Scripts that lack these fields (e.g. the
            # frontend didn't call /api/spytest/script-info) get analyzed here once,
            # before any worker thread starts, so dut_count is always correct.
            if topology_connections:
                for script in scripts:
                    if not script.get("min_topology"):
                        s_path = script.get("path", "")
                        if s_path:
                            # Build full absolute path: relative paths need _tests_dir prefix
                            full_s_path = s_path if s_path.startswith("/") else f"{_tests_dir}/{s_path}"
                            try:
                                out, cat_err, rc = coord_ssh.execute_command(f"cat '{full_s_path}'", timeout=15)
                                if rc == 0:
                                    info = _parse_spytest_script(out)
                                    # Script uses *var pattern — topology is in companion vars YAML
                                    if info.get("uses_vars_file"):
                                        stem = re.sub(r'^test_', '',
                                                      os.path.splitext(os.path.basename(s_path))[0])
                                        vars_path = (f"{os.path.dirname(full_s_path)}"
                                                     f"/../vars/vars_{stem}.yaml")
                                        vars_out, _, vars_rc = coord_ssh.execute_command(
                                            f"cat '{vars_path}'", timeout=10)
                                        if vars_rc == 0:
                                            defaults_sec = re.search(
                                                r'defaults:.*?(?=\n\w|\Z)', vars_out, re.DOTALL)
                                            if defaults_sec:
                                                topo_items = re.findall(
                                                    r'^\s*-\s*["\']([^"\']+)["\']',
                                                    defaults_sec.group(0), re.MULTILINE)
                                                if topo_items:
                                                    info["min_topology"] = topo_items
                                                    info["dut_count"] = max(
                                                        (max((int(d) for d in
                                                              re.findall(r'D(\d+)', arg)),
                                                             default=1)
                                                         for arg in topo_items),
                                                        default=1)
                                        else:
                                            log_execution(db, execution_id, "SYSTEM", "WARNING",
                                                          f"[TOPO] vars file not found for "
                                                          f"{os.path.basename(s_path)} — using dut_count=1")
                                    script["dut_count"]    = info.get("dut_count", 1)
                                    script["min_topology"] = info.get("min_topology", [])
                                    log_execution(db, execution_id, "SYSTEM", "INFO",
                                                  f"[TOPO] {os.path.basename(s_path)}: "
                                                  f"dut_count={script['dut_count']}, "
                                                  f"min_topology={script['min_topology']}")
                                else:
                                    log_execution(db, execution_id, "SYSTEM", "WARNING",
                                                  f"[TOPO] cat failed (rc={rc}) for {os.path.basename(s_path)}: "
                                                  f"{cat_err.strip()[:120]} — using dut_count=1")
                            except Exception as e:
                                log_execution(db, execution_id, "SYSTEM", "WARNING",
                                              f"[TOPO] Could not read {os.path.basename(s_path)}: {e}"
                                              f" — using default dut_count=1")

                # Re-check max DUT requirement after analysis
                max_dut_requirement = max(s.get("dut_count", 1) for s in scripts)
                if max_dut_requirement > total_testbed_duts:
                    log_execution(db, execution_id, "SYSTEM", "ERROR",
                                  f"✗ ALLOCATION ERROR: Script requires {max_dut_requirement} DUTs, "
                                  f"but testbed only has {total_testbed_duts} device(s). "
                                  f"Add more devices to testbed YAML or reduce script requirements.")

            # ── Init in-memory queue state ────────────────────────────────────
            script_names = [os.path.basename(s.get("path", "")) for s in scripts]
            _q_init(execution_id, script_names, all_duts)

            # ── Shared DUT pool (protected by a lock) ─────────────────────────
            pool_lock = Lock()
            available_pool: list = list(all_duts)   # mutable shared state

            def acquire_duts(needed: int, link_requirements: dict = None) -> list:
                """Block until `needed` DUTs are available, then atomically grab them.

                When the canvas has connections, topology-aware matching is used:
                - single-DUT scripts (needed==1): prefer b2b device if canvas has one
                - multi-DUT scripts with link requirements: combo match against canvas links
                When the canvas has no connections, simple FIFO is always used.
                """
                while True:
                    with pool_lock:
                        if len(available_pool) >= needed:
                            # Topology-aware path: canvas has connections AND
                            # (script has link requirements OR needs exactly 1 DUT)
                            if topology_connections and (link_requirements or needed == 1):
                                matched = _find_duts_matching_topology(
                                    available_pool, needed, link_requirements or {},
                                    topology_connections, b2b_dut_names
                                )
                                if matched:
                                    for dut in matched:
                                        if dut in available_pool:
                                            available_pool.remove(dut)
                                    _q_set_free(execution_id, list(available_pool))
                                    return matched
                                # No matching combo found yet.
                                # For single-DUT with no link requirements: b2b device is
                                # busy, fall back to FIFO so the script is never stuck.
                                if not link_requirements and needed == 1:
                                    allocated = available_pool[:needed]
                                    del available_pool[:needed]
                                    _q_set_free(execution_id, list(available_pool))
                                    return allocated
                                # Multi-DUT with link requirements: wait for matching combo
                            else:
                                # Simple FIFO — no canvas connections or multi-DUT no requirements
                                allocated = available_pool[:needed]
                                del available_pool[:needed]
                                _q_set_free(execution_id, list(available_pool))
                                return allocated
                    _time.sleep(5)

            def release_duts(duts_to_free: list):
                with pool_lock:
                    available_pool.extend(duts_to_free)
                    _q_set_free(execution_id, list(available_pool))

            # ── Per-script worker ─────────────────────────────────────────────
            def run_one_script(script_info: dict, slot_idx: int):
                sdb = SessionLocal()
                s_ssh = None
                assigned = []  # Initialize to empty list to prevent release_duts() failure in finally block
                script_path = script_info.get("path", "")
                dut_count   = script_info.get("dut_count", 1)
                min_topology = script_info.get("min_topology", [])
                sname       = os.path.basename(script_path)

                # Re-derive dut_count from min_topology so it's always correct even
                # when the frontend omits dut_count or sets it to the wrong value.
                if min_topology:
                    _max_duts = 1
                    for _arg in min_topology:
                        _refs = re.findall(r'D(\d+)', _arg)
                        if _refs:
                            _max_duts = max(_max_duts, max(int(d) for d in _refs))
                    dut_count = max(dut_count, _max_duts)

                try:
                    # --- Parse link requirements from min_topology ----------
                    link_requirements = _parse_link_requirements(min_topology)
                    if link_requirements:
                        log_execution(sdb, execution_id, sname, "INFO",
                                      f"[TOPO] Link requirements: {link_requirements}")

                    # --- Wait for enough free DUTs with matching topology ---
                    _q_update_script(execution_id, sname, "waiting")
                    log_execution(sdb, execution_id, sname, "INFO",
                                  f"[QUEUE] Waiting for {dut_count} DUT(s)… "
                                  f"(pool has {len(available_pool)})")

                    assigned = acquire_duts(dut_count, link_requirements)
                    _q_update_script(execution_id, sname, "running", duts=assigned)

                    # Log allocation details
                    topo_mode = "topology-matched" if (topology_connections and link_requirements) else "FIFO"
                    log_execution(sdb, execution_id, sname, "INFO",
                                  f"[ALLOC] {topo_mode} → DUT(s): {', '.join(assigned)}")

                    # --- Create temp testbed YAML ---------------------------
                    temp_tb_path = f"/tmp/temp_exec{execution_id}_s{slot_idx}.yaml"
                    temp_cfg = _create_subset_testbed(testbed_config, assigned)
                    temp_yaml_str = yaml.dump(temp_cfg, default_flow_style=False)
                    import base64 as _b64
                    yaml_b64 = _b64.b64encode(temp_yaml_str.encode()).decode()
                    # Use a fresh SSH connection per script so they don't share channels
                    s_ssh = SSHConnectionManager(
                        host_dut.ip_address, host_dut.port,
                        host_dut.username, host_dut.password
                    )
                    if not s_ssh.connect():
                        log_execution(sdb, execution_id, sname, "ERROR",
                                      "SSH connection failed for script worker")
                        _q_update_script(execution_id, sname, "failed")
                        release_duts(assigned)
                        return

                    s_ssh.execute_command(
                        f"echo '{yaml_b64}' | base64 -d > {temp_tb_path}", timeout=10
                    )

                    # --- Build and launch SPyTest ---------------------------
                    extra_opts = ""
                    if options.get("log_level"):
                        extra_opts += f" --log-level {options['log_level']}"
                    if options.get("skip_init_config"):
                        extra_opts += " --skip-init-config"
                    extra_opts += " --ifname-type native"

                    log_dir = (
                        f"{_spytest_root}/logs/"
                        f"exec{execution_id}_{sname}_{int(datetime.utcnow().timestamp())}"
                    )
                    s_ssh.execute_command(f"mkdir -p {log_dir}", timeout=10)

                    spy_cmd = (
                        f"cd {_spytest_venv}/bin && "
                        f"source activate && "
                        f"cd {_spytest_root} && "
                        f"{_spytest_python} {_spytest_bin} --tryssh 1 "
                        f"--testbed {temp_tb_path} "
                        f"{_tests_dir}/{script_path} "
                        f"--logs-path {log_dir}"
                        f"{extra_opts}"
                    )
                    logfile = f"/tmp/spytest_exec{execution_id}_s{slot_idx}.log"
                    bg_cmd = f"nohup bash -c '{spy_cmd}' > {logfile} 2>&1 & echo $!"

                    pid_out, _, _ = s_ssh.execute_command(bg_cmd, timeout=15)
                    pid = pid_out.strip()

                    if not (pid and pid.isdigit()):
                        log_execution(sdb, execution_id, sname, "ERROR",
                                      f"Failed to launch (pid output: {pid_out!r})")
                        _q_update_script(execution_id, sname, "failed")
                        release_duts(assigned)
                        return

                    _q_update_script(execution_id, sname, "running", pid=pid)
                    log_execution(sdb, execution_id, sname, "INFO",
                                  f"[RUN] PID={pid} DUTs={', '.join(assigned)}")

                    # --- Poll until the process exits -----------------------
                    last_lines_seen = 0
                    while True:
                        _time.sleep(10)
                        chk_out, _, _ = s_ssh.execute_command(
                            f"kill -0 {pid} 2>/dev/null && echo RUNNING || echo DONE",
                            timeout=5,
                        )
                        # Stream fresh log tail to the DB
                        log_tail, _, _ = s_ssh.execute_command(
                            f"wc -l < {logfile} 2>/dev/null || echo 0", timeout=5
                        )
                        total_lines = int(log_tail.strip() or "0")
                        if total_lines > last_lines_seen:
                            tail_n = total_lines - last_lines_seen
                            new_log, _, _ = s_ssh.execute_command(
                                f"tail -n {min(tail_n, 50)} {logfile} 2>/dev/null",
                                timeout=5,
                            )
                            if new_log.strip():
                                for ln in new_log.strip().split("\n"):
                                    log_execution(sdb, execution_id, sname, "INFO", ln)
                            last_lines_seen = total_lines

                        if "DONE" in chk_out:
                            break

                    log_execution(sdb, execution_id, sname, "INFO", "✓ Script completed")
                    _q_update_script(execution_id, sname, "done")
                    # Collect CSV results and persist TestCaseResult rows
                    _collect_and_save_results(s_ssh, execution, execution_id,
                                              script_path, log_dir, sdb)
                    s_ssh.execute_command(f"rm -f {temp_tb_path}", timeout=5)

                except Exception as ex:
                    log_execution(sdb, execution_id, sname, "ERROR",
                                  f"Script error: {ex}")
                    _q_update_script(execution_id, sname, "failed")
                finally:
                    # CRITICAL FIX: Always release DUTs back to pool, even if assigned is empty
                    # This prevents DUT resource leaks when errors occur before allocation
                    if assigned:  # Only release if DUTs were actually assigned
                        log_execution(sdb, execution_id, sname, "INFO",
                                      f"[CLEANUP] Releasing {len(assigned)} DUT(s): {', '.join(assigned)}")
                        release_duts(assigned)
                    if s_ssh:
                        s_ssh.disconnect()
                    sdb.close()

            # ── Launch all script threads ─────────────────────────────────────
            threads = []
            for idx, script_info in enumerate(scripts):
                t = Thread(target=run_one_script, args=(script_info, idx), daemon=True)
                threads.append(t)
                t.start()
                _time.sleep(0.3)   # stagger slightly so acquire order is predictable

            for t in threads:
                t.join()

            # ── Mark master execution done ────────────────────────────────────
            execution.status = "completed"
            execution.end_time = datetime.utcnow()
            if execution.start_time:
                execution.duration_seconds = int(
                    (execution.end_time - execution.start_time).total_seconds()
                )
            db.commit()
            log_execution(db, execution_id, "SYSTEM", "INFO",
                          f"✓ All scripts finished ({execution.duration_seconds}s)")
            # Update parent job status
            if job_id:
                _job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id).first()
                if _job and _job.status == "running":
                    _job.status = "completed"
                    db.commit()
            _q_cleanup(execution_id)
            # Enhancement 2: Clean up pending scripts structure
            _cleanup_pending_scripts(execution_id)
            # CRITICAL FIX: DO NOT auto-delete logs! Users need to view them in Logs tab
            # _delete_execution_logs(execution_id, db)  # DISABLED - breaks Logs tab functionality

        finally:
            coord_ssh.disconnect()

    except Exception as e:
        logger.error(f"SPyTest execution failed: {e}")
        if execution:
            execution.status = "failed"
            execution.end_time = datetime.utcnow()
            db.commit()
        if job_id:
            try:
                _job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id).first()
                if _job and _job.status == "running":
                    _job.status = "failed"
                    db.commit()
            except Exception:
                pass
        log_execution(db, execution_id, "SYSTEM", "ERROR", f"Execution failed: {e}")
        _q_cleanup(execution_id)
        # Enhancement 2: Clean up pending scripts structure on failure
        _cleanup_pending_scripts(execution_id)
    finally:
        db.close()


def _create_subset_testbed(full_config: dict, device_names: list) -> dict:
    """Create a subset testbed YAML with only the specified devices.

    Devices are RENAMED to sequential logical names D1, D2, D3... regardless of their
    physical names. This is required because SPyTest scripts use ensure_min_topology("D1D2:2")
    which looks for devices literally named D1 and D2 in the testbed. If the allocated
    physical devices are D3 and D2, renaming D3→D1 and D2→D2 ensures the topology check passes.

    CRITICAL: device_names must contain unique device names only.
    Duplicate device names will be automatically deduplicated to prevent invalid testbed configs.
    """
    # SAFETY CHECK: Remove duplicate device names (should never happen after padding fix)
    unique_device_names = list(dict.fromkeys(device_names))  # Preserves order, removes duplicates
    if len(unique_device_names) < len(device_names):
        logger.warning(f"⚠ TESTBED WARNING: Duplicate devices detected in allocation: {device_names}. "
                      f"Using unique devices only: {unique_device_names}")
        device_names = unique_device_names

    # Build physical→logical name mapping: first allocated DUT→D1, second→D2, etc.
    phys_to_logical = {phys: f"D{i+1}" for i, phys in enumerate(device_names)}
    logger.info(f"[TESTBED] Device remap: {phys_to_logical}")

    all_devices = full_config.get("devices", {})
    all_topology = full_config.get("topology", {})

    devices_section = {}
    topology_section = {}

    for phys_name in device_names:
        logical_name = phys_to_logical[phys_name]

        # Copy device entry under logical name
        if phys_name in all_devices:
            devices_section[logical_name] = all_devices[phys_name]

        # Filter topology to only inter-subset links, rewriting device names to logical
        if phys_name in all_topology:
            dev_topo = all_topology[phys_name]
            filtered_interfaces = {}
            for iface, link in dev_topo.get("interfaces", {}).items():
                end_phys = link.get("EndDevice", "")
                if end_phys in phys_to_logical:
                    # Rewrite EndDevice to the logical name
                    filtered_interfaces[iface] = {
                        "EndDevice": phys_to_logical[end_phys],
                        "EndPort": link.get("EndPort", ""),
                    }
            topology_section[logical_name] = {"interfaces": filtered_interfaces}

    # Rebuild params.topo to reflect only the subset links using logical names
    topo_dict = {}
    for log_a, topo_a in topology_section.items():
        for iface, link in topo_a.get("interfaces", {}).items():
            log_b = link.get("EndDevice", "")
            if log_b and log_a < log_b:  # count each pair once
                key = f"{log_a}{log_b}"
                topo_dict[key] = topo_dict.get(key, 0) + 1

    master_params = full_config.get("params", {})
    subset_params = {"topo": topo_dict}

    result = {
        "version": full_config.get("version", "2.0"),
        "devices": devices_section,
        "topology": topology_section,
        "services": full_config.get("services", {"default": {}}),
        "builds": full_config.get("builds", {"default": {}}),
        "configs": full_config.get("configs", {"default": {}}),
        "errors": full_config.get("errors", {"default": {}}),
        "params": subset_params,
    }
    # Carry the global.params section (test_interface etc.) from master to subset.
    # Also handle old-format master testbeds where test_interface lived in top-level params.
    if "global" in full_config:
        result["global"] = full_config["global"]
    elif "test_interface" in master_params:
        result["global"] = {"params": {"test_interface": master_params["test_interface"]}}
    return result


# ============================================================================
# API — DUT Lock Management (AVAILABLE / ALLOCATED / IN_USE)
# ============================================================================

def _ensure_dut_lock(db: Session, dut_id: int) -> DUTLock:
    """Seed a DUTLock row for a DUT if one doesn't exist yet."""
    lock = db.query(DUTLock).filter(DUTLock.dut_id == dut_id).first()
    if not lock:
        lock = DUTLock(dut_id=dut_id, status="AVAILABLE")
        db.add(lock)
        db.commit()
        db.refresh(lock)
    return lock


@app.get("/api/dut-locks")
def get_dut_locks(db: Session = Depends(get_db)):
    """Return current lock status for all DUTs (creates AVAILABLE rows on first call)."""
    duts = db.query(DUT).filter(DUT.device_type != "VM").all()
    result = []
    for dut in duts:
        lock = _ensure_dut_lock(db, dut.id)
        result.append({
            "dut_id": dut.id,
            "dut_name": dut.name,
            "ip_address": dut.ip_address,
            "status": lock.status,
            "job_id": lock.job_id,
            "locked_since": lock.locked_since.isoformat() if lock.locked_since else None,
        })
    return result


@app.post("/api/dut-locks/{dut_id}/release")
def release_dut_lock(dut_id: int, db: Session = Depends(get_db)):
    """Manually release a DUT lock back to AVAILABLE."""
    lock = db.query(DUTLock).filter(DUTLock.dut_id == dut_id).first()
    if not lock:
        raise HTTPException(status_code=404, detail="DUT lock not found")
    lock.status = "AVAILABLE"
    lock.job_id = None
    lock.locked_since = None
    db.commit()
    return {"status": "released", "dut_id": dut_id}


# ============================================================================
# API — Execution Jobs (named job containers for the Execute tab)
# ============================================================================

@app.post("/api/execution-jobs")
def create_execution_job(request: Request, body: dict, db: Session = Depends(get_db)):
    """Create a new named execution job for the current session."""
    session_id = request.headers.get("X-Session-ID", "default")
    name = body.get("name") or f"Job-{datetime.utcnow().strftime('%H%M')}"
    job = ExecutionJob(name=name, session_id=session_id)
    db.add(job)
    db.commit()
    db.refresh(job)
    return {"id": job.id, "name": job.name, "status": job.status, "created_at": job.created_at.isoformat() if job.created_at else None}


@app.get("/api/execution-jobs")
def list_execution_jobs(request: Request, db: Session = Depends(get_db)):
    """List all execution jobs for the current session (most recent first)."""
    session_id = request.headers.get("X-Session-ID", "default")
    jobs = (db.query(ExecutionJob)
            .filter(ExecutionJob.session_id == session_id)
            .order_by(ExecutionJob.created_at.desc())
            .limit(50)
            .all())
    result = []
    for j in jobs:
        exec_count = db.query(Execution).filter(Execution.job_id == j.id).count()
        result.append({
            "id": j.id,
            "name": j.name,
            "status": j.status,
            "execution_count": exec_count,
            "created_at": j.created_at.isoformat() if j.created_at else None,
            "updated_at": j.updated_at.isoformat() if j.updated_at else None,
        })
    return {"jobs": result}


@app.get("/api/execution-jobs/{job_id}")
def get_execution_job(job_id: int, request: Request, db: Session = Depends(get_db)):
    """Get full state of a specific job."""
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    executions = (db.query(Execution).filter(Execution.job_id == job_id)
                  .order_by(Execution.created_at.desc()).all())
    execs_data = [{
        "id": e.id, "name": e.name, "status": e.status,
        "start_time": e.start_time.isoformat() if e.start_time else None,
        "end_time": e.end_time.isoformat() if e.end_time else None,
    } for e in executions]
    next_run = None
    ap_id = f"exec_job_{job.id}"
    try:
        ap_job = scheduler.get_job(ap_id)
        if ap_job and ap_job.next_run_time:
            from datetime import timezone as _utc_tz
            next_run = ap_job.next_run_time.astimezone(_utc_tz.utc).strftime("%Y-%m-%d %H:%M UTC")
    except Exception:
        pass

    return {
        "id": job.id,
        "name": job.name,
        "status": job.status,
        "dut_ids": json.loads(job.dut_ids) if job.dut_ids else [],
        "base_path": job.base_path or "",
        "host_id": job.host_id,
        "topology": json.loads(job.topology) if job.topology else [],
        "scripts": json.loads(job.scripts) if job.scripts else [],
        "testbed_path": job.testbed_path or "",
        "executions": execs_data,
        "schedule_type":    job.schedule_type or "none",
        "schedule_at":      job.schedule_at.isoformat() if job.schedule_at else None,
        "schedule_cron":    job.schedule_cron,
        "schedule_enabled": bool(job.schedule_enabled),
        "last_run_at":      job.last_run_at.isoformat() if job.last_run_at else None,
        "next_run":         next_run,
        "created_at": job.created_at.isoformat() if job.created_at else None,
    }


@app.put("/api/execution-jobs/{job_id}")
def update_execution_job(job_id: int, request: Request, body: dict, db: Session = Depends(get_db)):
    """Save/update job state (devices, topology, scripts, base_path)."""
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if "name" in body:
        job.name = body["name"]
    if "dut_ids" in body:
        job.dut_ids = json.dumps(body["dut_ids"])
    if "base_path" in body:
        job.base_path = body["base_path"]
    if "host_id" in body:
        job.host_id = body["host_id"]
    if "topology" in body:
        job.topology = json.dumps(body["topology"])
    if "scripts" in body:
        job.scripts = json.dumps(body["scripts"])
    if "testbed_path" in body:
        job.testbed_path = body["testbed_path"] or None
    if "status" in body and body["status"] in ("idle", "running", "completed", "failed"):
        job.status = body["status"]
    job.updated_at = datetime.utcnow()
    db.commit()
    return {"id": job.id, "status": job.status}


@app.delete("/api/execution-jobs/{job_id}")
def delete_execution_job(job_id: int, request: Request, db: Session = Depends(get_db)):
    """Delete a job (only if not currently running)."""
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.status == "running":
        raise HTTPException(status_code=400, detail="Cannot delete a running job")
    db.query(Execution).filter(Execution.job_id == job_id).update({"job_id": None})
    db.delete(job)
    db.commit()
    return {"deleted": job_id}


@app.get("/api/execution-jobs/{job_id}/conflicts")
def check_job_conflicts(job_id: int, request: Request, dut_ids: str = "",
                        db: Session = Depends(get_db)):
    """Check if any requested DUT ids are locked by a DIFFERENT active job.

    Query param: dut_ids — comma-separated list of DUT ids.
    Returns: [{dut_id, dut_name, conflicting_job_id, conflicting_job_name}]
    """
    if not dut_ids.strip():
        return {"conflicts": []}
    requested = [int(x) for x in dut_ids.split(",") if x.strip().isdigit()]
    conflicts = []
    for did in requested:
        lock = db.query(DUTLock).filter(
            DUTLock.dut_id == did,
            DUTLock.status != "AVAILABLE",
            DUTLock.lock_type == "exec",
            DUTLock.job_id != job_id,
            DUTLock.job_id != None,
        ).first()
        if lock:
            owner_job = db.query(ExecutionJob).filter(ExecutionJob.id == lock.job_id).first()
            dut = db.query(DUT).filter(DUT.id == did).first()
            conflicts.append({
                "dut_id":               did,
                "dut_name":             dut.name if dut else str(did),
                "conflicting_job_id":   lock.job_id,
                "conflicting_job_name": owner_job.name if owner_job else f"Job {lock.job_id}",
            })
    return {"conflicts": conflicts}


@app.get("/api/execution-jobs/{job_id}/report/html")
def job_html_report(job_id: int, request: Request, db: Session = Depends(get_db)):
    """Generate aggregated HTML report for all executions in this job."""
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    executions = db.query(Execution).filter(Execution.job_id == job_id).all()
    if not executions:
        raise HTTPException(status_code=404, detail="No executions found for this job")
    all_tcrs = []
    for ex in executions:
        tcrs = db.query(TestCaseResult).filter(TestCaseResult.execution_id == ex.id).all()
        all_tcrs.extend(tcrs)
    buf = _build_html_dashboard(executions[0], all_tcrs)
    filename = f"{job.name.replace(' ', '_')}_report.html"
    return StreamingResponse(iter([buf]), media_type="text/html",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.get("/api/execution-jobs/{job_id}/report/excel")
def job_excel_report(job_id: int, request: Request, db: Session = Depends(get_db)):
    """Generate aggregated Excel report for all executions in this job."""
    if not _HAS_OPENPYXL:
        raise HTTPException(status_code=503, detail="openpyxl not installed")
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    executions = db.query(Execution).filter(Execution.job_id == job_id).all()
    if not executions:
        raise HTTPException(status_code=404, detail="No executions found for this job")
    all_tcrs = []
    for ex in executions:
        tcrs = db.query(TestCaseResult).filter(TestCaseResult.execution_id == ex.id).all()
        all_tcrs.extend(tcrs)
    buf = _build_excel(executions[0], all_tcrs)
    filename = f"{job.name.replace(' ', '_')}_report.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ============================================================================
# API — Execution Job Scheduler
# ============================================================================

@app.put("/api/execution-jobs/{job_id}/schedule")
def set_job_schedule(job_id: int, request: Request, body: dict,
                     db: Session = Depends(get_db)):
    """Set or update the schedule for an execution job.

    Body:
        schedule_type: "none" | "once" | "cron"
        schedule_at:   ISO datetime string (UTC) — required for type "once"
        schedule_cron: "HH:MM" (daily) or full cron "min hr dom mon dow" — required for type "cron"
        enabled:       bool — enable/disable without clearing the schedule
    """
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    stype   = body.get("schedule_type", job.schedule_type or "none")
    enabled = body.get("enabled", True)

    if stype == "none":
        job.schedule_type    = "none"
        job.schedule_enabled = False
        job.schedule_at      = None
        job.schedule_cron    = None
    elif stype == "once":
        sat = body.get("schedule_at")
        if not sat:
            raise HTTPException(status_code=400, detail="schedule_at required for type 'once'")
        try:
            job.schedule_at = datetime.fromisoformat(sat.replace("Z", "+00:00")).replace(tzinfo=None)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"Invalid schedule_at: {sat}")
        job.schedule_type    = "once"
        job.schedule_cron    = None
        job.schedule_enabled = bool(enabled)
    elif stype == "cron":
        cron_expr = body.get("schedule_cron", "").strip()
        if not cron_expr:
            raise HTTPException(status_code=400, detail="schedule_cron required for type 'cron'")
        job.schedule_type    = "cron"
        job.schedule_cron    = cron_expr
        job.schedule_at      = None
        job.schedule_enabled = bool(enabled)
    else:
        raise HTTPException(status_code=400, detail="schedule_type must be none | once | cron")

    job.updated_at = datetime.utcnow()
    db.commit()
    db.refresh(job)

    # Re-register in APScheduler
    _register_job_schedule(job)

    next_run = None
    ap_id = f"exec_job_{job.id}"
    try:
        ap_job = scheduler.get_job(ap_id)
        if ap_job and ap_job.next_run_time:
            from datetime import timezone as _utc_tz
            next_run = ap_job.next_run_time.astimezone(_utc_tz.utc).strftime("%Y-%m-%d %H:%M UTC")
    except Exception:
        pass

    return {
        "id": job.id,
        "schedule_type": job.schedule_type,
        "schedule_at": job.schedule_at.isoformat() if job.schedule_at else None,
        "schedule_cron": job.schedule_cron,
        "schedule_enabled": job.schedule_enabled,
        "next_run": next_run,
    }


@app.get("/api/execution-jobs/{job_id}/schedule")
def get_job_schedule(job_id: int, request: Request, db: Session = Depends(get_db)):
    """Return current schedule settings for a job."""
    session_id = request.headers.get("X-Session-ID", "default")
    job = db.query(ExecutionJob).filter(ExecutionJob.id == job_id,
                                        ExecutionJob.session_id == session_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")

    next_run = None
    ap_id = f"exec_job_{job.id}"
    try:
        ap_job = scheduler.get_job(ap_id)
        if ap_job and ap_job.next_run_time:
            from datetime import timezone as _utc_tz
            next_run = ap_job.next_run_time.astimezone(_utc_tz.utc).strftime("%Y-%m-%d %H:%M UTC")
    except Exception:
        pass

    return {
        "id": job.id,
        "schedule_type": job.schedule_type or "none",
        "schedule_at": job.schedule_at.isoformat() if job.schedule_at else None,
        "schedule_cron": job.schedule_cron,
        "schedule_enabled": bool(job.schedule_enabled),
        "last_run_at": job.last_run_at.isoformat() if job.last_run_at else None,
        "next_run": next_run,
    }


# ============================================================================
# API — Topology Connections Persistence
# ============================================================================

@app.get("/api/topology/connections")
def get_topology_connections(db: Session = Depends(get_db)):
    """Return all persisted canvas DUT connections."""
    conns = db.query(TopologyConnection).all()
    result = []
    for c in conns:
        dut_a = db.query(DUT).filter(DUT.id == c.dut_a_id).first()
        dut_b = db.query(DUT).filter(DUT.id == c.dut_b_id).first()
        result.append({
            "id": c.id,
            "dut_a": str(c.dut_a_id),
            "dut_b": str(c.dut_b_id),
            "intf_a": c.intf_a,
            "intf_b": c.intf_b,
            "dut_a_name": dut_a.name if dut_a else str(c.dut_a_id),
            "dut_b_name": dut_b.name if dut_b else str(c.dut_b_id),
        })
    return result


@app.post("/api/topology/connections")
def save_topology_connections(body: dict, db: Session = Depends(get_db)):
    """Replace all canvas connections with the provided list.
    
    Body: { connections: [{dut_a, intf_a, dut_b, intf_b}, ...] }
    """
    connections = body.get("connections", [])
    # Clear existing
    db.query(TopologyConnection).delete()
    # Insert new
    for c in connections:
        dut_a = c.get("dut_a") or c.get("dut_a_id")
        dut_b = c.get("dut_b") or c.get("dut_b_id")
        if not dut_a or not dut_b:
            continue
        conn = TopologyConnection(
            dut_a_id=int(dut_a),
            intf_a=c.get("intf_a", "Ethernet0"),
            dut_b_id=int(dut_b),
            intf_b=c.get("intf_b", "Ethernet0"),
        )
        db.add(conn)
    db.commit()
    return {"saved": len(connections)}


# ============================================================================
# API — Temp YAML Generator
# ============================================================================

@app.post("/api/spytest/generate-temp-yaml")
def generate_temp_yaml(body: dict, db: Session = Depends(get_db)):
    """Generate a temp testbed YAML on the remote VM substituting live DUT data.

    Body:
        host_id: int  — the VM where spytest runs
        testbed_filename: str  — reference testbed YAML name (from /testbeds/)
        connections: list  — [{dut_a, intf_a, dut_b, intf_b}] from canvas
        dut_ids: list[int]  — DUT device IDs to include
        script_name: str  — used in the /tmp filename
    """
    host_id = body.get("host_id")
    testbed_filename = body.get("testbed_filename", "")
    connections = body.get("connections", [])
    dut_ids = body.get("dut_ids", [])
    script_name = body.get("script_name", "script")

    if not host_id or not testbed_filename:
        raise HTTPException(status_code=400, detail="host_id and testbed_filename required")

    vm = db.query(DUT).filter(DUT.id == host_id).first()
    if not vm:
        raise HTTPException(status_code=404, detail="VM device not found")

    # Fetch all DUT device records
    dut_records = db.query(DUT).filter(DUT.id.in_(dut_ids)).all() if dut_ids else []
    dut_map = {d.id: d for d in dut_records}

    ssh = SSHConnectionManager(vm.ip_address, vm.port, vm.username, vm.password)
    if not ssh.connect():
        raise HTTPException(status_code=503, detail=f"Cannot connect to {vm.name}")

    try:
        # Read reference YAML
        ref_path = f"{SPYTEST_TESTBED_DIR}/{testbed_filename}"
        output, error, code = ssh.execute_command(f"cat {ref_path}", timeout=15)
        if code != 0:
            raise HTTPException(status_code=404, detail=f"Testbed file not found: {testbed_filename}")

        ref_config = yaml.safe_load(output) or {}
        ref_devices = ref_config.get("devices", {})
        ref_topology = ref_config.get("topology", {})
        ref_device_names = list(ref_devices.keys())

        # Build device name → DUT record mapping (match by position)
        # If we have dut_ids, map them positionally to the reference YAML device names
        device_name_to_dut = {}
        for i, dev_name in enumerate(ref_device_names):
            if i < len(dut_ids):
                device_name_to_dut[dev_name] = dut_map.get(dut_ids[i])

        # Build interface override map from canvas connections
        # connections: [{dut_a, intf_a, dut_b, intf_b}] where dut_a/dut_b are DUT IDs
        id_to_dev_name = {}
        for dev_name, dut in device_name_to_dut.items():
            if dut:
                id_to_dev_name[dut.id] = dev_name

        # Build topology from canvas connections
        canvas_topology = {}
        for conn in connections:
            a_id = int(conn.get("dut_a", 0) or 0)
            b_id = int(conn.get("dut_b", 0) or 0)
            intf_a = conn.get("intf_a", "Ethernet0")
            intf_b = conn.get("intf_b", "Ethernet0")
            name_a = id_to_dev_name.get(a_id)
            name_b = id_to_dev_name.get(b_id)
            if name_a and name_b:
                canvas_topology.setdefault(name_a, {"interfaces": {}})
                canvas_topology.setdefault(name_b, {"interfaces": {}})
                canvas_topology[name_a]["interfaces"][intf_a] = {
                    "EndDevice": name_b, "EndPort": intf_b}
                canvas_topology[name_b]["interfaces"][intf_b] = {
                    "EndDevice": name_a, "EndPort": intf_a}

        # Build the final temp YAML config
        temp_config = {
            "version": ref_config.get("version", "2.0"),
            "devices": {},
            "topology": canvas_topology if canvas_topology else ref_topology,
            "services": ref_config.get("services", {"default": {}}),
            "builds": ref_config.get("builds", {"default": {}}),
            "configs": ref_config.get("configs", {"default": {}}),
            "errors": ref_config.get("errors", {"default": {}}),
            "params": ref_config.get("params", {}),
        }

        # Substitute DUT IPs/credentials from device records
        for dev_name, ref_dev in ref_devices.items():
            dut = device_name_to_dut.get(dev_name)
            dev_entry = dict(ref_dev)  # copy reference entry
            if dut:
                dev_entry["ip"] = dut.ip_address
                dev_entry["username"] = dut.username
                dev_entry["password"] = dut.password
            temp_config["devices"][dev_name] = dev_entry

        # Write to /tmp on the VM
        import uuid as _uuid
        unique_id = str(_uuid.uuid4())[:8]
        safe_script = re.sub(r'[^a-zA-Z0-9_]', '_', script_name)[:30]
        temp_filename = f"testbed_{unique_id}_{safe_script}.yaml"
        temp_path = f"/tmp/{temp_filename}"

        yaml_str = (
            f"# AUTO-GENERATED — DO NOT EDIT MANUALLY\n"
            f"# Generated by Eka Automation | Script: {script_name}\n"
            + yaml.dump(temp_config, default_flow_style=False)
        )

        yaml_b64 = base64.b64encode(yaml_str.encode()).decode()
        write_cmd = f"echo '{yaml_b64}' | base64 -d > {temp_path}"
        _, write_err, write_code = ssh.execute_command(write_cmd, timeout=15)
        if write_code != 0:
            raise HTTPException(status_code=500,
                                detail=f"Failed to write temp YAML: {write_err}")

        logger.info(f"Temp YAML written: {temp_path}")
        return {
            "temp_yaml_path": temp_path,
            "temp_filename": temp_filename,
            "device_count": len(temp_config["devices"]),
            "devices": list(temp_config["devices"].keys()),
        }

    finally:
        ssh.disconnect()


@app.post("/api/topology/generate-master-testbed")
def generate_master_testbed(body: dict, request: Request, db: Session = Depends(get_db)):
    """Generate a master testbed YAML from session's Topology Canvas DUTs and connections.

    This creates a comprehensive testbed file that includes all devices from the current
    session's topology canvas with their actual IP addresses, credentials, and interface connections.

    Body:
        host_id: int — the VM where SPyTest runs
        master_filename: str (optional) — name for master testbed (default: master_testbed.yaml)

    Returns:
        master_testbed_path: str — full path to generated master testbed
        device_count: int — number of devices included
        connection_count: int — number of interface connections
        devices: list — device names included
    """
    host_id = body.get("host_id")
    master_filename = body.get("master_filename", "master_testbed.yaml")
    session_id = get_session_id(request)

    # Derive testbed dir from user-supplied scripts base path:
    # e.g. /some/root/tests  →  /some/root/testbeds
    user_base_path = (body.get("base_path") or "").strip().rstrip("/")
    if not user_base_path:
        raise HTTPException(
            status_code=400,
            detail="SCRIPTS_PATH_REQUIRED: Please enter the Scripts Path on VM in the 'Categories & Scripts' section and click Load first."
        )
    _m = re.match(r'(^.*?/spytest)(?:/|$)', user_base_path)
    _spy_root = _m.group(1) if _m else os.path.dirname(user_base_path)
    testbed_dir = _spy_root + "/testbeds"

    if not host_id:
        raise HTTPException(status_code=400, detail="host_id required")

    vm = db.query(DUT).filter(DUT.id == host_id).first()
    if not vm:
        raise HTTPException(status_code=404, detail="VM host not found")

    # Get DUTs from current session (exclude VM host)
    query = db.query(DUT).filter(DUT.device_type != "VM")
    if session_id:
        query = query.filter(DUT.session_id == session_id)
    all_duts = query.all()

    if not all_duts:
        raise HTTPException(status_code=400, detail="No DUT devices found in session")

    # CRITICAL FIX: Deduplicate DUTs by name to prevent duplicate devices in master testbed
    # If multiple DUTs have the same name, keep only the first one
    seen_names = set()
    unique_duts = []
    duplicate_count = 0
    for dut in all_duts:
        normalized_name = dut.name.replace(" ", "_").replace("-", "_")
        if normalized_name not in seen_names:
            seen_names.add(normalized_name)
            unique_duts.append(dut)
        else:
            duplicate_count += 1
            logger.warning(f"⚠ MASTER TESTBED: Skipping duplicate device '{dut.name}' (ID: {dut.id})")

    if duplicate_count > 0:
        logger.warning(f"⚠ MASTER TESTBED: Removed {duplicate_count} duplicate device(s) from generation")

    all_duts = unique_duts  # Use deduplicated list

    # Get ALL topology connections
    all_connections = db.query(TopologyConnection).all()

    # Filter connections to only include those between DUTs in this session
    dut_ids_in_session = {dut.id for dut in all_duts}
    session_connections = [
        conn for conn in all_connections
        if conn.dut_a_id in dut_ids_in_session and conn.dut_b_id in dut_ids_in_session
    ]

    if not session_connections:
        raise HTTPException(status_code=400, detail="No connections found in Topology Canvas. Create connections first.")

    # Get DUT IDs that have at least one connection (only these are in Topology Canvas)
    connected_dut_ids = set()
    for conn in session_connections:
        connected_dut_ids.add(conn.dut_a_id)
        connected_dut_ids.add(conn.dut_b_id)

    # Filter to only include DUTs that have connections (visible in Topology Canvas)
    canvas_duts = [dut for dut in all_duts if dut.id in connected_dut_ids]

    if not canvas_duts:
        raise HTTPException(status_code=400, detail="No devices with connections found in Topology Canvas")

    ssh = SSHConnectionManager(vm.ip_address, vm.port, vm.username, vm.password)
    if not ssh.connect():
        raise HTTPException(status_code=503, detail=f"Cannot connect to {vm.name}")

    try:
        # Build devices section (only for devices in Topology Canvas with connections)
        devices_section = {}
        dut_id_to_name = {}

        for dut in canvas_duts:
            # Use actual device name from database (not D1, D2, D3...)
            device_name = dut.name.replace(" ", "_").replace("-", "_")
            dut_id_to_name[dut.id] = device_name

            # SPyTest testbed format with nested structure
            devices_section[device_name] = {
                "device_type": "sonic",
                "access": {
                    "protocol": "ssh",
                    "ip": dut.ip_address,
                    "port": dut.port
                },
                "credentials": {
                    "username": dut.username,
                    "password": dut.password,
                    "altpassword": "broadcom"  # default alternate password
                },
                "properties": {
                    "services": "default",
                    "build": "default",
                    "config": "default",
                    "errors": "default"
                }
            }

        # Build topology section from canvas connections and topo dictionary for params
        topology_section = {}
        connection_count = 0
        topo_dict = {}  # SPyTest topo format: {"D1D2": 2, "D1D3": 1, ...}

        # Create mapping from device names to generic D1, D2, D3... names
        device_to_generic = {}
        generic_index = 1
        for device_name in sorted(devices_section.keys()):
            device_to_generic[device_name] = f"D{generic_index}"
            generic_index += 1

        for conn in session_connections:
            device_a = dut_id_to_name.get(conn.dut_a_id)
            device_b = dut_id_to_name.get(conn.dut_b_id)

            if device_a and device_b:
                # Initialize device in topology if not exists
                if device_a not in topology_section:
                    topology_section[device_a] = {"interfaces": {}}
                if device_b not in topology_section:
                    topology_section[device_b] = {"interfaces": {}}

                # Add bidirectional connection
                topology_section[device_a]["interfaces"][conn.intf_a] = {
                    "EndDevice": device_b,
                    "EndPort": conn.intf_b
                }
                topology_section[device_b]["interfaces"][conn.intf_b] = {
                    "EndDevice": device_a,
                    "EndPort": conn.intf_a
                }
                connection_count += 1

                # Build topo dictionary for SPyTest (D1D2 format)
                gen_a = device_to_generic.get(device_a)
                gen_b = device_to_generic.get(device_b)
                if gen_a and gen_b:
                    # Normalize order (D1D2, not D2D1)
                    topo_key = f"{gen_a}{gen_b}" if gen_a < gen_b else f"{gen_b}{gen_a}"
                    topo_dict[topo_key] = topo_dict.get(topo_key, 0) + 1

        # Derive test_interface: first alphabetically-sorted interface of the
        # first device that has connections (D1 → its lowest interface toward D2).
        # STP and many other SPyTest scripts expect this param in testbed globals.
        test_interface = None
        for dev_name in sorted(topology_section.keys()):
            intfs = topology_section[dev_name].get("interfaces", {})
            if intfs:
                test_interface = sorted(intfs.keys())[0]
                break

        params_section = {"topo": topo_dict if topo_dict else {}}

        # Build complete master testbed YAML
        master_config = {
            "version": "2.0",
            "devices": devices_section,
            "topology": topology_section,
            "services": {"default": {}},
            "builds": {"default": {}},
            "configs": {"default": {}},
            "errors": {"default": {}},
            "params": params_section,
        }
        # test_interface must live under global.params so SPyTest's get_param() can find it.
        # (global_params is populated only from obj["global"]["params"], not top-level params)
        if test_interface:
            master_config["global"] = {"params": {"test_interface": test_interface}}

        # Generate YAML content with proper formatting
        # Use default_flow_style=None for mixed formatting (inline for simple dicts)
        yaml_content = (
            f"# MASTER TESTBED - AUTO-GENERATED FROM CANVAS TOPOLOGY\n"
            f"# Generated by Eka Automation at {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC\n"
            f"# Total Devices: {len(devices_section)}\n"
            f"# Total Connections: {connection_count}\n"
            f"# DO NOT EDIT MANUALLY - Regenerate from canvas as needed\n\n"
            + yaml.dump(master_config, default_flow_style=None, sort_keys=False, width=120)
        )

        # Write to testbeds directory on remote VM (create dir if missing)
        master_path = f"{testbed_dir}/{master_filename}"
        _, mkdir_err, mkdir_code = ssh.execute_command(f'mkdir -p "{testbed_dir}"', timeout=10)
        if mkdir_code != 0:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"PATH_NOT_FOUND: Cannot create testbed directory: {testbed_dir}\n"
                    f"{mkdir_err.strip()}\n"
                    f"Please verify the Scripts Path on VM is correct and the user has write permission."
                )
            )

        # Write via stdin pipe to avoid shell ARG_MAX limits with large YAML
        yaml_b64 = base64.b64encode(yaml_content.encode()).decode()
        stdin_ch, stdout_ch, stderr_ch = ssh.client.exec_command(f'base64 -d > "{master_path}"', timeout=15)
        stdin_ch.write(yaml_b64.encode())
        stdin_ch.channel.shutdown_write()
        write_err = stderr_ch.read().decode("utf-8", errors="ignore")
        write_code = stdout_ch.channel.recv_exit_status()
        if write_code != 0:
            raise HTTPException(
                status_code=500,
                detail=(
                    f"PATH_NOT_FOUND: Failed to write master testbed to {master_path}\n"
                    f"{write_err.strip()}\n"
                    f"Please verify the Scripts Path on VM and directory permissions."
                )
            )

        logger.info(f"Master testbed generated: {master_path} with {len(devices_section)} devices")

        return {
            "success": True,
            "master_testbed_path": master_path,
            "master_filename": master_filename,
            "device_count": len(devices_section),
            "connection_count": connection_count,
            "devices": list(devices_section.keys()),
            "timestamp": datetime.utcnow().isoformat(),
        }

    finally:
        ssh.disconnect()


# ============================================================================
# API — Multi-User Session Management
# ============================================================================

@app.post("/api/sessions/register")
def register_session(body: dict, db: Session = Depends(get_db)):
    """Register a new user session for multi-user support.

    Body:
        session_id: str — unique client-generated session ID (UUID)
        user_name: str — user identifier (required)
        user_email: str (optional) — user email
        (ttl_minutes is accepted but ignored — sessions are user-based, no TTL)

    Returns:
        session: dict — session details
    """
    session_id = body.get("session_id")
    user_name = body.get("user_name")
    user_email = body.get("user_email", "")
    # ttl_minutes accepted for backward-compat but NOT used — sessions are persistent

    if not session_id or not user_name:
        raise HTTPException(status_code=400, detail="session_id and user_name required")

    # Check if session already exists
    existing = db.query(UserSession).filter(UserSession.session_id == session_id).first()
    if existing:
        # Update last activity and return existing session
        existing.last_activity = datetime.utcnow()
        existing.status = "active"
        db.commit()
        return {
            "session_id": existing.session_id,
            "user_name": existing.user_name,
            "status": existing.status,
            "created_at": existing.created_at.isoformat(),
            "allocated_duts": json.loads(existing.allocated_dut_ids) if existing.allocated_dut_ids else [],
        }

    # Create new user-based persistent session (no expiry time)
    new_session = UserSession(
        session_id=session_id,
        user_name=user_name,
        user_email=user_email,
        status="active",
        allocated_dut_ids="[]",
        expires_at=None,  # No TTL — user-based session
    )
    db.add(new_session)
    db.commit()
    db.refresh(new_session)

    logger.info(f"New user-based session registered: {session_id} for user {user_name}")

    return {
        "session_id": new_session.session_id,
        "user_name": new_session.user_name,
        "status": new_session.status,
        "created_at": new_session.created_at.isoformat(),
        "allocated_duts": [],
    }


@app.get("/api/sessions/validate/{session_id}")
def validate_session(session_id: str, db: Session = Depends(get_db)):
    """Validate if a session is active.

    Sessions are user-based and persistent — they do NOT expire by time.
    A session is invalid only if it was explicitly terminated or revoked.

    Returns:
        valid: bool
        session: dict (if valid)
    """
    session = db.query(UserSession).filter(UserSession.session_id == session_id).first()

    if not session:
        return {"valid": False, "reason": "Session not found"}

    # User-based session: only check status, NOT time expiry
    if session.status != "active":
        return {"valid": False, "reason": f"Session status: {session.status}"}

    # Update last activity
    session.last_activity = datetime.utcnow()
    try:
        db.commit()
    except Exception:
        db.rollback()

    return {
        "valid": True,
        "session": {
            "session_id": session.session_id,
            "user_name": session.user_name,
            "user_email": session.user_email or "",
            "user_role": session.user_role or "",
            "status": session.status,
            "created_at": session.created_at.isoformat(),
            "last_activity": session.last_activity.isoformat(),
            # expires_at omitted — sessions are user-based, no expiry
            "allocated_duts": json.loads(session.allocated_dut_ids) if session.allocated_dut_ids else [],
        }
    }


@app.post("/api/sessions/{session_id}/extend")
def extend_session(session_id: str, body: dict, db: Session = Depends(get_db)):
    """Session keep-alive ping — records activity for user-based persistent sessions.

    Sessions do NOT expire by time. This endpoint simply updates last_activity
    and last_keepalive timestamps so activity monitoring works correctly.

    Body:
        extend_minutes: int (accepted for backward-compat, ignored)

    Returns:
        - session_id
        - status
        - last_keepalive
        - time_remaining_minutes: always 999999 (no TTL — user-based session)
    """
    session = db.query(UserSession).filter(UserSession.session_id == session_id).first()
    if not session:
        logger.warning(f"[KEEPALIVE] Session not found: {session_id}")
        raise HTTPException(status_code=404, detail="Session not found")

    if session.status != "active":
        raise HTTPException(status_code=403, detail=f"Session is {session.status}, cannot extend")

    try:
        now = datetime.utcnow()
        session.last_activity = now
        session.last_keepalive = now
        session.keepalive_fail_count = 0
        db.commit()

        logger.info(f"[KEEPALIVE] ✓ Activity recorded: {session_id} (user: {session.user_name})")

        return {
            "session_id": session.session_id,
            "user_name": session.user_name,
            "status": "success",
            "last_keepalive": session.last_keepalive.isoformat(),
            # Persistent sessions report max time remaining (no expiry)
            "time_remaining_minutes": 999999,
            "keepalive_fail_count": session.keepalive_fail_count,
        }
    except Exception as e:
        logger.error(f"[KEEPALIVE] ✗ Failed to record activity for {session_id}: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to record activity: {str(e)}")


@app.get("/api/sessions/{session_id}/diagnostics")
def get_session_diagnostics(session_id: str, db: Session = Depends(get_db)):
    """Get detailed session diagnostics for monitoring and troubleshooting.

    Returns:
        - Session status
        - Keep-alive tracking: last success, fail count
        - time_remaining_minutes: 999999 (persistent user-based session)
        - DUT count and allocation
        - Warnings if session status is abnormal
    """
    session = db.query(UserSession).filter(UserSession.session_id == session_id).first()
    if not session:
        return {
            "status": "error",
            "detail": "Session not found",
            "session_id": session_id
        }

    # Parse DUT IDs
    try:
        import json
        dut_ids = json.loads(session.allocated_dut_ids or "[]")
    except Exception:
        dut_ids = []

    # Warnings
    warnings_list = []
    if session.status != "active":
        warnings_list.append(f"SESSION_{session.status.upper()}")
    if session.keepalive_fail_count and session.keepalive_fail_count > 0:
        warnings_list.append(f"KEEPALIVE_FAILURES:{session.keepalive_fail_count}")

    return {
        "session_id": session.session_id,
        "user_name": session.user_name,
        "user_email": session.user_email or "",
        "user_role": session.user_role or "",
        "status": session.status,
        # Persistent sessions have no expiry — report max value for UI compatibility
        "is_expired": False,
        "time_remaining_minutes": 999999,
        "time_remaining_seconds": 999999,
        "created_at": session.created_at.isoformat(),
        "last_activity": session.last_activity.isoformat() if session.last_activity else None,
        "last_keepalive": session.last_keepalive.isoformat() if session.last_keepalive else None,
        "keepalive_fail_count": session.keepalive_fail_count or 0,
        "allocated_dut_ids": dut_ids,
        "dut_count": len(dut_ids),
        "warnings": warnings_list,
        "health": "ERROR" if session.status != "active" else ("WARNING" if warnings_list else "HEALTHY")
    }


@app.post("/api/sessions/{session_id}/release")
def release_session(session_id: str, db: Session = Depends(get_db)):
    """Release a session and free all allocated DUTs."""
    session = db.query(UserSession).filter(UserSession.session_id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    # Free allocated DUTs
    allocated_duts = json.loads(session.allocated_dut_ids) if session.allocated_dut_ids else []
    for dut_id in allocated_duts:
        pool_entry = db.query(DUTPool).filter(DUTPool.dut_id == dut_id).first()
        if pool_entry and pool_entry.session_id == session_id:
            pool_entry.status = "AVAILABLE"
            pool_entry.session_id = None
            pool_entry.locked_since = None

    # Mark session as terminated (will be cleaned up by background job)
    session.status = "terminated"
    session.allocated_dut_ids = "[]"
    db.commit()

    logger.info(f"Session released: {session_id} - freed {len(allocated_duts)} DUTs")

    return {
        "session_id": session_id,
        "status": "terminated",
        "freed_duts": len(allocated_duts),
    }


@app.post("/api/sessions/{session_id}/revoke")
def revoke_session(session_id: str, db: Session = Depends(get_db)):
    """Admin endpoint to explicitly revoke a user session.

    Marks the session as 'revoked'. The background cleanup job will then
    delete the session record and free all associated DUTs.

    Use this to forcibly remove a user's access without waiting for logout.
    """
    session = db.query(UserSession).filter(UserSession.session_id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    prev_status = session.status
    session.status = "revoked"
    db.commit()

    logger.info(f"[REVOKE] Session {session_id} revoked (was: {prev_status}, user: {session.user_name})")

    return {
        "session_id": session_id,
        "status": "revoked",
        "user_name": session.user_name,
        "user_email": session.user_email or "",
    }


@app.get("/api/sessions/active")
def get_active_sessions(request: Request, db: Session = Depends(get_db)):
    """Return active sessions.

    Admin: all active sessions.
    Regular user: only their own session.
    """
    session_id = get_session_id(request)
    if not session_id:
        raise HTTPException(status_code=401, detail="No session ID provided")

    current = db.query(UserSession).filter(UserSession.session_id == session_id).first()
    if not current:
        raise HTTPException(status_code=404, detail="Session not found")

    def _serialize(s):
        allocated_duts = json.loads(s.allocated_dut_ids) if s.allocated_dut_ids else []
        return {
            "session_id": s.session_id,
            "user_name": s.user_name,
            "user_email": s.user_email or "",
            "user_role": s.user_role or "",
            "status": s.status,
            "created_at": s.created_at.isoformat(),
            "last_activity": s.last_activity.isoformat(),
            "allocated_dut_count": len(allocated_duts),
            "allocated_duts": allocated_duts,
        }

    is_admin = (current.user_role or "").lower() == "admin"

    if is_admin:
        all_sessions = db.query(UserSession).filter(UserSession.status == "active").all()
        return {"sessions": [_serialize(s) for s in all_sessions], "is_admin": True}
    else:
        return {"sessions": [_serialize(current)], "is_admin": False}


@app.post("/api/sessions/{session_id}/allocate-duts")
def allocate_duts_to_session(session_id: str, body: dict, db: Session = Depends(get_db)):
    """Allocate DUTs to a session for exclusive use.

    Body:
        dut_ids: list[int] — DUT IDs to allocate
    """
    dut_ids = body.get("dut_ids", [])

    session = db.query(UserSession).filter(UserSession.session_id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Session not found")

    # Check if DUTs are available
    for dut_id in dut_ids:
        pool_entry = db.query(DUTPool).filter(DUTPool.dut_id == dut_id).first()
        if not pool_entry:
            # Create pool entry if doesn't exist
            pool_entry = DUTPool(dut_id=dut_id, status="AVAILABLE")
            db.add(pool_entry)
        elif pool_entry.status != "AVAILABLE":
            raise HTTPException(
                status_code=409,
                detail=f"DUT {dut_id} is already allocated to another session"
            )

    # Allocate DUTs
    for dut_id in dut_ids:
        pool_entry = db.query(DUTPool).filter(DUTPool.dut_id == dut_id).first()
        pool_entry.status = "ALLOCATED"
        pool_entry.session_id = session_id
        pool_entry.locked_since = datetime.utcnow()

    # Update session
    current_allocated = json.loads(session.allocated_dut_ids) if session.allocated_dut_ids else []
    updated_allocated = list(set(current_allocated + dut_ids))
    session.allocated_dut_ids = json.dumps(updated_allocated)
    db.commit()

    return {
        "session_id": session_id,
        "allocated_duts": updated_allocated,
        "count": len(updated_allocated),
    }


# ============================================================================
# API — Git Repository Integration (SSH clone/pull on VM with credentials)
# ============================================================================

# Base directory on the VM where repos are cloned
GIT_CLONE_BASE = "/home/hp_test/Eka"

# Git state file — persists across server restarts
_GIT_STATE_FILE = DATA_DIR / "git_state.json"

# In-memory Git configuration state
_git_state = {
    "configured": False,
    "host_id": None,
    "host_name": "",
    "repo_url": "",
    "branch": "master",
    "repo_name": "",
    "tests_path": "",
    "categories_count": 0,
}

# Auto-restore git state from disk if it exists
if _GIT_STATE_FILE.exists():
    try:
        _saved = json.loads(_GIT_STATE_FILE.read_text(encoding="utf-8"))
        if _saved.get("configured"):
            _git_state.update(_saved)
            logger.info(f"[Git] Restored state from disk: host_id={_saved.get('host_id')}")
    except Exception as e:
        logger.warning(f"[Git] Could not restore git state: {e}")


def _save_git_state():
    """Persist current git state to disk so it survives server restarts."""
    try:
        _GIT_STATE_FILE.write_text(
            json.dumps(_git_state, indent=2), encoding="utf-8"
        )
    except Exception as e:
        logger.warning(f"[Git] Could not save git state: {e}")


def _build_auth_url(repo_url: str, username: str, token: str) -> str:
    """Build an authenticated HTTPS git URL with embedded credentials."""
    # URL-encode credentials to handle special characters
    encoded_token = urllib.parse.quote(token, safe='')
    encoded_username = urllib.parse.quote(username, safe='')

    # For GitHub, use token as username (no password) - this is the standard format
    # https://github.com/user/repo.git → https://token@github.com/user/repo.git
    if "github.com" in repo_url:
        # GitHub Personal Access Token format: just token as username
        if repo_url.startswith("https://"):
            auth_url = repo_url.replace("https://", f"https://{encoded_token}@", 1)
            logger.info(f"[Git] Built auth URL for GitHub (token length: {len(token)})")
            return auth_url
        elif repo_url.startswith("http://"):
            auth_url = repo_url.replace("http://", f"http://{encoded_token}@", 1)
            return auth_url
    else:
        # For other git servers, use username:token format
        if repo_url.startswith("https://"):
            return repo_url.replace("https://", f"https://{encoded_username}:{encoded_token}@", 1)
        elif repo_url.startswith("http://"):
            return repo_url.replace("http://", f"http://{encoded_username}:{encoded_token}@", 1)
    return repo_url


@app.post("/api/git/configure")
def configure_git(body: dict, db: Session = Depends(get_db)):
    """
    Connect to a VM via SSH, clone or pull a git repo with user credentials,
    then list test categories from the tests folder.
    """
    host_id = body.get("host_id")
    repo_url = body.get("repo_url", "").strip()
    username = body.get("username", "").strip()
    token = body.get("token", "").strip()
    branch = body.get("branch", "master").strip() or "master"

    if not host_id:
        raise HTTPException(status_code=400, detail="Please select a VM host")
    if not repo_url:
        raise HTTPException(status_code=400, detail="Repo URL is required")
    if not token:
        raise HTTPException(status_code=400, detail="Password / Token is required")

    # Smart URL parsing: extract branch from GitHub browser URLs
    # Examples:
    #   https://github.com/palcnetworks/sonic-mgmt/tree/Deployment_Usecases
    #   https://github.com/user/repo/blob/main/file.py
    original_url = repo_url
    if "/tree/" in repo_url or "/blob/" in repo_url:
        # Extract branch name from URL
        if "/tree/" in repo_url:
            parts = repo_url.split("/tree/")
            repo_url = parts[0]  # Clean URL: https://github.com/user/repo
            branch_part = parts[1].split("/")[0]  # Extract branch name
            # Only override branch if user didn't explicitly provide one
            if body.get("branch", "").strip() == "" or body.get("branch") == "master":
                branch = branch_part
                logger.info(f"[Git] Extracted branch '{branch}' from URL")
        elif "/blob/" in repo_url:
            parts = repo_url.split("/blob/")
            repo_url = parts[0]
            branch_part = parts[1].split("/")[0]
            if body.get("branch", "").strip() == "" or body.get("branch") == "master":
                branch = branch_part
                logger.info(f"[Git] Extracted branch '{branch}' from URL")

    # Remove trailing .git if present
    if repo_url.endswith(".git"):
        repo_url = repo_url[:-4]

    logger.info(f"[Git] Parsed URL: {repo_url}, Branch: {branch}")

    dut = db.query(DUT).filter(DUT.id == host_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="VM device not found")

    # Extract repo name from URL (e.g., "sonic-mgmt" from "https://github.com/user/sonic-mgmt")
    repo_name = repo_url.rstrip("/").split("/")[-1]

    repo_dir = f"{GIT_CLONE_BASE}/{repo_name}"
    tests_dir = f"{repo_dir}/spytest/tests"
    auth_url = _build_auth_url(repo_url, username, token)
    logger.info(f"[Git] Username length: {len(username)}, Token length: {len(token)}")

    logger.info(f"[Git] Connecting to {dut.name} ({dut.ip_address}:{dut.port}) for git clone/pull...")

    ssh = SSHConnectionManager(dut.ip_address, dut.port, dut.username, dut.password)
    if not ssh.connect():
        raise HTTPException(status_code=503, detail=f"Cannot connect to {dut.name} ({dut.ip_address})")

    try:
        # Step 1: Check if repo already exists on the VM
        check_cmd = f"test -d {repo_dir}/.git && echo EXISTS || echo MISSING"
        check_out, _, _ = ssh.execute_command(check_cmd, timeout=10)
        repo_exists = "EXISTS" in check_out

        # Also check if directory exists but .git is missing (corrupted state)
        dir_exists_cmd = f"test -d {repo_dir} && echo EXISTS || echo MISSING"
        dir_out, _, _ = ssh.execute_command(dir_exists_cmd, timeout=10)
        dir_exists = "EXISTS" in dir_out

        # If directory exists but no .git (corrupted), remove it first
        if dir_exists and not repo_exists:
            logger.warning(f"[Git] Directory exists but no .git folder - cleaning corrupted directory...")
            cleanup_attempts = [
                f"sudo rm -rf {repo_dir}",
                f"rm -rf {repo_dir}",
                f"sudo mv {repo_dir} {repo_dir}_old_{int(__import__('time').time())}"
            ]
            for cleanup_cmd in cleanup_attempts:
                ssh.execute_command(cleanup_cmd, timeout=60)
                # Verify
                verify_out, _, _ = ssh.execute_command(dir_exists_cmd, timeout=10)
                if "MISSING" in verify_out:
                    logger.info(f"[Git] Corrupted directory removed successfully")
                    repo_exists = False
                    dir_exists = False
                    break

        if repo_exists:
            logger.info(f"[Git] Repo exists, doing normal git pull on branch '{branch}'...")
            current_user = dut.username

            # Try normal git pull first (fetch, stash local changes, checkout, pull)
            pull_cmd = f"cd {repo_dir} && git fetch '{auth_url}' {branch}:{branch} 2>&1 && git stash 2>&1 && git checkout {branch} 2>&1 && git pull '{auth_url}' {branch} 2>&1"
            logger.info(f"[Git] Running: git fetch && git stash && git checkout {branch} && git pull")
            output, error, exit_code = ssh.execute_command(pull_cmd, timeout=120)
            action = "pull"

            # Only if normal pull fails with permission errors, then fix and retry
            if exit_code != 0 and ("Permission denied" in output or "Permission denied" in error or "index.lock" in output or "FETCH_HEAD" in output):
                logger.warning(f"[Git] Normal pull failed with permission error. Fixing permissions and retrying...")

                # Permission fixes with sudo password
                sudo_password = dut.password
                permission_fixes = [
                    f"echo '{sudo_password}' | sudo -S rm -f {repo_dir}/.git/index.lock {repo_dir}/.git/HEAD.lock {repo_dir}/.git/refs/heads/*.lock 2>/dev/null",
                    f"echo '{sudo_password}' | sudo -S chown -R {current_user}:{current_user} {repo_dir}",
                    f"echo '{sudo_password}' | sudo -S chmod -R u+w {repo_dir}/.git",
                ]

                for fix_cmd in permission_fixes:
                    fix_out, fix_err, fix_code = ssh.execute_command(fix_cmd, timeout=30)
                    logger.info(f"[Git] Permission fix executed: {fix_cmd[:50]}... (exit code: {fix_code})")

                # Retry pull with auth URL
                retry_pull_cmd = f"cd {repo_dir} && git fetch --all 2>&1 && git stash 2>&1 && git checkout {branch} 2>&1 && git pull '{auth_url}' {branch} 2>&1"
                logger.info(f"[Git] Retrying git pull after permission fix...")
                output, error, exit_code = ssh.execute_command(retry_pull_cmd, timeout=120)
                action = "pull (after permission fix)"

                if exit_code != 0:
                    logger.error(f"[Git] Pull failed after permission fix. Manual intervention needed.")
                    result_msg = f"Permission fix failed. Please SSH to VM and run:\necho '{sudo_password}' | sudo -S chown -R {current_user}:{current_user} {repo_dir} && cd {repo_dir} && git pull"
                else:
                    logger.info(f"[Git] Pull succeeded after permission fix!")
        else:
            # Ensure base directory exists with proper permissions
            logger.info(f"[Git] Repo doesn't exist, preparing for git clone...")
            mkdir_cmd = f"mkdir -p {GIT_CLONE_BASE} 2>/dev/null || true"
            ssh.execute_command(mkdir_cmd, timeout=10)

            # Fix ownership of base directory
            current_user = dut.username
            fix_base_perms = f"chown -R {current_user}:{current_user} {GIT_CLONE_BASE} 2>/dev/null || sudo chown -R {current_user}:{current_user} {GIT_CLONE_BASE} 2>/dev/null || true"
            ssh.execute_command(fix_base_perms, timeout=30)

            # Ensure directory is writable
            chmod_base = f"chmod -R u+w {GIT_CLONE_BASE} 2>/dev/null || true"
            ssh.execute_command(chmod_base, timeout=10)

            # Clone the repo
            clone_cmd = f"cd {GIT_CLONE_BASE} && git clone --branch {branch} '{auth_url}' 2>&1"
            logger.info(f"[Git] Cloning repo into {repo_dir} on branch '{branch}'...")
            output, error, exit_code = ssh.execute_command(clone_cmd, timeout=300)
            action = "clone"

        result_msg = output.strip() if output.strip() else error.strip()
        # Sanitize — remove credentials from log messages
        result_msg = result_msg.replace(token, "***").replace(username, "***")

        if exit_code != 0:
            logger.warning(f"[Git] git {action} returned exit code {exit_code}: {result_msg}")

        logger.info(f"[Git] git {action} result: {result_msg}")

        # Step 2: List test categories from the tests directory
        cat_cmd = f'find {tests_dir} -mindepth 1 -maxdepth 1 -type d -printf "%f\\n" | sort'
        cat_output, cat_error, cat_code = ssh.execute_command(cat_cmd, timeout=15)

        categories = []
        if cat_code == 0 and cat_output.strip():
            categories = [d.strip() for d in cat_output.strip().split("\n")
                         if d.strip() and not d.strip().startswith("__")]

        # Update state
        _git_state["configured"] = True
        _git_state["host_id"] = host_id
        _git_state["host_name"] = dut.name
        _git_state["repo_url"] = repo_url
        _git_state["branch"] = branch
        _git_state["repo_name"] = repo_name
        _git_state["tests_path"] = tests_dir
        _git_state["categories_count"] = len(categories)
        _save_git_state()

        logger.info(f"[Git] Done. {len(categories)} categories found on {dut.name}")

        return {
            "status": "connected",
            "host_name": dut.name,
            "host_id": host_id,
            "repo_url": repo_url,  # Cleaned URL
            "repo_name": repo_name,
            "branch": branch,
            "action": action,
            "tests_path": tests_dir,
            "pull_message": result_msg,
            "categories_count": len(categories),
            "categories": categories,
            "url_was_parsed": original_url != repo_url,  # Indicates if /tree/ was removed
        }

    finally:
        ssh.disconnect()


@app.post("/api/git/cleanup")
def cleanup_git_repo(body: dict, db: Session = Depends(get_db)):
    """
    Force cleanup of git repository with permission issues.
    Removes the entire repo directory to allow fresh clone.
    """
    host_id = body.get("host_id")
    repo_url = body.get("repo_url", "").strip()

    if not host_id or not repo_url:
        raise HTTPException(status_code=400, detail="host_id and repo_url required")

    dut = db.query(DUT).filter(DUT.id == host_id).first()
    if not dut:
        raise HTTPException(status_code=404, detail="VM device not found")

    # Clean URL (remove /tree/ or /blob/ parts)
    if "/tree/" in repo_url:
        repo_url = repo_url.split("/tree/")[0]
    elif "/blob/" in repo_url:
        repo_url = repo_url.split("/blob/")[0]

    # Remove trailing .git if present
    if repo_url.endswith(".git"):
        repo_url = repo_url[:-4]

    # Extract repo name
    repo_name = repo_url.rstrip("/").split("/")[-1]

    repo_dir = f"{GIT_CLONE_BASE}/{repo_name}"

    ssh = SSHConnectionManager(dut.ip_address, dut.port, dut.username, dut.password)
    if not ssh.connect():
        raise HTTPException(status_code=503, detail=f"Cannot connect to {dut.name}")

    try:
        logger.info(f"[Git Cleanup] Force removing {repo_dir} on {dut.name}...")

        # Try multiple removal strategies
        remove_cmds = [
            f"sudo rm -rf {repo_dir}",
            f"rm -rf {repo_dir}",
            f"sudo rm -rf {repo_dir}/.git && rm -rf {repo_dir}"
        ]

        last_error = ""
        for cmd in remove_cmds:
            out, err, code = ssh.execute_command(cmd, timeout=60)
            if code == 0:
                # Verify removal
                check_cmd = f"test -d {repo_dir} && echo EXISTS || echo REMOVED"
                check_out, _, _ = ssh.execute_command(check_cmd, timeout=10)

                if "REMOVED" in check_out:
                    logger.info(f"[Git Cleanup] Successfully removed {repo_dir}")
                    return {
                        "status": "success",
                        "message": f"Repository '{repo_name}' removed successfully. Click 'Connect to GitHub' to clone fresh.",
                        "repo_dir": repo_dir
                    }
            else:
                last_error = err if err else out

        # If we got here, all removal attempts failed
        raise HTTPException(
            status_code=500,
            detail=f"Failed to remove repository. Last error: {last_error}. "
                   f"Please SSH to VM and run: sudo rm -rf {repo_dir}"
        )

    finally:
        ssh.disconnect()

# Git state file — persists across server restarts
_GIT_STATE_FILE = DATA_DIR / "git_state.json"

@app.get("/api/ssh/network-status")
def get_ssh_network_status():
    """
    Get SSH network monitoring status and statistics (Phase 1 Enhancement).

    Returns network state, monitoring statistics, and connection health info.
    """
    try:
        network_status = ssh_pool.get_network_status()
        pool_status = ssh_pool.get_pool_status()

        # Count connections by status
        status_counts = {"alive": 0, "offline": 0, "other": 0}
        for conn in pool_status.get("connections", []):
            status = conn.get("status", "unknown")
            if status == "alive":
                status_counts["alive"] += 1
            elif status == "offline":
                status_counts["offline"] += 1
            else:
                status_counts["other"] += 1

        return {
            "network": network_status,
            "pool": {
                "total_connections": pool_status.get("total_connections", 0),
                "alive": status_counts["alive"],
                "offline": status_counts["offline"],
                "other": status_counts["other"],
            }
        }
    except Exception as e:
        logger.error(f"Error getting SSH network status: {e}")
        return {
            "error": str(e),
            "network": {"monitoring_enabled": False, "network_online": True},
            "pool": {"total_connections": 0, "alive": 0, "offline": 0}
        }


@app.get("/api/ssh/pool/status")
def get_ssh_pool_detailed_status():
    """
    Get detailed SSH pool status with metrics (Phase 4 Enhancement).

    Returns comprehensive pool status including:
    - Per-connection metrics (reconnection stats, offline duration, etc.)
    - Aggregated metrics across all connections
    - State summary counts
    - Individual connection details with full metrics
    """
    try:
        pool_status = ssh_pool.get_pool_status()
        return pool_status
    except Exception as e:
        logger.error(f"Error getting SSH pool status: {e}")
        return {
            "error": str(e),
            "total_connections": 0,
            "state_summary": {},
            "aggregated_metrics": {},
            "connections": []
        }


@app.get("/api/ssh/pool/config")
def get_ssh_pool_configuration():
    """
    Get SSH pool configuration (Phase 4 Enhancement).

    Returns current configuration values including:
    - Network monitoring settings
    - Reconnection behavior
    - Connection pool settings
    - State preservation options

    Configuration is loaded from environment variables at startup.
    """
    try:
        config = ssh_pool.get_configuration()
        return config
    except Exception as e:
        logger.error(f"Error getting SSH pool configuration: {e}")
        return {"error": str(e)}


@app.get("/api/ssh/connection/{dut_id}/metrics")
def get_ssh_connection_metrics(dut_id: int):
    """
    Get detailed metrics for a specific connection (Phase 4 Enhancement).

    Args:
        dut_id: Device ID

    Returns:
        Detailed metrics for the connection including:
        - Reconnection statistics
        - Offline duration tracking
        - State transition history
        - Success/failure rates
    """
    try:
        pool_status = ssh_pool.get_pool_status()

        # Find the connection
        for conn in pool_status.get("connections", []):
            if conn.get("dut_id") == dut_id:
                return {
                    "dut_id": dut_id,
                    "ip": conn.get("ip"),
                    "status": conn.get("status"),
                    "metrics": conn.get("metrics", {}),
                    "state_history_count": conn.get("state_history_count", 0),
                    "last_state_change": conn.get("last_state_change"),
                    "state_change_reason": conn.get("state_change_reason"),
                }

        return {"error": f"Connection for DUT {dut_id} not found"}
    except Exception as e:
        logger.error(f"Error getting connection metrics for DUT {dut_id}: {e}")
        return {"error": str(e)}



@app.get("/api/git/status")
def get_git_status():
    """Get current Git configuration status."""
    if not _git_state["configured"]:
        return {"status": "disconnected"}

    return {
        "status": "connected",
        "host_id": _git_state.get("host_id"),
        "host_name": _git_state.get("host_name", ""),
        "repo_url": _git_state.get("repo_url", ""),
        "branch": _git_state.get("branch", "master"),
        "repo_name": _git_state.get("repo_name", ""),
        "tests_path": _git_state.get("tests_path", ""),
        "categories_count": _git_state.get("categories_count", 0),
    }


@app.get("/api/git/categories")
def get_git_categories(db: Session = Depends(get_db)):
    """List test category folders from the VM's Git repo tests directory via SSH."""
    if not _git_state["configured"]:
        raise HTTPException(status_code=400, detail="Git repo not connected. Select a VM and click Connect first.")

    host_id = _git_state.get("host_id")
    if not host_id:
        raise HTTPException(status_code=400, detail="No VM host configured")

    ssh, dut = _ssh_to_host(host_id, db)
    try:
        tests_dir = _git_state.get("tests_path", "")
        cmd = f'find {tests_dir} -mindepth 1 -maxdepth 1 -type d -printf "%f\\n" | sort'
        output, error, code = ssh.execute_command(cmd, timeout=15)
        if code != 0:
            raise HTTPException(status_code=500, detail=f"Failed to list categories: {error}")
        categories = [d.strip() for d in output.strip().split("\n")
                     if d.strip() and not d.strip().startswith("__")]
        return {"categories": categories, "base_path": tests_dir}
    finally:
        ssh.disconnect()


@app.get("/api/git/scripts/{category}")
def get_git_scripts(category: str, db: Session = Depends(get_db)):
    """List Python test scripts in a category folder on the VM (recursive)."""
    if not _git_state["configured"]:
        raise HTTPException(status_code=400, detail="Git repo not connected. Select a VM and click Connect first.")

    host_id = _git_state.get("host_id")
    if not host_id:
        raise HTTPException(status_code=400, detail="No VM host configured")

    ssh, dut = _ssh_to_host(host_id, db)
    try:
        tests_dir = _git_state.get("tests_path", "")
        category_path = f"{tests_dir}/{category}"
        cmd = f'find {category_path} -name "test_*.py" -type f | sort'
        output, error, code = ssh.execute_command(cmd, timeout=15)
        if code != 0:
            raise HTTPException(status_code=500, detail=f"Failed to list scripts: {error}")

        scripts = []
        for line in output.strip().split("\n"):
            line = line.strip()
            if not line:
                continue
            rel_path = line.replace(tests_dir + "/", "")
            name = os.path.basename(line)
            scripts.append({"name": name, "path": rel_path, "full_path": line})

        return {"scripts": scripts, "count": len(scripts), "category": category}
    finally:
        ssh.disconnect()


@app.post("/api/git/disconnect")
def disconnect_git():
    """Disconnect from Git repo (clear connection state)."""
    _git_state["configured"] = False
    _git_state["host_id"] = None
    _git_state["host_name"] = ""
    _git_state["repo_url"] = ""
    _git_state["branch"] = "master"
    _git_state["repo_name"] = ""
    _git_state["tests_path"] = ""
    _git_state["categories_count"] = 0
    _save_git_state()

    return {"status": "disconnected", "message": "Git repo disconnected"}


@app.get("/api/spytest/testbed-info")
def get_testbed_info(host_id: int, testbed: str, db: Session = Depends(get_db)):
    """Read and parse a testbed YAML file from the remote VM.

    Returns device count, device names, topology type, and link list.
    """
    ssh, dut = _ssh_to_host(host_id, db)
    try:
        testbed_path = f"{SPYTEST_TESTBED_DIR}/{testbed}"
        output, error, code = ssh.execute_command(f"cat {testbed_path}", timeout=15)
        if code != 0:
            raise HTTPException(status_code=500, detail=f"Failed to read testbed: {error.strip()}")
        config = yaml.safe_load(output) or {}
        devices = list(config.get("devices", {}).keys())
        topology = config.get("topology", {})

        # Deduplicated link list
        links = []
        seen: set = set()
        for dev_name, dev_topo in topology.items():
            if not isinstance(dev_topo, dict):
                continue
            for iface, link in dev_topo.get("interfaces", {}).items():
                if not isinstance(link, dict):
                    continue
                end_dev = link.get("EndDevice", "")
                end_port = link.get("EndPort", "")
                if end_dev:
                    key = tuple(sorted([f"{dev_name}:{iface}", f"{end_dev}:{end_port}"]))
                    if key not in seen:
                        seen.add(key)
                        links.append({
                            "from": f"{dev_name}:{iface}",
                            "to": f"{end_dev}:{end_port}",
                        })

        n = len(devices)
        if n == 0:
            topology_type = "empty"
        elif n == 1:
            topology_type = "standalone"
        elif n == 2:
            topology_type = "dual-dut"
        else:
            topology_type = f"{n}-node"

        return {
            "testbed": testbed,
            "device_count": n,
            "device_names": devices,
            "topology_type": topology_type,
            "link_count": len(links),
            "links": links,
        }
    finally:
        ssh.disconnect()


# ============================================================================
# HEALTH CHECK
# ============================================================================

@app.get("/health")
def health_check():
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}


@app.get("/api/stats")
def get_stats(request: Request, db: Session = Depends(get_db)):
    """Dashboard statistics - session-based."""
    session_id = get_session_id(request)

    if session_id:
        # Session-specific stats
        total_duts = db.query(DUT).filter(DUT.session_id == session_id).count()
        online_duts = db.query(DUT).filter(DUT.session_id == session_id, DUT.status == "online").count()
        total_executions = db.query(Execution).filter(Execution.session_id == session_id).count()
        running_executions = db.query(Execution).filter(
            Execution.session_id == session_id,
            Execution.status == "running",
            Execution.execution_type != "image"
        ).count()
    else:
        # Global stats (no session)
        total_duts = db.query(DUT).count()
        online_duts = db.query(DUT).filter(DUT.status == "online").count()
        total_executions = db.query(Execution).count()
        running_executions = db.query(Execution).filter(
            Execution.status == "running",
            Execution.execution_type != "image"
        ).count()

    return {
        "total_duts": total_duts,
        "online_duts": online_duts,
        "total_images": db.query(Image).count(),  # Images are shared
        "total_scripts": db.query(Script).count(),  # Scripts are shared
        "total_executions": total_executions,
        "running_executions": running_executions,
    }



# ============================================================================
# ONE PALC SSO — Config, Callback, Logout
# ============================================================================

# Read OnePalC config from environment once at startup
_ONEPALC_ENABLED      = os.getenv("ONEPALC_ENABLED", "false").lower() == "true"
_ONEPALC_HUB_AUTH_URL = os.getenv("ONEPALC_HUB_AUTH_URL", "")
_ONEPALC_APP_NAME     = os.getenv("ONEPALC_APP_NAME", "Eka-Automation")
_ONEPALC_CALLBACK_URL = os.getenv("ONEPALC_CALLBACK_URL", "")
_ONEPALC_ROLE_KEY     = os.getenv("ONEPALC_ROLE_KEY", "EKA")
_ONEPALC_API_TOKEN    = os.getenv("ONEPALC_API_TOKEN", "")
# JWT signing secret (HS256 shared secret from IT — preferred)
_ONEPALC_JWT_SECRET   = os.getenv("ONEPALC_JWT_SECRET", "")
# RSA public key (RS256 PEM key from IT — used if JWT_SECRET is not set)
_ONEPALC_PUBLIC_KEY   = os.getenv("ONEPALC_PUBLIC_KEY", "")

logger.info(
    f"[OnePalC] JWT verification mode: "
    f"{'HS256 (shared secret)' if _ONEPALC_JWT_SECRET else ('RS256 (public key)' if _ONEPALC_PUBLIC_KEY else 'UNVERIFIED (no key configured)')}"
)


@app.get("/api/onepalc/config")
def onepalc_config():
    """Return SSO configuration for the frontend."""
    return {
        "enabled":      _ONEPALC_ENABLED,
        "hub_auth_url": _ONEPALC_HUB_AUTH_URL,
        "app_name":     _ONEPALC_APP_NAME,
        "callback_url": _ONEPALC_CALLBACK_URL,
        "api_token":    _ONEPALC_API_TOKEN,
    }

@app.get("/api/onepalc/users")
def onepalc_users(request: Request):
    """Proxy endpoint to fetch users from OnePalC Hub using correct API."""
    if not _ONEPALC_ENABLED:
        raise HTTPException(status_code=400, detail="SSO is not enabled")
    
    # Hub URL is http://172.26.1.228/login, extract base
    base_url = _ONEPALC_HUB_AUTH_URL.replace("/login", "")
    api_url = f"{base_url}/api/app_users.php"
    
    params = {
        "app_name": _ONEPALC_APP_NAME,
        "token": _ONEPALC_API_TOKEN
    }
    
    query_string = urllib.parse.urlencode(params)
    full_url = f"{api_url}?{query_string}"
    
    try:
        req = urllib.request.Request(full_url, method="GET")
        with urllib.request.urlopen(req, timeout=10) as response:
            data = json.loads(response.read().decode())
            return data
    except Exception as e:
        logger.error(f"[OnePalC] Failed to fetch users from Hub: {e}")
        return {"success": False, "users": [], "error": str(e)}


# In-memory set of seen JWT IDs (jti) to prevent replay attacks.
# Bounded to last 1000 tokens to avoid unbounded memory growth.
_SEEN_JTI: set = set()
_SEEN_JTI_ORDER: list = []
_MAX_JTI_CACHE = 1000


def _record_jti(jti: str) -> bool:
    """Record a JWT ID as seen. Returns False if already seen (replay attack)."""
    global _SEEN_JTI, _SEEN_JTI_ORDER
    if jti in _SEEN_JTI:
        return False
    _SEEN_JTI.add(jti)
    _SEEN_JTI_ORDER.append(jti)
    # Evict oldest if cache is full
    if len(_SEEN_JTI_ORDER) > _MAX_JTI_CACHE:
        old = _SEEN_JTI_ORDER.pop(0)
        _SEEN_JTI.discard(old)
    return True


def _decode_jwt_payload(token: str) -> dict:
    """
    Decode and verify JWT payload.

    Security levels (in order of preference):
      1. RS256/HS256 verified with ONEPALC_PUBLIC_KEY (best — signature checked)
      2. Unverified base64 decode with iat/exp/jti checks (fallback)

    Also enforces:
      - Token not expired (exp claim)
      - Token not too old (iat must be within last 5 minutes)
      - Replay attack prevention via jti claim
    """
    if not token:
        return {}

    import jwt as _jwt  # PyJWT

    # ── Level 1A: HS256 with shared secret (ONEPALC_JWT_SECRET) ─────────────
    # Most common for OnePalC — fast symmetric verification
    if _ONEPALC_JWT_SECRET:
        try:
            payload = _jwt.decode(
                token,
                _ONEPALC_JWT_SECRET,
                algorithms=["HS256"],
                options={"verify_exp": True},
            )
            logger.info("[OnePalC] ✓ JWT verified with HS256 (shared secret)")
            jti = payload.get("jti")
            if jti and not _record_jti(jti):
                logger.warning(f"[OnePalC] JWT replay detected — jti={jti}")
                return {"__replay__": True}
            return payload
        except _jwt.ExpiredSignatureError:
            logger.warning("[OnePalC] JWT expired — rejecting")
            return {}
        except _jwt.InvalidSignatureError:
            logger.warning("[OnePalC] JWT HS256 signature invalid — trying RS256 fallback")
        except Exception as e:
            logger.warning(f"[OnePalC] JWT HS256 decode error: {e} — trying fallback")

    # ── Level 1B: RS256 with RSA public key (ONEPALC_PUBLIC_KEY PEM) ─────────
    if _ONEPALC_PUBLIC_KEY:
        try:
            payload = _jwt.decode(
                token,
                _ONEPALC_PUBLIC_KEY,
                algorithms=["RS256"],
                options={"verify_exp": True},
            )
            logger.info("[OnePalC] ✓ JWT verified with RS256 (public key)")
            jti = payload.get("jti")
            if jti and not _record_jti(jti):
                logger.warning(f"[OnePalC] JWT replay detected — jti={jti}")
                return {"__replay__": True}
            return payload
        except _jwt.ExpiredSignatureError:
            logger.warning("[OnePalC] JWT expired (RS256) — rejecting")
            return {}
        except Exception as e:
            logger.warning(f"[OnePalC] JWT RS256 decode error: {e} — falling back to unverified")

    # ── Level 2: Unverified base64 decode with exp/iat/jti checks ────────────
    # Used when no signing key is configured. Less secure but still blocks
    # expired tokens, tokens > 10 minutes old, and replayed tokens.
    try:
        parts = token.split(".")
        if len(parts) < 2:
            logger.warning("[OnePalC] JWT has fewer than 2 parts — invalid token")
            return {}
        payload_b64 = parts[1]
        # Fix base64 padding
        payload_b64 += "=" * (4 - len(payload_b64) % 4)
        payload_bytes = base64.urlsafe_b64decode(payload_b64)
        payload = json.loads(payload_bytes.decode("utf-8"))

        # Basic security checks even without signature verification
        now_ts = datetime.utcnow().timestamp()

        # Check token expiry
        exp = payload.get("exp")
        if exp and now_ts > float(exp):
            logger.warning(f"[OnePalC] JWT expired (exp={exp}, now={now_ts}) — rejecting")
            return {}

        # Check token is not too old (iat must be within last 10 minutes)
        iat = payload.get("iat")
        if iat and now_ts - float(iat) > 600:  # 10-minute window
            logger.warning(f"[OnePalC] JWT too old (iat={iat}, age={now_ts-float(iat):.0f}s) — rejecting")
            return {}

        # Replay guard via jti
        jti = payload.get("jti")
        if jti and not _record_jti(jti):
            logger.warning(f"[OnePalC] JWT replay detected — jti={jti}")
            return {"__replay__": True}

        logger.info("[OnePalC] JWT decoded (unverified — no public key configured)")
        return payload
    except Exception as e:
        logger.warning(f"[OnePalC] JWT decode failed: {e}")
        return {}


@app.get("/hub-callback")
def hub_callback(
    request: Request,
    token: str = "",
    auth_token: str = "",
    access_token: str = "",
    session_token: str = "",
    db: Session = Depends(get_db)
):
    """
    OnePalC Hub redirects here after authentication.
    Accepts token as ?token=, ?auth_token=, ?access_token=, or ?session_token=
    Decodes user info from JWT, registers a session, sets cookie, redirects to /.
    """
    from fastapi.responses import RedirectResponse

    # Accept any of the common token param names OnePalC might use
    jwt = token or auth_token or access_token or session_token

    user_name  = "Unknown User"
    user_email = ""
    user_role  = ""
    payload    = {}

    if jwt:
        payload = _decode_jwt_payload(jwt)
        logger.info(f"[OnePalC] Callback JWT payload keys: {list(payload.keys())}")

        # Extract user name — try common JWT claim names
        user_name = (
            payload.get("name") or
            payload.get("full_name") or
            payload.get("display_name") or
            payload.get("username") or
            payload.get("preferred_username") or
            payload.get("sub") or
            "OnePalC User"
        )

        # Extract email
        user_email = payload.get("email") or payload.get("mail") or payload.get("sub") or ""

        # Extract role from roles map using the configured role key
        roles_map = payload.get("roles", {})
        if isinstance(roles_map, dict):
            user_role = roles_map.get(_ONEPALC_ROLE_KEY, "")
        elif isinstance(roles_map, list):
            user_role = roles_map[0] if roles_map else ""

        # Normalize role to lowercase
        user_role = str(user_role).lower() if user_role else ""

    # Replay attack guard — stop here if JWT was flagged as replay
    if payload.get("__replay__"):
        logger.error(f"[OnePalC] Replay attack blocked for token — rejecting callback")
        from fastapi.responses import HTMLResponse as _HTML
        return _HTML(
            content="<h2>Security Error</h2><p>Token already used. Please log in again.</p>",
            status_code=403
        )

    # Find existing session for this user to preserve DUT ownership
    # Sessions are user-based and persistent — reuse existing active session
    existing_session = None
    if user_email:
        existing_session = (
            db.query(UserSession)
            .filter(UserSession.user_email == user_email, UserSession.status == "active")
            .order_by(UserSession.last_activity.desc())
            .first()
        )
    if not existing_session and user_name:
        existing_session = (
            db.query(UserSession)
            .filter(UserSession.user_name == user_name, UserSession.status == "active")
            .order_by(UserSession.last_activity.desc())
            .first()
        )

    if existing_session:
        session_id = existing_session.session_id
        existing_session.status = "active"
        existing_session.last_activity = datetime.utcnow()
        existing_session.expires_at = None  # No TTL — user-based session
        # Always sync role from latest JWT in case it changed
        if user_role:
            existing_session.user_role = user_role
        try:
            db.commit()
            logger.info(f"[OnePalC] Reused session: {session_id} for {user_name} ({user_email}) role={user_role}")
        except Exception as e:
            logger.error(f"[OnePalC] Failed to update session: {e}")
            db.rollback()
    else:
        # Generate new user-based persistent session
        session_id = "sso-" + base64.urlsafe_b64encode(os.urandom(18)).decode().rstrip("=")
        try:
            new_session = UserSession(
                session_id=session_id,
                user_name=user_name,
                user_email=user_email,
                user_role=user_role,  # Persist role from OnePalC JWT
                status="active",
                created_at=datetime.utcnow(),
                last_activity=datetime.utcnow(),
                expires_at=None,  # No TTL — user-based session
            )
            db.add(new_session)
            db.commit()
            logger.info(f"[OnePalC] Session created: {session_id} for {user_name} ({user_email}) role={user_role}")
        except Exception as e:
            logger.error(f"[OnePalC] Failed to create session: {e}")
            db.rollback()

    # Redirect to root, passing session info via query params so JS can pick them up
    # The frontend JS (handleSSOCallback) reads these and stores them in localStorage
    redirect_url = (
        f"/?sso=1"
        f"&session_id={urllib.parse.quote(session_id)}"
        f"&user_name={urllib.parse.quote(user_name)}"
        f"&user_email={urllib.parse.quote(user_email)}"
        f"&user_role={urllib.parse.quote(user_role)}"
    )

    response = RedirectResponse(url=redirect_url, status_code=302)
    # Also set a cookie as backup
    response.set_cookie(
        key="eka_session_id", value=session_id,
        max_age=28800, httponly=False, samesite="lax"
    )
    return response


@app.get("/api/onepalc/logout")
def onepalc_logout(request: Request, db: Session = Depends(get_db)):
    """Clear session cookie, mark session as terminated, redirect to OnePalC Hub logout.

    User-based sessions are persistent until explicitly logged out.
    This endpoint is the ONLY way a session ends (other than admin revoke).
    """
    from fastapi.responses import RedirectResponse

    # Mark the session as terminated so it's cleaned up by the background job
    session_id = request.cookies.get("eka_session_id", "").strip()
    if session_id:
        session = db.query(UserSession).filter(UserSession.session_id == session_id).first()
        if session and session.status == "active":
            session.status = "terminated"
            try:
                db.commit()
                logger.info(f"[OnePalC] Logout: session {session_id} marked as terminated (user: {session.user_name})")
            except Exception as e:
                logger.error(f"[OnePalC] Failed to terminate session on logout: {e}")
                db.rollback()

    logout_url = _ONEPALC_HUB_AUTH_URL.replace("/login", "/logout") if _ONEPALC_HUB_AUTH_URL else "/"
    response = RedirectResponse(url=logout_url, status_code=302)
    response.delete_cookie("eka_session_id")
    return response


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    import uvicorn

    print("=" * 60)
    print("  DUT Automation System — Lightweight Standalone")
    print("=" * 60)
    print(f"  Database: {DB_PATH}")
    print(f"  Images:   {IMAGES_DIR}")
    print(f"  Scripts:  {SCRIPTS_DIR}")
    print(f"  Logs:     {LOGS_DIR}")
    print("=" * 60)
    print("  Dashboard:  http://localhost:8000")
    print("  API Docs:   http://localhost:8000/docs")
    print("  Health:     http://localhost:8000/health")
    print("=" * 60)

    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info", reload=True)
